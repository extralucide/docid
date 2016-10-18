# -*- coding: utf-8 -*-
__author__ = 'Olive'
import re # For regular expressions
from os.path import join
import warnings
#from tool import Tool,Style
import sys
import os
import time
from math import floor

# Abstract Syntax Tree
sys.path.append("pycparser")
from pycparser import c_ast,parse_file
from pycparser.plyparser import ParseError

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.styles import PatternFill

# For tool
try:
    from ConfigParser import ConfigParser
except ImportError as exception:
    from configparser import ConfigParser

class Tool():
    def __init__(self,config_filename="docid.ini"):
        '''
            get in file .ini information to access synergy server
            '''
        # Get config
        self.stack = []
        self.list_coverage = {}
        self.found_config = False
        self.config_parser = ConfigParser()
        config_file = join("conf",config_filename)
        result = self.config_parser.read(config_file)

    def getOptions(self,key,tag):
        if self.config_parser.has_option(key,tag):
            value = self.config_parser.get(key,tag)
        else:
            value = ""
        return value

    @staticmethod
    def getCoord(txt):
        coord = re.sub(r"^[\w\\_\.:]*:([0-9]*)$",r"\1",str(txt))
        return coord

    @staticmethod
    def getFileName(filename):
        #doc_name = re.sub(r"^(.*)(\/|\\)([A-Za-z ]*)\.(.*)$",r"\3",filename)
        doc_name = re.sub(r"^.*(\/|\\)(.*)\.([a-zA-Z]){1,6}$", r"\2", filename)
        return doc_name

# For  Style
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
from openpyxl.styles.borders import BORDER_THIN,BORDER_MEDIUM
try:
    from openpyxl.drawing.image import Image
except ImportError as e:
    warnings.warn(str(e))
try:
    from openpyxl.utils import get_column_letter,range_boundaries
except ImportError as e:
    warnings.warn(str(e))
class Style:
    def __init__(self,
                 border=None,
                 alignment=None,
                 fill=None,
                 font=None):
        self.border=border
        self.alignment=alignment
        self.font=font
        self.fill=fill

    @staticmethod
    def putLogo(ws,image="small_logo_zodiac.jpeg"):
        try:
            img = Image("img/{:s}".format(image))
            img.drawing.top = 1
            img.drawing.left = 20
            ws.add_image(img)
        except ImportError:
            pass

    @staticmethod
    def set_border(ws,
                   cell_range,
                   font = Font(name='Arial',size=12,bold=True),
                   border_style=BORDER_MEDIUM,
                   alignment_horizontal="center"):
        #font = Font(name='Arial',size=12,bold=True)
        border=Border(left=Side(border_style=border_style),
                                                top=Side(border_style=border_style),
                                                bottom=Side(border_style=border_style))
        alignment=Alignment(horizontal=alignment_horizontal,vertical='center',wrap_text=True)
        style_border_left = Style(border,alignment)
        border=Border(right=Side(border_style=border_style),
                                                top=Side(border_style=border_style),
                                                bottom=Side(border_style=border_style))
        alignment=Alignment(horizontal=alignment_horizontal,vertical='center',wrap_text=True)
        style_border_right = Style(border,alignment)
        border=Border(top=Side(border_style=border_style),
                                                bottom=Side(border_style=border_style))
        alignment=Alignment(horizontal=alignment_horizontal,vertical='center',wrap_text=True)
        style_border_middle = Style(border,alignment)
        #row = ws.iter_rows(cell_range)
        min_col, min_row, max_col, max_row = range_boundaries(cell_range.upper())
        #print "TEST:",min_col, min_row, max_col, max_row
        for index_row, rows in enumerate(ws.iter_rows(cell_range)):
        #for row in rows:
            index_column = 0
            for row in rows:
                #print "ROW:",index_row,index_column,row
                if index_column == 0:
                    Style.setStyleRow(row,style_border_left)
                elif index_column == max_col - min_col:
                    Style.setStyleRow(row,style_border_right)
                else:
                    Style.setStyleRow(row,style_border_middle)
                index_column +=1
    @staticmethod
    def setStyleRow(row,style):
        row.border = style.border
        row.alignment = style.alignment

    @staticmethod
    def setStyle(cell,style):
        cell.border = style.border
        cell.alignment = style.alignment

    @staticmethod
    def setCell(ws,line,row,col_idx,style=None,number_format=None):
        column = get_column_letter(col_idx)
        current_cell = ws['%s%s'%(column, row)]
        if col_idx > 0:
            current_cell.value = '%s' % (line[col_idx - 1])
        else:
            current_cell.value = '%s' % (line)
        if style:
            if style.border is not None:
                current_cell.border = style.border
            if style.alignment is not None:
                current_cell.alignment = style.alignment
            if style.font is not None:
                current_cell.font = style.font
            if style.fill is not None:
                current_cell.fill = style.fill
        if number_format is not None:
            current_cell.number_format=number_format

    @staticmethod
    def setHyperlink(ws,row,col_idx,hyperlink):
        column = get_column_letter(col_idx)
        current_cell = ws.cell('%s%s'%(column, row))
        #current_cell.value = '%s' % (line[col_idx - 1])
        current_cell.hyperlink = hyperlink
        current_cell.font = Font(color="0000BB",underline="single")

class FuncCallVisitor(c_ast.NodeVisitor):
    def __init__(self,dico_func_called,list_func_def):
        self.dico_func_called = dico_func_called
        self.list_func_def = list_func_def
        self.nb_func_called = 0

    def visit_FuncCall(self, node):
        coord = Tool.getCoord(node.name.coord)
        #print "self.list_func_def",self.list_func_def
        self.nb_func_called += 1
        found_func_def = False
        for func_def,coord_def in self.list_func_def:
            found_func_def = func_def
            if coord < coord_def:
                break
            else:
                found_func_def = func_def
        #self.list_func_called.append((found_func_def,node.name.name))
        if found_func_def not in self.dico_func_called:
            if node.name.name not in ("__asm","asm"):
                self.dico_func_called[found_func_def] = [node.name.name]
        else:
            if node.name.name not in self.dico_func_called[found_func_def] and node.name.name not in ("__asm","asm"):
                self.dico_func_called[found_func_def].append(node.name.name)
        #print('%s called at %s' % (node.name.name, coord))

class FuncDefVisitor(c_ast.NodeVisitor):
    def __init__(self,list_func_def):
        self.list_func_def = list_func_def
    def visit_FuncDef(self, node):
        coord = Tool.getCoord(node.decl.coord)
        self.list_func_def.append((node.decl.name,coord))
        #print node.decl.coord
        #print('%s defined at %s' % (node.decl.name, coord))

class Stack():

    def __init__(self,
                 master=None,
                 basename="",
                 config_file='docid.ini'):
        # Stack analysis
        self.master_ihm = master
        self.basename = basename
        self.depth = 0
        self.stack = []
        self.list_code = []
        tool = Tool(config_filename=config_file)
        self.root_user_dir  = tool.getOptions("Lifecycle","root")
        self.src_user_dir   = tool.getOptions("Lifecycle","src")
        self.build_user_dir = tool.getOptions("Lifecycle","build")
        self.editor = tool.getOptions("Stack","editor")
        if self.editor == "":
            self.editor = "notepad++"
        self.compiler = tool.getOptions("Stack","compiler")
        if self.compiler == "":
            self.compiler = "gcc"
        self.dico_func_called       = {}
        self.depth_func_call        = 0
        self.dico_file_vs_function  = {}
        self.dico_file_vs_link      = {}
        self.dico_functions_vs_file = {}

    def reset_basename(self,basename):
        self.basename = basename

    def listDir(self):
        """
        Recursive function to find files in directories.
        Treatment for Excel and Word file is different
        :param dirname:
        :param type:
        :return:
        """

        self.depth += 1
        new_concat_dirname = self.basename
        for dir in self.stack:
            new_concat_dirname = join(new_concat_dirname,dir)
            if sys.platform.startswith('win32'):
                new_concat_dirname = "{:s}\\".format(new_concat_dirname)
            else:
                new_concat_dirname = "{:s}/".format(new_concat_dirname)

        try:
            list_dir = os.listdir(new_concat_dirname)
        except OSError as e:
            try:
                self.log("{:s}".format(str(e)))
            except UnicodeEncodeError as exception:
                pass
            list_dir = []

        for found_dir in list_dir:
            path_dir = os.path.join(new_concat_dirname, found_dir)
            isdir = os.path.isdir(path_dir)
            if isdir:
                self.stack.append(found_dir)
                self.listDir()
                self.stack.pop()
            else:
                void = re.sub(r"(~\$)(.*)\.(.*)",r"\1",found_dir)
                name = re.sub(r"(.*)\.(.*)",r"\1",found_dir)
                extension = re.sub(r"(.*)\.(.*)",r"\2",found_dir)
                if extension in ("c","s","asm","vhd") and void != "~$":
                    self.log("Parse {:s}".format(found_dir),gui_display=True)
                    filename = join(new_concat_dirname,found_dir)
                    self.list_code.append(filename)
                else:
                    self.log("Discard {:s}".format(found_dir),gui_display=True)
                    # Wrong Word format, only openxml
                    text = "Unexpected format for {:s}, only ('c','s','asm','vhd') accepted".format(found_dir)
                    self.log(text)
        self.depth -= 1

    def log(self,text,gui_display=True):
        print(text)

    def _isSourceFile(self,filename):
        m = re.match("(.*)\.(c)",filename)
        if m:
            result = True
        else:
            result = False
        return result

    def _reccurFoundCalling(self,function,tbl):
        print ("TBL:",tbl)
        self.depth_func_call += 1
        if self.depth_func_call > 13:
            print ("function:",function)
            return False
        if function in self.dico_func_called:
            list_calling = self.dico_func_called[function]
            if list_calling is not []:
                result = True
                for sub_function in list_calling:
                    tbl[self.depth_func_call + 1] = sub_function
                    sub_result = self._reccurFoundCalling(sub_function,tbl)
                    if not sub_result:
                        copy_tbl = tbl[0:self.depth_func_call+2]
                        while len(copy_tbl) < 12:
                            copy_tbl.append("")
                        self.leaves.append(copy_tbl)
                        self.leaves_index += 1
            else:
                result = False
        else:
            result = False
        self.depth_func_call -= 1
        return result

    def _computeLeaves(self):
        self.leaves = []
        self.leaves_index = 0
        tbl = []
        try:
            # Python 2
            for callee,calling in self.dico_func_called.iteritems():
                if calling is not []:
                    for function in calling:
                        del(tbl[:])
                        tbl = [callee,function,"","","","","","","","","","","","",""]
                        sub_result = self._reccurFoundCalling(function,tbl)
                        if not sub_result:
                            self.leaves.append(tbl[:])
                            self.leaves_index += 1
                else:
                    tbl = [callee,"","","","","","","","","","","","","",""]
                    self.leaves.append(tbl[:])
        except AttributeError:
            print ("iteritems is a python 2 method")
            # Python 3
            for callee,calling in self.dico_func_called.items():
                if calling is not []:
                    for function in calling:
                        del(tbl[:])
                        tbl = [callee,function,"","","","","","","","","","","","",""]
                        sub_result = self._reccurFoundCalling(function,tbl)
                        if not sub_result:
                            self.leaves.append(tbl[:])
                            self.leaves_index += 1
                else:
                    tbl = [callee,"","","","","","","","","","","","","",""]
                    self.leaves.append(tbl[:])


    def _getStackFromAsm(self):
        # py2
        import codecs
        import warnings
        def open(file, mode='r', buffering=-1, encoding=None,
                 errors=None, newline=None, closefd=True, opener=None):
            if newline is not None:
                warnings.warn('newline is not supported in py2')
            if not closefd:
                warnings.warn('closefd is not supported in py2')
            if opener is not None:
                warnings.warn('opener is not supported in py2')
            return codecs.open(filename=file, mode=mode, encoding=encoding,
                        errors=errors, buffering=buffering)
        code_dir = join(self.root_user_dir,self.build_user_dir)
        self.reset_basename(code_dir)
        self.listDir()
        dico_source_files = {}
        index=0
        function_name=""
        for filename in self.list_code:
            index += 1
            print ("File reading:",filename)
            with open(filename, 'r',encoding='utf-8') as of:
                function_found = False
                try:
                    for line in of:
                        if function_found:
                            # stwu  r1, -X(r1) Store the stack pointer and update. create a frame of X bytes
                            m = re.search(r'stwu\t*r1,-([0-9]{1,4})\(r1\)',line)
                            if m:
                                stack_size = m.group(1)
                                print ("function_name",function_name)
                                dico_source_files[function_name]=stack_size
                                function_found = False
                        m = re.findall(r'^(\w*):',line)
                        if m:
                            function_name = m[0]
                            function_found = True
                except UnicodeDecodeError as exception:
                    warnings.warn(str(exception))
        return dico_source_files

    def _computeStackSize(self,
                          line,
                          dico_function_vs_stack_size):
        compute_stack = 0
        for function in line:
            if function in dico_function_vs_stack_size:
                compute_stack += int(dico_function_vs_stack_size[function])
        return compute_stack

    def _stackAnalysis(self):
        code_dir = join(self.root_user_dir,
                        self.src_user_dir)
        include=join(code_dir,"INCLUDE")
        self.reset_basename(code_dir)
        self.listDir()
        index=0
        max_stack_size = 0
        max_function_call_tree = ""
        for filename in self.list_code:
            index += 1
            if self._isSourceFile(filename):
                try:
                    ast = parse_file(filename,
                                     use_cpp=True,
                                     cpp_path=self.compiler,
                                     cpp_args=[r'-E', r'-I{:s}'.format(include)])

                    # List of defined functions and where
                    list_func_def    = []
                    del(list_func_def[:])
                    v = FuncDefVisitor(list_func_def)
                    v.visit(ast)

                    # List of called functions and where
                    v = FuncCallVisitor(self.dico_func_called,
                                        list_func_def)
                    v.visit(ast)

                    short_filename = Tool.getFileName(filename)
                    src_code_link="file:///{:s}".format(filename)
                    self.dico_file_vs_function[short_filename]=list_func_def
                    self.dico_file_vs_link[short_filename]=src_code_link

                    if v.nb_func_called < 2:
                        text = "function"
                    else:
                        text = "functions"
                    if self.master_ihm is not None:
                        self.master_ihm.log("Find {:s} ({:} {:s} called)".format(short_filename,v.nb_func_called,text))
                        for function in list_func_def:
                            self.master_ihm.log("{:s} => {:s}".format(short_filename,function))
                            if function in self.dico_func_called:
                                for func_called in self.dico_func_called[function]:
                                    self.master_ihm.log("=> {:s} => {:s}".format(function,func_called))
                    else:
                        print ("Find {:s} ({:} {:s} called)".format(short_filename,v.nb_func_called,text))
                        for function in list_func_def:
                            print("{:s} => {:s}".format(short_filename,function))
                            if function in self.dico_func_called:
                                for func_called in self.dico_func_called[function]:
                                    print("=> {:s} => {:s}".format(function,func_called))
                except ParseError as e:
                    short_filename = Tool.getFileName(filename)
                    if self.master_ihm is not None:
                        self.master_ihm.log("Find {:s} (AST failed: {:s})".format(short_filename,str(e)))
                    else:
                        print ("Find {:s} (AST failed: {:s})".format(short_filename,str(e)))
                    print (e)

            else:
                if self.master_ihm is not None:
                    self.master_ihm.log("Discard {:s} (header file)".format(short_filename))
                else:
                    print ("Discard {:s} (header file)".format(short_filename))
        # inverse dico file vs function
        print (self.dico_file_vs_function)
        try:
            # Python 2
            for file,functions in self.dico_file_vs_function.iteritems():
                for function,index in functions:
                    self.dico_functions_vs_file[function]=file
        except AttributeError:
            # Python 3
            for file,functions in self.dico_file_vs_function.items():
                for function,index in functions:
                    self.dico_functions_vs_file[function]=file
        #print "self.dico_functions_vs_file",self.dico_functions_vs_file
        dico_function_vs_stack_size = self._getStackFromAsm()
        if not dico_function_vs_stack_size:
            if self.master_ihm is not None:
                self.master_ihm.log("No assembler files found.")
            else:
                print ("No assembler files found.")
        wb = Workbook()
        if wb is not None:
            ws = wb.worksheets[0]
            if ws is not None:
                Style.putLogo(ws)
                Style.setCell(ws,["Functions call tree"],8,1)
                row = 9
                tbl = ("Stack","Depth 1","Depth 2","Depth 3","Depth 4","Depth 5","Depth 6","Depth 7","Depth 8","Depth 9","Depth 10","Depth 11")
                for col_idx in range(1,13):
                    Style.setCell(ws,tbl,row,col_idx)
                row += 1
                self._computeLeaves()
                index = 0
                style_border = Style(fill=PatternFill(patternType='solid',start_color='CCCCCCCC'))
                for line in self.leaves:
                    index += 1
                    compute_stack = self._computeStackSize(line,dico_function_vs_stack_size)
                    if max_stack_size < compute_stack:
                        max_stack_size = compute_stack
                        max_function_call_tree = " => ".join(line)
                    line.insert(0,compute_stack)
                    for col_idx in range(1,13):
                        if col_idx == 1:
                            Style.setCell(ws,line,row,col_idx,number_format='0.00E+00')
                        else:
                            Style.setCell(ws,line,row,col_idx)
                            function = line[col_idx-1]
                            if function in self.dico_functions_vs_file:
                                filename        = self.dico_functions_vs_file[function]
                                src_code_link   = self.dico_file_vs_link[filename]
                                Style.setHyperlink(ws,row,col_idx,src_code_link)
                            else:
                                Style.setCell(ws,line,row,col_idx,style_border)
                    row += 1
                # Autofilter
                ws.auto_filter.ref = "A9:L9"
                filename = "functions_call_tree_%d.xlsx" % floor(time.time())
                wb.save(join("result",filename))
                #self.master_ihm.resultHyperLink(filename,text="SCOD created.")
                if filename is not None and self.master_ihm is not None:
                    self.master_ihm.resultGenerateCID(filename,
                                                False,
                                                text="FUNCTIONS CALL TREE GENERATION")
        return max_stack_size,max_function_call_tree

if __name__ == '__main__':
    test = Stack(config_file='stack.ini')
    max_stack_size,max_function_call_tree = test._stackAnalysis()
    print ("Max stack depth found",max_stack_size)
    print ("Functions call tree",max_function_call_tree)
