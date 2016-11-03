__author__ = 'olivier.appere'
#!/usr/bin/env python 2.7.3
# -*- coding: utf-8 -*-
import warnings
from openpyxl import load_workbook,Workbook
#
# Class Style
#
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
from openpyxl.styles.borders import BORDER_THIN,BORDER_MEDIUM
from openpyxl.styles.colors import RED,YELLOW,WHITE,GREEN

try:
    from openpyxl.drawing.image import Image
except ImportError as e:
    warnings.warn(str(e))
try:
    from openpyxl.utils import get_column_letter,range_boundaries
except ImportError as e:
    warnings.warn(str(e))

class Style:
    def __init__(self,
                 border=None,
                 alignment=None,
                 fill=None,
                 font=None):
        self.border=border
        self.alignment=alignment
        self.font=font
        self.fill=fill

    @staticmethod
    def putLogo(ws,image="small_logo_zodiac.jpeg"):
        try:
            img = Image("img/{:s}".format(image))
            img.drawing.top = 1
            img.drawing.left = 20
            ws.add_image(img)
        except ImportError:
            pass

    @staticmethod
    def set_border(ws,
                   cell_range,
                   bg=WHITE,
                   font = Font(name='Arial',size=12,bold=True),
                   border_style=BORDER_MEDIUM,
                   alignment_horizontal="center"):
        #font = Font(name='Arial',size=12,bold=True)
        border=Border(left=Side(border_style=border_style),
                                                top=Side(border_style=border_style),
                                                bottom=Side(border_style=border_style))
        alignment=Alignment(horizontal=alignment_horizontal,vertical='center',wrap_text=True)
        style_border_left = Style(border,alignment)
        border=Border(right=Side(border_style=border_style),
                                                top=Side(border_style=border_style),
                                                bottom=Side(border_style=border_style))
        alignment=Alignment(horizontal=alignment_horizontal,vertical='center',wrap_text=True)
        style_border_right = Style(border,alignment)
        border=Border(top=Side(border_style=border_style),
                                                bottom=Side(border_style=border_style))
        alignment=Alignment(horizontal=alignment_horizontal,vertical='center',wrap_text=True)
        style_border_middle = Style(border,alignment)
        #row = ws.iter_rows(cell_range)
        min_col, min_row, max_col, max_row = range_boundaries(cell_range.upper())
        #print "TEST:",min_col, min_row, max_col, max_row
        for index_row, rows in enumerate(ws.iter_rows(cell_range)):
        #for row in rows:
            index_column = 0
            for row in rows:
                #print "ROW:",index_row,index_column,row
                if index_column == 0:
                    Style.setStyleRow(row,style_border_left)
                elif index_column == max_col - min_col:
                    Style.setStyleRow(row,style_border_right)
                else:
                    Style.setStyleRow(row,style_border_middle)
                index_column +=1

    @staticmethod
    def set_border_is(ws, cell_range,bg=WHITE,align=True):
        font = Font(name='Arial',size=10,bold=False)
        if align:
            alignment=Alignment(horizontal='center',vertical='center',wrap_text=False,shrink_to_fit=True)
        else:
            alignment=Alignment(wrap_text=False,shrink_to_fit=True)
        if bg != WHITE:
            style_border = Style(border=Border(left=Side(border_style=BORDER_MEDIUM),
                                               right=Side(border_style=BORDER_MEDIUM),
                                                    top=Side(border_style=BORDER_MEDIUM),
                                                    bottom=Side(border_style=BORDER_MEDIUM)),
                                      alignment=alignment,
                                      fill=PatternFill(patternType='solid',start_color=bg),
                                      font=font)
        else:
            style_border = Style(border=Border(left=Side(border_style=BORDER_MEDIUM),
                                               right=Side(border_style=BORDER_MEDIUM),
                                                    top=Side(border_style=BORDER_MEDIUM),
                                                    bottom=Side(border_style=BORDER_MEDIUM)),
                                      alignment=alignment,
                                      font=font)

        #row = ws.iter_rows(cell_range)
        min_col, min_row, max_col, max_row = range_boundaries(cell_range.upper())
        #print "TEST:",min_col, min_row, max_col, max_row
        for index_row, rows in enumerate(ws.iter_rows(cell_range)):
        #for row in rows:
            index_column = 0
            for row in rows:
                Style.setStyleRow(row,style_border)
                #row.style = style_border
                index_column +=1

    @staticmethod
    def setStyleRow(row,style):
        row.border = style.border
        row.alignment = style.alignment

    @staticmethod
    def setStyle(cell,style):
        cell.border = style.border
        cell.alignment = style.alignment

    @staticmethod
    def setCell(ws,line,row,col_idx,style=None,number_format=None):
        column = get_column_letter(col_idx)
        current_cell = ws['%s%s'%(column, row)]
        if col_idx > 0:
            x = line[col_idx - 1]
        else:
            x = line
        if type(x) is int:
            current_cell.value = x
        else:
            current_cell.value = '%s' % (x)
        if style:
            if style.border is not None:
                current_cell.border = style.border
            if style.alignment is not None:
                current_cell.alignment = style.alignment
            if style.font is not None:
                current_cell.font = style.font
            if style.fill is not None:
                current_cell.fill = style.fill
        if number_format is not None:
            current_cell.number_format=number_format

    @staticmethod
    def setHyperlink(ws,row,col_idx,hyperlink):
        column = get_column_letter(col_idx)
        current_cell = ws.cell('%s%s'%(column, row))
        #current_cell.value = '%s' % (line[col_idx - 1])
        current_cell.hyperlink = hyperlink
        current_cell.font = Font(color="0000BB",underline="single")

class Excel():
    def createWorkBook(self,title="",header=["UN","DEUX","TROIS"]):
        wb = Workbook()
        if wb is not None:
            ws = wb.worksheets[0]
            if ws is not None:
                Style.putLogo(ws)
                Style.setCell(ws,[title],8,1)
                row = 9
                for col_idx in range(1,len(header)+1):
                    Style.setCell(ws,header,row,col_idx)
                row += 1
        return ws,wb

    def loadWorkbook(self,filename,**kwargs):
        """
        Load an excel workbook
        :param filename_is:
        :return:
        """
        #for key in kwargs:
        #    print "another keyword arg: %s: %s" % (key, kwargs[key])
        try:
            wb = load_workbook(filename = filename,**kwargs)
            if wb is not None:
                self.debug("Workbook {:s}.".format(filename))
        except: # InvalidFileException:
            self.log("File {:s} opening error.".format(filename),gui_display=True)
            return False
        return wb