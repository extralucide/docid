#!/usr/bin/env python 2.7.3
# -*- coding: utf-8 -*-
import Queue
from Tkconstants import END, NORMAL, CURRENT
from math import floor
import re
import threading
import time
from export_doc import BuildDoc
from reviews import Review
from synergy import Synergy
from tool import Tool,Style
from ccb import CCB
import os
from os.path import join
import webbrowser
from openpyxl import load_workbook,Workbook
from openpyxl.compat import range
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
from openpyxl.styles.borders import BORDER_THIN,BORDER_MEDIUM
from check_llr import CheckLLR
from check_is import CheckIS
from export_derived import Derived
from get_ig_jquery import easyIG,getQA,exportIS_HTML,exportCR_HTML,exportSCOD_HTML
from datetime import datetime
from conf import VERSION
from convert_xml_html import HtmlConverter
import sys
sys.path.append("intelhex")
from intelhex import IntelHex,IntelHex16bit
from swrd import Swrd
import xml.etree.ElementTree as ET
try:
    import win32com.client as win32
    import pythoncom
except ImportError as e:
    print e
try:
    from blockdiag import parser, builder, drawer
    from blockdiag.imagedraw import png
    from blockdiag.noderenderer import box
except ImportError as e:
    print e

# Abstract Syntax Tree
sys.path.append("pycparser")
from pycparser import c_parser, c_ast,parse_file
from pycparser.plyparser import ParseError
from stack import Stack

__author__ = 'olivier'
class FuncCallVisitor(c_ast.NodeVisitor):
    def __init__(self,dico_func_called,list_func_def):
        self.dico_func_called = dico_func_called
        self.list_func_def = list_func_def
        self.nb_func_called = 0

    def visit_FuncCall(self, node):
        coord = Tool.getCoord(node.name.coord)
        #print "self.list_func_def",self.list_func_def
        self.nb_func_called += 1
        found_func_def = False
        for func_def,coord_def in self.list_func_def:
            found_func_def = func_def
            if coord < coord_def:
                break
            else:
                found_func_def = func_def
        #self.list_func_called.append((found_func_def,node.name.name))
        if found_func_def not in self.dico_func_called:
            if node.name.name not in ("__asm","asm"):
                self.dico_func_called[found_func_def] = [node.name.name]
        else:
            if node.name.name not in self.dico_func_called[found_func_def] and node.name.name != "__asm":
                self.dico_func_called[found_func_def].append(node.name.name)
        #print('%s called at %s' % (node.name.name, coord))

class FuncDefVisitor(c_ast.NodeVisitor):
    def __init__(self,list_func_def):
        self.list_func_def = list_func_def
    def visit_FuncDef(self, node):
        coord = Tool.getCoord(node.decl.coord)
        self.list_func_def.append((node.decl.name,coord))
        #print node.decl.coord
        #print('%s defined at %s' % (node.decl.name, coord))

# TODO: Attention a _getParentCR appellee par getParentCR dans la classe Synergy
# TODO: Il faut simplifier
class ThreadQuery(threading.Thread,Synergy):
    dico_cr_trans = {
                      "In Analysis":1,
                      "In Review":2,
                      "Under Modification":3,
                      "Under Verification":4,
                      "Fixed":5,
                      "Closed":6,
                      "Cancelled":7,
                      "Rejected":8,
                      "Postponed":9,
                      "Complementary Analysis":10
                      }
    dico_cr_status_relation_forbidden = {
        "In Review":("Under Modification","Under Verification","Complementary Analysis"),
        "Closed":("Fixed","Closed","Under Verification"),
        "Fixed":{"Fixed","Under Verification"},
        "Rejected":("Rejected","Cancelled"),
        "Postponed":("Postponed"),
        "Cancelled":("Cancelled")
    }
    def lock(self):
        # global count_baseline
        # count_baseline +=1
##        print "Wait lock release: " + str(count_baseline) + "\n"
##        print "amount of threads alive:" + str(threading.active_count()) + "\n"
        self.verrou.acquire()
    def unlock(self):
        self.verrou.release()
##        print "Release lock.\n"

    def __init__(self,
                 name_id="",
                 master=None,
                 queue="",
                 login="",
                 password="",
                 **kwargs):
        """
        :param name_id:
        :param master:
        :param queue:
        :param kwargs:
        :return:
        """

        for key in kwargs:
            self.__dict__[key] = kwargs[key]
        if "no_start_session" not in self.__dict__:
            self.no_start_session = False

        threading.Thread.__init__(self)
        # Create the queue
        self.queue = queue
        self.master_ihm = master
        self.dico_thread = {}
        Synergy.__init__(self,
                         session_started=False,
                         ihm=self.master_ihm)
        self.login = login
        self.password = password
        self.running = 1
        self.database = None
        if "system" not in self.__dict__:
            print "Missing system definition."
            self.system = ""
        if "item" in self.__dict__:
            # Get database name and aircraft name
            print "Get database name and aircraft name 1:",self.system,self.item
            self.database,self.aircraft = self.get_sys_item_database(self.system,
                                                                     self.item)
            print "Get database name and aircraft name 2:",self.database,self.aircraft
        else:
            print "Missing item definition."
            self.item = ""
        ci_id = self.get_ci_sys_item_identification(self.system,self.item)
        if self.master_ihm is not None:
            self.component = self.master_ihm.component
            # Display system name
            self.master_ihm.log("System: {:s}".format(self.system), False)
            # Display item name
            self.master_ihm.log("Item: {:s}".format(self.item), False)
            # Display configuration item ID
            #ci_id = self.get_ci_sys_item_identification(self.system,self.item)
            if ci_id is not None:
                self.master_ihm.log("CI ID: {:s}".format(ci_id), False)
            else:
                self.master_ihm.log("CI ID: Unknown", False)
        if self.database is None:
            self.database,self.aircraft = self.get_sys_database()
        self.author = ""
        self.reference = ""
        self.release = ""
        self.project = ""
        self.baseline = ""
        self.revision = ""
        # Recursive lock
        self.verrou = threading.RLock()
        self.name_id = name_id
        self.input_data_filter = ""
        self.peer_reviews_filter = ""
        self.list_projects = []
        self.export_cr_list_filename = "export_CR_list_template_2.xlsx"
        self.dico_cr_log = {}
        self.dico_cr_transition = {}
        self.easyig = easyIG()
        self.getqa = getQA()
        self.export_is_html     = exportIS_HTML()
        self.export_scod_html   = exportSCOD_HTML()
        # Lifecycle
        tool = Tool(config_filename="docid.ini")
        self.root_user_dir          = tool.getOptions("Lifecycle","root")
        self.src_user_dir           = tool.getOptions("Lifecycle","src")
        self.build_user_dir         = tool.getOptions("Lifecycle","build")
        swrd                        = tool.getOptionsTuple("Lifecycle","swrd")
        swdd            = tool.getOptions("Lifecycle","swdd")
        is_swdd         = tool.getOptions("Lifecycle","is_swdd")
        shlvcp          = tool.getOptions("Lifecycle","shlvcp")
        self.dir_swrd   = join(self.root_user_dir,swrd[0])
        self.is_swrd    = join(self.dir_swrd,swrd[1])
        self.dir_swdd   = join(self.root_user_dir,swdd)
        self.is_swdd    = join(self.root_user_dir,is_swdd)
        self.shlvcp     = join(self.root_user_dir,shlvcp)
        self.hsid       = join(self.root_user_dir,tool.getOptions("Lifecycle","hsid"))
        self.xml_csci   = join(self.root_user_dir,tool.getOptions("Lifecycle","xml_csci"))
        # Stack analysis
        #tool = Tool(config_filename="docid.ini")
        if self.config_parser.has_section("Stack"):
            self.editor         = tool.getOptions("Stack","editor")
            self.compiler       = tool.getOptions("Stack","compiler")
        else:
            self.editor         = "notepad++"
            self.compiler       = "gcc"
        self.dico_func_called       = {}
        self.depth_func_call        = 0
        self.dico_file_vs_function  = {}
        self.dico_file_vs_link      = {}
        self.dico_functions_vs_file = {}

    def stopSession(self):
        if self.session_started:
            stdout,stderr = self.ccm_query('stop','Stop Synergy session')
            if stdout != "":
                # remove \r
                text = re.sub(r"\r\n",r"\n",stdout)
                self.master_ihm.log(text, False)
            if stderr:
                 # remove \r
                text = re.sub(r"\r\n",r"\n",stderr)
                self.master_ihm.log(text, False)

    def processIncoming(self):
        """
        Handle all the messages currently in the queue (if any).
         - BUILD_DOCX
            . Store selection
         - START_SESSION
         - GET_BASELINES
         - GET_RELEASES
         - GET_PROJECTS
         - etc.
        """
        while self.queue.qsize():
            try:
                self.lock()
##                print threading.enumerate();
                # Check contents of message
                action = self.queue.get(0)
                print "ACTION:",action
                print time.strftime("%H:%M:%S", time.localtime()) + " Commmand: " + action
                if action == "BUILD_CID":
                    data = self.queue.get(1)
                    print "TEST_DATA",data
                    release = data[0]
                    project = data[1]
                    baseline = data[2]
                    self.release = release
                    self.project = project
                    self.baseline = baseline
                    implemented = data[3]
                    item = data[4]
                    previous_baseline = data[5]
                    detect = data[6]
                    cr_type = data[7]
                    component = data[8]
                    cr_domain = data[9]
                    cid_type = data[10]
                    header_image = data[11]
                    list_cr_type = data[12]
                    list_cr_status = data[13]
                    list_cr_doamin = data[14]
                    dico_parameters = data[15]
                    preview = data[16]
                    self.storeSelection(self.project,
                                        self.system,
                                        self.release,
                                        self.baseline)
                    self.dico_thread[action] = threading.Thread(None,self._generateCID,None,(release,
                                                                                          baseline,
                                                                                          project,
                                                                                          implemented,
                                                                                          item,
                                                                                          previous_baseline,
                                                                                          detect,
                                                                                          cr_type,
                                                                                          component,
                                                                                          cr_domain,
                                                                                          cid_type,
                                                                                          header_image,
                                                                                          list_cr_type,
                                                                                          list_cr_status,
                                                                                          list_cr_doamin,
                                                                                          dico_parameters,
                                                                                          preview))
                    self.dico_thread[action].start()
                elif action == "EASY_IG":
                    self.easy_ig_thread = threading.Thread(None,self._easyIG,None)
                    self.easy_ig_thread.start()
                elif action == "GET_QA_ACTIONS":
                    url_root = self.queue.get(1)
                    name,mail,tel,service,qams_user_id = self.get_user_infos(self.login)
                    self.get_qa_thread = threading.Thread(None,self._getQA,None,(qams_user_id,url_root,name))
                    self.get_qa_thread.start()
                elif action == "READ_BPROC":
                    bproc_filename = self.queue.get(1)
                    current_dir = os.getcwd()
                    xsl = join(current_dir,"template\\xsl_procedure_ece-1.0.xsl")
                    html_name = Tool.getFileName(bproc_filename)
                    html_filename = join(current_dir,"result\\" + html_name)
                    self.read_bproc_thread = threading.Thread(None,self._readBPROC,None,(bproc_filename,xsl,html_filename))
                    self.read_bproc_thread.start()
                elif action == "READ_GPROC":
                    gproc_filename = self.queue.get(1)
                    current_dir = os.getcwd()
                    xsl = join(current_dir,"template\\gproc_makefile.xsl")
                    html_name = Tool.getFileName(gproc_filename)
                    html_filename = join(current_dir,"result\\" + html_name)
                    self.read_gproc_thread = threading.Thread(None,self._readBPROC,None,(gproc_filename,xsl,html_filename))
                    self.read_gproc_thread.start()
                elif action == "READ_RTP":
                    filename = self.queue.get(1)
                    current_dir = os.getcwd()
                    xsl = join(current_dir,"template\\rtp.xsl")
                    html_name = Tool.getFileName(filename)
                    html_filename = join(current_dir,"result\\" + html_name)
                    self.read_gproc_thread = threading.Thread(None,self._readBPROC,None,(filename,xsl,html_filename))
                    self.read_gproc_thread.start()
                elif action == "GET_COVERAGE":
                    directory = self.queue.get(1)
                    self.get_coverage_thread = threading.Thread(None,self._getCoverage,None,(directory,))
                    self.get_coverage_thread.start()
                elif action == "READ_EOC":
                    eoc_filename = self.queue.get(1)
                    current_dir = os.getcwd()
                    if self.config_parser.has_section("EOC"):
                        addr_hw_sw_compatibility =  self.getOptions("EOC","addr_hw_sw_compatibility")
                        addr_pn =                   self.getOptions("EOC","addr_pn")
                        addr_checksum =             self.getOptions("EOC","addr_checksum")
                        dspic =                     self.getOptions("EOC","dspic")
                        addr_hw_sw_compatibility_range = addr_hw_sw_compatibility.split(",")
                        addr_pn_range = addr_pn.split(",")
                        addr_checksum_range = addr_checksum.split(",")
                        dico_addr={"hw_sw_compat":addr_hw_sw_compatibility_range,
                                   "pn":addr_pn_range,
                                   "checksum":addr_checksum_range}
                    else:
                        dico_addr={"hw_sw_compat":("0x400","0x402"),
                                   "pn":("0x400","0x424"),
                                   "checksum":("0x4DE8","0x4DEA")}
                        dspic = False
                    self.read_eoc_thread = threading.Thread(None,self.thread_readEOC,None,(eoc_filename,dico_addr,dspic))
                    self.read_eoc_thread.start()
                elif action == "GET_BASELINE_STATUS":
                    baseline = self.queue.get(1)
                    if baseline != "":
                        self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("GET_BASELINE_STATUS","",baseline))
                        self.send_cmd_thread.start()
                elif action == "GET_RELEASE_INFO":
                    release = self.queue.get(1)
                    if release != "":
                        self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("GET_RELEASE_INFO",release))
                        self.send_cmd_thread.start()
                elif action == "BUILD_SQAP":
                    data = self.queue.get(1)
                    author = data[0]
                    self.reference = data[1]
                    self.revision = data[2]
                    self.build_doc_thread = threading.Thread(None,self._generateSQAP,None,(author,self.reference,self.revision,self.aircraft,self.system,self.item))
                    self.build_doc_thread.start()
                elif action == "BUILD_CCB":
                    data = self.queue.get(1)
                    dico_parameters = data[0]
                    dico_parameters["login"] = self.login
                    cr_with_parent = data[1]
                    cr_workflow = data[2]
                    cr_domain = data[3]
                    log_on = data[4]
                    list_cr_for_ccb = data[5]
                    status_list = data[6]
                    ccb_time = data[7]
                    list_cr_type = data[8]
                    list_cr_status = data[9]
                    list_cr_doamin = data[10]
                    #print "dico_parameters in BUILD CCB",dico_parameters
                    self.dico_thread[action] = threading.Thread(None,self._generateCCB,None,(dico_parameters,
                                                                                          cr_with_parent,
                                                                                          cr_workflow,
                                                                                          cr_domain,
                                                                                          log_on,
                                                                                          list_cr_for_ccb,
                                                                                          status_list,
                                                                                          ccb_time,
                                                                                          list_cr_type,
                                                                                          list_cr_status,
                                                                                          list_cr_doamin))
                    self.dico_thread[action].start()

                elif action == "BUILD_REVIEW_REPORT":
                    review_id = self.queue.get(1)
                    empty = self.queue.get(2)
                    dico = self.queue.get(3)
                    project_list = self.queue.get(4)
                    list_cr_type = self.queue.get(5)
                    list_cr_status = self.queue.get(6)
                    list_cr_doamin = self.queue.get(7)
                    self.build_doc_thread = threading.Thread(None,self._generateReviewReport,None,(review_id,empty,dico,project_list,list_cr_type,list_cr_status,list_cr_doamin))
                    self.build_doc_thread.start()
                elif action == "BUILD_DELIVERY_SHEET":
                    type_sds = self.queue.get(1)
                    dico_tags = self.queue.get(2)
                    #print dico_tags
                    self.build_doc_thread = threading.Thread(None,self._generateDeliverySheet,None,(type_sds,dico_tags))
                    self.build_doc_thread.start()
                elif action == "START_SESSION":
                    # start synergy session
                    self.start_session_thread = threading.Thread(None,self._startSession,None,(self.system,
                                                                                               self.item,
                                                                                               self.database,
                                                                                               self.login,
                                                                                               self.password,
                                                                                               self.aircraft))
                    self.start_session_failed = False
                    self.start_session_thread.start()
                    self.launch_session = True
                    self.setSessionStarted()
                elif action == "GET_BASELINES":
                    if self.session_started:
                        #release = self.master_ihm.release
                        release = self.queue.get(1)
                        self.get_baselines_thread = threading.Thread(None,self._getBaselinesList,None,(release,))
                        self.get_baselines_thread.start()
                elif action == "GET_RELEASES":
                    if self.session_started:

                        regexp = self.queue.get(1)
                        active = self.master_ihm.getActive()
                        if active:
                            query = "release -active -u -l"
                        else:
                            query = "release -u -l"
                        self.master_ihm.log("ccm " + query)
                        self.get_releases_thread = threading.Thread(None,self._getReleasesList,None,(query,regexp))
                        self.get_releases_thread.start()
                elif action == "GET_PROJECTS":
                    if self.session_started:
                        baseline = self.master_ihm.baseline
                        release = self.master_ihm.release
                        query = self._defineProjectQuery(release,
                                                         baseline)
                        self.master_ihm.log("ccm " + query)
                        self.get_projects_thread = threading.Thread(None,self._getProjectsList,None,(query,release,baseline))
                        self.get_projects_thread.start()
                elif action == "READ_STATUS":
                    self.set_status_thread = threading.Thread(None,self._getSessionStatus,None)
                    self.set_status_thread.start()
                elif action == "CLOSE_SESSION":
                    self.set_status_thread = threading.Thread(None,self._closeSession,None)
                    self.set_status_thread.start()
                elif action == "MAKE_DIFF":
                    data = self.queue.get(1)
                    baseline_prev = data[0]
                    baseline_cur = data[1]
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("BASELINES_DIFF","","","",baseline_prev,baseline_cur))
                    self.send_cmd_thread.start()
                elif action == "SHOW_BASELINE":
                    data = self.queue.get(1)
                    baseline_cur = data[0]
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("BASELINES_SHOW","","","","",baseline_cur))
                    self.send_cmd_thread.start()
                elif action == "SEND_CMD":
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None)
                    self.send_cmd_thread.start()
                elif action == "EXPORT_CR":
                    cr_id = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._exportCR,None,(cr_id,))
                    self.send_cmd_thread.start()
                elif action == "LIST_ITEMS":
                    release = self.queue.get(1)
                    project = self.queue.get(2)
                    baseline = self.queue.get(3)
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("ITEMS",release,baseline,project))
                    self.send_cmd_thread.start()
                elif action == "SCOPE":
                    release = self.queue.get(1)
                    project = self.queue.get(2)
                    baseline = self.queue.get(3)
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("SCOPE",release,baseline,project))
                    self.send_cmd_thread.start()
                elif action == "LIST_TASKS":
                    release = self.queue.get(1)
                    baseline = self.queue.get(2)
                    with_cr = self.queue.get(3)
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("TASKS",release,baseline,with_cr))
                    self.send_cmd_thread.start()
                elif action == "LIST_HISTORY":
                    release = self.queue.get(1)
                    baseline = self.queue.get(2)
                    project = self.queue.get(3)
                    history_scope = self.queue.get(4)
                    cid_type = self.queue.get(5)
                    dico_parameters = self.queue.get(6)
                    self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("HISTORY",
                                                                                     release,
                                                                                     baseline,
                                                                                     project,
                                                                                     history_scope,
                                                                                     cid_type,
                                                                                     dico_parameters))
                    self.send_cmd_thread.start()
                elif action == "GET_RELEASE_VS_BASELINE":
                    if self.session_started:
                        baseline = self.queue.get(1)
                        self.send_cmd_thread = threading.Thread(None,self._sendCmd,None,("GET_RELEASE_VS_BASELINE","",baseline))
                        self.send_cmd_thread.start()
                elif action == "EXEC_USER_CMD":
                        tbl_user_cmd = self.queue.get(1)
                        self.send_cmd_thread = threading.Thread(None,self._execUserCmd,None,(tbl_user_cmd,))
                        self.send_cmd_thread.start()
                elif action == "PREVIEW_CR_QUERY":
                    self.send_cmd_thread = threading.Thread(None,self._preview_CR_Query,None)
                    self.send_cmd_thread.start()

                elif action == "GET_CR":
                    data = self.queue.get(1)
                    baseline = data[0]
                    ccb_type = data[1]
                    extension = True
                    for_review_on = data[2]
                    cr_with_parent = data[3]
                    log_on = data[4]
                    component_type = data[5]
                    detected_on = data[6]
                    implemented_for = data[7]
                    old_cr_workflow = data[8]
                    ccb_time = data[9]
                    children = data[10]
                    list_cr_type = data[11]
                    list_cr_status = data[12]
                    list_cr_doamin = data[13]
                    self.build_doc_thread = threading.Thread(None,self._getCR,None,(baseline,
                                                                                    extension,
                                                                                    for_review_on,
                                                                                    cr_with_parent,
                                                                                    log_on,
                                                                                    component_type,
                                                                                    detected_on,
                                                                                    implemented_for,
                                                                                    old_cr_workflow,
                                                                                    ccb_time,
                                                                                    children,
                                                                                    list_cr_type,
                                                                                    list_cr_status,
                                                                                    list_cr_doamin))
                    self.build_doc_thread.start()
                elif action == "TEST_BLOCKDIAG":
                    self.test_blockdiag_thread = threading.Thread(None,self._testBlockDiag,None)
                    self.test_blockdiag_thread.start()
                elif action == "GET_CR_MAPPING":
                    data = self.queue.get(1)
                    baseline = data[0]
                    ccb_type = data[1]
                    extension = True
                    for_review_on = data[2]
                    cr_with_parent = data[3]
                    log_on = data[4]
                    component_type = data[5]
                    detected_on = data[6]
                    implemented_for = data[7]
                    old_cr_workflow = data[8]
                    ccb_time = data[9]
                    children = data[10]
                    list_cr_type = data[11]
                    list_cr_status = data[12]
                    list_cr_doamin = data[13]
                    cr_mapping_direction = data[14]
                    list_cr_for_ccb = data[15]
                    self.dico_thread[action] = threading.Thread(None,self._getCR_Mapping,None,(baseline,
                                                                                    extension,
                                                                                    for_review_on,
                                                                                    cr_with_parent,
                                                                                    log_on,
                                                                                    component_type,
                                                                                    detected_on,
                                                                                    implemented_for,
                                                                                    old_cr_workflow,
                                                                                    ccb_time,
                                                                                    True,
                                                                                    list_cr_type,
                                                                                    list_cr_status,
                                                                                    list_cr_doamin,
                                                                                    cr_mapping_direction,
                                                                                    list_cr_for_ccb))
                    self.dico_thread[action].start()
                elif action == "ABORT_THREAD":
                    for key,thread in self.dico_thread.iteritems():
                        # Caution: _stop might disappear in next Python version
                        #thread.__stop() # missing stop method
                        self.master_ihm.log("Abort {:s} thread".format(key))

                elif action == "START_APACHE":
                    config= "httpd_ece.conf"
                    self.send_cmd_thread = threading.Thread(None,self.__apache_start,None,(config,))
                    self.send_cmd_thread.start()

                elif action == "CHECK_UPDATE":
                    self.check_update_thread = threading.Thread(None,self._checkUpdate,None,)
                    self.check_update_thread.start()

                elif action == "RELOAD_CONFIG":
                    # Get config
##                    self.__loadConfig()
                    self.ihm.popup("Config file docid.ini reloaded.")

                elif action == "RELOAD_BASELINEBOX":
                    if self.session_started:
                        stdout = self.queue.get(1)
                        if stdout != "":
                            self.master_ihm.baselinelistbox.configure(text="white")
                            self.master_ihm.log("Available baseline found:")
                            output = stdout.splitlines()
                            self.master_ihm.baselinelistbox.delete(0, END)
                            if len(output) > 1:
                                self.master_ihm.baselinelistbox.insert(END, "All")
                                self.master_ihm.baselinelistbox.selection_set(first=0)
                            for line in output:
                                line = re.sub(r"^ *[0-9]{1,3}\) ",r"",line)
                                self.master_ihm.baselinelistbox.insert(END, line)
                                self.master_ihm.baselinelistbox_1.insert(END, line)
                                self.master_ihm.baselinelistbox_2.insert(END, line)
                                self.master_ihm.log(line)
                            self.master_ihm.releaselistbox.selection_set(first=0)
                            self.master_ihm.baselinelistbox.configure(text="white")
                        else:
                            self.master_ihm.resetBaselineListbox()
                            self.master_ihm.log(" No available baselines found.")
                        #self.resetProjectListbox()
                        self.master_ihm.baselinelistbox.configure(text=NORMAL)
                        # Set scrollbar at the bottom
                        self.master_ihm.defill()
                elif action == "RELOAD_RELEASEBOX":
                    if self.session_started:
                        stdout = self.queue.get(1)
                        if stdout != "":
                            # List of releases found
                            output = stdout.splitlines()
                            # Populate release listbox
                            self.master_ihm.updateReleaseListBox(output)
                        else:
                            self.master_ihm.noneReleaseListBox()
                elif action == "RELOAD_PROJECTBOX":
                    if self.session_started:
                        stdout = self.queue.get(1)
                        release = self.queue.get(2)
                        baseline_selected = self.queue.get(3)
                        if stdout != "":
                            #self.master_ihm.projectlistbox.delete(0, END)
                            self.master_ihm.projectlistbox.clear()
                            output = stdout.splitlines()
                            # Here the list of projects is set
                            self.list_projects = []
                            if Tool.isAttributeValid(baseline_selected):
                                if Tool.isAttributeValid(release):
                                    for line in output:
                                        line = re.sub(r"^ *[0-9]{1,3}\) ",r"",line)
                                        m = re.match(r'(.*)-(.*);(.*|<void>)$',line)
                                        if m:
                                            project = m.group(1) + "-" + m.group(2)
                                            baseline_string = m.group(3)
                                            baseline_splitted = baseline_string.split(',')
                                            for baseline in baseline_splitted:
                                                baseline = re.sub(r".*#",r"",baseline)
                                                if baseline == baseline_selected:
                                                    self.list_projects.append(project)
                                                    break
                                        else:
                                            m = re.match(r'^Baseline(.*):$',line)
                                            if not m:
                                                project = line
                                                self.list_projects.append(project)
                                else:
                                    num = 0
                                    for project in output:
                                        if num > 0:
                                            self.list_projects.append(project)
                                        num += 1
                            else:
                                for line in output:
                                    line = re.sub(r"^ *[0-9]{1,3}\) ",r"",line)
                                    m = re.match(r'(.*)-(.*);(.*)$',line)
                                    if m:
                                        project = m.group(1) + "-" + m.group(2)
                                        #print "name " + m.group(1) + " version " + m.group(2)
                                    else:
                                        project = line
                                    self.list_projects.append(project)
                            # Update list of project of GUI
                            self.master_ihm.updateProjectListBox(self.list_projects)
                        else:
                            self.master_ihm.noneProjectListBox()
                elif action == "RELOAD_CRLISTBOX":
                    if self.session_started:
                        try:
                            print "Display CR RELOAD_CRLISTBOX"
                            list_cr = self.queue.get(1)
                            # Update list of project of GUI
                            crlistbox = self.master_ihm.crlistbox
                            crlistbox.configure(text=NORMAL)
                            crlistbox.delete(0, END)
                            inter = 0
                            for cr_description in list_cr:
                                crlistbox.insert(END, cr_description)
                                if inter % 2 == 0:
                                    crlistbox.itemconfig(inter,{'bg':'gray88','fg':'black'})
                                else:
                                    crlistbox.itemconfig(inter,{'bg':'lightgrey','fg':'black'})
                                inter += 1
                            crlistbox.configure(text="white")
                        except AttributeError:
                            pass
                elif action == "CHECK_LLR":
                    dirname = self.queue.get(1)
                    if dirname == "":
                        dirname = self.dir_swdd
                    hsid_dirname = self.queue.get(2)
                    if hsid_dirname == "":
                        hsid_dirname = self.hsid
                    self.send_cmd_thread = threading.Thread(None,self._checkLLRCmd,None,(dirname,False,("SWDD","PLDDD"),hsid_dirname))
                    self.send_cmd_thread.start()
                elif action == "CHECK_HLR":
                    dirname = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._checkLLRCmd,None,(dirname,True,("SWRD","PLDRD")))
                    self.send_cmd_thread.start()
                elif action == "CHECK_UPPER":
                    dirname = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._checkUpperCmd,None,(dirname,))
                    self.send_cmd_thread.start()
                elif action == "EXPORT_IS_HLR":
                    dirname_req = self.queue.get(1)
                    dirname_upper = self.queue.get(2)
                    reference = self.queue.get(3)
                    issue = self.queue.get(4)
                    release = self.queue.get(5)
                    hsid_dirname = self.queue.get(6)
                    reviewer_name = self.queue.get(7)
                    default_status = self.queue.get(8)
                    dico_parameters = self.queue.get(9)
                    self.send_cmd_thread = threading.Thread(None,self._exportIS,None,(dirname_req,
                                                                                      dirname_upper,
                                                                                      True,
                                                                                      reference,
                                                                                      issue,
                                                                                      release,
                                                                                      "",
                                                                                      reviewer_name,
                                                                                      default_status,
                                                                                      dico_parameters
                    ))
                    self.send_cmd_thread.start()
                elif action == "EXPORT_IS_LLR":
                    dirname_req = self.queue.get(1)
                    dirname_upper = self.queue.get(2)
                    reference = self.queue.get(3)
                    issue = self.queue.get(4)
                    release = self.queue.get(5)
                    hsid_dirname = self.queue.get(6)
                    reviewer_name = self.queue.get(7)
                    default_status = self.queue.get(8)
                    dico_parameters = self.queue.get(9)
                    self.send_cmd_thread = threading.Thread(None,self._exportIS,None,(dirname_req,
                                                                                      dirname_upper,
                                                                                      False,
                                                                                      reference,
                                                                                      issue,
                                                                                      release,
                                                                                      hsid_dirname,
                                                                                      reviewer_name,
                                                                                      default_status,
                                                                                      dico_parameters
                    ))
                    self.send_cmd_thread.start()
                elif action == "CHECK_IS_HLR":
                    dirname_upper = self.queue.get(1)
                    dirname_req = self.queue.get(2)
                    filename_is = self.queue.get(3)
                    component = self.queue.get(4)
                    skip_change_synergy_var = self.queue.get(5)
                    self.send_cmd_thread = threading.Thread(None,self._checkISCmd,None,(dirname_upper,
                                                                                        dirname_req,
                                                                                        filename_is,
                                                                                        component,
                                                                                        True,
                                                                                        skip_change_synergy_var
                    ))
                    self.send_cmd_thread.start()
                elif action == "EXPORT_IS_SYNTHESIS":
                    dirname_upper = self.queue.get(1)
                    dirname_req = self.queue.get(2)
                    filename_is = self.queue.get(3)
                    component = self.queue.get(4)
                    skip_change_synergy_var = self.queue.get(5)
                    self.send_cmd_thread = threading.Thread(None,self._checkISCmd,None,(dirname_upper,
                                                                                        dirname_req,
                                                                                        filename_is,
                                                                                        component,
                                                                                        True,
                                                                                        skip_change_synergy_var,
                                                                                        True
                    ))
                    self.send_cmd_thread.start()
                elif action == "CHECK_IS_LLR":
                    dirname_upper = self.queue.get(1)
                    if dirname_upper == "":
                        dirname_upper = self.dir_swrd
                    dirname_req = self.queue.get(2)
                    if dirname_req == "":
                        dirname_req = self.dir_swdd
                    filename_is = self.queue.get(3)
                    if filename_is == "":
                        filename_is = self.is_swdd
                    print "IS:",filename_is
                    component = self.queue.get(4)
                    skip_change_synergy_var = self.queue.get(5)
                    self.send_cmd_thread = threading.Thread(None,self._checkISCmd,None,(dirname_upper,
                                                                                        dirname_req,
                                                                                        filename_is,
                                                                                        component,
                                                                                        False,
                                                                                        skip_change_synergy_var))
                    self.send_cmd_thread.start()
                elif action == "CHECK_IS_DOC":
                    filename_is = self.queue.get(1)
                    cr_process_version = self.queue.get(2)
                    self.send_cmd_thread = threading.Thread(None,self._checkISDocCmd,None,(filename_is,))
                    self.send_cmd_thread.start()
                elif action == "GEN_DERIVED_HLR":
                    dirname = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._genHLRDerivedCmd,None,(dirname,))
                    self.send_cmd_thread.start()
                elif action == "GEN_DERIVED_LLR":
                    dirname = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._genLLRDerivedCmd,None,(dirname,))
                    self.send_cmd_thread.start()
                elif action == "GEN_DERIVED_UPPER":
                    dirname = self.queue.get(1)
                    upper = CheckLLR()
                    list_upper = upper.getListUpper()
                    self.send_cmd_thread = threading.Thread(None,self._genHLRDerivedCmd,None,(dirname,list_upper))
                    self.send_cmd_thread.start()
                elif action == "UPDATE_CHAPTER_HLR":
                    filename = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._getChapterDialogHLR,None,(filename,))
                    self.send_cmd_thread.start()
                elif action == "CHECK_SHLVCP":
                    dirname_shlvcp = self.queue.get(1)
                    #if dirname_shlvcp == "":
                    #    dirname_shlvcp = self.dir_shlvcp
                    dirname_swrd = self.queue.get(2)
                    #if dirname_swrd == "":
                    #    dirname_swrd = self.dir_shlvcp
                    self.send_cmd_thread = threading.Thread(None,self._checkSHLVCP,None,(dirname_shlvcp,dirname_swrd))
                    self.send_cmd_thread.start()
                elif action == "CHECK_SHLVCP_VS_SWRD":
                    dirname_shlvcp = self.queue.get(1)
                    self.send_cmd_thread = threading.Thread(None,self._checkSHLVCP_VS_SWRD,None,(dirname_shlvcp,))
                    self.send_cmd_thread.start()
                elif action == "EXPORT_SCOD":
                    self.send_cmd_thread = threading.Thread(None,self._buildSCOD,None,("ACEM",))
                    self.send_cmd_thread.start()
                elif action == "EXPORT_VHDL":
                    self.send_cmd_thread = threading.Thread(None,self._buildVHDL,None,("ACEM",))
                    self.send_cmd_thread.start()
                elif action == "EXPORT_FUNC_CALL_TREE":
                    stack = Stack()
                    self.send_cmd_thread = threading.Thread(None,stack._stackAnalysis,None)
                    self.send_cmd_thread.start()
                else:
                    pass
                self.unlock()
            except Queue.Empty:
                pass

    def thread_readEOC(self,
                       eoc_filename,
                       dico_addr,
                       dspic=False):
        hw_sw_compatibility,part_number,checksum,failed = self._readEOC(eoc_filename,
                                                                        dico_addr,
                                                                        dspic)
        self.master_ihm.displayEOC_Info((hw_sw_compatibility,
                                         part_number,
                                         checksum,
                                         failed))
        self.master_ihm.resultGenerateCID(not failed,
                                          None,
                                          text="EXECUTABLE OBJECT CODE READING")
        return hw_sw_compatibility,part_number,checksum,failed

    def periodicCall(self):
        """
        Check every 1000 ms if there is something new in the queue.
        """
##        print time.strftime("%H:%M:%S", time.localtime())
##        print time.strftime("PERIODIC CALL " + self.name_id)
        self.processIncoming()
        if not self.running:
            # This is the brutal stop of the system. You may want to do
            # some cleanup before actually shutting it down.
            import sys
            sys.exit(1)
        try:
            self.master_ihm.after(1000, self.periodicCall)
        except AttributeError:
            time.sleep(1)
            self.periodicCall

    def _setRelease(self):
        self.master_ihm.release = self.previous_release
        self.master_ihm.button_list_items.configure(text=NORMAL)
        self.master_ihm.button_list_tasks.configure(text=NORMAL)
        self.master_ihm.button_set_baselines.configure(text=NORMAL)
        self.master_ihm.setBaseline(self.master_ihm.release)

    def _setBaseline(self):
        self.master_ihm.baseline = self.previous_baseline
        self.master_ihm.setBaselineSynergy(self.master_ihm.baseline)
        self.master_ihm.projectlistbox.configure(text=NORMAL)
        self.master_ihm.button_find_projects.configure(text=NORMAL)
        self.master_ihm.button_list_items.configure(text=NORMAL)
        self.master_ihm.button_list_tasks.configure(text=NORMAL)
        executed = self._sendCmd("GET_RELEASE_VS_BASELINE","",self.master_ihm.baseline)
        if executed:
            pass
##            interface.button_select.configure(state=NORMAL)
        query = self._defineProjectQuery(self.master_ihm.release,self.master_ihm.baseline)
        self._getProjectsList(query,self.master_ihm.release,self.master_ihm.baseline)

    def _setProject(self):
        self.master_ihm.project = self.previous_project
        self.master_ihm.button_select.configure(text=NORMAL)
        self.master_ihm.button_create_delivery_sheet.configure(text=NORMAL)
        self.master_ihm.button_list_items.configure(text=NORMAL)
        self.master_ihm.button_list_tasks.configure(text=NORMAL)
        self.master_ihm.setProject(self.master_ihm.project)

    def _add(self, action):
        # add an action to the manager.  returns tags to use in
        # associated text widget
        tag = "hlink-%d" % len(self.links)
        self.links[tag] = action
        return "hlink", tag

    def _click(self, event):
        for tag in self.master_ihm.general_output_txt.tag_names(CURRENT):
            if tag[:6] == "hlink-":
                            self.links[tag]()

    def _startSession(self,
                      system,
                      item,
                      database,
                      login,
                      password,
                      aircraft="",
                      queue_thread_gui=None):
        """ Function to start Synergy session
             - invoke command ccm start ...
             - display synergy feedback
             - retrieve last session information
             - enable SELECT and REFRESH buttons
             - get list of releases
            called by the thread """
        # GUI/CLI
        print "_startSession"
        #self.master_ihm.put_in_gui_queue("Test on_main_thread")
        #self.lock()

        #self.previous_release = ""
        #self.previous_baseline = ""
        #self.previous_project = ""
        session_started = False
        if database is not None \
                and login != "":
            try:
                query = "start /nogui /q /d /usr/local/ccmdb/{:s} /u /usr/local/ccmdb/{:s} /s {:s} /n {:s} /pw {:s}".format(database,database,self.ccm_server,login,password)
                stdout,stderr = self.ccm_query(query,"Synergy session start")
            except UnicodeEncodeError:
                stdout = False
            self.master_ihm.resultStartSession(stdout,stderr)

        else:
            self.master_ihm.sayNoDatabase()
            self.start_session_failed = True
            stdout = ""

        return stdout

    def _easyIG(self):
        filename = self.easyig.get()
        self.master_ihm.displayHyperlink("hlink",filename,"Web page created.")
        self.easyig.start()

    def _getQA(self,qams_user_id,url_root,name=""):
        filename = self.getqa.get(qams_user_id,
                                  url_root=url_root,
                                  name=name)
        self.master_ihm.displayHyperlink("hlink_local_qams",filename,"Local web page created.")
        self.getqa.start()

    def _getReleasesList(self,query="cmd release -u -l",regexp=""):
        ''' get releases list '''
        self.lock()
        stdout,stderr = self.ccm_query(query,"Get releases")
        if regexp != "":
            output = stdout.splitlines()
            list_release = ""
            if stdout != "":
                for line in output:
                    m = re.match(regexp,line)
                    if m:
                        list_release += line + "\n"
            if list_release == "":
                self.master_ihm.log("Check release_regexp parameter in docid.ini which value is: " + regexp)
        else:
            list_release = stdout
        ##self.master_ihm.resultGenerateCID(docx_filename,
        #                                  False,
        #                                  text="SYNERGY GET RELEASES COMMAND")
        self.queue.put("RELOAD_RELEASEBOX") # action to reload release listbox
        self.queue.put(list_release)
        self.unlock()

    def _getBaselinesList(self,release):
        """
        get baseline list for Release/Baseline/Project window
        :param release:
        :return:
        """
        self.lock()
        list_baselines = []
        if Tool.isAttributeValid(release):
            query = 'baseline -l -release {:s} -f "%name"'.format(release)
            query_all = 'baseline -l -u -f "%name"'
            self.master_ihm.log("ccm " + query_all)
            stdout,stderr = self.ccm_query(query_all,"Get all baselines")
            if stdout != "":
                #print "TEST 1"
                #print stdout
                #print "TEST 2"
                output = stdout.splitlines()
                release_name = Tool.getReleaseName(release)
                for line in output:
                    m = re.match(r'^\s*({:s}.*)'.format(release_name), line)
                    if m:
                        list_baselines.append(line)
        else:
            query = 'baseline -l -f "%name"'
        self.master_ihm.log("ccm " + query)
        stdout,stderr = self.ccm_query(query,"Get baselines")
        if stdout != "":
            output = stdout.splitlines()
            self.master_ihm.updateBaselineListBox(output,list_baselines)
        else:
            self.master_ihm.noneBaselineListBox()
        self.unlock()

    def _getProjectsList(self,query,release,baseline_selected):
        """
        :param query:
        :param release:
        :param baseline_selected:
        :return:
        """
        self.lock()
        stdout,stderr = self.ccm_query(query,"Get projects")
        self.queue.put("RELOAD_PROJECTBOX") # action to get projects
        self.queue.put(stdout)
        self.queue.put(release)
        self.queue.put(baseline_selected)
        #self.master_ihm.success.config(fg='magenta',bg = 'green',text="COMMAND SUCCEEDED")
        self.unlock()

    def _exportCreateCR_Table(self,cr_id,type_cr="parent"):
        # Get parent CR
        tbl_parent_cr_id = self._getParentCR(cr_id,type_cr)
        if tbl_parent_cr_id:
            #
            # Get parent ID information
            #
            parent_cr = ""
            for parent_cr_id in tbl_parent_cr_id:
                res_parent_cr = self._getParentInfo(parent_cr_id)
                if res_parent_cr:
                    parent_cr += res_parent_cr
                    self.master_ihm.log("{:s} CR: {:s}".format(type_cr, res_parent_cr))
                #else:
                #    self.master_ihm.log("No result for _getParentInfo (twice).")
        else:
            parent_cr = "<td><IMG SRC=\"../img/changeRequestIcon.gif\">---</td><td>---</td><td>---</td><td>---</td><td>---</td>"
        return parent_cr

    def _exportCR(self,cr_id):
        """
        Function to export CR to web browser
        called by crlistbox_onselect
        """
        query = "query -t problem \"(problem_number='" + cr_id + "')\" -u -f \
                 \"<table border='1'>\
                 <cell name='CR_domain'>%CR_domain</cell>\
                 <cell name='CR_type'>%CR_type</cell>\
                 <cell name='crstatus'>%crstatus</cell>\
                 <cell name='problem_synopsis'>%problem_synopsis</cell>\
                 <cell name='SCR_In_Analysis_id'>%SCR_In_Analysis_id</cell>\
                 <cell name='create_time'>%create_time</cell>\
                 <cell name='CR_ECE_classification'>%CR_ECE_classification</cell>\
                 <cell name='CR_customer_classification'>%CR_customer_classification</cell>\
                 <cell name='CR_request_type'>%CR_request_type</cell>\
                 <cell name='CR_detected_on'>%CR_detected_on</cell>\
                 <cell name='CR_applicable_since'>%CR_applicable_since</cell>\
                 <cell name='CR_implemented_for'>%CR_implemented_for</cell>\
                 <cell name='CR_origin'>%CR_origin</cell>\
                 <cell name='CR_origin_desc'>%CR_origin_desc</cell>\
                 <cell name='CR_expected'>%CR_expected</cell>\
                 <cell name='CR_observed'>%CR_observed</cell>\
                 <cell name='CR_functional_impact'>%CR_functional_impact</cell>\
                 <cell name='CR_analysis'>%CR_analysis</cell>\
                 <cell name='CR_correction_description'>%CR_correction_description</cell>\
                 <cell name='CR_product_impact'>%CR_product_impact</cell>\
                 <cell name='CR_doc_impact'>%CR_doc_impact</cell>\
                 <cell name='CR_verif_impact'>%CR_verif_impact</cell>\
                 <cell name='impact_analysis'>%impact_analysis</cell>\
                 <cell name='functional_limitation_desc'>%functional_limitation_desc</cell>\
                 <cell name='implemented_modification'>%implemented_modification</cell>\
                 <cell name='CR_implementation_baseline'>%CR_implementation_baseline</cell>\
                 <cell name='SCR_Verif_Test_Bench'>%SCR_Verif_Test_Bench</cell>\
                 <cell name='SCR_Verif_Test_Procedure'>%SCR_Verif_Test_Procedure</cell>\
                 <cell name='CR_verification_activities'>%CR_verification_activities</cell>\
                 <cell name='functional_limitation'>%functional_limitation</cell>\
                 <cell name='SCR_Closed_id'>%SCR_Closed_id</cell>\
                 <cell name='SCR_Closed_time'>%SCR_Closed_time</cell>\
                 <cell name='problem_number'>%problem_number</cell>\
                 <cell name='modify_time'>%modify_time</cell>\
                 <cell name='SCR_Fixed_time'l>%SCR_Fixed_time</cell>\
                 </table>\""
##                 <cell name='transition_log'>%transition_log</cell>\
        executed = True
        filename = "log_SCR_" + cr_id + "_%d.html" % floor(time.time())
        #with open(self.gen_dir + filename, 'w') as of:
        if query != "":
            ccm_query = 'ccm ' + query + '\n'
            cmd_out = self._ccmCmd(query,False)
            # Replace STX and ETS and e cute characters
            char = {r'\x02':r'<',r'\x03':r'>',r'\xe9':r'e'}
            for before, after in char.iteritems():
                cmd_out = re.sub(before,after,cmd_out)
            if cmd_out == "":
                self.master_ihm.log("No result.")
                executed = False
            #
            # Get transition log
            #
            query = "query -t problem \"(problem_number='{:s}')\" -u -f \"%transition_log\"".format(cr_id)
            ccm_query = 'ccm ' + query + '\n'
            transi_log = self._ccmCmd(query,False)
            transi_log_filtered = self._filterASCII(transi_log)

            # Get parent CR
            if 0==1:
                tbl_parent_cr_id = self._getParentCR(cr_id)
                if tbl_parent_cr_id:
                    #
                    # Get parent ID information
                    #
                    parent_cr = ""
                    for parent_cr_id in tbl_parent_cr_id:
                        res_parent_cr = self._getParentInfo(parent_cr_id)
                        if res_parent_cr:
                            parent_cr += res_parent_cr
                            self.master_ihm.log("Parent CR:" + res_parent_cr)
                        #else:
                        #    self.master_ihm.log("No result for _getParentInfo (twice).")
                else:
                    parent_cr = "<td><IMG SRC=\"../img/changeRequestIcon.gif\">---</td><td>---</td><td>---</td><td>---</td><td>---</td>"
            parent_cr = self._exportCreateCR_Table(cr_id,"parent")
            # Get information CR
            information_cr = self._exportCreateCR_Table(cr_id,"information")
            print "information_cr",information_cr
            # Get child CR
            child_cr = self._exportCreateCR_Table(cr_id,"child")
            print "child_cr",child_cr
            self._parseCR(cmd_out,
                          transi_log_filtered,
                          parent_cr,
                          information_cr,
                          child_cr,
                          join(self.gen_dir,filename))
            # Get information CR
            #TODO:
        if executed:
            self.master_ihm.resultGenerateCID(filename,
                                              None,
                                              text="EXPORT CR")
            #self.master_ihm.log("Command executed.")
            #self.master_ihm.displayHyperlink("hlink",filename,"Log created.")
            url = join(os.getcwd(),"result")
            url = join(url,filename)
            print "URL",url
            webbrowser.open(url)
            #self.master_ihm.success.config(fg='magenta',bg = 'green',text="EXPORT SUCCEEDED")
        else:
            self.master_ihm.resultGenerateCID(False,
                                  None,
                                  text="EXPORT CR")
            #self.master_ihm.success.config(fg='yellow',bg = 'red',text="EXPORT FAILED")
        return executed

    def _getItems(self,release="",baseline="",project=""):
        global session_started
        output = ""
        output_format = "csv"
        release_name = re.sub(r"\/",r"",release)
        executed = False
        filename = "log_items_" + release_name + "_" + baseline + "_" + project + "_%d" % floor(time.time())
        filename += ".{:s}".format(output_format)
        if output_format == "csv":
            display_attr = ' -f "%release;%name;%version;%modify_time;%status;%task;%task_status;%change_request;%type" '
            show_header = "-nch"
        else:
            display_attr = ' -f "%release %name %version %modify_time %status %task %task_status %change_request %type" '
            show_header = "-ch"
        if baseline not in ("","All"):
            # Baseline
            # sh: show
            #  u: no number
            query = "baseline -sh objects  {:s} -u {:s}".format(baseline,display_attr)
            executed = True
        elif release not in ("","All"):
            # Query with a specifcic release
            #  ch: Specifies to use column headers for the output
            # nch: Specifies not to use column headers for the output
            #   u: no number
            query = 'query -sby name {:s} -n *.* -u -release {:s} '.format(show_header,release)
            if project not in ("*","All",""):
                # a project is selected
                # get sub-projects
                name, version = self.getProjectInfo(project)
                query += '"recursive_is_member_of(cvtype=\'project\' and name=\'{:s}\' and version=\'{:s}\' , \'none\')" {:s}'.format(name,version)
            query += display_attr
            executed = True
        elif project not in ("","All"):
            # No baseline, nor release selected but a project is
            query = 'query -sby name {:s} -n *.* -u "(is_member_of(\'{:s}\'))" {:s}'.format(show_header,release,display_attr)
            executed = True
        else:
            self.master_ihm.log("Please select a release or a baseline or a project.")
        if executed:
            self.master_ihm.log(" ccm " + query)
            self.master_ihm.defill()
            ccm_query = 'ccm ' + query + '\n\n'
            self.master_ihm.log("List objects (directories and executable objects are discarded).")
            cmd_out = self._ccmCmd(query)
            with open(join(self.gen_dir,filename), 'w') as of:
                if output_format == "csv":
                    header = "Release;Name;Version;Modify time;Status;Task;Task status;CR;Type\n"
                    of.write(header)
                else:
                    of.write(ccm_query)
                output = cmd_out.splitlines()
                for line in output:
                    # Skip directory or relocatable objects
                    # Skip automatic tasks and components tasks
                    # Remove Baseline info at the beginning
                    if output_format == "csv":
                        if not re.search("(dir|relocatable_obj)$",line) and not re.search("(task_automatic|component_task)",line) and not re.search("(^Baseline)",line):
                            # For CLI
                            print line
                            of.write(line)
                            of.write("\n")
                    else:
                        if not re.search("(dir|relocatable_obj)$",line) and not re.search("(task_automatic|component_task)",line):
                            of.write(line)
                            of.write("\n")
        if executed:
            self.master_ihm.log("Command executed.")
            try:
                self.master_ihm.displayHyperlink("hlink",filename,"Log created.")
            except AttributeError:
                pass
        # Set scrollbar at the bottom
        self.master_ihm.defill()
        return output

    def _execUserCmd(self,tbl_user_cmd):
        executed = True
        export = True
        filename = "log_%d.txt" % floor(time.time())
        with open(join(self.gen_dir,filename), 'w') as of:
            for query in tbl_user_cmd:
                if query != "":
                    self.master_ihm.log('ccm ' + query)
                    ccm_query = 'ccm ' + query + '\n'
                    cmd_out = self._ccmCmd(query)
                    if cmd_out == None:
                        executed = False
                        break
                    try:
                        of.write(ccm_query)
                    except UnicodeEncodeError as exception:
                        print "Character not supported:", exception
                    of.write(cmd_out)
                    if cmd_out == "":
                        self.master_ihm.log("No result.")
                        executed = False
        if executed:
            self.master_ihm.log("Command executed.")
            if export:
                 # Create hyperlink
                if filename is not None:
                    #self.master_ihm.displayHyperlink("hlink",filename,"Log created.")
                    self.master_ihm.resultGenerateCID(filename,
                                                      False,
                                                      text="USER COMMAND")
                else:
                    self.master_ihm.resultGenerateCID(False,
                                                      False,
                                                      text="USER COMMAND")
        else:
            self.master_ihm.success.config(fg='yellow',bg = 'red',text="COMMAND FAILED")

    def _sendCmd(self,cmd="",
                 release="",
                 baseline="",
                 project="",
                 baseline_prev="",
                 baseline_cur="",
                 dico_parameters={}):
        try:
            self.master_ihm.success.config(fg='red',bg = 'yellow',text="SYNERGY COMMAND IN PROGRESS")
        except AttributeError:
            pass
        export = True
        global session_started
        output_format = "csv"
        release_name = re.sub(r"\/",r"",release)
        filename = "log_" + release_name + "_" + baseline + "_" + str(project) + "_%d" % floor(time.time())
        if output_format == "txt":
            filename += ".txt"
        else:
            filename += ".csv"
        executed = False
        if cmd == "SCOPE":
            if release not in ("","All") and project not in ("*","All",""):
                test_string = "SQAP"
                text_found = ""
                project_name, project_version = self.getProjectInfo(project)
                query = "finduse -query \"release='" + release + "' and (cvtype='xls' or cvtype='doc' or cvtype='pdf' or cvtype='ascii' or cvtype='csrc') and recursive_is_member_of(cvtype='project' and name='"+ project_name +"' and version='"+ project_version +"' , 'none')\""
                self.master_ihm.log('ccm ' + query)
                self.master_ihm.general_output_txt.see(END)
                ccm_query = 'ccm ' + query + '\n\n'
                cmd_out = self._ccmCmd(query)
                output = cmd_out.splitlines()
##                test_string_1 = "Input_Data"
##                test_string_2 = "SQAP"
                list_items_skipped_1 = []
                list_items_skipped_2 = []
                regexp_1 = '^(.*)'+ project_name + '\\\\' + re.escape(self.input_data_filter) + '\\\\(.*)-(.*)@(.*)-(.*)$'
                regexp_2 = '^(.*)'+ project_name + '\\\\' + re.escape(self.peer_reviews_filter) + '\\\\(.*)-(.*)@(.*)-(.*)$'
                for line in output:
##                    print "Tested: " + line
                    # ex: SW_PLAN\SDP\IS_SDP_SW_PLAN_SQA.xlsm-1.7.0@SW_PLAN-1.3
                    m = re.match(regexp_1,line)
                    if m:
##                        text_found = m.group(2)
                        list_items_skipped_1.append(m.group(2))
                    else:
                        pass
                    m = re.match(regexp_2,line)
                    if m:
##                        text_found = m.group(2)
                        list_items_skipped_2.append(m.group(2))
                    else:
                        pass
##                print regexp_1
##                print regexp_2
                list_wo_doublons_1 = list(set(list_items_skipped_1))
                list_wo_doublons_2 = list(set(list_items_skipped_2))
##                print list_wo_doublons_1
##                print list_wo_doublons_2
                executed = True
            else:
                self.master_ihm.log("Please select a release and a project.")
        elif cmd == "GET_BASELINE_STATUS":
            executed = self._getBaselineInfo(baseline)
            export = False
        elif cmd == "GET_RELEASE_INFO":
            executed = self._getReleaseInfo(release)
            export = False
        elif cmd == "ITEMS":
            filename = "log_items_" + release_name + "_" + baseline + "_" + project + "_%d" % floor(time.time())
            if output_format == "csv":
                display_attr = ' -f "%release;%name;%version;%modify_time;%status;%task;%task_status;%change_request;%type" '
                show_header = "-nch"
                filename += ".csv"
            else:
                display_attr = ' -f "%release %name %version %modify_time %status %task %task_status %change_request %type" '
                show_header = "-ch"
                filename += ".txt"
            if Tool.isAttributeValid(project):
                # No baseline, nor release selected but a project is
                query  = 'query -sby name {:s} -n *.* -u '.format(show_header)
                #query += ' "(is_member_of(\'{:s}\'))" '.format(project)
                prj_name,prj_version = Tool.getProjectInfo(project)
                query += ' "(recursive_is_member_of(cvtype=\'project\' and name=\'{:s}\' and version=\'{:s}\' , \'none\'))" '.format(prj_name,prj_version)
                query += display_attr
                executed = True
            elif Tool.isAttributeValid(baseline):
                # Baseline
                # sh: show
                #  u: no number
                query = "baseline -sh objects {:s} -u ".format(baseline)
                query += display_attr
                executed = True
            elif Tool.isAttributeValid(release):
                # Query with a specifcic release
                #  ch: Specifies to use column headers for the output
                # nch: Specifies not to use column headers for the output
                #   u: no number
                query = 'query -sby name {:s} -n *.* -u -release {:s} '.format(show_header,release)
                if project not in ("*","All",""):
                    # a project is selected
                    # get sub-projects
                    name, version = self.getProjectInfo(project)
                    query += '"recursive_is_member_of(cvtype=\'project\' and name=\'' + name + '\' and version=\'' + version + '\' , \'none\')" '
                query += display_attr
                executed = True
            else:
                self.master_ihm.log("Please select a release or a baseline or a project.")
            if executed:
                self.master_ihm.log(" ccm " + query)
                ccm_query = 'ccm ' + query + '\n\n'
                self.master_ihm.log("List objects (directories and executable objects are discarded).")
                cmd_out = self._ccmCmd(query)
                with open(join(self.gen_dir,filename), 'w') as of:
                    if output_format == "csv":
                        header = "Release;Name;Version;Modify time;Status;Task;Task status;CR;Type\n"
                        of.write(header)
                    else:
                        of.write(ccm_query)
                    output = cmd_out.splitlines()
                    for line in output:
                        # Skip directory or relocatable objects
                        # Skip automatic tasks and components tasks
                        # Remove Baseline info at the beginning
                        if output_format == "csv":
                            if not re.search("(dir|relocatable_obj)$",line) and not re.search("(task_automatic|component_task)",line) and not re.search("(^Baseline)",line):
                                of.write(line)
                                of.write("\n")
                        else:
                            if not re.search("(dir|relocatable_obj)$",line) and not re.search("(task_automatic|component_task)",line):
                                of.write(line)
                                of.write("\n")
        elif cmd == "HISTORY":
            filename = "log_history_" + release_name + "_" + baseline + "_" + project + "_%d" % floor(time.time())
            filename += ".csv"
            self.display_attr = ' -f "<cell>%name</cell>' \
                         '<cell>%version</cell>' \
                         '<cell>%task</cell>' \
                         '<cell>%task_synopsis</cell>' \
                         '<cell>%change_request</cell>' \
                         '<cell>%change_request_synopsis</cell>' \
                         '<cell>%type</cell>' \
                         '<cell>%owner</cell>"'
            #self.display_attr = ' -f "%name|%version|%task|%task_synopsis|%change_request|%change_request_synopsis|%type" '
            header = ["Document","Issue","Tasks","Synopsis","CR","Synopsis","Owner"]

            # Patch
            source_only = baseline_prev
            cid_type = baseline_cur
            if source_only:
                list_type_src = ("csrc","incl") #self.master_ihm.log.list_type_src_sci
            else:
                list_type_src = ()
            self.object_released = False
            self.object_integrate = False
            output = self.getArticles(list_type_src,
                                      release,
                                      baseline,
                                      project,
                                      True,
                                      cid_type=cid_type)
            index_src = 0
            cid = BuildDoc(self.master_ihm,
                           session_started=self.getSessionStarted(),
                           dico_parameters=dico_parameters,
                           cid_type=cid_type)
            with open(join("result",filename), 'w') as of:
                header = "File;Version;Task;Synopsis;CR;Synopsis;Owner\n"
                of.write(header)
                for line in output:
                    line = re.sub(r"<void>",r"",line)
                    tbl_decod = self._parseCRCell(line)
                    #m = re.match(r'(.*)\|(.*)\|(.*)\|(.*)\|(.*)\|(.*)\|(.*)',line)
                    if tbl_decod:
                        result = cid._createTblSourcesHistory(tbl_decod,
                                                              source_only)
                        if result:
                            index_src +=1
                            # Remove Baseline info at the beginning
                            if not re.search("(^Baseline)",line):
                                for line_csv in result:
                                    of.write(line_csv)
                                    of.write("\n")
            print "Amount of source files found: " + str(index_src)# + "\n"
            executed = True
        elif cmd == "TASKS":
            with_cr = project
            filename = "log_tasks_" + release_name + "_" + baseline + "_%d" % floor(time.time())
            if output_format == "csv":
                display_attr = '"%displayname;%status;%task_synopsis"'
                show_header = "-nch"
                filename += ".csv"
            else:
                display_attr = '"%displayname %status %task_synopsis"'
                show_header = "-ch"
                filename += ".txt"
            if Tool.isAttributeValid(baseline):
                query = 'baseline -sh task ' + baseline + ' -u -f ' + display_attr + '\n'
                executed = True
            elif Tool.isAttributeValid(release):
                #   -u: is not numbered
                #  -qu: query
                # -rel: release
                query = 'task -u -qu -ts all_tasks ' + show_header + ' -rel ' + release + ' -f ' + display_attr + '\n'
                executed = True
            else:
                query = 'task -u -qu -ts all_tasks ' + show_header + ' -f ' + display_attr + '\n'
                executed = True
            if executed:
                ccm_query = 'ccm ' + query + '\n'
                self.master_ihm.log(ccm_query)
                cmd_out = self._ccmCmd(query)
                with open(join(self.gen_dir,filename), 'w') as of:
                    if output_format == "csv":
                        if not with_cr:
                            header = "Task ID;Status;Synopsis\n"
                        else:
                            header = "Task ID;Task status;Task synopsis;CR ID;CR status;CR synopsis;Owner\n"
                        of.write(header)
                    else:
                        of.write(ccm_query)
                    output = cmd_out.splitlines()
                    for line in output:
                        if with_cr:
                            # Add cr information
                            mtask = re.match(r'(.*);(.*);(.*)',line)
                            # Get task ID
                            if mtask:
                                task_id = mtask.group(1)
                                task_status = mtask.group(2)
                                task_synopsis = mtask.group(3)
                                query = 'task -u -show change_request ' + task_id + ' -f "CR %problem_number;;%problem_synopsis;;%crstatus" \n'
                                ccm_query = 'ccm ' + query + '\n'
                                self.master_ihm.log(ccm_query)
                                cmd_out = self._ccmCmd(query)
                                output_cr = cmd_out.splitlines()
                                cr_id_tbl = []
                                for line_cr in output_cr:
                                    mcr = re.match(r'^CR ([0-9]*);;(.*);;(.*)$',line_cr)
                                    # Get CR ID
                                    if mcr:
                                        cr_id = mcr.group(1)
                                        cr_synopsis = mcr.group(2)
                                        cr_status = mcr.group(3)
                                        print cr_id,cr_synopsis,cr_status
                                        #  Discard CR status prefix
                                        cr_status_lite = re.sub(r'(EXCR|SyCR|ECR|SACR|HCR|SCR|BCR|PLDCR)_(.*)', r'\2', cr_status)
                                        cr_id_tbl.append([task_id,task_status,task_synopsis,cr_id,cr_status_lite,cr_synopsis])
                                    else:
                                        cr_id_tbl.append([task_id,task_status,task_synopsis,"","",""])
                                self.master_ihm.defill()
                            for task_id,task_status,task_synopsis,cr_id,cr_synopsis,cr_status in cr_id_tbl:
                                text = ""
                                text += task_id + "; " + task_status + "; " + task_synopsis + "; "
                                text += cr_id + "; " + cr_synopsis + "; " + cr_status
                                line = text
                                # Remove Baseline info at the beginning
                                if output_format == "csv":
                                    # Skip automatic tasks and components tasks
                                    if not re.search("(task_automatic|component_task)",line) and not re.search("(^Baseline)",line):
                                        of.write(line)
                                        of.write("\n")
                                else:
                                    # Skip automatic tasks and components tasks
                                    if not re.search("(task_automatic|component_task)",line):
                                        of.write(line)
                                        of.write("\n")
                        else:
                            # Remove Baseline info at the beginning
                            if output_format == "csv":
                                # Skip automatic tasks and components tasks
                                if not re.search("(task_automatic|component_task)",line) and not re.search("(^Baseline)",line):
                                    of.write(line)
                                    of.write("\n")
                            else:
                                # Skip automatic tasks and components tasks
                                if not re.search("(task_automatic|component_task)",line):
                                    of.write(line)
                                    of.write("\n")
        elif cmd == "BASELINES_DIFF":
            #
            # - objects
            # all objects that are included in the baseline are displayed. The default format is:
            #   display name, status, owner, release, create time
            #
            # - tasks
            # all objects that are included in the baseline are displayed. The default format is:
            #   id, release, assignee, create time, description
            #
            # - changes requests
            #    include details about change requests (CRs) that are partially included and fully included in the two baselines.
            #    display name, problem synopsis
            #
            query = 'baseline -compare ' + baseline_prev + ' ' + baseline_cur + ' -tasks -objects -change_requests'
            ccm_query = 'ccm ' + query + '\n'
            self.master_ihm.log(ccm_query)
            cmd_out = self._ccmCmd(query)
            filename = "log_baseline_diff_" + release_name + "_" + baseline_prev + "_vs_" + baseline_cur + "_%d.txt" % floor(time.time())
            with open(join(self.gen_dir,filename), 'w') as of:
                of.write(ccm_query)
                of.write(cmd_out)
            executed = True
        elif cmd == "BASELINES_SHOW":
            #
            # - objects
            # all objects that are included in the baseline are displayed. The default format is:
            #   display name, status, owner, release, create time
            #
            # - tasks
            # all objects that are included in the baseline are displayed. The default format is:
            #   id, release, assignee, create time, description
            #
            # - changes requests
            #    include details about change requests (CRs) that are partially included and fully included in the two baselines.
            #    display name, problem synopsis
            #
            query = 'baseline -sh objects  ' + baseline_cur
            ccm_query = 'ccm ' + query + '\n'
            self.master_ihm.log(ccm_query)
            cmd_out = self._ccmCmd(query)
            filename = "log_baseline_show_" + baseline_cur + "_%d.txt" % floor(time.time())
            with open(join(self.gen_dir,filename), 'w') as of:
                of.write(ccm_query)
                of.write(cmd_out)
            executed = True
        elif cmd == "GET_RELEASE_VS_BASELINE":
            query = "baseline -show information {:s}".format(baseline)
            ccm_query = 'ccm ' + query + '\n'
            self.master_ihm.log(ccm_query)
            cmd_out = self._ccmCmd(query)
            if cmd_out is None:
                executed = False
            else:
                filename = "log_baseline_show_" + baseline + "_%d.txt" % floor(time.time())
                with open(join(self.gen_dir,filename), 'w') as of:
                    of.write(ccm_query)
                    of.write(cmd_out)
                output = cmd_out.splitlines()
                for line in output:
                    # Attention aux espaces a supprimer
                    m = re.match(r'^  Release:( *)([^ .]*)',line)
                    if m:
                        release = m.group(2)
                        self.master_ihm.log("Associated release is: " + release)
                executed = True
        else:
            # User command
            self.master_ihm.getUserCmd()
            #cmd_txt = self.master_ihm.command_txt.get(1.0,END)
            #output = cmd_txt.splitlines()
            return
        if executed:
            self.master_ihm.log("Command executed.")
            if export:
                 # Create hyperlink
                if filename is not None:
                    #self.master_ihm.displayHyperlink("hlink",filename,"Log created.")
                    self.master_ihm.resultGenerateCID(filename,
                                                      False,
                                                      text="COMMAND")
                else:
                    self.master_ihm.resultGenerateCID(False,
                                                      False,
                                                      text="COMMAND")
        else:
            self.master_ihm.success.config(fg='yellow',bg = 'red',text="COMMAND FAILED")
        return executed

    def _getCoverage(self,directory):
        tool = Tool()
        tool.basename = directory
        tool.listDir("")
        filename = "tu_coverage_%d.txt" % floor(time.time())
        with open(join("result",filename), 'w') as of:
            of.write("{:s};{:s};{:s};{:s};{:s}\n".format("File","Statements","Decisions","Basic conditions","Modified conditions"))
            for name,percentage in tool.list_coverage.iteritems():
                self.master_ihm.log("Read:" + name)
                self.master_ihm.log("Statement blocks:" + percentage["Statement blocks"])
                self.master_ihm.log("Decisions:" + percentage["Decisions"])
                self.master_ihm.log("Basic conditions:" + percentage["Basic conditions"])
                self.master_ihm.log("Modified conditions:" + percentage["Modified conditions"])
                of.write("{:s};{:s};{:s};{:s};{:s}\n".format(name,
                                                        percentage["Statement blocks"],
                                                        percentage["Decisions"],
                                                        percentage["Basic conditions"],
                                                        percentage["Modified conditions"],
                                                        ))

            self.master_ihm.resultGenerateCID(filename,
                                              False,
                                              text="COMMAND")
    def _readBPROC(self,
                   bproc_filename,
                   xsl,
                   html_filename,
                   display=True):
        if bproc_filename != "":
            htmlC = HtmlConverter(bproc_filename,xsl)
            print "bproc_filename",bproc_filename
            print "html_filename",html_filename
            html_final = htmlC.toHtml(html_filename + ".html")
            if display:
                os.startfile(html_filename + ".html")
            self.master_ihm.resultGenerateCID(True,
                                            None,
                                            text="G-PROC MAKEFILE READING")
    def _checkUpdate(self):
        self.master_ihm.log("Check for new version of doCID. Please wait ...")
        new_version = self.updateCheck()
        if new_version:
            self.master_ihm.log("A new version of doCID is available: v{:s}".format(new_version))
        else:
            self.master_ihm.log("You are already running the most up-to-date version of doCID v{:s}".format(VERSION))

    def _generateCID(self,
                     release="",
                     baseline="",
                     project="",
                     implemented="",
                     item="",
                     previous_baseline="",
                     detect="",
                     cr_type="",
                     component="",
                     cr_domain = "",
                     cid_type="SCI",
                     header_image=True,
                     list_cr_type=[],
                     list_cr_status=[],
                     list_cr_doamin=[],
                     dico_parameters={},
                     preview=False):
        """
        get items by invoking synergy command
        get sources by invoking synergy command
        get CR by invoking synergy command
        """
        # Create CID
        object_released = self.master_ihm.status_released
        object_integrate = self.master_ihm.status_integrate
        list_projects_set = self.master_ihm.project_set_list

        # BuildDoc instance
        cid = BuildDoc(self.master_ihm,
                       session_started=self.getSessionStarted(),
                       dico_parameters=dico_parameters,
                       preview=preview)
        docx_filename,exception = cid.createCID(list_projects_set,
                                                header_image=header_image,
                                                object_released=object_released,
                                                object_integrate=object_integrate,
                                                cid_type=cid_type,
                                                ccb_type=cr_domain,
                                                item=item,
                                                release=release,
                                                baseline=baseline,
                                                project=project,
                                                target_release=implemented,
                                                previous_baseline=previous_baseline,
                                                detect=detect,
                                                cr_type=cr_type,
                                                component=component,
                                                cr_domain=cr_domain,
                                                list_cr_type=list_cr_type,
                                                list_cr_status=list_cr_status,
                                                list_cr_doamin=list_cr_doamin
                                                )
        self.master_ihm.resultGenerateCID(docx_filename,
                                          exception,
                                          text="CONFIGURATION INDEX DOCUMENT")

    def _generateSQAP(self,
                    author,
                    reference,
                    revision,
                    aircraft,
                    system,
                    item):
        '''
        '''
        sqap = BuildDoc(author,reference,revision,aircraft,system,item)
        self.master_ihm.general_output_txt.insert(END, time.strftime("%H:%M:%S", time.localtime()) + " Creation doc in progress...\n")
        # Create docx
        docx_filename,exception = sqap.createSQAP()
        if docx_filename == False:
            self.master_ihm.general_output_txt.insert(END, time.strftime("%H:%M:%S", time.localtime()) + " " + exception.strerror + ", document not saved.\n")
        else:
            # Create hyperlink
            if docx_filename is not None:
                self.master_ihm.displayHyperlink("hlink",docx_filename,"Software Quality Assurance Plan in Word format.")

    def _generateCCB(self,
                     dico_parameters={"author":"",
                                      "login":"",
                                      "reference":"",
                                      "issue":"",
                                      "release":"",
                                      "baseline":"",
                                      "system":"",
                                      "item":"",
                                      "component":"",
                                      "project":"",
                                      "detect":"",
                                      "implemented":"",
                                      "cr_type":""},
                     cr_with_parent=False,
                     cr_workflow=False,
                     cr_domain=["SCR"],
                     log_on=False,
                     list_cr_for_ccb=[],
                     status_list=False,
                     ccb_time=False,
                     list_cr_type=[],
                     list_cr_status=[],
                     list_cr_doamin=[]):
        """
        To generate a CCB report
        """
        # Get action items
        action = self.master_ihm.action
        db_exist = action.isFilenameDbExist()
        if db_exist:
            list_action_items = action.getActionItem("",1) # Only action items open
        else:
            list_action_items = []
        ccb = CCB(self.master_ihm,
                  system=dico_parameters["system"],
                  item=dico_parameters["item"],
                  detect=dico_parameters["detect"],
                  implemented=dico_parameters["implemented"],
                  cr_domain=cr_domain)
        ccb.setWorkflow(cr_workflow)
        # <NEW>
        #ccb.old_cr_workflow = ccb.get_sys_item_old_workflow(dico_parameters["system"],
        #                                                        dico_parameters["item"])
        #ccb.setDetectRelease(dico_parameters["detect"])
        #ccb.setImplRelease(dico_parameters["implemented"])
        #ccb.setDomain(cr_domain)

        ccb.setListCR(list_cr_for_ccb,
                       status_list)
        # CR list created based on list self.tableau_pr
        tableau_pr_unsorted,found_cr = ccb.getPR_CCB(cr_with_parent=cr_with_parent,
                                                      cr_type=dico_parameters["cr_type"],
                                                      list_cr_type=list_cr_type,
                                                      list_cr_status=list_cr_status,
                                                      list_cr_doamin=list_cr_doamin)
        # </NEW>
        dico_former_cr_status_list = {}
        ccb_time_obj = False
        if log_on:
            if ccb_time:
                ccb_time_obj = datetime.strptime(ccb_time, '%Y/%m/%d %H:%M:%S')
            else:
                ccb_time_obj = False
            dico_cr_log = {}
            dico_cr_transition = {}
            for cr_id in list_cr_for_ccb:
                # Get transition log
                query = "query -t problem \"(problem_number='{:s}')\" -u -f \"%transition_log\"".format(cr_id)
                transi_log = self._ccmCmd(query,False)
                found_status = self.parseLog(cr_id,
                              transi_log,
                              dico_cr_transition,
                              dico_cr_log,
                              ccb_time_obj)
                if found_status:
                    #cr_id_int = int(cr_id)
                    dico_former_cr_status_list[cr_id]=found_status

        #for key,value in dico_former_cr_status_list.iteritems():
        #    print "CR ID: {:s} {:s}".format(key,value)
        docx_filename,exception = ccb.createCCB(self.list_projects,
                                                cr_domain, # deprecated to be removed
                                                list_action_items,
                                                cr_with_parent,
                                                dico_parameters,
                                                list_cr_for_ccb,
                                                status_list,
                                                ccb_time,
                                                dico_former_cr_status_list,
                                                tableau_pr_unsorted,
                                                found_cr,
                                                str(ccb_time_obj))
        self.queue.put("RELOAD_CRLISTBOX") # action to get projects
        self.queue.put(ccb.list_change_requests)
        self.master_ihm.resultGenerateCID(docx_filename,
                                          exception,
                                          text="CHANGE CONTROL BOARD REPORT")

    def _generateDeliverySheet(self,
                              type_sds="SDS",
                              dico_tags={}
                              ):
        """
        Generate Software Delivery Sheet
        Seek .hex or .srec files to extract Part Number, Hw/Sw compatibility index and checksum
        :param type_sds:
        :param dico_tags:
        :return:
        """
        def _setOuptutFilename(template_type,dico_tags):
            """
            :return:
            """
            docx_filename = "{:s}_".format(dico_tags["system"])
            if self.item != "":
                docx_filename += "{:s}_".format(dico_tags["item"])
            if self.component != "":
                docx_filename += "{:s}_".format(dico_tags["component"])
            docx_filename += template_type + "_" + self.reference + "_%d" % floor(time.time()) + ".docx"
            self.ihm.log("Preparing " + docx_filename + " document.")
            return docx_filename

        list_projects_set = self.master_ihm.project_set_list
        if list_projects_set == [] and not Tool.isAttributeValid(dico_tags["project"]):
            # No project set list and no project selected ?
            list_projects = self._getProjectsList_wo_ihm(dico_tags["release"],
                                                         dico_tags["baseline"])
        if list_projects_set != []:
            # Projects are available in GUI
            self.master_ihm.log("Use project set list to create CID for documents", False)
            # Project set in GUI
            list_projects = self.master_ihm.project_set_list
            # List of projects from GUI
            release_text,baseline_text,project_text = self.getContext(list_projects)
        else:
            if Tool.isAttributeValid(dico_tags["project"]):
                find_sub_projects = True
                list_projects = [[dico_tags["release"],
                                  dico_tags["baseline"],
                                  dico_tags["project"]]]
                prj_name, prj_version = self.getProjectInfo(dico_tags["project"])
                self.findSubProjects(prj_name,
                                     prj_version,
                                     list_projects)
                #print "TBL",list_projects
                for sub_release,sub_baseline,sub_project in list_projects:
                    if dico_tags["project"] != sub_project:
                        self.ihm.log("Find sub project {:s}".format(sub_project))
            else:
                list_projects = [[self.release,self.baseline,""]]

        cid = BuildDoc(self.master_ihm,
                       session_started=self.getSessionStarted())
        dico_tags["eoc_id"] = ""
        tbl_bin = []
        list_found_items = []
        # Find Executable Object Code
        for release,baseline,project in list_projects:
            l_tbl_program_file = cid.getSpecificBuild(release,
                                                      baseline,
                                                      project,
                                                      filters=["BIN"],
                                                      list_found_items=list_found_items)
            self.get_eoc_infos(list_found_items,dico_tags)

            tbl_bin.extend(l_tbl_program_file)
        print "tbl_bin",tbl_bin
        pn = self.getComponentPartNumber(dico_tags["component"])
        list_tags = {
                    'CI_ID':{'type':'str','text':pn,'fmt':{}},
                    'REFERENCE':{'type':'str','text':dico_tags["reference"],'fmt':{}},
                    'ISSUE':{'type':'str','text':dico_tags["issue"],'fmt':{}},
                    'ITEM':{'type':'str','text':dico_tags["item"],'fmt':{}},
                    'COMPONENT':{'type':'str','text':dico_tags["component"],'fmt':{}},
                    'DATE':{'type':'str','text':time.strftime("%d %b %Y", time.localtime()),'fmt':{}},
                    'PROJECT':{'type':'str','text':dico_tags["project"],'fmt':{}},
                    'RELEASE':{'type':'str','text':dico_tags["release"],'fmt':{}},
                    'BASELINE':{'type':'str','text':dico_tags["baseline"],'fmt':{}},
                    'WRITER':{'type':'str','text':dico_tags["author"],'fmt':{}},
                    'PART_NUMBER':{'type':'str','text':dico_tags["part_number"],'fmt':{}},
                    'CHECKSUM':{'type':'str','text':dico_tags["checksum"],'fmt':{}},
                    'PROTOCOL_COMPAT':{'type':'str','text':"",'fmt':{}},
                    'DATA_COMPAT':{'type':'str','text':"",'fmt':{}},
                    'HW_COMPAT':{'type':'str','text':dico_tags["hw_sw_compatibility"],'fmt':{}},
                    'EOC_ID':{'type':'str','text':dico_tags["eoc_id"],'fmt':{}}}
        template_dir = join(os.path.dirname("."), 'template')
        template_name = self.getOptions("Template","SDS")
        template = join(template_dir, template_name)
        # Prepare output file
        docx_filename = _setOuptutFilename("SDS",dico_tags)
        self.ihm.docx_filename = docx_filename
        docx_filename,exception = self._createDico2Word(list_tags,
                                                             template,
                                                             docx_filename)
        if not docx_filename:
            self.master_ihm.log(exception + ": document not saved.")
        else:
            try:
                self.master_ihm.cid_word_img_can.itemconfigure(self.master_ihm.cid_word_img,state='normal')
                self.master_ihm.success.config(fg='magenta',bg = 'green',text="SOFTWARE DELIVERY SHEET GENERATION SUCCEEDED")
            except AttributeError:
                pass
            self.master_ihm.displayHyperlink("hlink",docx_filename,"Software Delivery Sheet in Word format.")
        # Set scrollbar at the bottom
        self.master_ihm.defill()

    def _generateReviewReport(self,
                              review_number,
                              empty=False,
                              dico={},
                              project_set_list=[],
                              list_cr_type=[],
                              list_cr_status=[],
                              list_cr_doamin=[]):
        """
        :param review_number:
        :param empty:
        :return:
        """
        # Create docx
        cr_type = dico["cr_type"]
        release = dico["release"]
        baseline = dico["baseline"]
        project = dico["project"]
        if not Tool.isAttributeValid(project):
            project = ""
        if not Tool.isAttributeValid(baseline):
            baseline = ""

        project_list = []
        if not empty:
            if project_set_list == []:
                if Tool.isAttributeValid(project):
                    project_list = [[release,baseline,project]]
                    prj_name, prj_version = self.getProjectInfo(project)
                    self.findSubProjects(prj_name,
                                         prj_version,
                                         project_list,
                                         mute = True)
                    #print "TBL",list_projects
                    for sub_release,sub_baseline,sub_project in project_list:
                        if project != sub_project:
                            self.master_ihm.log("Find sub project {:s}".format(sub_project))
            else:
                project_list = project_set_list
        else:
            pass
        print "DICO before Review init",dico
        review = Review(review_number,
                        detect_release=dico["detect"],
                        impl_release=dico["implemented"],
                        session_started=self.session_started,
                        project_list=project_list,
                        author=dico["author"],
                        system=dico["system"],
                        item=dico["item"],
                        component=dico["component"],
                        part_number=dico["part_number"],
                        checksum=dico["checksum"],
                        reference=dico["reference"],
                        issue=dico["issue"],
                        review_qams_id=dico['review_qams_id'],
                        conformity_level=dico['conformity_level'],
                        cr_type = cr_type,
                        sw_level = dico["dal"],
                        ihm=self.master_ihm)
        review_name = review.getName(review_number)
        self.master_ihm.log(("Creation {:s} review report in progress...").format(review_name))
        docx_filename,exception = review.createReviewReport(empty,
                                                            review_number,
                                                            detect_release=dico["detect"],
                                                            impl_release=dico["implemented"],
                                                            list_cr_type=list_cr_type,
                                                            list_cr_status=list_cr_status,
                                                            list_cr_doamin=list_cr_doamin)
        #old_review.docx_filename = docx_filename
        if not docx_filename:
            self.master_ihm.log(exception + ": document not saved.")
        else:
            self.master_ihm.resultGenerateCID(docx_filename,
                                              exception,
                                              text="REVIEW REPORT")

    def _generateDocument(self,template_key,list_tags):
        """
        generat generic document with tag list input
        """
        self.master_ihm.log("Creation document in progress...")
        # Create docx
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="DOCUMENT GENERATION IN PROGRESS")
        docx_filename,exception = generic_doc.create(list_tags,
                                                    template_key)
        if not docx_filename:
            self.master_ihm.success.config(fg='yellow',bg = 'red',text="DOCUMENT GENERATION FAILED")
            self.master_ihm.log(exception.strerror + ", document not saved.")
        else:
            self.master_ihm.success.config(fg='magenta',bg = 'green',text="DOCUMENT GENERATION SUCCEEDED")
            self.master_ihm.displayHyperlink("hlink",docx_filename,"Document in Word format.")
        # Set scrollbar at the bottom
        self.master_ihm.defill()

    def _closeSession(self):
        '''
        Close session
        '''
        global session_started
        query = "stop"
        self.master_ihm.log("ccm " + query)
        stdout = self._ccmCmd(query)
        # Set scrollbar at the bottom
        #self.master_ihm.general_output_txt.see(END)

    def _getSessionStatus(self):
        '''
        Retrieve database used
        '''
        global session_started
        query = "status"
        self.master_ihm.log("ccm " + query)
        stdout = self._ccmCmd(query)
        # Set scrollbar at the bottom
        try:
            self.master_ihm.general_output_txt.see(END)
        except AttributeError:
            pass
        output = stdout.splitlines()
        for line in output:
            m = re.search(r'Database:(.*)',line)
            if m:
                database = m.group(1)
                self.master_ihm.log("Database used is:" + database)
                return_code = database
                break;
            else:
                return_code = False
        if not return_code:
            print "No Synergy database found"
        return return_code

    def _checkSHLVCP(self,
                     dirname,
                     dirname_upper=""):
        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        # Get SwRD requirements
        upper = CheckLLR(dirname_upper,
                         hlr_selected=True,
                         callback = self.master_ihm.log)

        upper.openLog("SWRD")
        upper.extract(dirname_upper,
                      type=("SWRD",))
        print "upper.tbl_list_llr",upper.tbl_list_llr
        # Get Test Cases
        shlvcp = CheckLLR(dirname,
                       hlr_selected = True,
                       callback = self.master_ihm.log)
        shlvcp.openLog("SHLVCP")
        attr_check_filename,file_check_filename = shlvcp.extract(dirname=dirname,
                                                                 type=("SHLVCP",))
        list_upper_req = []
        shlvcp.getUpperReqList(list_upper_req)
        nb_upper_reqs = len(list_upper_req)
        print "list_upper_req",list_upper_req
        print "nb_upper_reqs",nb_upper_reqs
        shlvcp.logErrors()
        shlvcp.logWarnings()
        shlvcp.closeLog()
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))
        if attr_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink1",attr_check_filename,"List of requirements with attributes.")
        if file_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink3",file_check_filename,"List of files with amount of requirenents per file.")
        if shlvcp.log_filename is not None:
            self.master_ihm.displayHyperlink("hlink2",shlvcp.log_filename,"Log created.")
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="SHLVCP CHECK SUCCEEDED")

    def _checkSHLVCP_VS_SWRD(self,
                     dirname_shlvcp):
        print "TEST extract tables"
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        #component = "SW_BITE" #"BITE" # "ACLOG_SW"
        #self.setF5X(component=component)
        spec = CheckLLR(dirname_shlvcp,
                         hlr_selected=True,
                         nogencache = True,
                         callback=self.master_ihm.log)
        spec.openLog("TEST")
        spec.use_full_win32com = True
        self.master_ihm.log(text="", color="white")
        spec.listDir(tbl_type=("SWRD",))
        print "Extract result"
        print "Found {:d} tables at the beginning".format(spec.nb_tables)
        self.master_ihm.log("Found {:d} tables at the beginning".format(spec.nb_tables))
        nb_reqs = len(spec.tbl_list_llr)
        print "NB Requirements:",nb_reqs
        swrd = Swrd(spec.list_tbl_tables_begin,
                    callback = self.master_ihm.log)
        nb_ext_signals = swrd.populateDicoExtSignal()
        self.master_ihm.log("{:d} external signals found.".format(nb_ext_signals))
        nb_int_signals = swrd.populateDicoIntSignal()
        self.master_ihm.log("{:d} internal signals found.".format(nb_int_signals))
        nb_alias = swrd.populateDicoAlias()
        self.master_ihm.log("{:d} alias found.".format(nb_alias))
        nb_errors = 0
        self.dico_missing_signals = {}
        for req,value in spec.tbl_list_llr.iteritems():
            if "table" in value and value["table"] is not None:
                header = True
                for line in value["table"]:
                    # Monitoring
                    # |{SW_MODE} | [xH_LANDING_LIGHT_FAILURE] | Tempo(ms) | {xH_LANDING_LIGHT_FAILURE_MON} |

                    # Messages
                    # | Signal name	| LSB |	MSB | LSDSIO | RSDSIO |

                    # etc ..
                    if header:
                        # Look for alias or signal
                        for cell in line:
                            data = False
                            alias = swrd.isAlias(cell)
                            if alias:
                                data = alias
                            else:
                                signal = swrd.isSignal(cell)
                                if signal:
                                    data = signal
                            if data:
                                if swrd.signalExists(data) or swrd.aliasExists(data):
                                    pass #print "Signal {:s} found.".format(signal)
                                else:
                                    # Check boaard side Right or Left
                                    lh_signal,rh_signal = swrd.allocateSide(data)
                                    # Check signal exists
                                    if lh_signal:
                                        if swrd.signalExists(lh_signal) or swrd.aliasExists(lh_signal):
                                            pass #print "Signal {:s} found.".format(lh_signal)
                                        else:
                                            swrd.traceError(lh_signal,req)
                                        if swrd.signalExists(rh_signal) or swrd.aliasExists(rh_signal):
                                            pass #print "Signal {:s} found.".format(lh_signal)
                                        else:
                                            swrd.traceError(rh_signal,req)
                                    else:
                                        # Check phases A,B or C
                                        pha_signal,phb_signal,phc_signal = swrd.allocatePhase(data)
                                        if pha_signal:
                                            if swrd.signalExists(pha_signal) or swrd.aliasExists(pha_signal):
                                                pass
                                            else:
                                                swrd.traceError(pha_signal,req)
                                        if phb_signal:
                                            if swrd.signalExists(phb_signal) or swrd.aliasExists(phb_signal):
                                                pass
                                            else:
                                                swrd.traceError(phb_signal,req)
                                        if phc_signal:
                                            if swrd.signalExists(phc_signal) or swrd.aliasExists(phc_signal):
                                                pass
                                            else:
                                                swrd.traceError(phc_signal,req)
                                        else:
                                            if data != "TBD":
                                                swrd.traceError(data,req)
                        header = False
                #print line
            else:
                pass
                #print "Pas de table"
            #break
        for sig,reqs in sorted(swrd.dico_missing_signals.iteritems()):
            list_reqs = ",".join(reqs)
            self.master_ihm.log("Signal/Alias {:s} used in:".format(sig))
            for req in reqs:
                self.master_ihm.log("{:s}".format(req))
            self.master_ihm.log("not found in interface chapter")
        self.master_ihm.log("NB signals/alias not found: {:d}".format(len(swrd.dico_missing_signals)))
        #print "DICO EXT:",swrd.dico_ext_signal
        res = swrd.sqlite_connect()
        if res:
            print "TEST_Z"
            res = swrd.sqlite_create()
            if res:
                #db_hlr.sqlite_insert(req_id,chapter)
                print "TEST_A"
                nb = swrd.sqlite_insert_many(swrd.dico_ext_signal,"EXT")
                print "Insert {:d} lines in SQL database".format(nb)
                nb = swrd.sqlite_insert_many(swrd.dico_int_signal,"INT")
                print "Insert {:d} lines in SQL database".format(nb)
                nb = swrd.sqlite_insert_many(swrd.dico_alias,"ALIAS")
                print "Insert {:d} lines in SQL database".format(nb)
            print "TEST_B"
            swrd.sqlite_close()
        self.master_ihm.resultGenerateCID("",
                                            False,
                                            text="SHLVCP CHECK")
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        print "Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds)

    def _getDesign(self,
                   xml_csci,
                   csci_name="NONE",
                   dico_llr_code={},
                   dico_code_llr={}):
        tree = ET.parse(xml_csci)
        root = tree.getroot()
        for layer in root.findall('LAYER'):
            #csc = layer.find('CSC').text
            layer_func_name = layer.get('func_name')
            layer_reg_name = layer.get('reg_name')
            #print func_name, reg_name
            for csc in layer.findall('CSC'):
                csc_id = csc.get('id')
                csc_func_name = csc.get('func_name')
                csc_reg_name = csc.get('reg_name')
                #print func_name, csc_reg_name
                for csu in csc.findall('CSU'):
                    csu_id = csu.get('id')
                    csu_func_name = csu.get('func_name')
                    csu_reg_name = csu.get('reg_name')
                    llr_name =  "CSC{:s}_CSU{:s}".format(csc_id.zfill(3),csu_id.zfill(3))
                    #llr_name =  "SWDD_G7000_PPDS_{:s}_CSC{:s}_CSU{:s}".format(csci_name,csc_id.zfill(3),csu_id.zfill(3))
                    func_name = "{:s}/{:s}/{:s}.docx".format(layer_func_name,csc_func_name,csu_func_name)
                    source_code = "{:s}_{:s}".format(csc_reg_name,csu_reg_name)
                    source_file = "{:s}/{:s}{:s}/{:s}{:s}_{:s}.c".format(layer_reg_name.upper(),layer_reg_name,csc_reg_name,layer_reg_name,csc_reg_name,csu_reg_name)
                    dico_llr_code[llr_name] = (source_file,func_name,"{:s}{:s}_{:s}.c".format(layer_reg_name,csc_reg_name,csu_reg_name),csc_func_name,csu_func_name)
                    dico_code_llr[source_code] = (llr_name,func_name,csu_func_name,"{:s}{:s}_{:s}.c".format(layer_reg_name,csc_reg_name,csu_reg_name))
                    #print "{:s}/{:s}/{:s}.docx => {:s}/{:s}/{:s}_{:s}.c".format(layer_func_name,csc_func_name,csu_func_name,layer_reg_name,csc_reg_name,csc_reg_name,csu_reg_name)
        #dico_code_llr.sort()
    def exportXlsScod(self,tbl_list_req=[],hlr_vs_llr={},llr_vs_code={}):
        sheet_name = 'Top Bottom'
        wb = Workbook() #load_workbook(filename = join('template','Trace_template.xlsx'))
        if wb is not None:
            ws = wb.create_sheet(title=sheet_name)
            # ws = wb.get_sheet_by_name(name = sheet_name)
            filename = None
            if ws is not None:
                Style.putLogo(ws)

                border=Border(left=Side(border_style=BORDER_THIN),
                              right=Side(border_style=BORDER_THIN),
                              top=Side(border_style=BORDER_THIN),
                              bottom=Side(border_style=BORDER_THIN))
                alignment=Alignment(wrap_text=True,shrink_to_fit=True)
                style_border = Style(border,alignment)

                row = 9
                user_dir=(self.root_user_dir,self.src_user_dir,self.dir_swdd,self.dir_swrd)
                for req_id in tbl_list_req:
                    if req_id in hlr_vs_llr:
                        for low_level_req_id in hlr_vs_llr[req_id]:
                            csu_id = re.sub(r"\w*(CSC[0-9]{3}_CSU[0-9]{3})_[0-9]{3}",r"\1",low_level_req_id)
                            if csu_id in llr_vs_code:
                                func_name = llr_vs_code[csu_id][1]
                                source_code = llr_vs_code[csu_id][0]
                                source_code_name = llr_vs_code[csu_id][2]
                                csc_name = llr_vs_code[csu_id][3]
                                link_code = join(user_dir[0],user_dir[1])
                                link_code = join(link_code,source_code)
                                link_llr = join(user_dir[0],user_dir[2])
                                link_llr = join(link_llr,func_name)
                                llr_link="file:///{:s}".format(link_llr)
                                src_code_link="file:///{:s}".format(link_code)
                                tbl = (req_id,low_level_req_id,source_code_name)
                                #print "TBL",tbl
                                for col_idx in range(1,4):
                                    Style.setCell(ws,tbl,row,col_idx,style_border)
                                Style.setHyperlink(ws,row,2,llr_link)
                                Style.setHyperlink(ws,row,3,src_code_link)
                                row += 1

                filename = "traca_scod_%d.xlsx" % floor(time.time())
                wb.save(join("result",filename))

    def parseVHDL(self,
                  vhdl_file="",
                  dico_process_vs_req={}):
        with open(vhdl_file, 'r') as of:
            list_reqs = []
            for line in of:
                #print "LINE:",line
                m = re.findall(r'Implements \[(PLDDD_\w*)\]', line)
                if m:
                    trace_found = True
                    print "TAG:",m
                    list_reqs.extend(m)
                m = re.findall(r'(\w*) ?: ?PROCESS\w*', line)
                if m:
                    process = m[0]
                    dico_process_vs_req[process] = list(list_reqs)
                    #print list_reqs
                    del(list_reqs[:])
                    print "PROCESS:",m

    def _isSourceFile(self,filename):
        m = re.match("(.*)\.(c)",filename)
        if m:
            result = True
        else:
            result = False
        return result

    def _reccurFoundCalling(self,function,tbl):
        self.depth_func_call += 1
        if function in self.dico_func_called:
            list_calling = self.dico_func_called[function]
            if list_calling is not []:
                result = True
                for sub_function in list_calling:
                    tbl[self.depth_func_call + 1] = sub_function
                    sub_result = self._reccurFoundCalling(sub_function,tbl)
                    if not sub_result:
                        print "TBL",tbl[:]
                        print "self.depth_func_call",self.depth_func_call
                        copy_tbl = tbl[0:self.depth_func_call+2]
                        while len(copy_tbl) < 12:
                            copy_tbl.append("")
                        self.leaves.append(copy_tbl)
                        self.leaves_index += 1
                        # clear last index
                        #tbl[self.depth_func_call] = ""
                        #tbl[self.depth_func_call + 1] = ""
            else:
                result = False
        else:
            result = False
        self.depth_func_call -= 1
        return result

    def _computeLeaves(self):
        self.leaves = []
        self.leaves_index = 0
        tbl = []
        for callee,calling in self.dico_func_called.iteritems():
            if calling is not []:
                for function in calling:
                    del(tbl[:])
                    tbl = [callee,function,"","","","","","","","","",""]
                    sub_result = self._reccurFoundCalling(function,tbl)
                    if not sub_result:
                        self.leaves.append(tbl[:])
                        self.leaves_index += 1
            else:
                tbl = [callee,"","","","","","","","","","",""]
                self.leaves.append(tbl[:])

    def _getStackFromAsm(self):
        code_dir = join(self.root_user_dir,self.build_user_dir)
        if self.master_ihm is not None:
            code = CheckLLR(code_dir,
                            callback = self.master_ihm.log)
        else:
            code = CheckLLR(code_dir)
        code.listDir()
        dico_source_files = {}
        #dico_func_called = {}
        index=0
        function_name=""
        for filename in code.list_code:
            index += 1
            print filename
            with open(filename, 'r') as of:
                function_found = False
                for line in of:
                    if function_found:
                        # stwu  r1, -X(r1) Store the stack pointer and update. create a frame of X bytes
                        m = re.search(r'stwu\t*r1,-([0-9]{1,4})\(r1\)',line)
                        if m:
                            stack_size = m.group(1)
                            print "function_name",function_name
                            dico_source_files[function_name]=stack_size
                            function_found = False
                    m = re.findall(r'^(\w*):',line)
                    if m:
                        function_name = m[0]
                        function_found = True
            #break
        return dico_source_files

    def _computeStackSize(self,
                          line,
                          dico_function_vs_stack_size):
        compute_stack = 0
        for function in line:
            if function in dico_function_vs_stack_size:
                compute_stack += int(dico_function_vs_stack_size[function])
        return compute_stack

    def _stackAnalysis(self):
        code_dir = join(self.root_user_dir,self.src_user_dir)
        include=join(code_dir,"INCLUDE")

        if self.master_ihm is not None:
            code = CheckLLR(code_dir,
                            callback = self.master_ihm.log)
        else:
            code = CheckLLR(code_dir)
        code.listDir()

        index=0

        for filename in code.list_code:
            index += 1
            if self._isSourceFile(filename):
                try:
                    ast = parse_file(filename,
                                     use_cpp=True,
                                     cpp_path=self.compiler,
                                     cpp_args=['-E ', r'-I{:s}'.format(include)])

                    # List of called functions and where
                    list_func_def    = []
                    del(list_func_def[:])
                    v = FuncDefVisitor(list_func_def)
                    v.visit(ast)

                    # List of defined functions and where
                    v = FuncCallVisitor(self.dico_func_called,
                                        list_func_def)
                    v.visit(ast)

                    short_filename = Tool.getFileName(filename)
                    src_code_link="file:///{:s}".format(filename)
                    self.dico_file_vs_function[short_filename]=list_func_def
                    self.dico_file_vs_link[short_filename]=src_code_link

                    if v.nb_func_called < 2:
                        text = "function"
                    else:
                        text = "functions"
                    if self.master_ihm is not None:
                        self.master_ihm.log(
                            "Find {:s} ({:} {:s} called)".format(short_filename, v.nb_func_called, text))
                    else:
                        print "Find {:s} ({:} {:s} called)".format(short_filename,v.nb_func_called,text)
                except ParseError,e:
                    short_filename = Tool.getFileName(filename)
                    if self.master_ihm is not None:
                        self.master_ihm.log("Find {:s} (AST failed: {:s})".format(short_filename, str(e)))
                    else:
                        print "Find {:s} (AST failed: {:s})".format(short_filename,str(e))
                    print e

            else:
                if self.master_ihm is not None:
                    self.master_ihm.log("Discard {:s} (header file)".format(short_filename))
                else:
                    print "Discard {:s} (header file)".format(short_filename)
        # inverse dico file vs function
        for file,functions in self.dico_file_vs_function.iteritems():
            for function,index in functions:
                #print "TESTC",function,file
                self.dico_functions_vs_file[function]=file
        #print "self.dico_functions_vs_file",self.dico_functions_vs_file
        dico_function_vs_stack_size = self._getStackFromAsm()
        wb = Workbook()
        if wb is not None:
            ws = wb.worksheets[0]
            if ws is not None:
                Style.putLogo(ws)
                Style.setCell(ws,["Functions call tree"],8,1)
                row = 9
                tbl = ("Stack","Depth 1","Depth 2","Depth 3","Depth 4","Depth 5","Depth 6","Depth 7","Depth 8","Depth 9","Depth 10","Depth 11")
                for col_idx in range(1,13):
                    Style.setCell(ws,tbl,row,col_idx)
                row += 1
                self._computeLeaves()
                index = 0
                style_border = Style(fill=PatternFill(patternType='solid',start_color='CCCCCCCC'))
                for line in self.leaves:
                    index += 1
                    compute_stack = self._computeStackSize(line,dico_function_vs_stack_size)
                    line.insert(0,compute_stack)
                    for col_idx in range(1,13):
                        if col_idx == 1:
                            Style.setCell(ws,line,row,col_idx,number_format='0.00E+00')
                        else:
                            Style.setCell(ws,line,row,col_idx)
                            function = line[col_idx-1]
                            if function in self.dico_functions_vs_file:
                                filename        = self.dico_functions_vs_file[function]
                                src_code_link   = self.dico_file_vs_link[filename]
                                Style.setHyperlink(ws,row,col_idx,src_code_link)
                            else:
                                Style.setCell(ws,line,row,col_idx,style_border)
                    row += 1
                # Autofilter
                ws.auto_filter.ref = "A9:L9"
                filename = "functions_call_tree_%d.xlsx" % floor(time.time())
                wb.save(join("result",filename))
                #self.master_ihm.resultHyperLink(filename,text="SCOD created.")
                if filename is not None and self.master_ihm is not None:
                    self.master_ihm.resultGenerateCID(filename,
                                                False,
                                                text="FUNCTIONS CALL TREE GENERATION")
    def _buildVHDL(self,
                  csci_name):
        # Get traceability information from vhd files
        dico_filename_process_vs_req = {}
        code_dir = join(self.root_user_dir,self.src_user_dir)
        code = CheckLLR(code_dir,
                        callback = self.master_ihm.log)
        code.listDir()
        for filename in code.list_code:
            short_filename = Tool.getFileName(filename)
            self.master_ihm.log("Find {:s}".format(short_filename))
            dico_process_vs_req = {}
            self.parseVHDL(filename,
                           dico_process_vs_req)
            dico_filename_process_vs_req[short_filename] = (dico_process_vs_req,filename)
        # Get traceability link from PLDDD
        code.basename = self.dir_swdd
        code.hlr_selected = False
        code.tbl_list_llr.clear()
        code.extract(dirname=self.dir_swdd,
                         type=("PLDDD",))
        hlr_vs_llr = {}
        llr_vs_hlr = {}
        for req,value in code.tbl_list_llr.iteritems():
            source_code = CheckLLR.getAtribute(value,"source_code")
            list_source_code = source_code.split(",")
            #print "X:",req,list_source_code
            list_refer,list_constraints = code.getLLR_Trace(value)
            # source code => PLDDD
            for src in list_source_code:
                short_src = Tool.getFileNameAlone(src)
                if short_src not in hlr_vs_llr:
                    hlr_vs_llr[short_src]=[str(req)]
                else:
                    hlr_vs_llr[short_src].append(str(req))
            # PLDDD => source code
            if req not in llr_vs_hlr:
                llr_vs_hlr[req]=list_refer,list_source_code
            else:
                llr_vs_hlr[req].extend(list_refer,list_source_code)
            #print "REQ",req,value
        wb = Workbook()
        #wb = load_workbook(filename = join('template','Trace_template.xlsx'))
        if wb is not None:
            #ws = wb.get_sheet_by_name(name = sheet_name)
            ws = wb.worksheets[0]
            if ws is not None:
                Style.putLogo(ws)
                row = 9
                tbl = ("File","Process","PLDDD Req","PLDRD Req","Comment")
                for col_idx in range(1,6):
                    Style.setCell(ws,tbl,row,col_idx)
                row += 1
                llr_link="file:///{:s}".format(self.dir_swdd)
                hlr_link="file:///{:s}".format(self.dir_swrd)
                nb_error = 0
                for filename,dico_process_vs_req in dico_filename_process_vs_req.iteritems():
                    modified_filename =filename.upper()
                    src_code_link="file:///{:s}".format(dico_process_vs_req[1])
                    max_len_comment = 0
                    for process,reqs in dico_process_vs_req[0].iteritems():
                        print "Filename from src:",filename,reqs
                        if filename in hlr_vs_llr:
                            print "src from swdd:",hlr_vs_llr[filename]
                            nb_error += code.cmpList(hlr_vs_llr[filename], # src from swdd
                                                  reqs,   # Requirements in IS excel file
                                                  cmp_one=filename,
                                                  cmp_two="swdd from src"
                                                )
                        #TODO inverser vhd -> PLDDD => PLDDD -> vhd pour detecter lien en trop
                        for req in reqs:
                            if req in llr_vs_hlr:
                                list_hlr_req,list_src = llr_vs_hlr[req]
                                modified_list_src = map(lambda i: str(re.sub(r"(.*)\.(.*)",r"\1",Tool.removeBlankSpace(i)).upper()), list_src)
                                for req_hlr in list_hlr_req:
                                    if modified_filename in modified_list_src:
                                        comment = "OK"
                                    else:
                                        print "UN",filename
                                        print "UN",list_src
                                        print "TROIS",modified_list_src
                                        #exit()
                                        list_src_text = ", ".join(map(str, modified_list_src))
                                        comment = "NOK, expecting {:s}".format(list_src_text)
                                    if len(str(comment)) > max_len_comment:
                                        max_len_comment = len(str(comment))
                                    tbl = [filename,process,req,req_hlr,comment]
                                    for col_idx in range(1,6):
                                        Style.setCell(ws,tbl,row,col_idx)
                                    Style.setHyperlink(ws,row,1,src_code_link)
                                    Style.setHyperlink(ws,row,2,src_code_link)
                                    Style.setHyperlink(ws,row,3,llr_link)
                                    Style.setHyperlink(ws,row,4,hlr_link)
                                    row += 1
                            else:
                                tbl = ["No link","",req,req_hlr,list_src]
                                for col_idx in range(1,6):
                                        Style.setCell(ws,tbl,row,col_idx)
                                row += 1
        #for x,y in hlr_vs_llr.iteritems():
        #    print "src from swdd:",x,y
        #ws.column_dimensions['D'].width = max_len_comment
        ws.auto_filter.ref = "A9:E9"
        filename = "traca_vhdl_%d.xlsx" % floor(time.time())
        wb.save(join("result",filename))
        #self.master_ihm.resultHyperLink(filename,text="SCOD created.")
        if filename is not None:
            self.master_ihm.resultGenerateCID(filename,
                                        False,
                                        text="VHDL TRACE GENERATION")
    def _buildSCOD(self,
                  csci_name):
        llr_vs_code ={}
        code_vs_llr ={}
        self._getDesign(self.xml_csci,
                        csci_name,
                        llr_vs_code,
                        code_vs_llr)
        #print "llr_vs_code",llr_vs_code
        nb_llr = len(llr_vs_code)
        # get HLR first
        extract_req = CheckLLR(basename=self.dir_swrd,
                           hlr_selected = True,
                           callback = self.master_ihm.log)

        extract_req.openLog("RD")
        export_scod_html = exportSCOD_HTML()
        extract_req.extract(dirname=self.dir_swrd,
                         type=("SWRD",))
        #print "tbl_req_vs_section 1",extract_req.tbl_req_vs_section
        #extract_req.tbl_list_llr = {u'SWRD_GLOBAL-ACENM_0008': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.2', 'refer': u'[CAN-IRD-346]'}, u'SWRD_GLOBAL-ACENM_0551': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0009': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[CAN-IRD-350]'}, u'SWRD_GLOBAL-ACENM_0361': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Initial value is provided at start-up.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0360': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-857]'}, u'SWRD_GLOBAL-ACENM_0523': {'body': '', 'status': u'MATURE', 'additional': u'The preliminary tests results of the second execution override the preliminary tests results of the first execution.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1415]'}, u'SWRD_GLOBAL-ACENM_0522': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Default values are specified for data which are extracted from CAN bus when data are not available or not valid on CAN bus.', 'issue': u'1.7', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0027': {'body': '', 'status': u'MATURE', 'additional': u'One packet contains 6 bytes of data.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[CAN-IRD-636]'}, u'SWRD_GLOBAL-ACENM_0366': {'body': '', 'status': u'MATURE', 'additional': u'Each ACMP is commanded only by one CAN bus, commands of an ACMP cannot be split between two CAN busses. If closed command of an ACMP is invalid, all the commands of this ACMP are switched to the other CAN bus. ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_722],[SSCS_ACLog_725],[SSCS_ACLog_731],[SSCS_ACLog_693],[SSCS_ACLog_692]'}, u'SWRD_GLOBAL-ACENM_0497': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0496': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0495': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0494': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0493': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0492': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0491': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-727]'}, u'SWRD_GLOBAL-ACENM_0490': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-727]'}, u'SWRD_GLOBAL-ACENM_0499': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0498': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0189': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1058]'}, u'SWRD_GLOBAL-ACENM_0445': {'body': '', 'status': u'MATURE', 'additional': u'DSI AC EP overvoltage is defined in HSID.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1506]'}, u'SWRD_GLOBAL-ACENM_0127': {'body': '', 'status': u'MATURE', 'additional': u'The maximum opening time of the contactor (20ms) is included in the tolerance.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_846],[SSCS_ACLog_459],[SSCS_ACLog_1097],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0126': {'body': '', 'status': u'MATURE', 'additional': u'At start-up these data are initialized first with values from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1584],[SSCS_ACLog_1142],[SSCS_ACLog_1119],[SSCS_ACLog_1280]'}, u'SWRD_GLOBAL-ACENM_0125': {'body': '', 'status': u'MATURE', 'additional': u'At start-up these data are initialized first with values from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1584],[SSCS_ACLog_1142],[SSCS_ACLog_1119],[SSCS_ACLog_1280]'}, u'SWRD_GLOBAL-ACENM_0124': {'body': '', 'status': u'MATURE ', 'additional': u'Need to have a global protection status to compute RCCB states.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_459],[SSCS_ACLog_867],[SSCS_ACLog_1114]'}, u'SWRD_GLOBAL-ACENM_0123': {'body': '', 'status': u'MATURE', 'additional': u'Need to have a global protection status to compute RCCB states. ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_459],[SSCS_ACLog_867],[SSCS_ACLog_1114]'}, u'SWRD_GLOBAL-ACENM_0089': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1149],[SSCS_ACLog_922],[SSCS_ACLog_1181],[SSCS_ACLog_883]'}, u'SWRD_GLOBAL-ACENM_0121': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_1326],[SSCS_ACLog_1337],[SSCS_ACLog_1336],[SSCS_ACLog_1345],[CAN-IRD-426],[CAN-IRD-427],[CAN-IRD-534]'}, u'SWRD_GLOBAL-ACENM_0120': {'body': '', 'status': u'MATURE', 'additional': u'The maximum opening time of the contactor (20ms) is included in the tolerance.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_445],[SSCS_ACLog_437],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0084': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_883],[SSCS_ACLog_875],[SSCS_ACLog_1541]'}, u'SWRD_GLOBAL-ACENM_0085': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0086': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0087': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_883],[SSCS_ACLog_874]'}, u'SWRD_GLOBAL-ACENM_0080': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_1519],[SSCS_ACLog_610]'}, u'SWRD_GLOBAL-ACENM_0081': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_610],[SSCS_ACLog_1407]'}, u'SWRD_GLOBAL-ACENM_0129': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'GFI protection has been removed from SSCS', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0083': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_883],[SSCS_ACLog_874]'}, u'SWRD_GLOBAL-ACENM_0066': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_614],[SSCS_ACLog_609]'}, u'SWRD_GLOBAL-ACENM_0067': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_610],[SSCS_ACLog_1407]'}, u'SWRD_GLOBAL-ACENM_0064': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_614],[SSCS_ACLog_609]'}, u'SWRD_GLOBAL-ACENM_0065': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_609],[SSCS_ACLog_1407]'}, u'SWRD_GLOBAL-ACENM_0062': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_596],,[SSCS_ACLog_595]'}, u'SWRD_GLOBAL-ACENM_0063': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1149],[SSCS_ACLog_922],[SSCS_ACLog_1181],[SSCS_ACLog_596]'}, u'SWRD_GLOBAL-ACENM_0060': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0308': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Default value NOT_FAILED is specified during SW initialization phase.', 'issue': u'1.0', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0307': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure computation is defined at HSID level.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0306': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure computation is defined at HSID level.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0305': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure computation is defined at HSID level.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0304': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure computation is defined at HSID level.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0303': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure computation is defined at HSID level.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0302': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure computation is defined at HSID level.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0068': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1519],[SSCS_ACLog_610]'}, u'SWRD_GLOBAL-ACENM_0069': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_610],[SSCS_ACLog_1407]'}, u'SWRD_GLOBAL-ACENM_0398': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Default value NOT_FAILED is specified during SW initialization phase.', 'issue': u'1.7', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0399': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Converter failure computation is defined at HSID level.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0468': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0546': {'body': '', 'status': u'MATURE', 'additional': u'At start-up, the ACMPx tripped states are restored from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_721],[SSCS_ACLog_1119],[SSCS_ACLog_1280],[SSCS_ACLog_1581],[SSCS_ACLog_1583]'}, u'SWRD_GLOBAL-ACENM_0017': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_667],[SSCS_ACLog_897]'}, u'SWRD_GLOBAL-ACENM_0016': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Not defined in SSCS', 'issue': u'1.0', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0015': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'The figure included in this requirement has been moved outside of a requirement.', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0014': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[CAN-IRD-352]'}, u'SWRD_GLOBAL-ACENM_0462': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0463': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0460': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-845]'}, u'SWRD_GLOBAL-ACENM_0461': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0466': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0467': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0464': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0465': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0264': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_920],[SSCS_ACLog_1196],[SSCS_ACLog_1612],[SSCS_ACLog_1613]'}, u'SWRD_GLOBAL-ACENM_0265': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Upper requirement SSCS_ACLog_1313 has been deleted.', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0266': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'The figure included in this requirement has been moved outside of a requirement (refer to Figure 10).', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0267': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_701]'}, u'SWRD_GLOBAL-ACENM_0260': {'body': '', 'status': u'MATURE', 'additional': u'At start-up this data is initialized first with value from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_920],[SSCS_ACLog_1196],[SSCS_ACLog_1610]'}, u'SWRD_GLOBAL-ACENM_0261': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Merged in with the ATCX failed open management requirement', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0262': {'body': '', 'status': u'MATURE', 'additional': u'At start-up this data is initialized first with value from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_920],[SSCS_ACLog_1196],[SSCS_ACLog_1606]'}, u'SWRD_GLOBAL-ACENM_0263': {'body': '', 'status': u'MATURE', 'additional': u'This failure is not latched in context NVM and is reset at each software start-up.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_919],[SSCS_ACLog_1196],[SSCS_ACLog_1612],[SSCS_ACLog_1613]'}, u'SWRD_GLOBAL-ACENM_0268': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A ', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_667]'}, u'SWRD_GLOBAL-ACENM_0269': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.2', 'refer': u'[SSCS_ACLog_667]'}, u'SWRD_GLOBAL-ACENM_0391': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_663]'}, u'SWRD_GLOBAL-ACENM_0392': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1140]'}, u'SWRD_GLOBAL-ACENM_0393': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'FO/FC monitoring of the opposite contactor has been removed in SSCS', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0150': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_683],[SSCS_ACLog_1040],[SSCS_ACLog_1038]'}, u'SWRD_GLOBAL-ACENM_0169': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_1068],[SSCS_ACLog_1121],[SSCS_ACLog_721],[SSCS_ACLog_661],[SSCS_ACLog_464],[SSCS_ACLog_869],[SSCS_ACLog_1122],[SSCS_ACLog_1320],[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0151': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_683],[SSCS_ACLog_1040],[SSCS_ACLog_1038]'}, u'SWRD_GLOBAL-ACENM_0394': {'body': '', 'status': u'MATURE', 'additional': u'Each fault has a unique fault code. The power supplies presence failures are not stored in NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1188],[SSCS_ACLog_1380],[SSCS_ACLog_1400],[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0163': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_557],[SSCS_ACLog_627],[SSCS_ACLog_828],[SSCS_ACLog_1214],[SSCS_ACLog_1315],[SSCS_ACLog_1462],[SSCS_ACLog_1463],[SSCS_ACLog_1483],[SSCS_ACLog_1515],[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0162': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_557],[SSCS_ACLog_1228],[SSCS_ACLog_639],[SSCS_ACLog_827],[SSCS_ACLog_566],[SSCS_ACLog_1222],[SSCS_ACLog_1315],[SSCS_ACLog_1462],[SSCS_ACLog_1463],[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0161': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_557],[SSCS_ACLog_1228],[SSCS_ACLog_638],[SSCS_ACLog_828],[SSCS_ACLog_566],[SSCS_ACLog_1222],[SSCS_ACLog_1315],[SSCS_ACLog_1462],[SSCS_ACLog_1463],[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0160': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_557],[SSCS_ACLog_1228],[SSCS_ACLog_637],[SSCS_ACLog_566],[SSCS_ACLog_1222],[SSCS_ACLog_1315],[SSCS_ACLog_1462],[SSCS_ACLog_1463],[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0167': {'body': '', 'status': u'MATURE', 'additional': u'Closed state corresponds to a GCU "acknowledged" and Open state corresponds to a GCU "not acknowledged"', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_557],[SSCS_ACLog_627],[SSCS_ACLog_1217],[SSCS_ACLog_1462],[SSCS_ACLog_1611]'}, u'SWRD_GLOBAL-ACENM_0166': {'body': '', 'status': u'MATURE', 'additional': u'Closed state corresponds to a GCU "acknowledged" and Open state corresponds to a GCU "not acknowledged"', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_557],[SSCS_ACLog_624],[SSCS_ACLog_1216],[SSCS_ACLog_1462],[SSCS_ACLog_1611]'}, u'SWRD_GLOBAL-ACENM_0165': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1055],[SSCS_ACLog_1228],[SSCS_ACLog_566],[SSCS_ACLog_1462],[SSCS_ACLog_1463]'}, u'SWRD_GLOBAL-ACENM_0164': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'The DSI Emerlog AEC open is no more used.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0185': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_663]'}, u'SWRD_GLOBAL-ACENM_0184': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_663]'}, u'SWRD_GLOBAL-ACENM_0028': {'body': '', 'status': u'MATURE', 'additional': u'Each time slot contains 20 messages including 6 bytes of NVM data.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1077],[CAN-IRD-633],[CAN-IRD-643],[CAN-IRD-641],[CAN-IRD-644],[CAN-IRD-642],[CAN-IRD-868],[CAN-IRD-1034]'}, u'SWRD_GLOBAL-ACENM_0395': {'body': '', 'status': u'MATURE', 'additional': u'The procedure to test the AC TIE current transformer is described in HSID.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1141],[SSCS_ACLog_1419]'}, u'SWRD_GLOBAL-ACENM_0181': {'body': '', 'status': u'MATURE', 'additional': u'For ATC1, ATC2 and AEC contactors, the validity of XFR is checked in the STEP5.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_558],[SSCS_ACLog_563],[SSCS_ACLog_561]'}, u'SWRD_GLOBAL-ACENM_0180': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Failed open/failed closed failures have no impact on network re-configuration', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0183': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_563]'}, u'SWRD_GLOBAL-ACENM_0182': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_561]'}, u'SWRD_GLOBAL-ACENM_0022': {'body': '', 'status': u'MATURE', 'additional': u'The cold start phase includes preliminary tests and PBIT tests.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_669]'}, u'SWRD_GLOBAL-ACENM_0023': {'body': '', 'status': u'MATURE', 'additional': u'The warm start phase includes preliminary tests.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_672]'}, u'SWRD_GLOBAL-ACENM_0020': {'body': '', 'status': u'MATURE', 'additional': u'DSI_5S_POWER_CUT is ACTIVE when a power interrupt greater than 5s has occurred.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_664]'}, u'SWRD_GLOBAL-ACENM_0021': {'body': '', 'status': u'MATURE', 'additional': u'PBIT is not performed if there is a IBIT request.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_668],[SSCS_ACLog_896]'}, u'SWRD_GLOBAL-ACENM_0026': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Provide the initialization value of NVM data to transmit to EDMU.', 'issue': u'1.12', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0158': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_575],[SSCS_ACLog_927]'}, u'SWRD_GLOBAL-ACENM_0024': {'body': '', 'status': u'MATURE', 'additional': u'ACTIVE corresponds to Ground in SSCS, INACTIVE corresponds to Open in SSCS. BOARD_ERROR state is used at software level to catch all wrong pin programming combination.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_886]'}, u'SWRD_GLOBAL-ACENM_0025': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1084],[SSCS_ACLog_1174]'}, u'SWRD_GLOBAL-ACENM_0509': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0508': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0159': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO\t', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_936],[SSCS_ACLog_927]'}, u'SWRD_GLOBAL-ACENM_0118': {'body': '', 'status': u'MATURE', 'additional': u'The maximum opening time of the contactor (20ms) is included in the tolerance.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_443],[SSCS_ACLog_437],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0119': {'body': '', 'status': u'MATURE', 'additional': u'The maximum opening time of the contactor (20ms) is included in the tolerance.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_444],[SSCS_ACLog_437],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0116': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Unbalanced protection has been removed from SSCS.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0117': {'body': '', 'status': u'MATURE', 'additional': u'The frequency range is defined in HSID', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.10', 'refer': u'[SSCS_ACLog_1089]'}, u'SWRD_GLOBAL-ACENM_0114': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_437],[SSCS_ACLog_440]'}, u'SWRD_GLOBAL-ACENM_0115': {'body': '', 'status': u'MATURE', 'additional': u'The maximum opening time of the contactor (20ms) is included in the tolerance.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_441],[SSCS_ACLog_888],[SSCS_ACLog_437],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0112': {'body': '', 'status': u'MATURE', 'additional': u'The maximum opening time of the contactor (20ms) is included in the tolerance.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_437],[SSCS_ACLog_438],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0113': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_437],[SSCS_ACLog_439]'}, u'SWRD_GLOBAL-ACENM_0110': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_437]'}, u'SWRD_GLOBAL-ACENM_0111': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'The AC EP protection are no more latched', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0396': {'body': '', 'status': u'MATURE', 'additional': u'Computed SW checksum is the same as in ROM integrity test', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.8', 'refer': u'[CAN-IRD-211]'}, u'SWRD_GLOBAL-ACENM_0326': {'body': '', 'status': u'MATURE', 'additional': u'The complement value will be used for data integrity check in static area.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_1261]'}, u'SWRD_GLOBAL-ACENM_0019': {'body': '', 'status': u'MATURE', 'additional': u'A critical software error occurs in case of an unexpected interruption, an exception (address error, trap error,...), a CPU overload, ....  ', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Prevent SW from an unexpected behavior in case of a critical SW error', 'issue': u'1.4', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0336': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_1224],[SSCS_ACLog_1373],[SSCS_ACLog_1372]'}, u'SWRD_GLOBAL-ACENM_0337': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1224],[SSCS_ACLog_1373]'}, u'SWRD_GLOBAL-ACENM_0334': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1230]'}, u'SWRD_GLOBAL-ACENM_0335': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1230]'}, u'SWRD_GLOBAL-ACENM_0332': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1601],[SSCS_ACLog_1602],[SSCS_ACLog_1605],[SSCS_ACLog_1609]'}, u'SWRD_GLOBAL-ACENM_0333': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1443],[SSCS_ACLog_1288]'}, u'SWRD_GLOBAL-ACENM_0330': {'body': '', 'status': u'MATURE', 'additional': u'Even if the NVM compatibility is declared as failed, the HW data are restored from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_1119],[SSCS_ACLog_1550],[SSCS_ACLog_1581],[SSCS_ACLog_1583],[SSCS_ACLog_1585],[SSCS_ACLog_1587]'}, u'SWRD_GLOBAL-ACENM_0331': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'IBIT command is saved in static area in order to take into account the request for the next SW start-up.', 'issue': u'1.4', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0237': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Merged in SWRD_GLOBAL-ACENM_0539', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0236': {'body': '', 'status': u'MATURE', 'additional': u'SW is protected against SEU/MBU to avoid unexpected behavior.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1300],[SSCS_ACLog_1301],[SSCS_ACLog_1302]'}, u'SWRD_GLOBAL-ACENM_0235': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1138],[SSCS_ACLog_1345],[CAN-IRD-426],[CAN-IRD-534]'}, u'SWRD_GLOBAL-ACENM_0234': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1137],[SSCS_ACLog_1336],[CAN-IRD-426],[CAN-IRD-427]'}, u'SWRD_GLOBAL-ACENM_0233': {'body': '', 'status': u'MATURE', 'additional': u'Other combinations (data are invalid) are managed in CAN bus management. If a data is invalid on one CAN bus, the data used is taken on the other bus. If a data is invalid on the two CAN busses, a default value is used. At start-up, the ACMPx open locked states are restored from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_721],[SSCS_ACLog_1585],[SSCS_ACLog_1586],[SSCS_ACLog_1587],[SSCS_ACLog_1588]'}, u'SWRD_GLOBAL-ACENM_0232': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1121],[SSCS_ACLog_721],[SSCS_ACLog_1582],[SSCS_ACLog_1586]'}, u'SWRD_GLOBAL-ACENM_0338': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1224],[SSCS_ACLog_1079],[SSCS_ACLog_1379]'}, u'SWRD_GLOBAL-ACENM_0339': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1394]'}, u'SWRD_GLOBAL-ACENM_0435': {'body': '', 'status': u'MATURE', 'additional': u'The IBIT tests results of the second execution override the IBIT results of the first execution.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_955],[SSCS_ACLog_1178]'}, u'SWRD_GLOBAL-ACENM_0434': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Not defined in SSCS', 'issue': u'1.4', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0341': {'body': '', 'status': u'MATURE', 'additional': u'Write NVM current LEG even If no failure is detected on this current LEG.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_1378]'}, u'SWRD_GLOBAL-ACENM_0436': {'body': '', 'status': u'MATURE', 'additional': u'Command DSO for RCCB/contactor will be override if a IBIT is requested. ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_898],[SSCS_ACLog_663],[SSCS_ACLog_1084],[SSCS_ACLog_1174]'}, u'SWRD_GLOBAL-ACENM_0347': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Merged in requirement SWRD_GLOBAL-ACENM_0237', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0430': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'The AC EP protection are no more latched', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0433': {'body': '', 'status': u'MATURE', 'additional': u'NO', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Copy of the anti-paralleling calculated protection status in the associated anti-paralleling global protections status.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0432': {'body': '', 'status': u'MATURE', 'additional': u'NO', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Copy of the differential calculated protection status in the associated differential global protections status.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0246': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1127]'}, u'SWRD_GLOBAL-ACENM_0349': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_692]'}, u'SWRD_GLOBAL-ACENM_0348': {'body': '', 'status': u'MATURE', 'additional': u'CAN Data with validities are only extracted from one and only one CAN bus.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_722],[SSCS_ACLog_725],[SSCS_ACLog_728],[SSCS_ACLog_729],[SSCS_ACLog_730],[SSCS_ACLog_731],[SSCS_ACLog_693],[SSCS_ACLog_694],[CAN-IRD-857]'}, u'SWRD_GLOBAL-ACENM_0439': {'body': '', 'status': u'MATURE', 'additional': u'Roll-over of fault index will be managed at design level (maximum 191 faults can be stored in NVM).', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Manage the restoration of the faults and the flight leg at power-up.', 'issue': u'1.4', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0438': {'body': '', 'status': u'MATURE', 'additional': u'When the maximum number of different faults for one flight leg is reached, the new faults are not registered. When the maximum occurrence of a given fault for one flight leg is reached, this fault is updated when it occurs again (but the number of occurrence remains to 255). A fault is updated only when its state switch from NOT_FAILED to FAILED.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1382],[SSCS_ACLog_1395]'}, u'SWRD_GLOBAL-ACENM_0247': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1128]'}, u'SWRD_GLOBAL-ACENM_0168': {'body': '', 'status': u'MATURE', 'additional': u'Closed state corresponds to a GCU "acknowledged" and Open state corresponds to a GCU "not acknowledged"', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_557],[SSCS_ACLog_624],[SSCS_ACLog_629],[SSCS_ACLog_1220],[SSCS_ACLog_1221],[SSCS_ACLog_1462]'}, u'SWRD_GLOBAL-ACENM_0378': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-636]'}, u'SWRD_GLOBAL-ACENM_0379': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Unlatch failure request is no more used', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0350': {'body': '', 'status': u'MATURE', 'additional': u'A communication failure can be a loss of the bus, a loss of one or several messages in reception, an issue to send message(s) on CAN bus or a protocol error.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1404],[SSCS_ACLog_1405],[SSCS_ACLog_693],[SSCS_ACLog_694],[SSCS_ACLog_1496],[CAN-IRD-857],[CAN-IRD-1017]'}, u'SWRD_GLOBAL-ACENM_0351': {'body': '', 'status': u'MATURE', 'additional': u'CAN Data without validities are extracted from the first CAN bus which provides the data.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1404],[SSCS_ACLog_1405],[SSCS_ACLog_693],[SSCS_ACLog_694],[CAN-IRD-857]'}, u'SWRD_GLOBAL-ACENM_0376': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-215]'}, u'SWRD_GLOBAL-ACENM_0377': {'body': '', 'status': u'MATURE', 'additional': u'If NVM download request, NVM erase or IBIT request are sent at the same time, only the first command received will be taken into account. A download request is ignored if the requested NVM block size is not consistent (greater than the maximum size of NVM). ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A ', 'issue': u'1.12', 'refer': u'[CAN-IRD-643],[CAN-IRD-641],[SSCS_ACLog_1531]'}, u'SWRD_GLOBAL-ACENM_0188': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1056],[SSCS_ACLog_1058],[SSCS_ACLog_1098]'}, u'SWRD_GLOBAL-ACENM_0187': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1056],[SSCS_ACLog_1058],[SSCS_ACLog_1206]'}, u'SWRD_GLOBAL-ACENM_0152': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Step 2 of test 2 has been removed.', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0153': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1040],[SSCS_ACLog_1038]'}, u'SWRD_GLOBAL-ACENM_0099': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_717],[SSCS_ACLog_718]'}, u'SWRD_GLOBAL-ACENM_0098': {'body': '', 'status': u'MATURE', 'additional': u'This failure is not latched in context NVM and is reset at each software start-up.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_717],[SSCS_ACLog_1072],[SSCS_ACLog_939],[SSCS_ACLog_1370],[SSCS_ACLog_1576],[SSCS_ACLog_1612],[SSCS_ACLog_1613],[SSCS_ACLog_1614]'}, u'SWRD_GLOBAL-ACENM_0156': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_575],[SSCS_ACLog_927]'}, u'SWRD_GLOBAL-ACENM_0157': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_575],[SSCS_ACLog_927]'}, u'SWRD_GLOBAL-ACENM_0154': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'PBIT execution time has been removed in SSCS.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0155': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_903],[SSCS_ACLog_840]'}, u'SWRD_GLOBAL-ACENM_0093': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'ACMP data about opposite side are no more used', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0092': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'ACMP data about opposite side are no more used', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0091': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'ACMP data about opposite side are no more used', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0090': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_717],[SSCS_ACLog_713]'}, u'SWRD_GLOBAL-ACENM_0097': {'body': '', 'status': u'MATURE', 'additional': u'This failure is not latched in context NVM and is reset at each software start-up. A failed open failure (due to chattering) clears a failed closed failure.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_717],[SSCS_ACLog_1071],[SSCS_ACLog_939],[SSCS_ACLog_1370],[SSCS_ACLog_1612],[SSCS_ACLog_1613],[SSCS_ACLog_1614]'}, u'SWRD_GLOBAL-ACENM_0096': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_717],[SSCS_ACLog_1070],[SSCS_ACLog_940],[SSCS_ACLog_1371]'}, u'SWRD_GLOBAL-ACENM_0095': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_717],[SSCS_ACLog_714]'}, u'SWRD_GLOBAL-ACENM_0094': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_717],[SSCS_ACLog_714]'}, u'SWRD_GLOBAL-ACENM_0075': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0074': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0077': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_609],[SSCS_ACLog_1407]'}, u'SWRD_GLOBAL-ACENM_0076': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_614],[SSCS_ACLog_609]'}, u'SWRD_GLOBAL-ACENM_0071': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0070': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0073': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0072': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0372': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Merged in requirement SWRD_GLOBAL-ACENM_0373', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0373': {'body': '', 'status': u'MATURE', 'additional': u'If several protections are active, the trip cause is set with the first protection which occurred. Each EDMU_ACMPX_TRIPPED_CMD is associated with one ACMP. An ACMP can receive a TRIP reset independently from other ACMPs.  At start-up, the ACMPx trip causes are restored from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_721],[SSCS_ACLog_1119],[SSCS_ACLog_1280],[SSCS_ACLog_1581],[SSCS_ACLog_1582],[SSCS_ACLog_1583],[SSCS_ACLog_1584]'}, u'SWRD_GLOBAL-ACENM_0370': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Default values are specified for data which are extracted from CAN bus when data are not available or not valid on CAN bus.', 'issue': u'1.4', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0371': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1405]'}, u'SWRD_GLOBAL-ACENM_0079': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_610],[SSCS_ACLog_1407]'}, u'SWRD_GLOBAL-ACENM_0078': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_614],[SSCS_ACLog_609]'}, u'SWRD_GLOBAL-ACENM_0374': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'If the protocol version is different for EDMU and ACLOG, the ACLOG continues to answer to all message from EDMU.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0375': {'body': '', 'status': u'MATURE', 'additional': u'If NVM download request and IBIT request are sent at the same time, only the first command received will be taken into account. While the IBIT has not been fully performed the IBIT request are ignored. IBIT request are ignored if a network reconfiguration is in progress.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_1259],[SSCS_ACLog_1260]'}, u'SWRD_GLOBAL-ACENM_0479': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0478': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0471': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0470': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0473': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0472': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0475': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0474': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0477': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0476': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0387': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_663]'}, u'SWRD_GLOBAL-ACENM_0386': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_932],[SSCS_ACLog_1194],[SSCS_ACLog_944],[SSCS_ACLog_1191],[SSCS_ACLog_1457],[SSCS_ACLog_1546]'}, u'SWRD_GLOBAL-ACENM_0385': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1080]'}, u'SWRD_GLOBAL-ACENM_0384': {'body': '', 'status': u'MATURE', 'additional': u'GREEN and RED are described in HSID. ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.2', 'refer': u'[SSCS_ACLog_701]'}, u'SWRD_GLOBAL-ACENM_0383': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Removed in SSCS', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0382': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Trip reset request is no more used', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0381': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Trip reset request is no more used', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0380': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Unlatch failure request is no more used', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0389': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1247],[SSCS_ACLog_437],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0388': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_1224],[SSCS_ACLog_1374]'}, u'SWRD_GLOBAL-ACENM_0206': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_753]'}, u'SWRD_GLOBAL-ACENM_0207': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_754]'}, u'SWRD_GLOBAL-ACENM_0204': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_536]'}, u'SWRD_GLOBAL-ACENM_0205': {'body': '', 'status': u'MATURE', 'additional': u'The ground service mode cannot be active if the ground servicing request is open.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_552],[SSCS_ACLog_554],[SSCS_ACLog_1504],[SSCS_ACLog_1058]'}, u'SWRD_GLOBAL-ACENM_0202': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_537],[SSCS_ACLog_1471]'}, u'SWRD_GLOBAL-ACENM_0203': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1593]'}, u'SWRD_GLOBAL-ACENM_0200': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_526]'}, u'SWRD_GLOBAL-ACENM_0201': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_537],[SSCS_ACLog_1471]'}, u'SWRD_GLOBAL-ACENM_0208': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_755]'}, u'SWRD_GLOBAL-ACENM_0209': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_735]'}, u'SWRD_GLOBAL-ACENM_0273': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'CPU test is not required during preliminary tests', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0272': {'body': '', 'status': u'MATURE', 'additional': u'Hardware/software compatibility index is defined in HSID', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1535],[SSCS_ACLog_1550]'}, u'SWRD_GLOBAL-ACENM_0271': {'body': '', 'status': u'MATURE', 'additional': u'A 32bits checksum is computed by an additional tool. In ROM, a 32bits constant (initially equal to 0x00000000) is replaced by the complemented value of the computed checksum. It is why ACENM software will get 0x00000000 as result of ROM checksum.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_667]'}, u'SWRD_GLOBAL-ACENM_0270': {'body': '', 'status': u'MATURE', 'additional': u'RAM integrity test algorithm is defined at design level.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_667]'}, u'SWRD_GLOBAL-ACENM_0277': {'body': '', 'status': u'MATURE', 'additional': u'CT tests are not included because these tests are linked to an external failure. ', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Used to have a global result of the PBIT.', 'issue': u'1.7', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0276': {'body': '', 'status': u'MATURE', 'additional': u'The PBIT tests results of the second execution override the PBIT results of the first execution.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1178]'}, u'SWRD_GLOBAL-ACENM_0275': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_899],[SSCS_ACLog_900],[SSCS_ACLog_904],[SSCS_ACLog_1141],[SSCS_ACLog_1176],[SSCS_ACLog_897],[SSCS_ACLog_1419],[SSCS_ACLog_1535]'}, u'SWRD_GLOBAL-ACENM_0274': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_897]'}, u'SWRD_GLOBAL-ACENM_0319': {'body': '', 'status': u'MATURE', 'additional': u'Improve lifetime of NVM device by not writing the same value several times.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1582],[SSCS_ACLog_1584]'}, u'SWRD_GLOBAL-ACENM_0279': {'body': '', 'status': u'MATURE', 'additional': u'The procedure to test the hardware overvoltage protection function is described in HSID.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_904]'}, u'SWRD_GLOBAL-ACENM_0278': {'body': '', 'status': u'MATURE', 'additional': u'MAX and MIN value are defined in HSID', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_900],[SSCS_ACLog_832],[SSCS_ACLog_833],[SSCS_ACLog_1084]'}, u'SWRD_GLOBAL-ACENM_0501': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-870]'}, u'SWRD_GLOBAL-ACENM_0500': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0503': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0502': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-871]'}, u'SWRD_GLOBAL-ACENM_0505': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0504': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-879]'}, u'SWRD_GLOBAL-ACENM_0507': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0506': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0437': {'body': '', 'status': u'MATURE', 'additional': u'Write NVM last LEG with fault.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1377]'}, u'SWRD_GLOBAL-ACENM_0039': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_507],[SSCS_ACLog_1208],[SSCS_ACLog_1481],[SSCS_ACLog_1482]'}, u'SWRD_GLOBAL-ACENM_0038': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_534],[SSCS_ACLog_1474],[SSCS_ACLog_1475]'}, u'SWRD_GLOBAL-ACENM_0031': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Filter on WOW DSI has been removed in SSCS (req SSCS_ACLog_662 removed).', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0030': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Filter on WOW DSI has been removed in SSCS.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0033': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_548],[SSCS_ACLog_953]'}, u'SWRD_GLOBAL-ACENM_0032': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_548],[SSCS_ACLOG_1507]'}, u'SWRD_GLOBAL-ACENM_0035': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_578]'}, u'SWRD_GLOBAL-ACENM_0034': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1469],[SSCS_ACLog_1470]'}, u'SWRD_GLOBAL-ACENM_0037': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Functionalities deleted in SSCS.', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0036': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_578],[SSCS_ACLog_925]'}, u'SWRD_GLOBAL-ACENM_0109': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[CAN-IRD-180]'}, u'SWRD_GLOBAL-ACENM_0108': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-215]'}, u'SWRD_GLOBAL-ACENM_0105': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'No external fuse to monitor.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0104': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1227]'}, u'SWRD_GLOBAL-ACENM_0107': {'body': '', 'status': u'MATURE', 'additional': u'This failure is not latched in context NVM and is reset at each software start-up.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_919],[SSCS_ACLog_1196],[SSCS_ACLog_1612]'}, u'SWRD_GLOBAL-ACENM_0106': {'body': '', 'status': u'MATURE', 'additional': u'Only EDMU trip cause on phase A is used (the ACMP protections are not computed for each phase).', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_721]'}, u'SWRD_GLOBAL-ACENM_0101': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_636],[SSCS_ACLog_632],[SSCS_ACLog_1450],[SSCS_ACLog_1451]'}, u'SWRD_GLOBAL-ACENM_0100': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_717],[SSCS_ACLog_719]'}, u'SWRD_GLOBAL-ACENM_0103': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_636],[SSCS_ACLog_634],[SSCS_ACLog_1450],[SSCS_ACLog_1451]'}, u'SWRD_GLOBAL-ACENM_0102': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_636],[SSCS_ACLog_632],[SSCS_ACLog_1450],[SSCS_ACLog_1451]'}, u'SWRD_GLOBAL-ACENM_0534': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-219]'}, u'SWRD_GLOBAL-ACENM_0535': {'body': '', 'status': u'MATURE', 'additional': u'Write accesses to BITE NVM are not authorized during BITE NVM reset but write accesses to CONTEXT NVM are still authorized. If an erase command occurs during a NVM writing in progress, the writing operation is finished before taking into account the erase command.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1364]'}, u'SWRD_GLOBAL-ACENM_0536': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-499],[CAN-IRD-501],[CAN-IRD-505],[CAN-IRD-506],[CAN-IRD-331]'}, u'SWRD_GLOBAL-ACENM_0537': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_559]'}, u'SWRD_GLOBAL-ACENM_0530': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_629]'}, u'SWRD_GLOBAL-ACENM_0531': {'body': '', 'status': u'MATURE', 'additional': u'A timer is allocated for each phase for each step. If a fault condition comes back, the associated step timer start from its last saved value.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1487]'}, u'SWRD_GLOBAL-ACENM_0532': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1533]'}, u'SWRD_GLOBAL-ACENM_0533': {'body': '', 'status': u'MATURE', 'additional': u'The procedure to test DSI multiplexer is described in HSID.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1107]'}, u'SWRD_GLOBAL-ACENM_0088': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_883],[SSCS_ACLog_875],[SSCS_ACLog_1541]'}, u'SWRD_GLOBAL-ACENM_0538': {'body': '', 'status': u'MATURE', 'additional': u'The AC_EP_PINF_STATE is initialized at ACTIVE state at start-up ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_839],[SSCS_ACLog_526],[SSCS_ACLog_1396],[SSCS_ACLog_1577]'}, u'SWRD_GLOBAL-ACENM_0539': {'body': '', 'status': u'MATURE', 'additional': u'During the 5s, the active AC EP protection BITE failure(s) are stored in NVM and the active AC EP protection BITE failure(s) are sent on CAN busses.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_447],[SSCS_ACLog_1359]'}, u'SWRD_GLOBAL-ACENM_0122': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1327],[SSCS_ACLog_1338],[SSCS_ACLog_1336],[SSCS_ACLog_1345],[CAN-IRD-426],[CAN-IRD-427],[CAN-IRD-534]'}, u'SWRD_GLOBAL-ACENM_0325': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1230]'}, u'SWRD_GLOBAL-ACENM_0324': {'body': '', 'status': u'MATURE', 'additional': u'N/A  ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_676],[SSCS_ACLog_1117],[SSCS_ACLog_1115],[SSCS_ACLog_1262],[SSCS_ACLog_1119],[SSCS_ACLog_1142]'}, u'SWRD_GLOBAL-ACENM_0327': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1230],[SSCS_ACLog_1119]'}, u'SWRD_GLOBAL-ACENM_0249': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Upper requirement has been deleted.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0321': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1601],[SSCS_ACLog_1602],[SSCS_ACLog_1605],[SSCS_ACLog_1609]'}, u'SWRD_GLOBAL-ACENM_0320': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'IBIT command is saved in static area in order to take into account the request for the next SW start-up.', 'issue': u'1.4', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0323': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_830],[SSCS_ACLog_1063],[SSCS_ACLog_1262],[SSCS_ACLog_1142],[SSCS_ACLog_1119]'}, u'SWRD_GLOBAL-ACENM_0322': {'body': '', 'status': u'MATURE', 'additional': u'The complement value will be used for data integrity check in NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1261]'}, u'SWRD_GLOBAL-ACENM_0242': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_504],[SSCS_ACLog_1485]'}, u'SWRD_GLOBAL-ACENM_0243': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_504],[SSCS_ACLog_1485]'}, u'SWRD_GLOBAL-ACENM_0240': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Only one software is used for the two ACLog. One unique part number is used.', 'issue': u'1.0', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0241': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'YES', 'safety': u'YES', 'rationale': u'Defined in PSAC.', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1085],[SSCS_ACLog_491]'}, u'SWRD_GLOBAL-ACENM_0329': {'body': '', 'status': u'MATURE', 'additional': u'All these data are initialized first with values from NVM at start-up. Then, others treatments at start-up can modify the value of these data.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1443]'}, u'SWRD_GLOBAL-ACENM_0328': {'body': '', 'status': u'MATURE', 'additional': u'NVM compatibility algorithm is defined at SwDD level', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'NVM data compatibility is needed to avoid restoration of wrong data', 'issue': u'1.7', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0244': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_935]'}, u'SWRD_GLOBAL-ACENM_0245': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_935]'}, u'SWRD_GLOBAL-ACENM_0400': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[CAN-IRD-211]'}, u'SWRD_GLOBAL-ACENM_0401': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Not defined in SSCS/IRD CAN', 'issue': u'1.3', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0402': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Used in order to ignore a IBIT command when the network reconfiguration is in progress.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0403': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Used in order to ignore a IBIT command when the network reconfiguration is in progress.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0404': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'This alias is no more useful (it is not relevant to merge all the CTC anti-paralleling protection).', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0405': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_845],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0406': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1479]'}, u'SWRD_GLOBAL-ACENM_0407': {'body': '', 'status': u'MATURE', 'additional': u'SW Part number is built according to PSAC', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'A software PN is defined for each software build.', 'issue': u'1.11', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0408': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-784]'}, u'SWRD_GLOBAL-ACENM_0409': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[CAN-IRD-315]'}, u'SWRD_GLOBAL-ACENM_0082': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1149],[SSCS_ACLog_922],[SSCS_ACLog_1181],[SSCS_ACLog_614]'}, u'SWRD_GLOBAL-ACENM_0128': {'body': '', 'status': u'MATURE', 'additional': u'The maximum opening time of the contactor (20ms) is included in the tolerance.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_850],[SSCS_ACLog_459],[SSCS_ACLog_1097],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0248': {'body': '', 'status': u'MATURE', 'additional': u'The preliminary tests are already taken into account in the failed mode management.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1107]'}, u'SWRD_GLOBAL-ACENM_0239': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1087]'}, u'SWRD_GLOBAL-ACENM_0238': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1087]'}, u'SWRD_GLOBAL-ACENM_0231': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1334],[SSCS_ACLog_1343],[CAN-IRD-180]'}, u'SWRD_GLOBAL-ACENM_0230': {'body': '', 'status': u'MATURE', 'additional': u'CT AC EP failure is computed only on ACLog2 board.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_574],[SSCS_ACLog_845],[SSCS_ACLog_576],[SSCS_ACLog_937],[SSCS_ACLog_1041],[SSCS_ACLog_957],[SSCS_ACLog_901],[SSCS_ACLog_1187],[SSCS_ACLog_913],[SSCS_ACLog_921],[SSCS_ACLog_926],[SSCS_ACLog_929],[SSCS_ACLog_928],[SSCS_ACLog_941],[SSCS_ACLog_1070],[SSCS_ACLog_1071],[SSCS_ACLog_1072],[SSCS_ACLog_1397],[SSCS_ACLog_1453],[SSCS_ACLog_1545],[SSCS_ACLog_1598],[SSCS_ACLog_1599],[SSCS_ACLog_1576],[CAN-IRD-180],[CAN-IRD-216],[CAN-IRD-870],[CAN-IRD-871],[CAN-IRD-1001]'}, u'SWRD_GLOBAL-ACENM_0343': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1368],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0342': {'body': '', 'status': u'MATURE', 'additional': u'Roll-over of fault index will be managed at design level (maximum 191 faults can be stored in NVM).', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Define LEG transition according to EDMU flight leg status.', 'issue': u'1.4', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0545': {'body': '', 'status': u'MATURE', 'additional': u'At start-up, the ACMPx open locked states are restored from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_721],[SSCS_ACLog_1585],[SSCS_ACLog_1587]'}, u'SWRD_GLOBAL-ACENM_0340': {'body': '', 'status': u'MATURE', 'additional': u'When the maximum number of fault stored in NVM is reached, there is a roll-over of the fault buffer. Roll-over of fault index will be managed at design level (maximum 191 faults can be stored in NVM).', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_1376],[SSCS_ACLog_1381]'}, u'SWRD_GLOBAL-ACENM_0431': {'body': '', 'status': u'MATURE', 'additional': u'NO', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Copy of the ACMP calculated protection statuses in the associated ACMP global protections statuses.', 'issue': u'1.7', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0346': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1369],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0345': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1367],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0344': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1366],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0368': {'body': '', 'status': u'MATURE', 'additional': u'{CAN_X_DATAX_VALIDITY} of ACMPX_CMD are managed in another requirement.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_693]'}, u'SWRD_GLOBAL-ACENM_0309': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Synthesis of all the contactor DSO failures used to compute global contactor DSO failure.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0547': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-636]'}, u'SWRD_GLOBAL-ACENM_0061': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_598],[SSCS_ACLog_596],[SSCS_ACLog_1517]'}, u'SWRD_GLOBAL-ACENM_0141': {'body': '', 'status': u'MATURE', 'additional': u'At start-up these data are initialized first with values from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1142],[SSCS_ACLog_1262],[SSCS_ACLog_1119],[SSCS_ACLog_1503],[SSCS_ACLog_1443]'}, u'SWRD_GLOBAL-ACENM_0140': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_826]'}, u'SWRD_GLOBAL-ACENM_0143': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1329],[SSCS_ACLog_1331],[SSCS_ACLog_1333],[SSCS_ACLog_1340],[SSCS_ACLog_1342],[SSCS_ACLog_1336],[SSCS_ACLog_1345],[CAN-IRD-426],[CAN-IRD-427],[CAN-IRD-534]'}, u'SWRD_GLOBAL-ACENM_0142': {'body': '', 'status': u'MATURE', 'additional': u'The DSI linked to anti-paralleling are defined in HSID.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1063],[SSCS_ACLog_832],[SSCS_ACLog_833],[SSCS_ACLog_837]'}, u'SWRD_GLOBAL-ACENM_0145': {'body': '', 'status': u'MATURE', 'additional': u"The computation is done on the same frequencies sent on CAN busses [EXT_AC_FREQUENCY]. That's why there is no tolerance. ", 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1350],[SSCS_ACLog_1353],[CAN-IRD-792],[CAN-IRD-879]'}, u'SWRD_GLOBAL-ACENM_0301': {'body': '', 'status': u'MATURE', 'additional': u'Fuse failure computation is defined at HSID level.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1181],[SSCS_ACLog_1455]'}, u'SWRD_GLOBAL-ACENM_0147': {'body': '', 'status': u'MATURE', 'additional': u'10ms timing is defined in HSID', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_683],[SSCS_ACLog_1040],[SSCS_ACLog_1038]'}, u'SWRD_GLOBAL-ACENM_0146': {'body': '', 'status': u'MATURE', 'additional': u"The computation is done on the same power sent on CAN busses [EXT_AC_LOAD]. That's why there is no tolerance.", 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1351],[SSCS_ACLog_1353],[CAN-IRD-792],[CAN-IRD-879]'}, u'SWRD_GLOBAL-ACENM_0149': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_683]'}, u'SWRD_GLOBAL-ACENM_0148': {'body': '', 'status': u'MATURE', 'additional': u'TCB status is sent every 250ms to EDMU through CAN bus.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_682]'}, u'SWRD_GLOBAL-ACENM_0300': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure computation is defined at HSID level.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0013': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[CAN-IRD-365],[CAN-IRD-525],[CAN-IRD-466],[CAN-IRD-366],[CAN-IRD-526],[CAN-IRD-367],[CAN-IRD-671],[CAN-IRD-501],[CAN-IRD-506]'}, u'SWRD_GLOBAL-ACENM_0541': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0012': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[CAN-IRD-365],[CAN-IRD-525],[CAN-IRD-466],[CAN-IRD-366],[CAN-IRD-526],[CAN-IRD-367],[CAN-IRD-671],[CAN-IRD-499],[CAN-IRD-505]'}, u'SWRD_GLOBAL-ACENM_0011': {'body': '', 'status': u'MATURE', 'additional': u'ACLog 1 is identified either by XLOG1 or ACLOG1. ACLog 2 is identified either by XLOG2 or ACLOG2.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[CAN-IRD-347]'}, u'SWRD_GLOBAL-ACENM_0369': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'This requirement is redundant with requirement SWRD_GLOBAL-ACENM_0383', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0041': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Source availability is no more impacted by failed open/closed states.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0042': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_571]'}, u'SWRD_GLOBAL-ACENM_0469': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0044': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_571]'}, u'SWRD_GLOBAL-ACENM_0045': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Source availability is no more impacted by failed open/closed states.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0046': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_571],[SSCS_ACLog_924]'}, u'SWRD_GLOBAL-ACENM_0047': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_571],[SSCS_ACLog_924]'}, u'SWRD_GLOBAL-ACENM_0048': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_571],[SSCS_ACLog_924]'}, u'SWRD_GLOBAL-ACENM_0049': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_557],[SSCS_ACLog_1225],[CAN-IRD-182]'}, u'SWRD_GLOBAL-ACENM_0363': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Initial value is provided at start-up.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0362': {'body': '', 'status': u'MATURE', 'additional': u'All messages are sent by EDMU every 1s.No valid CAN message means bad CRC on CAN message or bad CAN identifier or no message received or CAN HW error.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_693]'}, u'SWRD_GLOBAL-ACENM_0365': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_692]'}, u'SWRD_GLOBAL-ACENM_0364': {'body': '', 'status': u'MATURE', 'additional': u'A message is considered as failed if this message is not received on the CAN bus during 3 times its period.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_693]'}, u'SWRD_GLOBAL-ACENM_0367': {'body': '', 'status': u'MATURE', 'additional': u'Each ACMP is commanded only by one CAN bus, commands of an ACMP cannot be split between two CAN busses. If closed command of an ACMP is invalid, all the commands of this ACMP are switches to the other CAN bus.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_728],[SSCS_ACLog_730],[SSCS_ACLog_731],[SSCS_ACLog_693]'}, u'SWRD_GLOBAL-ACENM_0540': {'body': '', 'status': u'MATURE', 'additional': u'At start-up these data are initialized first with value from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_1576],[SSCS_ACLog_1119],[SSCS_ACLog_1142],[SSCS_ACLog_1262],[SSCS_ACLog_1443]'}, u'SWRD_GLOBAL-ACENM_0543': {'body': '', 'status': u'MATURE', 'additional': u'There is one engineering data (containing 128 bytes) for each active failure', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1391]'}, u'SWRD_GLOBAL-ACENM_0197': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1058],[SSCS_ACLog_1098]'}, u'SWRD_GLOBAL-ACENM_0542': {'body': '', 'status': u'MATURE', 'additional': u'CAN Data without validities are extracted from the first CAN bus which provides the data.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1404],[SSCS_ACLog_1405],[SSCS_ACLog_693],[SSCS_ACLog_694],[CAN-IRD-857]'}, u'SWRD_GLOBAL-ACENM_0444': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1516]'}, u'SWRD_GLOBAL-ACENM_0390': {'body': '', 'status': u'MATURE', 'additional': u'The frequency range is defined in HSID. ', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Validity bit is computed for CAN bus.', 'issue': u'1.7', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0446': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0447': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0440': {'body': '', 'status': u'MATURE', 'additional': u'Write NVM first fault index.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_1375]'}, u'SWRD_GLOBAL-ACENM_0441': {'body': '', 'status': u'MATURE', 'additional': u'The command of the AEC contactor is not sequenced. The command is directly applied independently from the other contactors.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1055]'}, u'SWRD_GLOBAL-ACENM_0442': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'FO/FC monitoring of the opposite contactor has been removed in SSCS', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0443': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'FO/FC monitoring of the opposite contactor has been removed in SSCS', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0448': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0449': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0193': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1056]'}, u'SWRD_GLOBAL-ACENM_0018': {'body': '', 'status': u'MATURE', 'additional': u'AEC is a normally closed contactor.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_892],[SSCS_ACLog_1308],[SSCS_ACLog_1412],[SSCS_ACLog_1363]'}, u'SWRD_GLOBAL-ACENM_0190': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1058]'}, u'SWRD_GLOBAL-ACENM_0549': {'body': '', 'status': u'MATURE', 'additional': u'At start-up this data is initialized first with value from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_920],[SSCS_ACLog_1196],[SSCS_ACLog_1607]'}, u'SWRD_GLOBAL-ACENM_0191': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Failed open/failed closed failures have no impact on network re-configuration', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0548': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1149],[SSCS_ACLog_922],[SSCS_ACLog_1181],[SSCS_ACLog_583],[SSCS_ACLog_1518]'}, u'SWRD_GLOBAL-ACENM_0215': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_744]'}, u'SWRD_GLOBAL-ACENM_0214': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_743]'}, u'SWRD_GLOBAL-ACENM_0217': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1335],[SSCS_ACLog_1344]'}, u'SWRD_GLOBAL-ACENM_0216': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_745]'}, u'SWRD_GLOBAL-ACENM_0211': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_737]'}, u'SWRD_GLOBAL-ACENM_0210': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_736]'}, u'SWRD_GLOBAL-ACENM_0213': {'body': '', 'status': u'MATURE', 'additional': u'For this contactor the hardware logic is inverted.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_739]'}, u'SWRD_GLOBAL-ACENM_0212': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_738]'}, u'SWRD_GLOBAL-ACENM_0397': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-215]'}, u'SWRD_GLOBAL-ACENM_0219': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_701]'}, u'SWRD_GLOBAL-ACENM_0218': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_701]'}, u'SWRD_GLOBAL-ACENM_0054': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_583]'}, u'SWRD_GLOBAL-ACENM_0288': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Global synthesis of all the CBIT failures is not used. ', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0289': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_909],[SSCS_ACLog_1083],[SSCS_ACLog_912],[SSCS_ACLog_919],[SSCS_ACLog_920],[SSCS_ACLog_1149],[SSCS_ACLog_924],[SSCS_ACLog_927],[SSCS_ACLog_932],[SSCS_ACLog_1194],[SSCS_ACLog_939],[SSCS_ACLog_940],[SSCS_ACLog_944],[SSCS_ACLog_1038],[SSCS_ACLog_840],[SSCS_ACLog_953],[SSCS_ACLog_902],[SSCS_ACLog_905]'}, u'SWRD_GLOBAL-ACENM_0286': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_955],[SSCS_ACLog_1141]'}, u'SWRD_GLOBAL-ACENM_0287': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'This requirement is not allocated to software.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0284': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1230]'}, u'SWRD_GLOBAL-ACENM_0285': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_955],[SSCS_ACLog_1141],[SSCS_ACLog_1615]'}, u'SWRD_GLOBAL-ACENM_0282': {'body': '', 'status': u'MATURE', 'additional': u'NO', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_899],[SSCS_ACLog_1084]'}, u'SWRD_GLOBAL-ACENM_0283': {'body': '', 'status': u'MATURE', 'additional': u'Write accesses to NVM are not authorized during NVM download to avoid inconsistency in NVM (checksum issue).', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1364]'}, u'SWRD_GLOBAL-ACENM_0280': {'body': '', 'status': u'MATURE', 'additional': u'The procedure to test the AC EP current transformer is described in HSID. This test is only performed on ACLog2 board.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1141],[SSCS_ACLog_1419]'}, u'SWRD_GLOBAL-ACENM_0281': {'body': '', 'status': u'MATURE', 'additional': u'The procedure to test the 5s power cut function is described in HSID.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_902]'}, u'SWRD_GLOBAL-ACENM_0051': {'body': '', 'status': u'MATURE', 'additional': u'N/A ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_583]'}, u'SWRD_GLOBAL-ACENM_0199': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_839],[SSCS_ACLog_526]'}, u'SWRD_GLOBAL-ACENM_0427': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_722],[SSCS_ACLog_725],[SSCS_ACLog_728],[SSCS_ACLog_729],[SSCS_ACLog_730],[SSCS_ACLog_731],[SSCS_ACLog_693],[SSCS_ACLog_694],[SSCS_ACLog_1496],[CAN-IRD-857],[CAN-IRD-1017]'}, u'SWRD_GLOBAL-ACENM_0138': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1328],[SSCS_ACLog_1330],[SSCS_ACLog_1332],[SSCS_ACLog_1339],[SSCS_ACLog_1341],[SSCS_ACLog_1336],[SSCS_ACLog_1345],[CAN-IRD-426],[CAN-IRD-427],[CAN-IRD-534]'}, u'SWRD_GLOBAL-ACENM_0139': {'body': '', 'status': u'MATURE', 'additional': u'At start-up these data are initialized first with values from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_831],[SSCS_ACLog_1142],[SSCS_ACLog_1262],[SSCS_ACLog_1119],[SSCS_ACLog_1443]'}, u'SWRD_GLOBAL-ACENM_0130': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'GFI protection has been removed from SSCS', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0131': {'body': '', 'status': u'MATURE', 'additional': u'The maximum opening time of the contactor (20ms) is included in the tolerance.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_865],[SSCS_ACLog_866],[SSCS_ACLog_867],[SSCS_ACLog_1499],[SSCS_ACLog_1457],[SSCS_ACLog_1546]'}, u'SWRD_GLOBAL-ACENM_0132': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_865],[SSCS_ACLog_866],[SSCS_ACLog_1594]'}, u'SWRD_GLOBAL-ACENM_0133': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_865],[SSCS_ACLog_866],[SSCS_ACLog_1594]'}, u'SWRD_GLOBAL-ACENM_0134': {'body': '', 'status': u'MATURE', 'additional': u'Converter failure computation is defined at HSID level.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1181],[SSCS_ACLog_1455]'}, u'SWRD_GLOBAL-ACENM_0135': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_912],[SSCS_ACLog_1181],[SSCS_ACLog_1195],[SSCS_ACLog_1084]'}, u'SWRD_GLOBAL-ACENM_0136': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_846],[SSCS_ACLog_1114],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0137': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_850],[SSCS_ACLog_1114],[SSCS_ACLog_1499]'}, u'SWRD_GLOBAL-ACENM_0488': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-727]'}, u'SWRD_GLOBAL-ACENM_0489': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0521': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1282]'}, u'SWRD_GLOBAL-ACENM_0520': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_501],[SSCS_ACLog_527]'}, u'SWRD_GLOBAL-ACENM_0527': {'body': '', 'status': u'MATURE', 'additional': u'The power cut test result of the second execution override the result of the first execution.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_902]'}, u'SWRD_GLOBAL-ACENM_0526': {'body': '', 'status': u'MATURE', 'additional': u'If NVM download request or NVM reset request or IBIT request are sent at the same time, only the first command received will be taken into account', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_1528]'}, u'SWRD_GLOBAL-ACENM_0525': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1257]'}, u'SWRD_GLOBAL-ACENM_0524': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1507],[SSCS_ACLog_548]'}, u'SWRD_GLOBAL-ACENM_0480': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0481': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0482': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-772]'}, u'SWRD_GLOBAL-ACENM_0483': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0484': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0485': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0486': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-727]'}, u'SWRD_GLOBAL-ACENM_0487': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-727]'}, u'SWRD_GLOBAL-ACENM_0310': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Default value NOT_FAILED is specified during SW initialization phase.', 'issue': u'1.0', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0311': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Synthesis of all the other DSO failures used to compute global other DSO failure.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0312': {'body': '', 'status': u'MATURE', 'additional': u'HW keeps transparency during 5ms and SW read DSI every 1ms. SW uses transparency information to avoid NVM corruption at start up.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1261]'}, u'SWRD_GLOBAL-ACENM_0313': {'body': '', 'status': u'MATURE', 'additional': u'A mapping of NVM is defined in order to specify block of memory allocated by functionalities.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_1491]'}, u'SWRD_GLOBAL-ACENM_0314': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'NVM storage is managed at design level.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0315': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Upper requirement has been removed. HW and SW PN are stored in NVM during ATP.', 'issue': u'1.5', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0316': {'body': '', 'status': u'MATURE', 'additional': u'Improve lifetime of NVM device by not writing the same value several times.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_1119]'}, u'SWRD_GLOBAL-ACENM_0317': {'body': '', 'status': u'MATURE', 'additional': u'Improve lifetime of NVM device by not writing the same value several times.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1582],[SSCS_ACLog_1584]'}, u'SWRD_GLOBAL-ACENM_0251': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'YES', 'safety': u'NO', 'rationale': u'Hardware component used is compliant with the ARINC-825 standard.', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1044],[CAN-IRD-312]'}, u'SWRD_GLOBAL-ACENM_0250': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[CAN-IRD-324],[SSCS_ACLog_706]'}, u'SWRD_GLOBAL-ACENM_0253': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1530]'}, u'SWRD_GLOBAL-ACENM_0252': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1363]'}, u'SWRD_GLOBAL-ACENM_0255': {'body': '', 'status': u'MATURE', 'additional': u'At start-up this data is initialized first with value from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1216],[SSCS_ACLog_1217],[SSCS_ACLog_919],[SSCS_ACLog_1196],[SSCS_ACLog_1603],[SSCS_ACLog_1604]'}, u'SWRD_GLOBAL-ACENM_0254': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1363],[SSCS_ACLog_892]'}, u'SWRD_GLOBAL-ACENM_0257': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Management of all the failed open failures for GLC1, GLC2 and ALC are merged in one requirement', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0256': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Failed closed failure is no more used', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0004': {'body': '', 'status': u'MATURE', 'additional': u'To be compliant with the protection timing constraint', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_695]'}, u'SWRD_GLOBAL-ACENM_0005': {'body': '', 'status': u'MATURE', 'additional': u'Maximum processing time are dedicated to HW. The "Tbit" data is computed from CAN refresh rate (1/500).', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_657],[CAN-IRD-314]'}, u'SWRD_GLOBAL-ACENM_0006': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[CAN-IRD-345],[CAN-IRD-313]'}, u'SWRD_GLOBAL-ACENM_0007': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[CAN-IRD-344],[CAN-IRD-313],[CAN-IRD-331],[CAN-IRD-332]'}, u'SWRD_GLOBAL-ACENM_0552': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0001': {'body': '', 'status': u'MATURE', 'additional': u'Acquisition frequency should be twice higher than maximum ASI frequency to measure (650Hz x 2)  ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A  ', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_695]'}, u'SWRD_GLOBAL-ACENM_0002': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'To be compliant with the protection measurement accuracy constraint', 'issue': u'1.6', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0003': {'body': '', 'status': u'MATURE', 'additional': u'To be compliant with the protection timing constraint', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_695]'}, u'SWRD_GLOBAL-ACENM_0417': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-338],[CAN-IRD-869]'}, u'SWRD_GLOBAL-ACENM_0416': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[CAN-IRD-336]'}, u'SWRD_GLOBAL-ACENM_0415': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[CAN-IRD-335],[CAN-IRD-364]'}, u'SWRD_GLOBAL-ACENM_0414': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[CAN-IRD-334]'}, u'SWRD_GLOBAL-ACENM_0413': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[CAN-IRD-333],[CAN-IRD-107],[CAN-IRD-108],[CAN-IRD-109],[CAN-IRD-110]'}, u'SWRD_GLOBAL-ACENM_0412': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'BNR format constraint is now traced on each HLR using BNR data ', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0411': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[CAN-IRD-317]'}, u'SWRD_GLOBAL-ACENM_0410': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[CAN-IRD-316]'}, u'SWRD_GLOBAL-ACENM_0544': {'body': '', 'status': u'MATURE', 'additional': u'There is one engineering data (containing 128 bytes) for each active failure', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1391]'}, u'SWRD_GLOBAL-ACENM_0529': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1455]'}, u'SWRD_GLOBAL-ACENM_0528': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1041],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0029': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Managed in the NVM acceptance.', 'issue': u'1.7', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0010': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[CAN-IRD-351]'}, u'SWRD_GLOBAL-ACENM_0144': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Redundant with requirement SWRD_GLOBAL-ACENM_0143', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0426': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'CPU margin is defined in order to keep resource for future evolutions.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0355': {'body': '', 'status': u'MATURE', 'additional': u'This alias is used to check if there is no communication on the two CAN busses.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_701]'}, u'SWRD_GLOBAL-ACENM_0356': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_663]'}, u'SWRD_GLOBAL-ACENM_0425': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_1087]'}, u'SWRD_GLOBAL-ACENM_0422': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Each contactor failure management is specific and has been split.', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0423': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'ACMPX failed open and failed closed failures are no more latched in context NVM', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0259': {'body': '', 'status': u'MATURE', 'additional': u'This failure is not latched in context NVM and is reset at each software start-up.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_919],[SSCS_ACLog_1196],[SSCS_ACLog_1612]'}, u'SWRD_GLOBAL-ACENM_0352': {'body': '', 'status': u'MATURE', 'additional': u'Checksum computation will be defined at design level.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1079],[SSCS_ACLog_1384],[SSCS_ACLog_1385],[SSCS_ACLog_1388],[SSCS_ACLog_1389],[SSCS_ACLog_1390],[SSCS_ACLog_1391]'}, u'SWRD_GLOBAL-ACENM_0258': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Failed closed failure is no more used', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0353': {'body': '', 'status': u'MATURE', 'additional': u'Checksum computation will be defined at design level. If the number of occurrence maximum is reached, the fault occurrence remains to 255.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1079],[SSCS_ACLog_1383],[SSCS_ACLog_1386],[SSCS_ACLog_1387],[SSCS_ACLog_1389],[SSCS_ACLog_1390],[SSCS_ACLog_1391]'}, u'SWRD_GLOBAL-ACENM_0419': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Merged in SWRD_GLOBAL-ACENM_0013', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0178': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_559]'}, u'SWRD_GLOBAL-ACENM_0179': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_561],[SSCS_ACLog_585],[SSCS_ACLog_589],[SSCS_ACLog_591],[SSCS_ACLog_1219]'}, u'SWRD_GLOBAL-ACENM_0174': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_559],[SSCS_ACLog_585],[SSCS_ACLog_565],[SSCS_ACLog_1462]'}, u'SWRD_GLOBAL-ACENM_0175': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_559],[SSCS_ACLog_585],[SSCS_ACLog_565],[SSCS_ACLog_1462]'}, u'SWRD_GLOBAL-ACENM_0176': {'body': '', 'status': u'MATURE', 'additional': u"The computation is done on the same voltage sent on CAN busses [EXT_AC_PHX_VOLTAGE]. That's why there is no tolerance.", 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1352],[SSCS_ACLog_1353],[CAN-IRD-792],[CAN-IRD-879]'}, u'SWRD_GLOBAL-ACENM_0177': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_559],[SSCS_ACLog_585],[SSCS_ACLog_561]'}, u'SWRD_GLOBAL-ACENM_0170': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_1068],[SSCS_ACLog_1121],[SSCS_ACLog_721],[SSCS_ACLog_661],[SSCS_ACLog_466],[SSCS_ACLog_869],[SSCS_ACLog_1122],[SSCS_ACLog_1320],[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0171': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.12', 'refer': u'[SSCS_ACLog_1068],[SSCS_ACLog_1121],[SSCS_ACLog_721],[SSCS_ACLog_661],[SSCS_ACLog_463],[SSCS_ACLog_868],[SSCS_ACLog_1122],[SSCS_ACLog_1320],[SSCS_ACLog_1576]'}, u'SWRD_GLOBAL-ACENM_0172': {'body': '', 'status': u'MATURE', 'additional': u'During IBIT contactor commands to apply are restored from static memory', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_663]'}, u'SWRD_GLOBAL-ACENM_0173': {'body': '', 'status': u'MATURE', 'additional': u'During IBIT contactor commands to apply are restored from static memory', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_663]'}, u'SWRD_GLOBAL-ACENM_0358': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Used in order to ignore a IBIT command when the network reconfiguration is in progress.', 'issue': u'1.0', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0418': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[CAN-IRD-365],[CAN-IRD-525],[CAN-IRD-466],[CAN-IRD-366],[CAN-IRD-526],[CAN-IRD-367],[CAN-IRD-671],[CAN-IRD-331],[CAN-IRD-858],[CAN-IRD-712]'}, u'SWRD_GLOBAL-ACENM_0186': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.2', 'refer': u'[SSCS_ACLog_1056],[SSCS_ACLog_1058],[SSCS_ACLog_1206]'}, u'SWRD_GLOBAL-ACENM_0359': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Initial value is provided at start-up.', 'issue': u'1.1', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0196': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1058],[SSCS_ACLog_1098]'}, u'SWRD_GLOBAL-ACENM_0318': {'body': '', 'status': u'MATURE', 'additional': u'Improve lifetime of NVM device by not writing the same value several times.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1586],[SSCS_ACLog_1588]'}, u'SWRD_GLOBAL-ACENM_0194': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1056],[SSCS_ACLog_1098]'}, u'SWRD_GLOBAL-ACENM_0195': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1056],[SSCS_ACLog_1098]'}, u'SWRD_GLOBAL-ACENM_0192': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Failed open/failed closed failures have no impact on network re-configuration', 'issue': u'1.11', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0428': {'body': '', 'status': u'MATURE', 'additional': u'NO', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1176]'}, u'SWRD_GLOBAL-ACENM_0059': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Contactor status validity deleted from SSCS', 'issue': u'1.9', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0058': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_596],[SSCS_ACLog_595]'}, u'SWRD_GLOBAL-ACENM_0057': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_598],[SSCS_ACLog_596],[SSCS_ACLog_1517]'}, u'SWRD_GLOBAL-ACENM_0056': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_583]'}, u'SWRD_GLOBAL-ACENM_0055': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1518]'}, u'SWRD_GLOBAL-ACENM_0292': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure has no impact on power supply monitoring.', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0053': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_583]'}, u'SWRD_GLOBAL-ACENM_0052': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1518]'}, u'SWRD_GLOBAL-ACENM_0198': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.1', 'refer': u'[SSCS_ACLog_912],[SSCS_ACLog_1181],[SSCS_ACLog_1195]'}, u'SWRD_GLOBAL-ACENM_0050': {'body': '', 'status': u'MATURE', 'additional': u'NO', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A ', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1111],[SSCS_ACLog_1126],[SSCS_ACLog_1466]'}, u'SWRD_GLOBAL-ACENM_0518': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1449],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0519': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1452],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0512': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-180]'}, u'SWRD_GLOBAL-ACENM_0513': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_829],[SSCS_ACLog_1062],[SSCS_ACLog_933],[SSCS_ACLog_1192],[SSCS_ACLog_1193],[SSCS_ACLog_449],[SSCS_ACLog_943],[SSCS_ACLog_1118],[SSCS_ACLog_1116],[SSCS_ACLog_1256],[SSCS_ACLog_1458],[SSCS_ACLog_1545],[CAN-IRD-216]'}, u'SWRD_GLOBAL-ACENM_0510': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0511': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-180]'}, u'SWRD_GLOBAL-ACENM_0516': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_1261]'}, u'SWRD_GLOBAL-ACENM_0517': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1477],[SSCS_ACLog_1478]'}, u'SWRD_GLOBAL-ACENM_0514': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'A flight leg fault counter is defined to manage the storage management of the failure.', 'issue': u'1.4', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0515': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[SSCS_ACLog_1261]'}, u'SWRD_GLOBAL-ACENM_0453': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'YES', 'safety': u'NO', 'rationale': u'This requirement only depends on the CAN protocol defined in CAN IRD. ', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1494],[CAN-IRD-848]'}, u'SWRD_GLOBAL-ACENM_0452': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0451': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1506]'}, u'SWRD_GLOBAL-ACENM_0450': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0457': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0456': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0455': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0454': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0459': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.4', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394]'}, u'SWRD_GLOBAL-ACENM_0458': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[CAN-IRD-315],[CAN-IRD-392],[CAN-IRD-394],[CAN-IRD-845]'}, u'SWRD_GLOBAL-ACENM_0429': {'body': '', 'status': u'MATURE', 'additional': u'Timeout is defined in HSID. Watchdog needs to be periodically refreshed in order to avoid a CPU reset. ', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_1103]'}, u'SWRD_GLOBAL-ACENM_0220': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_701]'}, u'SWRD_GLOBAL-ACENM_0221': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[CAN-IRD-194],[CAN-IRD-767],[CAN-IRD-195],[CAN-IRD-422],[CAN-IRD-423],[CAN-IRD-198]'}, u'SWRD_GLOBAL-ACENM_0222': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_597],[SSCS_ACLog_615],[SSCS_ACLog_879],[SSCS_ACLog_1226],[CAN-IRD-185],[CAN-IRD-186],[CAN-IRD-1001]'}, u'SWRD_GLOBAL-ACENM_0223': {'body': '', 'status': u'MATURE', 'additional': u'Trip cause is the same for each phase (protections are not computed by phase).', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_677],[SSCS_ACLog_718],[SSCS_ACLog_719],[SSCS_ACLog_1582],[SSCS_ACLog_1584],[SSCS_ACLog_1586],[SSCS_ACLog_1588],[CAN-IRD-194],[CAN-IRD-767],[CAN-IRD-195]'}, u'SWRD_GLOBAL-ACENM_0224': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[CAN-IRD-731],[CAN-IRD-1001],[SSCS_ACLog_1460],[SSCS_ACLog_1542]'}, u'SWRD_GLOBAL-ACENM_0225': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_506],[CAN-IRD-731],[CAN-IRD-1001],[SSCS_ACLog_1460]'}, u'SWRD_GLOBAL-ACENM_0226': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1073],[SSCS_ACLog_1074],[CAN-IRD-464],[CAN-IRD-688],[CAN-IRD-689],[CAN-IRD-205],[CAN-IRD-207],[SSCS_ACLog_1360],[SSCS_ACLog_1362],[SSCS_ACLog_1393],[CAN-IRD-1001]'}, u'SWRD_GLOBAL-ACENM_0227': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_550],[CAN-IRD-215],[CAN-IRD-1001]'}, u'SWRD_GLOBAL-ACENM_0228': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_686],[CAN-IRD-201],[CAN-IRD-1001]'}, u'SWRD_GLOBAL-ACENM_0229': {'body': '', 'status': u'MATURE', 'additional': u'Network is always considered as valid.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.9', 'refer': u'[SSCS_ACLog_1127],[SSCS_ACLog_1128],[SSCS_ACLog_1226],[SSCS_ACLog_1227],[SSCS_ACLog_1228],[CAN-IRD-180],[CAN-IRD-181],[CAN-IRD-1001]'}, u'SWRD_GLOBAL-ACENM_0354': {'body': '', 'status': u'MATURE', 'additional': u'Default values are specified for data which are extracted from CAN bus when data are not available or not valid on the two CAN busses or when software is in INIT mode.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_1068]'}, u'SWRD_GLOBAL-ACENM_0040': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_571]'}, u'SWRD_GLOBAL-ACENM_0424': {'body': '', 'status': u'MATURE', 'additional': u'Each EDMU_ACMPX_TRIPPED_CMD is associated with one ACMP. An ACMP can receive a TRIP reset independently from other ACMPs. At start-up, the ACMPx tripped states are restored from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.7', 'refer': u'[SSCS_ACLog_721],[SSCS_ACLog_1119],[SSCS_ACLog_1280],[SSCS_ACLog_1581],[SSCS_ACLog_1582],[SSCS_ACLog_1583],[SSCS_ACLog_1584]'}, u'SWRD_GLOBAL-ACENM_0357': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'YES', 'rationale': u'N/A', 'issue': u'1.0', 'refer': u'[SSCS_ACLog_663]'}, u'SWRD_GLOBAL-ACENM_0299': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Default value NOT_FAILED is specified during SW initialization phase.', 'issue': u'1.7', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0298': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure has no impact on power supply monitoring.', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0420': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Merged in SWRD_GLOBAL-ACENM_0012', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0421': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.6', 'refer': u'[SSCS_ACLog_1486],[CAN-IRD-182]'}, u'SWRD_GLOBAL-ACENM_0295': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Fuse failure has no impact on power supply monitoring.', 'issue': u'1.4', 'refer': 'EMPTY'}, u'SWRD_GLOBAL-ACENM_0294': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1183],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0297': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1183],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0296': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Default value NOT_FAILED is specified during SW initialization phase.', 'issue': u'1.0', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0291': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.5', 'refer': u'[SSCS_ACLog_1183],[SSCS_ACLog_1181]'}, u'SWRD_GLOBAL-ACENM_0290': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Default value NOT_FAILED is specified during SW initialization phase.', 'issue': u'1.0', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0293': {'body': '', 'status': u'MATURE', 'additional': u'N/A', 'derived': u'YES', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'Default value NOT_FAILED is specified during SW initialization phase.', 'issue': u'1.0', 'refer': u'N/A'}, u'SWRD_GLOBAL-ACENM_0550': {'body': '', 'status': u'MATURE', 'additional': u'At start-up this data is initialized first with value from NVM.', 'derived': u'NO', 'terminal': u'NO', 'safety': u'NO', 'rationale': u'N/A', 'issue': u'1.11', 'refer': u'[SSCS_ACLog_920],[SSCS_ACLog_1196],[SSCS_ACLog_1608]'}, u'SWRD_GLOBAL-ACENM_0043': {'body': '', 'status': u'DELETED', 'additional': 'EMPTY', 'derived': 'EMPTY', 'terminal': 'EMPTY', 'safety': 'EMPTY', 'rationale': u'Source availability is no more impacted by failed open/closed states.', 'issue': u'1.5', 'refer': 'EMPTY'}}
        tbl_list_req={}
        for req,value in extract_req.tbl_list_llr.iteritems():
            status = CheckLLR.getAtribute(value,"status")
            if status != "DELETED":
                issue = CheckLLR.getAtribute(value,"issue")
                tbl_list_req[str(req)]=issue
        # Get LLR now
        extract_req.basename = self.dir_swdd
        extract_req.hlr_selected = False
        extract_req.tbl_list_llr.clear()
        extract_req.extract(dirname=self.dir_swdd,
                         type=("SWDD",))
        hlr_vs_llr = {}
        llr_vs_hlr = {}
        # Low Level Requirements
        for req,value in extract_req.tbl_list_llr.iteritems():
            issue = CheckLLR.getAtribute(value,"issue")
            list_refer,list_constraints = extract_req.getLLR_Trace(value)
            #print "list_refer:",list_refer
            for refer in list_refer:
                if refer not in hlr_vs_llr:
                    hlr_vs_llr[refer]=[[req,issue]]
                else:
                    hlr_vs_llr[refer].append([req,issue])
            for refer in list_constraints:
                #print "refer",refer
                if refer not in tbl_list_req:
                    tbl_list_req[refer]=""
                if refer not in hlr_vs_llr:
                    hlr_vs_llr[refer]=[[req,issue]]
                else:
                    hlr_vs_llr[refer].append([req,issue])
            # redondant with get_ig_jquey.py line 671
            csu_id = re.sub(r"\w*(CSC[0-9]{3}_CSU[0-9]{3})_[0-9]{3}",r"\1",req)
            if csu_id not in llr_vs_hlr:
                llr_vs_hlr[csu_id]=list_refer
            else:
                llr_vs_hlr[csu_id].extend(list_refer)
        #tbl_list_req.sort()
        # Debug
        #for x,y in llr_vs_hlr.iteritems():
        #    print "llr_vs_hlr:",x,y
        #for x in e:tract_req.tbl_list_llr:
        #    print "check_is.tbl_list_llr",x
        # Excel
        #print "exportXlsScod",tbl_list_req
        #self.exportXlsScod(tbl_list_req=tbl_list_req,
        #                   hlr_vs_llr=hlr_vs_llr,
        #                   llr_vs_code=llr_vs_code)
        sheet_name = 'Top Bottom'
        wb = Workbook()
        #wb = load_workbook(filename = join('template','Trace_template.xlsx'))
        if wb is not None:
            #ws = wb.get_sheet_by_name(name = sheet_name)
            ws = wb.worksheets[0]
            if ws is not None:
                Style.putLogo(ws)
                #if 0==1:
                #    style_border = Style(border=Border(
                #        left=Side(border_style=BORDER_THIN),
                #        right=Side(border_style=BORDER_THIN),
                #        top=Side(border_style=BORDER_THIN),
                #        bottom=Side(border_style=BORDER_THIN)),
                #                         alignment=Alignment(wrap_text=True,shrink_to_fit=True))
                row = 9
                user_dir=(self.root_user_dir,self.src_user_dir,self.dir_swdd,self.dir_swrd)
                tbl = ("Section","HLR/Constraints","HLR version","CSC","CSU","LLR","LLR version","Source Code","Src version")
                for col_idx in range(1,10):
                    Style.setCell(ws,tbl,row,col_idx)
                row += 1
                #print "tbl_req_vs_section",extract_req.tbl_req_vs_section
                for req_id,hlr_version in tbl_list_req.iteritems():
                    if req_id in hlr_vs_llr:
                        for low_level_req_id,llr_version in hlr_vs_llr[req_id]:
                            tiny_llr_id = re.sub(r"\w*(CSC[0-9]{3}_CSU[0-9]{3}_[0-9]{3})",r"\1",low_level_req_id)
                            csu_id      = re.sub(r"\w*(CSC[0-9]{3}_CSU[0-9]{3})_[0-9]{3}",r"\1",low_level_req_id)
                            if csu_id in llr_vs_code:
                                func_name = llr_vs_code[csu_id][1]
                                source_code = llr_vs_code[csu_id][0]
                                source_code_name = llr_vs_code[csu_id][2]
                                csc_name = llr_vs_code[csu_id][3]
                                csu_name = llr_vs_code[csu_id][4]
                                link_code = join(user_dir[0],user_dir[1])
                                link_code = join(link_code,source_code)
                                link_llr = join(user_dir[0],user_dir[2])
                                link_llr = join(link_llr,func_name)
                                llr_link="file:///{:s}".format(link_llr)
                                src_code_link="file:///{:s}".format(link_code)
                                #print "REQ_ID",req_id
                                if req_id in extract_req.tbl_req_vs_section:
                                    req_section = extract_req.tbl_req_vs_section[req_id]
                                else:
                                    req_section = ""
                                source_code_version = Tool.get_source_code_version(link_code)
                                tbl = (req_section,req_id,hlr_version,csc_name,csu_name,tiny_llr_id,llr_version,source_code_name,source_code_version)
                                for col_idx in range(1,10):
                                    Style.setCell(ws,tbl,row,col_idx)

                                hlr_link="file:///{:s}".format(self.dir_swrd)
                                hsid_link="file:///{:s}".format(self.hsid)
                                m = re.search(r'SWRD',req_id)
                                if m:
                                    Style.setHyperlink(ws,row,2,hlr_link)
                                else:
                                    Style.setHyperlink(ws,row,2,hsid_link)
                                Style.setHyperlink(ws,row,5,llr_link)
                                Style.setHyperlink(ws,row,8,src_code_link)

                                row += 1
                ws.auto_filter.ref = "A9:H9"
                wb.create_sheet(title = 'Bottom-up')
                ws = wb.worksheets[1]
                tbl_src_files = []
                for source_file,values in code_vs_llr.iteritems():
                    llr_name = values[0] #CSC001_CSU002
                    func_name = values[1]
                    csu_func_name = values[2]
                    source_code_name = values[3]
                    if source_code_name not in tbl_src_files:
                        tbl_src_files.append(source_code_name)
                tbl_src_files.sort()
                row = 9
                for src in tbl_src_files:
                    for col_idx in range(1,2):
                        Style.setCell(ws,(src,""),row,col_idx)
                    row +=1
                filename = "traca_scod_%d.xlsx" % floor(time.time())
                wb.save(join("result",filename))
                #self.master_ihm.resultHyperLink(filename,text="SCOD created.")
                if filename is not None:
                    self.master_ihm.resultGenerateCID(filename,
                                                False,
                                                text="SCOD GENERATION")
        # HTML
        if 0==1:
            report_filename = export_scod_html.exportHTML(list_reqs_spec = tbl_list_req,
                                                          list_llr_per_hlr = hlr_vs_llr,
                                                          list_hlr_per_llr = llr_vs_hlr,
                                                          list_code_per_llr = llr_vs_code,
                                                          list_llr_per_code = code_vs_llr,
                                                          user_dir=(self.root_user_dir,
                                                                    self.src_user_dir,
                                                                    self.dir_swdd,
                                                                    self.dir_swrd))
            export_scod_html.start()

    def _checkISCmd(self,
                    dirname_upper="",
                    dirname_req="",
                    filename_is="",
                    component="",
                    hlr_selected=False,
                    skip_change_synergy_var=0,
                    exportHTML=False):
        """
        This function checks Inspection Sheet document for specification
        :param dirname_upper:
        :param dirname_req:
        :param filename_is:
        :param hlr_selected:
        :return:
        """

        #skip_change_synergy_var = self.master_ihm.skip_change_synergy_var.get()
        #print "SKIP:",skip_change_synergy_var
        check_is = CheckIS(basename=dirname_req,
                           hlr_selected = hlr_selected,
                           callback = self.master_ihm.log,
                           session_started=True)
        if hlr_selected:
            tbl_type=["SWRD","PLDRD"]
            check_is.openLog("RD")
        else:
            tbl_type=("SWDD",)
            check_is.openLog("DD")
        #print "_checkISCmd:filename_is",filename_is
        doc_upper,doc_inspected,filename_is_short,attr_check_filename,file_check_filename = check_is.checkISForSpec(filename_is = filename_is,
                                                                  dirname_req = dirname_req,
                                                                  dirname_upper = dirname_upper,
                                                                  type = tbl_type,
                                                                  skip_change_synergy_var=skip_change_synergy_var,
                                                                  component=component)

        # Export results of analysis in an Excel workbook
        if dirname_req != "":
            spec_available = True
        else:
            spec_available = False
        if not exportHTML:
            report_filename = check_is.export(doc_upper = doc_upper,
                                              doc_inspected = doc_inspected,
                                              filename_is = filename_is_short,
                                              spec_available=spec_available
                                              )
        else:
            # export HTML
            # get information for CR from Change database
            ccb = CCB(self.master_ihm)
            dico_tableau_pr = {"all":[],
                               "open":[],
                               "closed":[]}
            ccb.getPR(dico_tableau_pr,
                      cr_with_parent = True,
                       list_cr = check_is.list_cr,
                       no_header=True)
            check_is.dico_errors.update(check_is.dico_warnings)
            if dico_tableau_pr["all"] == [] and check_is.list_cr != []:
                # Problem, no Synergy session started.
                for cr in check_is.list_cr:
                    dico_tableau_pr["all"].append([cr,"No synergy session started.","","","","","CR",""])
            # generate HTML
            report_filename = self.export_is_html.exportHTML(doc_upper = doc_upper,
                                                  doc_inspected = doc_inspected,
                                                  filename_is = filename_is_short,
                                                  spec_available=spec_available,
                                                  list_reqs_is=check_is.dico_results,
                                                  dico_errors=check_is.dico_errors,
                                                  list_reqs_spec = check_is.tbl_list_llr,
                                                  list_cr=dico_tableau_pr["all"],
                                                  list_cr_not_found=check_is.list_cr_not_found,
                                                  target_release=check_is.is_release,
                                                  dico_list_applicable_docs = check_is.applicable_docs)
            #self.master_ihm.displayHyperlink("hlink",report_filename,"Web page created.")
            self.master_ihm.resultHyperLink(report_filename,text="Web page created.")
            self.export_is_html.start()
        check_is.closeLog()
        if check_is.log_filename is not None:
            self.master_ihm.resultHyperLink(check_is.log_filename,hyperlink_tag="hlink1",text="log")
        if attr_check_filename is not None:
            self.master_ihm.resultHyperLink(attr_check_filename,hyperlink_tag="hlink2",text="List of requirements with attributes.")
        if file_check_filename is not None:
            self.master_ihm.resultHyperLink(file_check_filename,hyperlink_tag="hlink3",text="List of files with amount of requirenents per file.")
        if report_filename is not None:
            self.master_ihm.resultGenerateCID(report_filename,
                                                False,
                                                text="INSPECTION CHECK")
        check_is.word.Application.Quit(-1)
        pythoncom.CoUninitialize()

        return check_is

    def _checkISDocCmd(self,
                       filename_is,
                       skip_change_synergy_var=False,
                       verif_issue_cr_process_start=999):
        """
        This function checks Inspection Sheet document for any other documents
        :param filename_is:
        :param verif_issue_cr_process_start:
        :return:
        """
        check_is = CheckIS("",
                           callback = self.master_ihm.log)
        check_is.openLog("Generic")
        result = check_is.CheckISGeneric(filename_is,
                                         skip_change_synergy_var,
                                         verif_issue_cr_process_start)
        check_is.logErrors()
        check_is.logWarnings()
        check_is.closeLog()
        if check_is.log_filename is not None:
            self.master_ihm.displayHyperlink("hlink",check_is.log_filename)
        if result:
            self.master_ihm.success.config(fg='magenta',bg = 'green',text="INSPECTION CHECK SUCCEEDED")
        else:
            self.master_ihm.success.config(fg='yellow',bg = 'red',text="INSPECTION CHECK FAILED")

    def _getChapterDialogHLR(self,filename):
        CheckLLR.getChapterReq(filename)

    def _exportIS(self,
                  dirname_req="",
                  dirname_upper="",
                  hlr_selected=False,
                  reference="",
                  issue="",
                  release="",
                  hsid_dirname="",
                  reviewer_name="",
                  default_status="",
                  dico_parameters={}
                  ):

        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        export_is = CheckIS(dirname_req,
                           hlr_selected = hlr_selected,
                           callback = self.master_ihm.log,
                           session_started=True)

        if not hlr_selected:
            type = "SWDD"
            export_is.openLog(type)
            export_is.getHSID(hsid_dirname)
        else:
            type = "SWRD"
            export_is.openLog(type)
        spec_ref = "{:s}_{:s}-{:s}".format(type,reference,issue)
        # Extract requirements from document
        export_is.extract(dirname_req,
                            (type,))
        export_is.closeLog()
        print "tbl_file_llr_wo_del",export_is.tbl_file_llr_wo_del
        print "tbl_list_llr",export_is.tbl_list_llr
        # Extract requirements from upper specifications.
        upper = CheckLLR(dirname_upper,
                         hlr_selected=True)
        if hlr_selected:
            upper.openLog("SSCS")
            list_upper = upper.getListUpper()
            upper.extract(dirname_upper,
                          type=list_upper)
        else:
            upper.openLog("SWRD")
            list_upper = ("SWRD",)
            upper.extract(dirname_upper,
                          type=list_upper)
        #print "list_upper",list_upper
        #print "upper.tbl_list_llr",upper.tbl_list_llr
        #for x in upper.tbl_list_llr:
        #    print "upper:",x
        # Create excel workbook

        # CR list
        #dico_parameters = self.master_ihm.getParameters()
        ccb = CCB(self.master_ihm)
        dico_tableau_pr = {"all":[],
                           "open":[],
                           "closed":[]}
        ccb.getPR(dico_tableau_pr,
                   dico_parameters["detect"],
                   dico_parameters["implemented"],
                   dico_parameters["cr_type"],
                   False)

        for pr,list in dico_tableau_pr.iteritems():
            print pr,list
        print "upper.tbl_list_llr",upper.tbl_list_llr
        filename_is = export_is.exportIS(spec_ref,
                                         reference=reference,
                                         issue=issue,
                                         release=release,
                                         reviewer_name = reviewer_name,
                                         default_status = default_status,
                                         author=dico_parameters["author"],
                                         project=dico_parameters["system"],
                                         dico_upper=upper.tbl_list_llr,
                                         tbl_cr=dico_tableau_pr["all"],
                                         type=type)
        upper.closeLog()

        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))
        if filename_is is not None:
            self.master_ihm.resultGenerateCID(filename_is,
                                    False,
                                    text="INSPECTION SHEET")
            #self.master_ihm.displayHyperlink("hlink",filename_is,"Inspection Sheet created.")
        #self.master_ihm.success.config(fg='magenta',bg = 'green',text="INSPECTION SHEET EXPORT SUCCEEDED")

    def _checkLLRCmd(self,
                     dirname="",
                     hlr_selected=False,
                     list_spec=("SWRD","PLDRD"),
                     hsid_dirname=""):

        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        llr = CheckLLR(dirname,
                       hlr_selected = hlr_selected,
                       callback = self.master_ihm.log)
        if hlr_selected:
            llr.openLog("SwRD")
        else:
            llr.openLog("SwDD")
            llr.getHSID(hsid_dirname)
        attr_check_filename,file_check_filename = llr.extract(dirname,
                                                              list_spec)
        llr.logErrors()
        llr.logWarnings()
        llr.closeLog()
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))
        #print "attr_check_filename:",attr_check_filename
        if attr_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink1",attr_check_filename,"List of requirements with attributes.")
        if file_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink3",file_check_filename,"List of files with amount of requirenents per file.")
        if llr.log_filename is not None:
            self.master_ihm.displayHyperlink("hlink2",llr.log_filename,"Log created.")
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="SPECIFICATION CHECK SUCCEEDED")

    def _checkUpperCmd(self,dirname):
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="SPECIFICATION CHECK IN PROGRESS ...")
        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        upper = CheckLLR(dirname,hlr_selected=True)
        upper.openLog("SSCS")
        list_upper = upper.getListUpper()
        print "LIST UPPER",list_upper
        attr_check_filename,file_check_filename = upper.extract(dirname,list_upper)
        upper.getAllocation()
        #print "TEST",upper.dico_alloc_vs_req
        upper.logErrors()
        upper.logWarnings()
        upper.closeLog()
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))
        self.master_ihm.log("{:d} requirements found.".format(upper.nb_reqs))
        if attr_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink1",attr_check_filename,"List of requirements with attributes.")
        if file_check_filename is not None:
            self.master_ihm.displayHyperlink("hlink3",attr_check_filename,"List of files with amount of requirenents per file.")
        if upper.log_filename is not None:
            self.master_ihm.displayHyperlink("hlink2",upper.log_filename)
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="SPECIFICATION CHECK SUCCEEDED")

    def _genLLRDerivedCmd(self,dirname,tbl_type=("SWDD",)):
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="DERIVED REQUIREMENTS EXPORT IN PROGRESS ...")
        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        hlr = Derived(dirname,
                      hlr_selected=False,
                      general_output_txt = self.master_ihm.general_output_txt)
        self.master_ihm.log(text="", color="white")
        hlr.listDir(dirname,tbl_type)
        hlr.invert()
        hlr.countDerived()
        filename = hlr.export()
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))

        if filename is not None:
            self.master_ihm.displayHyperlink("hlink",filename,"xlsx document created.")
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="DERIVED REQUIREMENTS EXPORT SUCCEEDED")

    def _genHLRDerivedCmd(self,dirname,tbl_type=["SWRD","PLDRD"]):
        self.master_ihm.success.config(fg='red',bg = 'yellow',text="DERIVED REQUIREMENTS EXPORT IN PROGRESS ...")
        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        hlr = Derived(dirname,
                      hlr_selected=True,
                      general_output_txt = self.master_ihm.general_output_txt)
        self.master_ihm.log(text="", color="white")
        hlr.listDir(dirname,tbl_type)
        hlr.invert()
        if "SWRD" in tbl_type:
            hlr.countDerived()
        else:
            hlr.countDerived("SSCS")
        filename = hlr.export()
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))

        if filename is not None:
            self.master_ihm.displayHyperlink("hlink",filename,"xlsx document created.")
        self.master_ihm.success.config(fg='magenta',bg = 'green',text="DERIVED REQUIREMENTS EXPORT SUCCEEDED")

    @staticmethod
    def extractCR(tbl_cr):
        cr_id = tbl_cr[0]
        cr_synopsis = tbl_cr[2]
        cr_status = tbl_cr[3]
        return cr_id,cr_synopsis,cr_status

    @staticmethod
    def cleanImpactAnalysis(impact_analysis):
        #impact_analysis = re.sub(r", ?$",r"",impact_analysis)
        import html2text
        impact_analysis_plain_txt = html2text.html2text(Tool.removeNonAscii(impact_analysis))
        #impact_analysis_plain_txt = re.sub(r"\r",r" ",impact_analysis_plain_txt)
        #impact_analysis_plain_txt = re.sub(r"\n",r" ",impact_analysis_plain_txt)
        #print "impact_analysis_plain_txt",impact_analysis_plain_txt
        return impact_analysis_plain_txt

    def export(self,
               tbl_cr,
               for_review_on=False):
        """
        export CRs list in Excel file
        :param tbl_cr:
        :param cr_type:
        :param for_review_on:
        :return:
        """
        sheet_name = 'Change Requests'
        wb = load_workbook(filename = join('template',self.export_cr_list_filename))
        #sheet_name = 'Change Requests'
        #wb = Workbook()
        if wb is not None:
            #ws = wb.create_sheet(title = sheet_name)
            ws = wb.get_sheet_by_name(name = sheet_name)
            #ws = wb.worksheets(sheet_name)
            filename = None
            if ws is not None:
                Style.putLogo(ws)
                border=Border(
                    left=Side(border_style=BORDER_THIN),
                    right=Side(border_style=BORDER_THIN),
                    top=Side(border_style=BORDER_THIN),
                    bottom=Side(border_style=BORDER_THIN))
                alignment=Alignment(wrap_text=True,shrink_to_fit=True)
                style_border = Style(border,alignment)
                row = 9
                if not for_review_on:
                    debug_tbl_cr = []
                    for cr in tbl_cr:
                        # ID
                        cr_id = cr[0]
                        # Patch to get CR domain from CR status
                        cr_domain = self.getStatusPrefix(cr[3])
                        cr[0] = "{:s} {:s}".format(cr_domain,cr[0].zfill(4))
                        # Synopsis
                        cr[2] = self.replaceNonASCII(cr[2])
                        # Status
                        cr[3] = self.removeStatusPrefix(cr[3])
                        # Impact analysis
                        impact_analysis = self.replaceBeacon(cr[9])
                        impact_analysis = self.cleanImpactAnalysis(impact_analysis)
                        #impact_analysis = Tool.adjustCR(impact_analysis)
                        cr[9] = impact_analysis
                        #cr[9] = self.cleanImpactAnalysis(cr[9])
                        # CR Functional Impact/Limitation
                        cr[11] = self.cleanImpactAnalysis(cr[11])
                        #print cr
                        hyperlink = "http://spar-syner1.in.com:8600/change/PTweb?ACTION_FLAG=frameset_form&TEMPLATE_FLAG=ProblemReportView&database=%2Fusr%2Flocal%2Fccmdb%2Fdb_sms_pds&role=User&problem_number={:s}".format(cr_id)
                        #hyperlink = "http://www.lemonde.fr"
                        #CheckLLR.setCell(ws,cr,row,1,style_border)
                        Style.setHyperlink(ws,row,1,hyperlink)
                        # TODO: factoriser avec ccb.py
                        impact_match = re.findall(r'(SW|HW|PLD) impact ?: ?([Y|y][E|e][S|s]|[N|n][O|o]n?e?)',cr[9])
                        #print "IMPACT List CR",impact_match
                        CCB.addImpactAnalysis(cr,impact_match)
                        debug_tbl_cr.append(cr)
                        for col_idx in range(1,21):
                            Style.setCell(ws,cr,row,col_idx,style_border)
                        row += 1
                    # Autofilter
                    ws.auto_filter.ref = "A8:T8"
                    Style.set_border(ws, "M7:T7")
                    print "CR",debug_tbl_cr
                else:
                    for cr in tbl_cr:
                        for col_idx in range(1,6):
                            Style.setCell(ws,cr,row,col_idx,style_border)
                        row += 1
                    # Autofilter
                    ws.auto_filter.ref = "A8:T8"
                    Style.set_border(ws, "M7:T7")
                # save the file
                filename = "Change_Requests_List_%d.xlsx" % floor(time.time())
                wb.save(join("result",filename))
            else:
                print "WorkSheet \"{:s}\" access failed.".format(sheet_name)
                ws_found = wb.get_sheet_names()
                print "Found:",ws_found
        else:
            print "WorkBook \"{:s}\" access failed.".format(self.export_cr_list_filename)
        return filename

    @staticmethod
    def getTransition(line):
        match_transition = re.match(r'^_TRANSITION_ (Submitted to|Transitioned to) (.*) by (.*) on ([0-9]{4}/[0-9]{2}/[0-9]{2} [0-9]{2}:[0-9]{2}:[0-9]{2})',line)
        if match_transition:
            transition = ThreadQuery.removeStatusPrefix(match_transition.group(2))
            date_change = match_transition.group(4)
            datetime_obj = datetime.strptime(date_change, '%Y/%m/%d %H:%M:%S')
            date = datetime_obj.strftime('%A %d %b %Y')
            time = datetime_obj.strftime('%H:%M:%S')
        else:
            transition = False
            date = False
            time = False
            datetime_obj = False
        return transition,date,time,datetime_obj

    def parseLog(self,
                 cr_id,
                 transi_log,
                 dico_cr_transition,
                 dico_cr_log,
                 ccb_time_obj=False):
        """

        :param cr_id:
        :param transi_log:
        :param dico_cr_transition:
        :param dico_cr_log:
        :param ccb_time:
        :return: found_status
        """
        # Replace FS and RS characters
        char = {r'\x1c':'_TRANSITION_ ',
                r'\x1e':'_UPDATE_ '}
        print "TRANSI_LOG",transi_log
        for before, after in char.iteritems():
            transi_log = re.sub(before,after,transi_log)
        if transi_log is not None:
            transi_log_filtered = self.replaceNonASCII(transi_log)
            #transi_log_filter.decode('latin1') #filter(string.printable[:-5].__contains__,transi_log_filter)
        else:
            transi_log_filtered = transi_log
        #transi_log_filtered = self._filterASCII(transi_log)
        dico_cr_log[cr_id] = transi_log_filtered
        tbl_log = transi_log_filtered.splitlines()
        list_transitions = []
        check_comment = False
        transition = False
        found_status = False
        for line in tbl_log:
            if transition:
                if check_comment:
                    # Get comment for transition
                    list_transitions.append((transition,date,time,line))
                    check_comment = False
                else:
                    list_transitions.append((transition,date,time,""))
                transition = False
            else:
                transition,date,time,datetime_obj = self.getTransition(line)
                if datetime_obj and ccb_time_obj:
                    if datetime_obj > ccb_time_obj:
                        if not found_status:
                            found_status = transition
                    elif datetime_obj < ccb_time_obj:
                        found_status = transition
                    else:
                        pass
                if transition in ("Under_Modification",
                                  "Closed",
                                  "Rejected",
                                  "Postponed",
                                  "Complementary_Analysis"):
                    # According to chapter 3.3.3.3 of SCMP, CCB minutes reference shall be documented on Comments field
                    # Missing "incomplete analysis" transition
                    check_comment = True
        dico_cr_transition[cr_id] = list_transitions
        return found_status

    def _preview_CR_Query(self):
        self.master_ihm.preview_CR_Query()

    def createChildrenDiagLink(self,
                               tbl_input_crs={},        # Input
                               dico_cr_children = {},   # Output
                               recur=True,
                               cr_direction="child"):
        # Get children of SACR
        self.children_cr_path += 1
        #dico_cr_children={}
        list_children_cr = []
        del(list_children_cr[:])
        for cr_tag in tbl_input_crs:
            # Get only number
            cr_id = self.getCR_ID(cr_tag)
            # Get child CRs
            tbl_children = self._getParentCR(cr_id,
                                                 type_cr = cr_direction,
                                                 full_info = True)
            if tbl_children:
                for cr_info,parent_cr_status,cr_synopsis,cr_implemented_for,cr_ac_milestone,cr_classif in tbl_children:
                    list_children_cr.append(cr_info)
                dico_cr_children[cr_tag] = tbl_children
        if recur and list_children_cr != [] and self.children_cr_path < self.max_cr_depth:
            print "list_children_cr:",list_children_cr
            print "self.children_cr_path",self.children_cr_path
            self.createChildrenDiagLink(sorted(set(list_children_cr)),
                                        dico_cr_children=dico_cr_children,
                                        cr_direction=cr_direction)
    def _buildCR_Diagram(self,
                         list_sycr=[],
                         tbl_crs = [],              # OUTPUT
                         tbl_cr_children=[],        # OUTPUT
                         cr_direction="child",
                         first_layer=True):

        """

        :type first_layer: object
        """

        def createFileDiag(body_txt,index):
            begin_txt = "diagram {orientation = landscape;" \
                        "class software [icon = \"img/SW.png\"];" \
                        "class pld [icon = \"img/fpga.png\"];" \
                        "class hardware [icon = \"img/HW.png\"]" #default_fontsize = 20;default_shape = roundedbox;"
            source_txt = begin_txt + body_txt + "}"
            tree = parser.parse_string(source_txt)
            filename = join("result","cr_link_{:d}.png".format(index))
            try:
                diagram = builder.ScreenNodeBuilder.build(tree)
                draw = drawer.DiagramDraw('PNG',
                                          diagram,
                                          filename=filename)
                draw.draw()
                draw.save()
            except AttributeError,e:
                print "Exception:",e
                print "SOURCE:",source_txt

            return filename

        # begin hack: https://groups.google.com/forum/#!topic/blockdiag-discuss/9rUQwZXay0k
        png.setup(png)
        box.setup(box)
        # end hack

        self.children_cr_path = 0
        if cr_direction == "child":
            edge_direction = "forward"
        else:
            edge_direction = "back"
        for sycr_id,type,sycr_synopsis,sycr_status,classif,detected_on,x,y,z,a,sycr_domain,sycr_type,syrc_classif,list_sacrs in list_sycr:
            #print "TRACE:",sycr_id,type,sycr_synopsis,sycr_status,classif,detected_on,x,y,z,a,sycr_domain,sycr_type,syrc_classif,list_sacrs
            dico_cr_children={}
            self.children_cr_path = 0
            if first_layer:
                parent_cr_id = sycr_id #.zfill(4)
                #REplace space bby underscore otherwise getCR_ID does not work properly
                sycr_type = re.sub(r' ',r'_',sycr_type)
                sycr_tag = sycr_domain + " " + sycr_type + " " + sycr_id.zfill(4)
                sycr_tag_no_zero_pad = sycr_domain + " " + sycr_type + " " + sycr_id
                body_txt = self.create_node(sycr_tag,
                                            sycr_status,
                                            parent_cr_id,
                                            icon=self.getIcon(sycr_tag))
                tbl_all_sacrs = []
                del(tbl_all_sacrs[:])
                if list_sacrs is not None:
                    #print "DEBUG:",list_sacrs
                    for sacr_tag,sacr_status,sacr_synopsis,sacr_implm,cr_ac_milestone,cr_classif in list_sacrs:
                        tbl_all_sacrs.append(sacr_tag)
                        parent_cr_id = sycr_id
                        child_cr_id = self.getCR_ID(sacr_tag)
                        body_txt += self.create_edge(parent_cr_id,
                                                     child_cr_id,
                                                     edge_direction=edge_direction)

                        body_txt += self.create_node(sacr_tag,
                                                     sacr_status,
                                                     child_cr_id,
                                                     icon=self.getIcon(sacr_tag))
                    dico_cr_children[sycr_tag_no_zero_pad] = list_sacrs
                list_cr_input = sorted(set(tbl_all_sacrs))
            else:
                sycr_tag = sycr_id
                parent_cr_id = self.getCR_ID(sycr_tag)
                body_txt = self.create_node(sycr_tag,
                                            sycr_status,
                                            parent_cr_id,
                                            icon=self.getIcon(sycr_tag))
                list_cr_input = [sycr_tag]
            #print "list_cr_input",list_cr_input
            self.createChildrenDiagLink(list_cr_input,              # Input
                                        dico_cr_children,           # Output
                                        recur=True,                 # Recursive function
                                        cr_direction=cr_direction)  # CR direction
            # root CR
            if dico_cr_children != {}:
                #print "dico_cr_children",dico_cr_children
                for parent_cr_tag,list_child_crs in dico_cr_children.iteritems():
                    # 2nd CRs layer
                    for child_cr_tag,cr_status,cr_synopsis,cr_impl,cr_ac_milestone,cr_classif in list_child_crs:
                        #self.checkStatusSequence(parent_cr_status,
                        #                         cr_status,
                        #                         cr_tag,
                        #                         cr_direction,
                        #                         tbl_list_cr_errors)
                        parent_cr_id = self.getCR_ID(parent_cr_tag)
                        child_cr_id = self.getCR_ID(child_cr_tag)
                        body_txt += self.create_edge(parent_cr_id,
                                                 child_cr_id,
                                                 edge_direction=edge_direction)

                        body_txt += self.create_node(child_cr_tag,
                                                 cr_status,
                                                 child_cr_id,
                                                 icon=self.getIcon(child_cr_tag))
                        #body_txt += self.create_node_edge(child_cr_tag,
                        #                             cr_status,
                        #                             parent_cr_id, #CR_parent_cr_id -> CR_child_cr_id
                        #                             child_cr_id,
                        #                             icon=self.getIcon(child_cr_tag))
                        cr_domain = "TBD"
                        tbl_cr_children.append([sycr_tag,
                                                child_cr_tag,
                                                self.removeQuotes(cr_synopsis),
                                                child_cr_id,
                                                cr_domain,
                                                cr_status])
            #print "BODY_TXT",body_txt
            #exit(0)
            filename = createFileDiag(body_txt,self.sycr_index)
            tbl_crs.append((self.sycr_index,
                            sycr_tag,
                            sycr_status,
                            filename,
                            self.removeQuotes(sycr_synopsis)))
            self.sycr_index += 1

    def _importCRs(self,
                   for_review_on=False,
                   component_type="",
                   detected_on="",
                   implemented_for="",
                   old_cr_workflow=False,
                   list_cr_type=[],
                   list_cr_status=[],
                   list_cr_doamin=[],
                   cr_direction="child",
                   tbl_cr_export=[],                 # OUTPUT
                   list_cr_selected_by_user=[],
                   tbl_list_cr_errors=[],
                   dico_log_errors={}
                   ):

        condition,detect_attribut = self.new_createConditionStatus(old_cr_workflow=old_cr_workflow,
                                                                   detect_release = detected_on,
                                                                   impl_release = implemented_for,
                                                                   cr_type = component_type,
                                                                   list_cr_type=list_cr_type,
                                                                   list_cr_status=list_cr_status,
                                                                   list_cr_doamin=list_cr_doamin,
                                                                   list_cr_selected_by_user=list_cr_selected_by_user
                                                                   )

        classification = CCB.getClassif(old_cr_workflow)
        if for_review_on:
            attributes = '-f "<cell>%problem_number</cell>' \
                         '<cell>%problem_synopsis</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>%crstatus</cell>' \
                         '<cell>%void</cell>"'.format(classification)
        else:
            #if not old_cr_workflow:
            # New problem report workflow
            implementation_baseline_f = "%CR_implementation_baseline"
            # new with tags
            detect_attribut_tag = re.sub(r";","</cell><cell>",detect_attribut)
            attributes = '-f "<cell>%problem_number</cell>' \
                         '<cell>%CR_request_type</cell>' \
                         '<cell>%problem_synopsis</cell>' \
                         '<cell>%crstatus</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>%modify_time</cell>' \
                         '<cell>%impact_analysis</cell>' \
                         '<cell>%CR_domain</cell>' \
                         '<cell>%CR_type</cell>' \
                         '<cell>%CR_customer_classification</cell>"'.format(classification,detect_attribut_tag,implementation_baseline_f)
        # query with no numbering of the line and sorted by problem_number
        query = "query -u -sby problem_number {:s} {:s} ".format(condition,attributes)
        self.master_ihm.log('ccm ' + query)
        # remove \n
        text = re.sub(r"\n",r"",query)
        stdout,stderr = self.ccm_query(text,'Get CRs mapping')
        #tbl_cr_export = []
        if stdout != "":
            #test = stdout.splitlines()
            #for x in test:
            #    print "X:",x
            stdout = self.replaceBeacon(stdout)
            #result = self.cleanImpactAnalysis(stdout)
            result = Tool.adjustCR(stdout)
            output = result.splitlines()
            #dico_log_errors = {}
            for line in output:
                #print "LINE:",line
                cr_decod = self._parseCRCell(line)
                cr_id,cr_synopsis,cr_status = self.extractCR(cr_decod)
                cr_decod[3] = Tool.discardCRPrefix(cr_decod[3])
                #print "CR STATUS:",cr_decod
                cr_status = cr_decod[3]
                cr_classif = cr_decod[12]
                tbl_parent_cr_id = self._getParentCR(cr_id,
                                                     type_cr = cr_direction,
                                                     full_info = True)
                if tbl_parent_cr_id:
                    # Get parent ID pieces of information
                    found_parent_cr_info = []
                    found_parent_cr_status = []
                    found_parent_cr_synopsis = []
                    found_parent_cr_implemented_for = []
                    list_cr = []
                    cr_tag = cr_decod[10] + " " + cr_decod[11] + " " + cr_id
                    dico_log_errors[cr_tag] = []
                    log = dico_log_errors[cr_tag]
                    print "tbl_parent_cr_id",tbl_parent_cr_id
                    for parent_decod in tbl_parent_cr_id:
                        print "PARENT_DECOD:",parent_decod
                        cr_info = parent_decod[0]
                        parent_cr_status =  parent_decod[1]
                        cr_synopsis = parent_decod[2]
                        cr_implemented_for = parent_decod[3]
                        found_parent_cr_info.append(cr_info)
                        found_parent_cr_status.append(parent_cr_status)
                        found_parent_cr_synopsis.append(cr_synopsis)
                        found_parent_cr_implemented_for.append(cr_implemented_for)
                        list_cr.append(parent_decod)
                        parent_cr_classif = parent_decod[5]
                        result_compare_status = self.checkStatusSequence(parent_cr_status,
                                                 cr_status,
                                                 cr_tag,
                                                 cr_direction,
                                                 tbl_list_cr_errors)
                        result_compare_classif = self.checkCustomerClassif(cr_status,
                                                                           parent_cr_classif,
                                                         cr_classif,
                                                         cr_tag,
                                                         cr_direction,
                                                         tbl_list_cr_errors)
                        if result_compare_status:
                            log.append((cr_info,"Status mismatch:CR status is <b>{:s}</b> whereas parent/child status is <b>{:s}</b>.".format(cr_status,parent_cr_status)))
                        if result_compare_classif:
                            if cr_classif != "" and parent_cr_classif != "":
                                log.append((cr_info,"Classification mismatch:CR is classified <b>{:s}</b> whereas parent/child is classified <b>{:s}</b>.".format(cr_classif,parent_cr_classif)))
                            elif cr_classif != "":
                                log.append((cr_info,"Parent/child classification is missing.".format(cr_classif,parent_cr_classif)))
                            elif parent_cr_classif != "":
                                log.append((cr_info,"CR classification is missing.".format(cr_classif,parent_cr_classif)))
                            else:
                                pass
                    cr_decod.append(list_cr)
                else:
                    cr_decod.append(None)
                tbl_cr_export.append(cr_decod)
            print "DICO_LOG",dico_log_errors
        if stderr:
            print time.strftime("%H:%M:%S", time.localtime()) + " " + stderr
             # remove \r
            result = re.sub(r"\r\n",r"\n",stderr)
            self.master_ihm.log(result)

    def checkCustomerClassif(self,cr_status,parent_cr_classif,cr_classif,cr_tag,cr_direction,tbl_list_cr_errors):
        result = False
        if cr_status not in ("Cancelled","Rejected","Closed"):
            if parent_cr_classif != cr_classif:
                print "Erreur:",cr_tag, cr_classif, parent_cr_classif
                tbl_list_cr_errors.append(cr_tag)
                result = True
            else:
                print "Normal:",cr_tag, cr_classif, parent_cr_classif
                result = False
        return result

    def checkStatusSequence(self,parent_cr_status,cr_status,cr_tag,cr_direction,tbl_list_cr_errors):
        result = False
        if 0==0:
            if cr_direction == "parent":
                if cr_status in self.dico_cr_status_relation_forbidden and parent_cr_status not in self.dico_cr_status_relation_forbidden[cr_status]:
                    #if cr_tag not in tbl_list_cr_errors:
                    tbl_list_cr_errors.append(cr_tag)
                    result = True
            else:
                if parent_cr_status in self.dico_cr_status_relation_forbidden and cr_status not in self.dico_cr_status_relation_forbidden[parent_cr_status]:
                    #if cr_tag not in tbl_list_cr_errors:
                    tbl_list_cr_errors.append(cr_tag)
                    result = True
        else:
            if parent_cr_status in self.dico_cr_trans and cr_status in self.dico_cr_trans:
                if cr_direction == "parent":
                    #print "PARENT CR STATUS:",parent_cr_status
                    if self.dico_cr_trans[parent_cr_status] < self.dico_cr_trans[cr_status]:
                        #print "Erreur:",cr_tag, parent_cr_status
                        if cr_tag not in tbl_list_cr_errors:
                            tbl_list_cr_errors.append(cr_tag)
                            result = True
                else:
                    print "CHILD CR STATUS:",parent_cr_status
                    if self.dico_cr_trans[parent_cr_status] >  self.dico_cr_trans[cr_status]:
                        #print "Erreur:",cr_tag, parent_cr_status
                        if cr_tag not in tbl_list_cr_errors:
                            tbl_list_cr_errors.append(cr_tag)
                            result = True
        return result

    def get_CR_Color(self,status):
        dico_color = {"In Analysis":("red","yellow"),
                        "Under Modification":("orange","green"),
                        "Fixed":("yellow","black"),
                        "Closed":("green","pink"),
                }
        status = re.sub(r'_',r' ',status)
        #print "STATUS",status
        if status in dico_color:
            color,text_color = dico_color[status]
        else:
            color = "white"
            text_color = "black"
        return color,text_color

    def removeQuotes(self,text):
        text = re.sub(r'([\"])',    r'\\\1', text)
        return text

    def getIcon(self,cr_id):
        if re.match(r'^SCR .*',cr_id):
            icon = ",class=\"software\""
        elif re.match(r'^PLDCR .*',cr_id):
            icon = ",class=\"pld\""
        elif re.match(r'^HCR .*',cr_id):
            icon = ",class=\"hardware\""
        else:
            icon = ""
        return icon

    def getBottomCR(self,cr_tag):
        if re.match(r'^SCR .*',cr_tag):
            result = True
        elif re.match(r'^PLDCR .*',cr_tag):
            result = True
        elif re.match(r'^HCR .*',cr_tag):
            result = True
        elif re.match(r'^BCR .*',cr_tag):
            result = True
        else:
            result = False
        return result

    def getCR_ID(self,cr_tag):
        cr_id = re.sub(r"^[A-Za-z0-9_ ]* ([0-9]*)$", r"\1",cr_tag)
        return cr_id

    def create_node(self,cr_id,
                     cr_status,
                     index_child,
                     domain="CR",
                     icon=""):
        cr_color,cr_text_color = self.get_CR_Color(cr_status)
        data = 'CR_{:s} [label="{:s}\n[{:s}]",width = 180,color = {:s},textcolor={:s}{:s}];'.format(index_child,
                                                                                                             cr_id,
                                                                                                             cr_status,
                                                                                                             cr_color,
                                                                                                             cr_text_color,
                                                                                                             icon)
        return data

    def create_edge(self,index_parent,
                         index_child,
                         domain="CR",
                         folded = "",
                         edge_direction="forward"):
        if index_child != index_parent:
            data = "{:s}_{:s} -> CR_{:s}  [dir={:s}{:s}];".format(domain,
                                                  index_parent,
                                                  index_child,
                                                  folded,
                                                  edge_direction)
        else:
            data = ""
        return data

    def create_node_edge(self,cr_id,
                         cr_status,
                         index_parent,
                         index_child,
                         domain="CR",
                         icon="",
                         folded=""):
        cr_color,cr_text_color = self.get_CR_Color(cr_status)
        if index_child != index_parent:
            data = "{:s}_{:s} -> CR_{:s} {:s};".format(domain,
                                                  index_parent,
                                                  index_child,
                                                  folded)
        else:
            data = ""
        data += 'CR_{:s} [label="{:s}\n[{:s}]",width = 180,color = {:s},textcolor={:s}{:s}];'.format(index_child,
                                                                                                             cr_id,
                                                                                                             cr_status,
                                                                                                             cr_color,
                                                                                                             cr_text_color,
                                                                                                             icon)
        return data

    def _getCR_Mapping(self,
               baseline="",
               extension=True,
               for_review_on=False,
               cr_with_parent = False,
               log_on = False,
               component_type="",
               detected_on="",
               implemented_for="",
               old_cr_workflow=False,
               ccb_time=False,
               children=True,
               list_cr_type=[],
               list_cr_status=[],
               list_cr_doamin=[],
               cr_mapping_direction={0:1,1:0},
               list_cr_selected_by_user=[]):
        '''
            List CR
            Generate an Excel file at the end
            get
                variables
                    previous_release,
                    impl_release,
                    baseline,
                    project
                    attribute ? Encore utilise ? pas sur.
                methods
                    getTypeWorkflow
                    cr_for_review_var
            from ThreadQuery <=== Interface class
            set
                methods
                    setPreviousRelease
                    setRelease
                    setBaseline
                    setProject
            to BuildDoc

            Note: baseline is used only fo A/C standard
            return: tbl_cr_export
        '''
        def remouve_doublons(tbl_input,tbl_output=[]):
            save_cr_tag = ""
            for cr in tbl_input:
                if cr[1] != save_cr_tag:
                    tbl_output.append(cr)
                    save_cr_tag = cr[1]
        # Launch clock
        dico_timestamp={}
        dico_timestamp["begin_script"] = datetime.now()
        tbl_cr_export = []
        tbl_list_cr_errors = []
        dico_log_errors = {}
        if cr_mapping_direction[0] and cr_mapping_direction[1]:
            cr_direction = "both"
            cr_direction_import = "child"
        elif cr_mapping_direction[0]:
            cr_direction = "child"
            cr_direction_import = "child"
        elif cr_mapping_direction[1]:
            cr_direction = "parent"
            cr_direction_import = "parent"
        else:
            cr_direction = "child"
            cr_direction_import = "child"
        print "cr_direction",cr_direction
        self._importCRs(for_review_on=for_review_on,
                        component_type=component_type,
                        detected_on=detected_on,
                        implemented_for=implemented_for,
                        old_cr_workflow=old_cr_workflow,
                        list_cr_type=list_cr_type,
                        list_cr_status=list_cr_status,
                        list_cr_doamin=list_cr_doamin,
                        cr_direction=cr_direction_import,
                        tbl_cr_export=tbl_cr_export,
                        list_cr_selected_by_user=list_cr_selected_by_user,
                        tbl_list_cr_errors=tbl_list_cr_errors,
                        dico_log_errors=dico_log_errors
                        )

        self.tbl_cr_export = tbl_cr_export
        tbl_crs_root_children = []
        tbl_crs_root_parents = []
        tbl_cr_parents = []
        tbl_cr_children = []
        self.sycr_index = 1 # used by _buildCR_Diagram
        self.max_cr_depth = 5 # used by _buildCR_Diagram
        # tbl_cr_export
        # Exemple: [['1232', 'Defect', 'PWM validity not correct at power-up', 'SyCR_In_Analysis', 'LRU analysis', 'H1.4/S1.6', '', '', '20/04/15 09:33', '[[SW ENM]]? [[SW BITE]]?', None],
        # tbl_cr_children
        # Exemple: [['463', 'SCR SW_ENM 1643', 'Modification of TRU undervoltage protection'],'1463', ...]
        if cr_direction == "child" or cr_direction == "both":
            self._buildCR_Diagram(list_sycr         = tbl_cr_export,
                                  cr_direction      = "child",
                                  tbl_crs           = tbl_crs_root_children,    # OUTPUT
                                  tbl_cr_children   = tbl_cr_children   # OUTPUT
                                  )
            if cr_direction == "both":
                # Need to get SCR, PLDCR and HCR and to make a list of parents
                #print "tbl_cr_export - both",tbl_cr_export
                tbl_scr_export = []
                for cr_parent_tag,cr_tag,synopsis,cr_id,cr_domain,cr_status in tbl_cr_children:
                    print "TAG:",cr_tag
                    #cr_id = self.getCR_ID(cr_tag)
                    # cr_id.zfill(4)
                    if self.getBottomCR(cr_tag):
                        tbl_scr_export.append((cr_tag,"",synopsis,cr_status,"","","","","","","","",[]))
                #print "tbl_scr_export",tbl_scr_export
                self._buildCR_Diagram(list_sycr         = tbl_scr_export,
                                      cr_direction      = "parent",
                                      tbl_crs           = tbl_crs_root_parents,     # OUTPUT
                                      tbl_cr_children   = tbl_cr_parents,           # OUTPUT
                                      first_layer = False
                                      )
        elif cr_direction == "parent":
            #print "tbl_cr_export",tbl_cr_export
            self._buildCR_Diagram(list_sycr      = tbl_cr_export,
                                  cr_direction    = cr_direction,
                                  tbl_crs         = tbl_crs_root_parents,    # OUTPUT
                                  tbl_cr_children = tbl_cr_parents   # OUTPUT
                                  )
        #print "tbl_crs_root_children",tbl_crs_root_children
        #print "tbl_cr_children",tbl_cr_children
        #print "tbl_crs_root_parents",tbl_crs_root_parents
        #print "tbl_cr_parents",tbl_cr_parents
        self.export_cr_html = exportCR_HTML()
        #tbl_cr_children_sorted = sorted(tbl_cr_children,key=lambda x: x[1])
        tbl_crs_root_parents_sorted = sorted(tbl_crs_root_parents,key=lambda x: x[1])
        # Exemple: [(1, 'SyCR PDS 174', 'SyCR_Under_Modification', 'result\\cr_link_1.png', 'Power up sequence')
        tbl_crs_root_parents_sorted_wo_doublons = []
        remouve_doublons(tbl_crs_root_parents_sorted,
                         tbl_crs_root_parents_sorted_wo_doublons)
        tbl_cr_parents_wo_doublons = []
        remouve_doublons(tbl_cr_parents,
                         tbl_cr_parents_wo_doublons)
        #tbl_cr_parents_sorted = sorted(tbl_cr_parents,key=lambda x: x[1])
        report_filename = self.export_cr_html.exportHTML(list_cr=tbl_crs_root_children,
                                                         list_cr_children=tbl_cr_children,
                                                         list_cr_bottom_up=tbl_crs_root_parents_sorted_wo_doublons,
                                                         list_cr_parent=tbl_cr_parents_wo_doublons,
                                                         list_cr_errors=tbl_list_cr_errors,
                                                         dico_log_errors=dico_log_errors,
                                                         database=self.database)
        self.export_cr_html.start()
        self.master_ihm.resultHyperLink(report_filename,None,"CHANGE REQUEST MAPPING")
        dico_timestamp["end_script"] = datetime.now()
        duree_execution_script = dico_timestamp["end_script"] - dico_timestamp["begin_script"]
        self.master_ihm.log("Temps d'exécution du script complet: {:d} seconds".format(duree_execution_script.seconds))
        return ""

    def _testBlockDiag(self):
        source_txt = 'blockdiag {B [label = "foo\ntoto"];C [label = "bar"];''A -> B -> C;}'
        #source_txt = "diagram {orientation = landscape;A -> B -> C;B -> D;)"
        print "diag:",source_txt
        # hack: https://groups.google.com/forum/#!topic/blockdiag-discuss/9rUQwZXay0k
        from blockdiag.imagedraw import png
        png.setup(png)

        from blockdiag.noderenderer import box
        box.setup(box)
        tree = parser.parse_string(source_txt)
        diagram = builder.ScreenNodeBuilder.build(tree)
        filename = join("result","cr_link_test.png")
        draw = drawer.DiagramDraw('PNG', diagram, filename=filename)
        draw.draw()
        draw.save()

    def _getCR(self,
               baseline="",
               extension=True,
               for_review_on=False,
               cr_with_parent = False,
               log_on = False,
               component_type="",
               detected_on="",
               implemented_for="",
               old_cr_workflow=False,
               ccb_time=False,
               children=False,
               list_cr_type=[],
               list_cr_status=[],
               list_cr_doamin=[]):
        '''
            List CR
            Generate an Excel file at the end
            get
                variables
                    previous_release,
                    impl_release,
                    baseline,
                    project
                    attribute ? Encore utilise ? pas sur.
                methods
                    getTypeWorkflow
                    cr_for_review_var
            from ThreadQuery <=== Interface class
            set
                methods
                    setPreviousRelease
                    setRelease
                    setBaseline
                    setProject
            to BuildDoc

            Note: baseline is used only fo A/C standard
            return: tbl_cr_export
        '''
        # Create CR list
        if ccb_time:
            ccb_time_obj = datetime.strptime(ccb_time, '%Y/%m/%d %H:%M:%S')
        else:
            ccb_time_obj = False
        output = ""
        log_filename = "log_list_crs_%d" % floor(time.time()) + ".txt"
        # Domain
        #self.ccb_type = cr_type

        if Tool.isAttributeValid(baseline):
            # get standard
            list_sub_std = []
            if self.master_ihm.dico_std.has_key(baseline):
                #
                # Cette partie ne marche plus, a checker
                #
                condition = '"(cvtype=\'problem\') '
                filter_cr = ""
                list_sub_std = self.master_ihm.dico_std[baseline]
                find_std = False
                implemented = ""
                num = 0
                if self.master_ihm.dico_list_std.has_key(baseline):
                    # Est-ce un standard avion ou un sous-standard projet ?
                    pass
                else:
                    delta_implemented,find_std = self.createCrImplemented(baseline,
                                                                          find_std,
                                                                          filter_cr)
                    implemented += delta_implemented
                for sub_std in list_sub_std:
                        delta_implemented,find_std = self.createCrImplemented(sub_std,
                                                                              find_std,
                                                                              filter_cr)
                        implemented += delta_implemented
                if find_std == True:
                    implemented +=  ') '
                condition += implemented
            else:
                pass

        condition,detect_attribut = self.new_createConditionStatus(old_cr_workflow=False,
                                                                   detect_release = detected_on,
                                                                   impl_release = implemented_for,
                                                                   cr_type = component_type,
                                                                   list_cr_type=list_cr_type,
                                                                   list_cr_status=list_cr_status,
                                                                   list_cr_doamin=list_cr_doamin
                                                                   )

        classification = CCB.getClassif(old_cr_workflow)
        if for_review_on:
            attributes = '-f "<cell>%problem_number</cell>' \
                         '<cell>%problem_synopsis</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>%crstatus</cell>' \
                         '<cell>%void</cell>"'.format(classification)
        else:
            #if not old_cr_workflow:
            # New problem report workflow
            implementation_baseline_f = "%CR_implementation_baseline"
            # new with tags
            detect_attribut_tag = re.sub(r";","</cell><cell>",detect_attribut)
            attributes = '-f "<cell>%problem_number</cell>' \
                         '<cell>%CR_request_type</cell>' \
                         '<cell>%problem_synopsis</cell>' \
                         '<cell>%crstatus</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>{:s}</cell>' \
                         '<cell>%modify_time</cell>' \
                         '<cell>%impact_analysis</cell>' \
                         '<cell>%CR_AC_milestones</cell>' \
                         '<cell>%CR_functional_impact</cell>' \
                         '"'.format(classification,detect_attribut_tag,implementation_baseline_f)
        # query with no numbering of the line and sorted by problem_number
        query = "query -u -sby problem_number {:s} {:s} ".format(condition,attributes)
        self.master_ihm.log('ccm ' + query)
        # remove \n
        text = re.sub(r"\n",r"",query)
        stdout,stderr = self.ccm_query(text,'Get CRs for CCB minutes')
        list_change_requests = []
        tbl_cr_export = []
        dico_cr_log = {}
        dico_cr_transition = {}
        filename = None
        if stdout != "":
            test = stdout.splitlines()
            #for x in test:
            #    print "X:",x
            result = self.replaceBeacon(stdout)
            #result = self.cleanImpactAnalysis(stdout)
            #result = Tool.adjustCR(stdout)
            output = result.splitlines()
            for line in output:
                #print "LINE:",line
                cr_decod = self._parseCRCell(line)
                cr_id,cr_synopsis,cr_status = self.extractCR(cr_decod)
                #  Used to fill UI CR list box
                CCB.createCRlist(cr_id,
                                 cr_synopsis,
                                 list_change_requests)
                # For CLI
                #print line
                if cr_with_parent or children:
                    if children:
                        tbl_parent_cr_id = self._getParentCR(cr_id,type_cr ="child")
                    else:
                        tbl_parent_cr_id = self._getParentCR(cr_id)
                    if tbl_parent_cr_id:
                        # Get parent ID pieces of information
                        found_parent_cr_info = []
                        found_parent_cr_status = []
                        found_parent_cr_synopsis = []
                        found_parent_cr_implemented_for = []
                        found_parent_cr_ac_milestone = []

                        for parent_cr_id in tbl_parent_cr_id:
                            parent_cr = self._getParentInfo(parent_cr_id)
                            parent_cr = re.sub("&"," and ",parent_cr)
                            if parent_cr:
                                parent_decod = self._parseCRParent(parent_cr)
                                # Parent CR;Parent CR status;Parent CR synopsis
                                print "parent/child_decod",parent_decod
                                cr_info = parent_decod[0] + " " + parent_decod[1] + " " + parent_decod[2]
                                parent_cr_status = Tool.discardCRPrefix(parent_decod[3])
                                cr_synopsis = Tool.replaceNonASCII(parent_decod[4])
                                cr_implemented_for = parent_decod[5]
                                if parent_decod[6] != "":
                                    cr_ac_milestone = parent_decod[6]
                                else:
                                    cr_ac_milestone = "None"
                                found_parent_cr_info.append(cr_info)
                                found_parent_cr_status.append(parent_cr_status)
                                found_parent_cr_synopsis.append(cr_synopsis)
                                found_parent_cr_implemented_for.append(cr_implemented_for)
                                found_parent_cr_ac_milestone.append(cr_ac_milestone)
                        found_parent_cr_info_str = ",\n".join(map(str, found_parent_cr_info))
                        found_parent_cr_status_str = ",\n".join(map(str, found_parent_cr_status))
                        found_parent_cr_synopsis_str = ",\n".join(map(str, found_parent_cr_synopsis))
                        found_parent_cr_implemented_for_str = ",\n".join(map(str, found_parent_cr_implemented_for))
                        found_parent_cr_ac_milestone_str = ",\n".join(map(str, found_parent_cr_ac_milestone))
                        cr_decod.extend([found_parent_cr_info_str,
                                         found_parent_cr_status_str,
                                         found_parent_cr_synopsis_str,
                                         found_parent_cr_implemented_for_str,
                                         found_parent_cr_ac_milestone_str])
                    else:
                        cr_decod.extend(["","","","",""])
                else:
                    cr_decod.extend(["","","","",""])
                tbl_cr_export.append(cr_decod)
                if log_on:
                    # Get transition log
                    query = "query -t problem \"(problem_number='{:s}')\" -u -f \"%transition_log\"".format(cr_id)
                    transi_log = self._ccmCmd(query,False)


                    found_status = self.parseLog(cr_id,
                                  transi_log,
                                  dico_cr_transition,
                                  dico_cr_log,
                                  ccb_time_obj)
                    if found_status:
                        print "CR ID: {:s} {:s} <-- {:s}".format(cr_id,found_status,cr_status)

            # end loop for CR parsing
            # Create Excel file with CRs listing
            filename = self.export(tbl_cr_export,
                                   for_review_on=for_review_on)
            list_change_requests.sort()
        self.master_ihm.reloadCR_ListBox(list_change_requests)
        #self.queue.put("RELOAD_CRLISTBOX") # action to get projects
        #self.queue.put(list_change_requests)
        if stderr:
            print time.strftime("%H:%M:%S", time.localtime()) + " " + stderr
             # remove \r
            result = re.sub(r"\r\n",r"\n",stderr)
            self.master_ihm.log(result)
        with open(join(self.gen_dir,log_filename), 'w') as of:
            ccm_query = 'ccm ' + query + '\n\n'
            of.write(ccm_query)
            for cr_id,log in dico_cr_log.iteritems():
                txt = "Full log for {:4s} {:5s}:\n".format("CR",cr_id)
                of.write(txt)
                of.write("-----------------------\n\n")
                txt = "{:s}\n".format(log)
                of.write(txt)
            of.write("\n\n--------------------------------------------------------------------------------------------------------------\n\n")
            for cr_id,transitions in dico_cr_transition.iteritems():
                txt = "Transitions timeline for {:4s} {:5s}:\n".format("CR",cr_id)
                of.write(txt)
                of.write("-----------------------------------\n\n")
                for transition,date,hour,comment in transitions:
                    if comment != "":
                        txt = "   Status set to {:20s} on {:25s} at {:15s} with comment: {:s}\n".format(transition,date,hour,comment)
                    else:
                        txt = "   Status set to {:20s} on {:25s} at {:15s} with no comment.\n".format(transition,date,hour)
                    of.write(txt)
                of.write("\n")
            #of.write(result)
        #self.master_ihm.log("Command executed.")
        self.master_ihm.resultGenerateCID(filename,None,"CHANGE REQUEST LISTING")
        try:
            if log_filename is not None:
                self.master_ihm.resultHyperLink(log_filename,None,"Log of CR and transitions summary.","hlink2")
                #self.master_ihm.displayHyperlink("hlink2",log_filename,"Log of CR and transitions summary.")
        except AttributeError:
            pass
        # Set scrollbar at the bottom
        #self.master_ihm.defill()
        # For debug purpose
        self.tbl_cr_export = tbl_cr_export
        return output

    def run(self):
        # sleep to enables the GUI to finish its setting
        import time
##        print time.strftime("%H:%M:%S", time.localtime()) + " Start thread " + self.name_id + "\n"
        time.sleep(2)
        self.periodicCall()

    def stop(self):
##        print time.strftime("%H:%M:%S", time.localtime()) + " Stop thread " + self.name_id + "\n"
        self.terminated = True

if __name__ == '__main__':
    test = ThreadQuery()
    #result = test._getStackFromAsm()
    #print result
    #exit()
    result = test._stackAnalysis()
    print "RESULT",result
    exit()
    txt = "SEQ.vhd"
    new = Tool.getFileNameAlone(txt)
    print "TXT",new
    exit(0)
    txt = u'SEQ_base_time '
    new = Tool.removeBlankSpace(txt)
    print "NEW:",new
    exit(0)
    if 0==1:
        sheet_name = 'Change Requests'
        wb = load_workbook(filename = join('template','export_CR_list_template_2.xlsx'))
        if wb is not None:
            ws = wb[sheet_name]
            filename = None
            if ws is not None:
                CheckLLR.putLogo(ws)
                style_border = Style(border=Border(
                    left=Side(border_style=BORDER_THIN),
                    right=Side(border_style=BORDER_THIN),
                    top=Side(border_style=BORDER_THIN),
                    bottom=Side(border_style=BORDER_THIN)),
                                     alignment=Alignment(wrap_text=True,shrink_to_fit=True))
                row = 9
                hyperlink = "http://www.lemonde.fr"
                cr = ("TEST","BETA")
                Style.setCell(ws,cr,row,1,style_border)
                Style.setHyperlink(ws,row,1,hyperlink)
                filename = "Test_Hyperlink_%d.xlsx" % floor(time.time())
                wb.save(join("result",filename))
        exit()
    test = ThreadQuery()
    if 0==0:
        tbl_cr = [['SACR 0168', 'Evolution', 'Safety attribute update according to SAQ 345', 'Under_Verification', 'Minor', 'S0', 'S1', 'EQT_ACLOG_02_04', '28/01/16 15:18', u'REQ_SAFETY attribute of derived requirements.  \n\n', '2nd Release ISTCR', u'N/A\n\n', '', '', '', '', '', '', '', ''], ['SACR 0179', 'Defect', 'Inverted logic for AEC contactor', 'Under_Verification', 'Major', 'S0', 'S0', 'Power On ESTR : EQT_ACLOG_02_01', '28/01/16 11:11', u'**The modification will impact:  \n  \nSpecifications:  \n\n- SSCS,\n\n- SWRD,\n\n- HWRD (will be take into acount in the issue 2),\n\n- HSID (will be take into acount in the issue 1)  \n  \nVerification:  \n-Informal test (log book)  \n  \n  \n\nSIRD should be updated to (SIRD 2303)\n\n', 'Power On ESTR', u'The AEC contactor is OPEN when Closure is needed and CLOSED when opening is\nneeded.\n\n', '', '', '', '', '', '', '', ''], ['SACR 0181', 'Evolution', 'AC External power function modification (removal WOW condition and PINF bug correction)', 'Under_Verification', 'Minor', 'S0', 'S0', 'Power On ISTCR : EQT_ACLOG_02_02', '28/01/16 11:17', u'**Assembly Board impact : 400CE06L01Y02\n\n> SW impact : yes  \nHW impact : none  \n  \n**SSCS impact : ET2923-S issue 3\n>\n\n>  \n\n>\n\n> **HSID impact : none\n\n>\n\n> **ETPR impact verification : none\n\n>\n\n>  \n**ATP impact : none\n\n', 'Power On ISTCR', u'functional impact:\n\n- AC External power cannot be used aircraft on jack\n\n- DSO_AC_EP_PINF information is not stable\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0196', 'Evolution', 'TCB statuses inverted on the CAN', 'Under_Verification', 'Minor', 'S0', 'S0', 'Power On ISTCR : EQT_ACLOG_02_02', '28/01/16 11:42', u'**Assembly Board impact : 400CE06L01Y03\n\n> SW impact : yes  \nHW impact : none\n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : ET2982-S issue 1\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **ATP impact : none\n\n', 'Power On ISTCR', u'Wrong TCB statuses on the CAN from ACLOG  \n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0209', 'Defect', 'ATCs Lockout', 'Under_Verification', 'Major', 'S0', 'S0', 'Power On ISTCR : EQT_ACLOG_02_02', '28/01/16 11:43', u'ACLOG will compute a wrong lockout status from GCU and the command of ATC\ncontactors will be impacted.\n\n  \n**Specification:\n\n- SSCS impact (SSCS_ACLog_632, SSCS_ACLog_634, SSCS_ACLog_636)\n\n- SWRD impact : requirements using this logic  \n  \n**Verification:  \n- informal test (log book)  \n  \n  \n\n', 'Power On ISTCR', u'ACLOG will compute a wrong lockout status from GCU and the command of ATC\ncontactors will be impacted.\n\n', '', '', '', '', '', '', '', ''], ['SACR 0223', 'Evolution', 'AC external Power protections impact and reset', 'Under_Verification', 'Major', 'S0', 'S0', 'Power On ESTR :EQT_ACLOG_02_01', '28/01/16 11:14', u'**Assembly Board impact : 400CE06L01Y03\n\n> SW impact : yes  \nHW impact : none\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : ET2982-S issue 1\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact only P/N change)\n\n', 'Power On ESTR', u'So, if an ACEP protection occurs,\n\n  \n\n- We will put AC External Power = NOT Power Ready.\n\n- This will open ELC and switch the AC External Power Mode (or Ground Servicing Mode) from ACTIVE to NOT ACTIVE.\n\n- So, we will reset the AC External Power protection.\n\n- And, if we reset the AC External Power protection, AC External Power will switch to POWER READY state.\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0247', 'Evolution', 'Voltage drop in the DSO 28V/OPEN in the 28V state', 'Under_Verification', 'Minor', 'S0', 'S0', 'Power On ISTCR : EQT_ACLOG_02_02', '28/01/16 11:44', u'**Assembly Board impact : 400CE06L01Y03\n\n> SW impact : none  \nHW impact : yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', 'Power On ISTCR', u'The minimum voltage of 17V at the output of the MDSO and HDSO may not be met\n(because of the voltage drop in the diodes and MOS @200mA - MDSO 28V/OPEN and\n4.5V - HDSO 28V/OPEN)\n\n', '', '', '', '', '', '', 'Yes', ''], ['SACR 0272', 'Defect', 'ASI_RCCBx_CURRENT_PHy interfaces update due to transformer burden resistors update', 'Under_Verification', 'Major', 'S0', 'S0', 'EQT_ACLOG_02_02 : Batch Power On ISTCR Delivery - EQT_ACLOG_02_04 : implemented on G7000_SSCS_ACLOG_ET2923_S issue 4', '28/01/16 11:10', u'**Assembly Board impact : 400CE06L01Y02\n\n> SW impact : yes (only software prototype impacted : ECE35-A447-0101 / CRC\n0x5cadf84e --> No SCR attached because this software is not under DO-178\nprocess)\n\n>\n\n>> HW impact : none\n\n>>\n\n>>  \n\n>>\n\n>> **SSCS impact specification : ET2923-S issue 3\n\n>>\n\n>>  \n\n>>\n\n>> **HSID impact : ET2982-S issue 1\n\n>>\n\n>> **\n\n>>\n\n>> **ETPR impact verification : none  \n\n>>\n\n>> **\n\n>>\n\n>> **ATP impact : none\n\n', 'Power On ISTCR', u'ACMP i2t timing cannot trip a quick as specifed for high current\n\n', '', '', '', '', '', 'Yes', '', ''], ['HCR 0275', 'Defect', 'ACLOG : Contactors TVS unidirectional diodes forbids the good discharge of the contactor coils.', 'Under_Verification', 'Major', 'S0', 'S0', 'HW_ACLOG_02_00 (Power On ISTCR)', '18/03/16 10:43', u'**Harware PN impact :\n\n400CE06L00Y02\n\n400CE06L01Y02\n\n674CE06Y02\n\n  \n\n**HwRD Impact :\n\nNone\n\n  \n\n**HSID Impact :\n\nNone\n\n  \n\n**HwDD Impact :\n\nNone\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y00\n\n  \n\n**HWVPR Impact :\n\nNone\n\n', '<void>', u'the contactors are not drive correctlly\n\n', '', '', '', '', '', '', '', ''], ['SACR 0285', 'Defect', 'HW time response to include in confirmation time for protections', 'Under_Verification', 'Major', 'S0', 'S0', 'Power On ISTCR : EQT_ACLOG_02_02', '28/01/16 12:26', u'**Assembly Board impact : 400CE06L01Y02\n\n> SW impact : yes  \nHW impact : none\n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **ATP impact : none\n\n', 'Power On ISTCR', u'protections delay time should be out of range\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0286', 'Evolution', 'ACLOG Compatibility HW/SW', 'Under_Verification', 'Major', 'S0', 'S0', 'Power On ISTCR : EQT_ACLOG_02_02', '28/01/16 12:28', u'**Assembly Board impact : 400CE06L01Y03\n\n> SW impact : yes  \nHW impact : none  \n  \n**SSCS impact : ET2923-S issue 3\n>\n\n>  \n\n>\n\n> **HSID impact : ET2982-S issue 1\n\n>\n\n> **ETPR impact verification : none\n\n>\n\n>  \n**ATP impact : ATP100841 issue 3\n\n', 'Power On ISTCR', u'1) SW should change is P/N root change\n\n  \n\n2) ACLOG HW/SW compatibility is too restrictive\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0292', 'Defect', 'AC EPC1&2 ATCx contactor cannot be closed when 115V is present (antiparalleling protection trouble)', 'Under_Verification', 'Major', 'S0', 'S0', 'Power On ISTCR : EQT_ACLOG_02_02', '18/03/16 12:25', u'**Assembly Board impact : 400CE06L01Y02\n\n> SW impact : yes  \nHW impact : none  \n  \n**SSCS impact : ET2923-S issue 3\n>\n\n>  \n\n>\n\n> **HSID impact : none\n\n>\n\n> **ETPR impact verification : none\n\n>\n\n>  \n**ATP impact : none\n\n', 'Power On ISTCR', u'Unable to connect External Power when One Gen is ON  \n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0302', 'Evolution', 'DCLOG NBPT timing not compatible with network reconfiguration (Break power transfer)', 'Under_Verification', 'Minor', 'S0', 'S0', 'Power On ISTCR : EQT_ACLOG_02_02', '28/01/16 12:42', u'**Assembly Board impact : 400CE06L01Y03\n\n> SW impact : yes  \nHW impact : none\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', 'Power On ISTCR', u'the AC configuration changes occurs then DC contactor are not closed for NBPT\n\n', '', '', '', '', '', 'Yes', '', ''], ['HCR 0323', 'Evolution', 'Voltage drop in the DSO 28V/OPEN in the 28V state', 'Under_Verification', 'Minor', 'S0', 'S0', 'HW_ACLOG_02_05 (EQT_ACLOG_02_08)', '27/01/16 11:01', u'**Harware PN impact : None\n\n  \n\n**HWRD Impact : ET3598-S issue 1\n\n  \n\n**HSID Impact :** None\n\n  \n\n**HWDD Impact :** None\n\n  \n\n**Schematics Impact :** None\n\n  \n\n**HWVPR Impact : ET3938-V issue 1D2\n\n', '<void>', u'The minimum voltage of 17V at the output of the MDSO and HDSO may not be met\n(because of the voltage drop in the diodes and MOS @200mA - MDSO 28V/OPEN and\n4.5V - HDSO 28V/OPEN)\n\n', '', '', '', '', '', '', '', ''], ['SACR 0397', 'Evolution', 'ACLOG AC EXT power overcurrent trip curve', 'Under_Verification', 'Major', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '28/01/16 14:42', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes  \nHW impact : none\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No impact except the P/N change)\n\n', '2nd Release ISTCR', u'Evolution of the function (therefore no limitation)\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0398', 'Evolution', 'ACLOG ACMP GFI protection', 'Under_Verification', 'Major', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '29/02/16 14:08', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes  \nHW impact : yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : ET2982_S issue 1\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No impact except the P/N change)\n\n', '2nd Release ISTCR', u'No limitation\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 0399', 'Evolution', 'ACLOG AC EXT phase open', 'Under_Verification', 'Major', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '28/01/16 14:45', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes  \nHW impact : none\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No impact except the P/N change)\n\n', '2nd Release ISTCR', u'TBD\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0402', 'Evolution', 'Chattering', 'Under_Verification', 'Major', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '28/01/16 14:47', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes  \nHW impact : none\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No impact except the P/N change)\n\n', '2nd Release ISTCR', u'The contactor will chatter without restriction\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0432', 'Defect', 'ASI Voltage ACEP precision is not conform', 'Under_Modification', 'Minor', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '01/02/16 10:18', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes (only software prototype impacted : ECE35-A447-0102 / CRC\n0xbbcc173d --> No SCR attached because this software is not under DO-178\nprocess)  \nHW impact : yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : ET2982_S issue 1 (no HW/SW compatibility impact ; only\nvoltage ratio improvement)\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No impact except the P/N change)\n\n', '2nd Release ISTCR', u'Voltage send on CAN busses is out of range\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 0433', 'Defect', 'PINF is not maintained to into active state during PBIT', 'Under_Verification', 'Minor', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '18/03/16 12:04', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes  \nHW impact : yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : ET2982-S issue 1\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No impact except the P/N change)\n\n', '2nd Release ISTCR', u'AC External power Push Button of the ground service panel shall be kept\npressed during more than 700ms to have the external power available.\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 0434', 'Defect', 'NBPT is generated upon ground / flight transition', 'Under_Verification', 'Minor', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '28/01/16 15:00', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes (only software prototype impacted : ECE35-A447-0102 / CRC\n0xbbcc173d --> No SCR attached because this software is not under DO-178\nprocess)  \nHW impact : none\n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No impact except the P/N change)\n\n', '2nd Release ISTCR', u'all DC CTC are closed because of the NBPT signal generation on ground / flight\ntransition\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0446', 'Evolution', 'Overfrequency and Underfrequency AC EXT', 'Under_Verification', 'Major', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '28/01/16 15:07', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes  \nHW impact : none\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No impact except the P/N change)\n\n', '2nd Release ISTCR', u'performance impact\n\n', '', '', '', '', '', 'Yes', '', ''], ['HCR 0455', 'Defect', 'ASI Voltage ACEP precision is not conform', 'Under_Verification', 'Minor', 'S0', 'S0', 'HW_ACLOG_02_01 (2nd release ISTCR)', '01/02/16 10:09', u'**Harware PN impact :\n\n400CE06L00Y03\n\n674CE06Y03\n\n  \n\n**HwRD Impact :\n\nNone\n\n  \n\n**HwDD Impact :\n\nNone\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y00\n\n  \n\n**HWVPR Impact :\n\nNone\n\n', '<void>', u'Voltage send on CAN busses is out of range\n\n', '', '', '', '', '', '', '', ''], ['SACR 0457', 'Evolution', 'Removal of AC External Power phase-to-phase short protection (unbalance)', 'Under_Verification', 'Minor', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '28/01/16 15:10', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes  \nHW impact : yes, but no HCR associeted because no requirement are impacted\nonly synoptic and board description\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No impact except the P/N change)\n\n', '2nd Release ISTCR', u'N/A\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 0460', 'Evolution', 'CB mapping on CAN network shall be clearly defined and homogenous in all the reference documents.', 'Under_Verification', 'Minor', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '28/01/16 15:12', u"**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes (only software prototype impacted : ECE35-A447-0102 / CRC\n0xbbcc173d --> No SCR attached because this software is not under DO-178\nprocess, in SWRD it's already taken into account)  \nHW impact : none\n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n", '2nd Release ISTCR', u'Evolution of requirement from SyCR and software prototype error to correct\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0461', 'Evolution', 'AC protection functional reaction time delayed 50ms', 'Under_Verification', 'Minor', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '28/01/16 15:14', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : yes  \nHW impact : none\n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release ISTCR', u'50ms added in all the reaction time of the protections.\n\n', '', '', '', '', '', 'Yes', '', ''], ['HCR 0464', 'Evolution', 'ACLOG ACMP GFI protection', 'Under_Verification', 'Major', 'S0', 'S0', 'HW_ACLOG_02_05 (EQT_ACLOG_02_08)', '28/01/16 15:56', u'**Harware PN impact :\n\n400CE06L00Y03\n\n674CE06Y03\n\n675CE06Y02\n\n  \n\n**HwRD Impact :\n\nET3598-S issue 1\n\n  \n\n**HwDD Impact :\n\nNone\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y00\n\nF0214-CE0005354Y00\n\n  \n\n**HWVPR Impact :\n\nNone\n\n  \n\n', '<void>', u'No limitation\n\n', '', '', '', '', '', '', '', ''], ['HCR 0465', 'Defect', 'PINF is not maintained to into active state during PBIT', 'Under_Verification', 'Minor', 'S0', 'S0', 'HW_ACLOG_02_03 (EQT_ACLOG_02_06)', '01/02/16 12:03', u'**Harware PN impact :\n\nnone (pull-up already present)\n\n  \n\n**HwRD Impact : ET3598-S issue 2\n\n  \n\n**HwDD Impact :\n\nnone\n\n  \n\n**Schematics Impact :\n\nnone\n\n  \n\n**HWVPR Impact :\n\nnone\n\n', '<void>', u'A pull up is present on the ACLOG PCB, but nothing to specify it\n\n', '', '', '', '', '', '', '', ''], ['SACR 0489', 'Defect', 'conducted Emission not compliant', 'Under_Verification', 'Major', 'S0', 'S0', '2nd release ISTCR : EQT_ACLOG_02_03', '18/03/16 12:24', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : none  \nHW impact : yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', '2nd Release ISTCR', u'The ACLOG power supply is not complient to the requirements\n\n', '', '', '', '', '', '', 'Yes', ''], ['HCR 0490', 'Defect', 'conducted Emission not compliant', 'Under_Verification', 'Major', 'S0', 'S0', 'HW_ACLOG_02_01 (2nd release ISTCR)', '10/03/16 12:08', u'**Harware PN impact :\n\n400CE06L00Y03\n\n674CE06Y03\n\n  \n\n**HwRD Impact :\n\nnone\n\n  \n\n**HwDD Impact :\n\nNone\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y00\n\n  \n\n**HWVPR Impact :\n\nNone\n\n', '<void>', u'The ACLOG power supply is not complient to the requirements\n\n', '', '', '', '', '', '', '', ''], ['SACR 0492', 'Defect', 'CT PBIT fail', 'Complementary_Analysis', 'Major', 'S1', 'S2', '<void>', '24/05/16 12:34', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : ET2982-S issue 1\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n>\n\n>  \n\n>\n\n>  \n\n>\n\n>  \n\n>\n\n> **_Complementary analysis :\n\n>\n\n> **\n\n>\n\n> **Assembly Board impact : 400CE06L07Y05\n\n>\n\n>> SW impact : YES  \nHW impact : YES\n\n>>\n\n>>  \n\n>>\n\n>> **SSCS impact specification : none\n\n>>\n\n>> **HSID impact : ET2982-S issue 2D4\n\n>>\n\n>> **\n\n>>\n\n>> **ETPR impact verification : none  \n\n>>\n\n>> **\n\n>>\n\n>> **ATP impact : ATP101005-09 issue 2\n\n', 'Power On FTV4', u'CT PBIT is failed even if CT is present\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 0517', 'Defect', 'Pin E voltage activation at 16V', 'Under_Verification', 'Minor', 'S0', 'S1', 'EQT_ACLOG_02_06', '18/03/16 13:17', u'**Assembly Board impact : 400CE06L04Y04\n\n> SW impact : yes  \nHW impact : yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : ET2982_S issue 1\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power ON FTV1 VFG', u'- PINE is declared present out of tolerence\n\n- PINE overvoltage is to much selective and out of range\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 0527', 'Evolution', 'Check the MCU silicon revision', 'Under_Verification', 'Minor', 'S0', 'S1', 'EQT_ACLOG_02_05', '18/03/16 13:28', u'**Assembly Board impact : 400CE06L02Y03\n\n> SW impact : none  \nHW impact : none\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : ET2982_S issue 1\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP100841 issue 4\n\n', 'Power ON Batch 3', u'No compatibility between the SW and the HW if a wrong silicon revision is\npresent\n\n', '', '', '', '', '', '', '', ''], ['SACR 0530', 'Defect', 'P/N send on CAN BUS from NVM is wrong after SW update', 'Under_Verification', 'Minor', 'S0', 'S1', 'EQT_ACLOG_02_05', '18/03/16 13:32', u'**Assembly Board impact : 400CE06L03Y04\n\n> SW impact : None  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : None\n\n>\n\n> **HSID impact : None\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : None\n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP100841 issue 5\n\n', 'Power ON Batch 3', u'The ACLOG part number send on CAN busses is not conform to the definition\nafter SW update\n\n', '', '', '', '', '', '', '', ''], ['SACR 0532', 'Defect', 'CE0007617Y00 sticker not conform to the temperature requirement', 'In_Analysis', '<void>', 'S0', '<void>', '<void>', '11/05/15 10:27', '\n\n', '<void>', u'TBD\n\n', '', '', '', '', '', '', '', ''], ['SACR 0537', 'Defect', 'AC External Power Not conform to SES-ACEPC-1017 : ACEP reset protection by the cockpit EXT AC selector switch', 'In_Analysis', 'Major', 'S1', 'S2', '<void>', '24/03/16 10:47', u'Cf. SACR 1457\n\n', 'Power On FTV4', u'Not conform to the behavior defined in SIRD\n\n', '', '', '', '', '', '', '', ''], ['SACR 0540', 'Evolution', 'Clarification of the reset of ACMP protections', 'Rejected', '<void>', 'S0', 'S1', '<void>', '26/05/15 10:57', u'rejected\n\n', 'Power On FTV1', u'evolution\n\n', '', '', '', '', '', '', '', ''], ['SACR 0541', 'Defect', 'ACMP RCCB computed EDMU command is not correct', 'Under_Verification', 'Major', 'S0', 'S1', 'EQT_ACLOG_02_04 (Baseline for ISTCR batch 2 + ACMP evolution (SACR 541))', '18/03/16 12:21', u'**Assembly Board impact : 400CE06L03Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 3\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power ON Batch 3', u'The RCCB ACMP command is not possible in this state\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0576', 'Evolution', 'AC Source paralleling condition with single or multiple failures', 'Under_Verification', 'Blocking', 'S0', 'S1', 'EQT_ACLOG_02_05', '18/03/16 13:38', u'**Assembly Board impact : 400CE06L03Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power ON Batch 3', u'Possible antiparraleling conditions with single or multiple failures\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0619', 'Evolution', 'Validity and refresh time of CB matrix', 'Under_Verification', 'Minor', 'S0', 'S1', 'EQT_ACLOG_02_09', '18/03/16 13:48', u'**Assembly Board impact : 400CE06L04Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : ET2982_S issue 1\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power ON FTV1 VFG', u'ACEPC TCB status could be invalid\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0643', 'Evolution', 'Pin E presence logic', 'Under_Modification', 'Major', 'S1', 'S1', '<void>', '30/11/15 15:49', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D2\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release First Flight FTV1', u'External AC power can be disconnected during ACMP start.\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0649', 'Evolution', 'NBPT activation following ACMP modification', 'Under_Verification', 'Major', 'S0', 'S1', 'EQT_ACLOG_02_06', '18/03/16 14:09', u'**Assembly Board impact : 400CE06L04Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power ON FTV1 VFG', u'DC might not be supplied during ACMP modifications.\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0652', 'Evolution', 'GLC AUX Failures to be sent on CAN', 'Under_Verification', 'Minor', 'S0', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:11', u'**Assembly Board impact : 400CE06L04Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'NA\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0671', 'Evolution', 'Informations concerning Ext AC source on CAN buses are not correct', 'Under_Verification', 'Minor', 'S0', 'S1', 'EQT_ACLOG_02_06', '18/03/16 14:13', u'**Assembly Board impact : 400CE06L04Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power ON FTV1 VFG', u'Synoptic display is not correct\n\n', '', '', '', '', '', 'Yes', '', ''], ['HCR 0680', 'Evolution', 'Pin E voltage activation at 16V', 'Under_Verification', 'Minor', 'S0', 'S1', 'HW_ACLOG_02_03 (EQT_ACLOG_02_06)', '01/02/16 16:33', u'**Harware PN impact :\n\nnone\n\n**HwRD Impact :\n\nET3598_S Issue2\n\n  \n\n**HwDD Impact :\n\nNone\n\n  \n\n**Schematics Impact :\n\nnone\n\n  \n\n**HWVPR Impact :\n\nNone\n\n', '<void>', u'- PINE is declared present out of tolerence\n\n', '', '', '', '', '', '', '', ''], ['SACR 0711', 'Evolution', 'Not compliant and TBD/TBC requirements from HWRD', 'In_Analysis', '<void>', 'S1', '<void>', '<void>', '16/07/15 13:47', '\n\n', '<void>', u'Cf. attach file\n\n', '', '', '', '', '', '', '', ''], ['SACR 0717', 'Evolution', 'AC EP 115V presence threshold', 'Under_Verification', 'Major', 'S1', 'S1', 'EQT_ACLOG_02_06', '18/03/16 14:13', u'**Assembly Board impact : 400CE06L04Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power ON FTV1 VFG', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0722', 'Defect', 'HW overvoltage stimuli are wrong in HSID ', 'Under_Verification', 'Minor', 'S1', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:12', u'**Assembly Board impact : 400CE06L04Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : ET2982-S issue 1\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'HW overvoltage is not correctely defined\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0723', 'Evolution', 'SN of boards is reported on CAN bus', 'Under_Modification', 'Minor', 'S0', 'S2', '<void>', '13/05/16 10:45', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : None  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : ET2982-S issue 2D4\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP101005_09 issue 2\n\n', 'Power On FTV4', u'SN of boards is not reported to the avionic.\n\n', '', '', '', '', '', '', '', ''], ['SACR 0727', 'Evolution', 'CTC command failure', 'In_Analysis', 'Major', 'S0', 'S1', '<void>', '28/04/16 12:04', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '<void>', u'none\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0745', 'Evolution', 'Ground / Flight transition when all inputs are open', 'Under_Verification', 'Major', 'S1', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:07', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'N/A\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0751', 'Evolution', 'AC BUS is detected Not Avail even if the fault time is too short', 'Under_Verification', 'Minor', 'S0', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:07', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'N/A\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0770', 'Evolution', 'behaviour when PBIT failed', 'Under_Verification', 'Major', 'S0', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:15', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'The Log does not start when the PBIT is failed : No communication performed\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0896', 'Evolution', 'timing protections evolution', 'Under_Verification', 'Major', 'S0', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:01', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'N/A\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 0898', 'Evolution', 'Aux Chattering failure shall only be raised when the CTC is commanded Closed', 'Under_Verification', 'Major', 'S1', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:15', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'None\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0900', 'Evolution', 'Supply source of bus bar computation', 'Under_Verification', 'Major', 'S1', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:16', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'Shedding not performed with the right information\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0903', 'Evolution', 'Several part of the Electrical Synoptic displayed in Magenta on the ISTCR', 'Under_Verification', 'Major', 'S0', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:06', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'Synoptic displayed incorrectly\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0910', 'Evolution', 'RCCB "open lock" switch to "not open lock" after a restart of the system', 'Under_Modification', 'Major', 'S1', 'S1', '<void>', '10/09/15 10:14', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : None\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'The lock open of the RCCBs has to be done at each restart of the EPDS.\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0911', 'Defect', 'Correct the ACMP1B command', 'Under_Modification', 'Major', 'S1', 'S1', '<void>', '10/09/15 16:18', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'ACMP1B is not correctly comanded\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 0943', 'Evolution', 'Update specified maximum weight of ACLOG', 'Under_Verification', 'Minor', 'S1', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:04', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : None  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP101005-03 issue 1\n\n', 'First Flight FTV1', u'Evolution\n\n', '', '', '', '', '', '', 'Yes', ''], ['HCR 0958', 'Evolution', 'update specified maximum weight of ACLog', 'Under_Verification', 'Minor', 'S1', 'S1', 'HW_ACLOG_02_06 (EQT_ACLOG_02_09)', '18/03/16 11:05', u'**Harware PN impact : None\n\n  \n\n**HWRD Impact : ET3598-S issue 3\n\n  \n\n**HSID Impact :** None\n\n  \n\n**HWDD Impact :** None\n\n  \n\n**Schematics Impact :** None\n\n  \n\n**HWVPR Impact : None\n\n', '<void>', u'N/A\n\n', '', '', '', '', '', '', '', ''], ['HCR 0960', 'Evolution', 'Timing protections evolution - ACLOG', 'Under_Verification', 'Major', 'S0', 'S1', 'HW_ACLOG_02_06 (EQT_ACLOG_02_09)', '18/03/16 11:05', u'**Harware PN impact : None\n\n  \n\n**HWRD Impact : ET3598-S issue 3\n\n  \n\n**HSID Impact :** None\n\n  \n\n**HWDD Impact :** None\n\n  \n\n**Schematics Impact :** None\n\n  \n\n**HWVPR Impact : None\n\n', '<void>', u'Timing protections evolution on the Hardware specification.\n\n', '', '', '', '', '', '', '', ''], ['SACR 0988', 'Evolution', 'CAN IRD applicable is now G7000_EPGDS_CAN_IRD_7N-16560_2.1', 'Under_Verification', 'Major', 'S1', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:17', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1019', 'Evolution', 'CAN IRD 3.0', 'Under_Modification', 'Enhancement', 'S1', 'S1', '<void>', '30/11/15 16:01', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D2\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release First Flight FTV1', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1046', 'Evolution', 'IBIT request shall not be accepted during a reconfiguration', 'Under_Verification', 'Minor', 'S1', 'S1', 'EQT_ACLOG_02_08', '18/03/16 14:18', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : None (SW is already complient with this evolution)  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 4\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'First Flight FTV1', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['SACR 1074', 'Evolution', 'ACMP CLOSURE SEQUENCE AND TIMING NOK & ACMP default state in ground servicing mode and at power up', 'In_Review', 'Major', 'S1', 'S2', '<void>', '19/05/16 09:12', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1075', 'Evolution', 'Contactor fail open and fail closed reaction', 'Under_Modification', 'Major', 'S1', 'S1', '<void>', '24/02/16 13:48', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D2\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release First Flight FTV1', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1076', 'Evolution', 'Confirmation time of the CAS messages BUS AVAIL and BUS FAIL', 'Under_Modification', 'Minor', 'S1', 'S1', '<void>', '02/12/15 12:03', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D2\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release First Flight FTV1', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1079', 'Evolution', 'ELC opening in case of differential protection activation', 'Under_Modification', 'Major', 'S1', 'S1', '<void>', '30/11/15 16:16', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D2\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release First Flight FTV1', u'The differential leakage is still powered by the GPU.\n\n', '', '', '', '', '', 'Yes', '', ''], ['HCR 1097', 'Defect', 'DSO 28v/Open commanded when supplied with a high overvoltage power source', 'Under_Verification', 'Major', 'S1', 'S1', 'HW_ACLOG_02_05 (2nd release First Flight FTV1)', '18/03/16 10:44', u'**Harware PN impact : \n\n400CE06L00Y04\n\n674CE06Y04\n\n  \n\n**HWRD Impact : None\n\n  \n\n**HSID Impact :** None\n\n  \n\n**HWDD Impact :** None\n\n  \n\n**Schematics Impact : \n\nF0214-CE0005344Y02\n\n  \n\n## **HWVPR Impact : None\n\n', '<void>', u'During overvoltage, all DSO 28V/Open are close with a command from CPU at\nOpen.\n\n', '', '', '', '', '', '', '', ''], ['SACR 1155', 'Evolution', 'IBIT and timing before contactor closure', 'Under_Modification', 'Minor', 'S1', 'S1', '<void>', '02/12/15 12:05', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D2\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release First Flight FTV1', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1159', 'Evolution', 'ACMP timings when out of sequencing', 'Under_Modification', 'N/A', 'S1', 'S1', '<void>', '02/12/15 12:12', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D2\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release First Flight FTV1', u'Doc only\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1173', 'Evolution', 'Differential protection tolerance', 'Under_Modification', 'Enhancement', 'S1', 'S1', '<void>', '30/11/15 16:23', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D2\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release First Flight FTV1', u'differential protection testability is not consistent\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1174', 'Defect', 'AClog reaction when supplied with a high overvoltage power source', 'Under_Modification', 'Major', 'S1', 'S1', '<void>', '30/11/15 16:31', u'**Assembly Board impact : 400CE06L06Y04\n\n> SW impact : No  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP101005-05 issue 2\n\n', '2nd Release First Flight FTV1', u'During overvoltage, all HDSO 28V/Open are close with a command from CPU at\nOpen.\n\n', '', '', '', '', '', '', 'Yes', ''], ['HCR 1290', 'Defect', 'DSI GND/OPEN function improvement', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '05/04/16 14:26', u'**Harware PN impact :\n\n400CE06L00Y05\n\n674CE06Y05\n\n675CE06Y02\n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\n  \n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y02  \nF0214-CE0005354Y00\n\n  \n\n**HWVPR Impact :\n\nHWVPR_G7000_ACLog_ET3938-V issue 1 (procedure not impacted, only results)\n\n', '<void>', u'  * The DSI could be damaged in case of lightning.\n  * The DSI could perform false detection due to the lack of precision.\n\n', '', '', '', '', '', '', '', ''], ['SACR 1292', 'Evolution', 'N/A', 'In_Analysis', 'Minor', 'S1', 'S2', '<void>', '25/05/16 16:44', u'N/A\n\n', 'Power On FTV4', u'N/A\n\n', '', '', '', '', '', '', '', ''], ['HCR 1294', 'Defect', 'MDSO GND/OPEN RESET_TRIP signal improvement', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '05/04/16 14:28', u'### **Harware PN impact :  \n675CE06Y02\n\n###  \n**HWRD Impact :   \nNone  \n  \n**HWDD Impact :**   \nNone\n\n###  \n**Schematics Impact :  \nF0214-CE0005354Y00\n\n###  \n**HWVPR Impact :  \nNone\n\n', '<void>', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['HCR 1298', 'Defect', 'Internal Power supply monitoring precision improvement', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '18/05/16 15:01', u'**Harware PN impact :\n\n400CE06L00Y05\n\n674CE06Y05\n\n675CE06Y02\n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\n  \n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y02  \nF0214-CE0005354Y00\n\n  \n\n**HWVPR Impact :\n\nHWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'It can cause false failure detection\n\n', '', '', '', '', '', '', '', ''], ['HCR 1314', 'Defect', 'BITE Loss 5s not conform (Timing higher than 6s)', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '11/03/16 11:02', u'### **Harware PN impact :  \n400CE06L00Y05  \n674CE06Y05  \n675CE06Y02  \n  \n**HWRD Impact :   \nNone  \n  \n**HWDD Impact :**   \nG7000_ACLOG_HWDD_ET4115_E_Issue1D2  \n  \n**Schematics Impact :  \nF0214-CE0005344Y02  \n  \n**HWVPR Impact :  \nNone  \n\n', '<void>', u'The Loss 5s PBIT not conform\n\n', '', '', '', '', '', '', '', ''], ['HCR 1320', 'Defect', 'Evolution of ACLOG Internal power supplies', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '04/05/16 11:41', u'### **Harware PN impact :  \n910CE06L00Y00  \n911CE06Y00  \n  \n**HWRD Impact :   \nNone  \n  \n**HWDD Impact :**   \nG7000_ACLOG_HWDD_ET4115_E_Issue1D3  \n  \n**Schematics Impact :  \nF0214-CE0007720Y00  \n  \n**HWVPR Impact :  \nNone\n\n', '<void>', u'- None\n\n', '', '', '', '', '', '', '', ''], ['SACR 1333', 'Evolution', 'Hardware overvoltage protection of external power threshold', 'In_Review', 'Major', 'S1', 'S2', '<void>', '17/05/16 16:52', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP101005_09 issue 2\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['HCR 1344', 'Evolution', 'DSI GND/OPEN withstand continuously -30V', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '05/04/16 14:28', u'**Harware PN impact :\n\nNone\n\n**HWRD Impact : \n\nG7000_HWRD_ACLOG_ET3598_S_Issue 4\n\n**HWDD Impact :**\n\n  \n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n  \n\n**Schematics Impact :\n\nNone\n\n**HWVPR Impact :\n\nHWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'NA\n\n', '', '', '', '', '', '', '', ''], ['SACR 1345', 'Evolution', 'ACLog DSI GND/OPEN capacity to withstand minus 30Vdc', 'Under_Modification', 'Minor', 'S1', 'S2', '<void>', '03/05/16 14:20', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : None  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', '', 'Yes', ''], ['SACR 1346', 'Defect', 'Internal Power supply monitoring modification', 'Complementary_Analysis', 'Minor', 'S1', 'S2', '<void>', '19/05/16 09:21', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : YES  \nHW impact : YES\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : ET2982-S issue 2\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n**   \n**ATP impact : ATP101005_09 issue 2\n\n', 'Power On FTV4', u'It can cause false failure detection\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['HCR 1347', 'Defect', 'Internal fuse monitoring function precision imrpovement', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '05/04/16 14:28', u'**Harware PN impact :\n\n400CE06L00Y05\n\n674CE06Y05\n\n  \n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\n  \n\nG7000_ACLOG_HWDD_ET4115_E Issue1D2\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n**HWVPR Impact :\n\nHWVPR_G7000_ACLog_ET3938-V issue 1 (procedure not impacted, only results)\n\n', '<void>', u'It can cause false failure detection.\n\n', '', '', '', '', '', '', '', ''], ['HCR 1348', 'Defect', 'LED Power Supply Indicator', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '05/04/16 14:56', u'**Harware PN impact :\n\n400CE06L00Y05\n\n674CE06Y05\n\n  \n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\n  \n\nG7000_ACLOG_HWDD_ET4115_E Issue1D2\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n**HWVPR Impact :\n\nHWVPR_G7000_ACLog_ET3938-V issue 1 (procedure not impacted, only results)\n\n', '<void>', u'No really impact.When the Primary Power supply is shutdown, the LED stays ON\nuntil the power supply tank circuit is empty.\n\n', '', '', '', '', '', '', '', ''], ['HCR 1349', 'Evolution', 'Signal "TEMP_ADR02" Analog Input Microcontroller.', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '05/04/16 14:57', u'**Harware PN impact :\n\n400CE06L00Y05\n\n675CE06Y02\n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\n  \n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005354Y00\n\n  \n\n**HWVPR Impact :\n\nNone\n\n', '<void>', u'No really impact just clean up of the schematic\n\n', '', '', '', '', '', '', '', ''], ['HCR 1352', 'Evolution', 'Hardware overvoltage protection of external power threshold', 'In_Analysis', 'Major', 'S1', 'S2', '<void>', '23/05/16 16:29', u'**Harware PN impact :\n\n400CE06L00Y05\n\n675CE06Y02\n\n  \n\n**HWRD Impact : \n\nET3598-S issue 4\n\n**HWDD Impact :**\n\n  \n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005354Y00\n\n  \n\n**HWVPR Impact :\n\nHWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['HCR 1353', 'Evolution', 'Input EMI filter', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '05/04/16 14:57', u'**Harware PN impact :\n\n400CE06L00Y05\n\n674CE06Y05\n\n  \n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\n  \n\nG7000_ACLOG_HWDD_ET4115_E_Issue1\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n**HWVPR Impact :\n\nHWVPR_G7000_ACLog_ET3938-V issue 1 (procedure not impacted, only results)\n\n', '<void>', u'none.\n\n  \n\nWaiting for EMI test on new 910CE06L00Y00 board\n\n', '', '', '', '', '', '', '', ''], ['HCR 1354', 'Defect', 'Flyback Power Supply improvement', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '05/04/16 14:57', u'**Harware PN impact :\n\n400CE06L00Y05\n\n674CE06Y05\n\n  \n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\n  \n\nG7000_ACLOG_HWDD_ET4115_E_Issue1\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n**HWVPR Impact :\n\nHWVPR_G7000_ACLog_ET3938-V issue 1 (procedure not impacted, only results)\n\n', '<void>', u'None\n\n', '', '', '', '', '', '', '', ''], ['SACR 1356', 'Evolution', 'AC overvoltage protection on pin E of the External power', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '17/05/16 12:01', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n>  \n\n>\n\n> **ICD impact : G7000_ICD_ACLOG_Standard 1 issue 1.1\n\n>\n\n> **HSID impact : ET2982-S issue 2\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP101005-09 issue 2\n\n', 'Power On FTV4', u'No AC overvoltage protection on pin E of the External power\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 1358', 'Evolution', 'BITE Circuitry independance', 'Under_Modification', 'Minor', 'S1', 'S1', '<void>', '25/03/16 09:44', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : None  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D3\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', '2nd Release First Flight FTV1', u'Evolution\n\n', '', '', '', '', '', '', 'Yes', ''], ['SACR 1359', 'Defect', 'Difference between ACLOG P/N 400CE06L00Y05 and 910CE06L00Y00', 'In_Analysis', 'Major', 'S1', 'S1', '<void>', '25/05/16 18:01', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : None  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5 (ICD impact)\n\n>\n\n> **ICD impact : ICD G7K_ACLOG_Standard 1.1\n\n>\n\n>  \n\n>\n\n> **HSID impact : ET2982-S issue 2\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP101005_09 issue 2\n\n', 'Power On FTV4', u'**_Fonction impacted :\n\n  * ADUM is not conform to the recommendation of the datasheet --> No functional impact\n  * DSI GND/OPEN 1mA --> No functional impact (performance)\n  * DSI GND/OPEN 5mA --> No functional impact (performance)\n  * Internal fuse monitoring function precision --> No functional impact\n  * +12V and -12V Secondary Buck power supplies capacitors --> No functional impact\n  * LED Power Supply Indicator stays ON until the power supply tank circuit is empty : no functional impact\n  * Signal TEMP_ADR02 Analog Input Microcontroller should shows an instability (but not used) : no functional impact \n  * EMI filter spare capacitors and inrush limiter transistor change : no functional impact\n  * Flyback Power Supply improvement : no impact (Conducted emissions results need to be confirm)\n  * Front panel connector changed --> No functional impact\n  * LDSO GND connection --> No functional impact\n  * Inhib signal HDSO 28V/OPEN RCCB1 --> No functional impact (performance)\n  * Nexus (Pull Down) and debug connector not cable --> No functional impact\n  * ASI AC accuracy --> No functional impact (evolution)\n  * ASI DC accuracy --> No functional impact (evolution)\n  * HDSO/MDSO feedback lightning protection improvement --> No functional impact (performance)\n  * Spare lightning diode +28V power supply --> No functional impact (evolution)\n  * ASI_DC_spare1 need to be deleted on 910CE06L00Y00 (lake of place)\n  * Spare DSI_1mA_spare11 need to be deleted (lake of place)\n  * ASI_spare7 and ASI_spare8 need to be deleted (lake of place)\n\n', '', '', '', '', '', '', 'Yes', ''], ['HCR 1360', 'Evolution', 'BIT Circuitry independance', 'Under_Modification', 'Minor', 'S1', 'S1', '<void>', '25/03/16 09:44', u'### **Harware PN impact : None  \n  \n**HWRD Impact : ET3598-S issue 4  \n  \n**HWDD Impact :** None  \n  \n**Schematics Impact :** None  \n  \n**HWVPR Impact :   \n**HWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['SACR 1364', 'Evolution', 'Selectivity issue between ACMP Unbalanced and I2t protection', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '03/05/16 14:39', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'N/A\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1366', 'Evolution', 'CAN IRD 4.0', 'In_Review', 'Major', 'S1', 'S2', '<void>', '19/05/16 09:16', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1368', 'Evolution', 'Protection reset after cold start in flight', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '13/05/16 09:47', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'No protection reset after cold start in flight\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1372', 'Defect', 'The Overload AC external protection maximum current', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '19/05/16 09:17', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : None\n\n>\n\n> **HSID impact : ET2982-S issue 2\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP101005_09 issue 2\n\n', 'Power On FTV4', u'accuracy is not conform\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['HCR 1382', 'Defect', 'LDSO GND/OPEN lighting protection', 'In_Analysis', 'Major', 'S0', 'S2', '<void>', '05/04/16 14:58', u'**Harware PN impact :\n\n675CE06Y02\n\n674CE06Y05\n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\nNone\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n  \n\n**HWVPR Impact :\n\nNone\n\n', '<void>', u'The ligthning test not passed.\n\n', '', '', '', '', '', '', '', ''], ['HCR 1384', 'Defect', 'Inhib signal HDSO 28V/OPEN RCCB1', 'In_Analysis', 'Minor', 'S0', 'S2', '<void>', '06/04/16 11:10', u'**Harware PN impact : \n\n674CE06Y05\n\n  \n\n**HWRD Impact :\n\nNone\n\n  \n\n**HWDD Impact :**\n\nNone\n\n  \n\n**Schematics Impact : \n\nF0214-CE0005344Y02\n\n  \n\n**HWVPR Impact : \n\nNone\n\n', '<void>', u'No impact (Spare)\n\n', '', '', '', '', '', '', '', ''], ['HCR 1385', 'Evolution', 'Connector Front Panel board change and Mechanical modifications', 'In_Analysis', 'Minor', 'S0', 'S1', '<void>', '26/02/16 11:54', u'**Hardware PN impact :\n\n400CE06L00Y05\n\n674CE06Y05\n\n  \n\n**HWRD Impact : ET3598-S issue 4\n\n  \n\n**HWDD Impact :**\n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n**Mechanical Specification :\n\n[F0214-CE0005340Y01\n\n**HWVPR Impact :\n\n**HWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'- Not protected against condensation and humidity.\n\n- Spacers board evolution.\n\n', '', '', '', '', '', '', '', ''], ['HCR 1386', 'Evolution', 'Nexus (Pull Down) and debug connector', 'In_Analysis', 'Minor', 'S0', 'S1', '<void>', '29/02/16 08:59', u'**Harware PN impact : \n\n674CE06Y05\n\n  \n\n**HWRD Impact :\n\nNone\n\n  \n\n**HWDD Impact :**\n\nNone\n\n  \n\n**Schematics Impact : \n\nF0214-CE0005344Y02\n\n  \n\n**HWVPR Impact : \n\nNone\n\n', '<void>', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['HCR 1387', 'Defect', 'ASI AC evolution', 'In_Analysis', 'Minor', 'S0', 'S2', '<void>', '04/05/16 09:42', u'**Hardware PN impact :\n\n400CE06L00Y05\n\n674CE06Y05\n\n675CE06Y02\n\n  \n\n**HWRD Impact : \n\nNone\n\n  \n\n**HWDD Impact :**\n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n### **HWVPR Impact :  \n**HWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'The 115VAC not respect the limite on the requirement.\n\n', '', '', '', '', '', '', '', ''], ['HCR 1389', 'Evolution', 'ASI DC evolution', 'In_Analysis', '<void>', 'S0', 'S2', '<void>', '03/05/16 14:53', u'### **Hardware PN impact :  \n674CE06Y05  \n  \n**HWRD Impact :   \nET3598-S issue 5  \n  \n**HWDD Impact :**   \nG7000_ACLOG_HWDD_ET4115_E_Issue1D3  \n  \n**Schematics Impact :  \nF0214-CE0005344Y02  \n  \n**HWVPR Impact :  \n**HWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'Evolution.\n\n', '', '', '', '', '', '', '', ''], ['HCR 1390', 'Evolution', 'Feedback Diode HDSO/MDSO', 'In_Analysis', 'Minor', 'S0', 'S2', '<void>', '05/04/16 14:59', u'**Hardware PN impact :\n\n400CE06L00Y05\n\n674CE06Y05\n\n675CE06Y02\n\n  \n\n**HWRD Impact : \n\nNone\n\n  \n\n**HWDD Impact :**\n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n### **HWVPR Impact :  \n**None\n\n', '<void>', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['HCR 1392', 'Evolution', 'Spare lightning diode +28V power supply', 'In_Analysis', 'Minor', 'S0', 'S2', '<void>', '05/04/16 14:59', u'**Hardware PN impact :\n\n674CE06Y05\n\n  \n\n**HWRD Impact : \n\nNone\n\n  \n\n**HWDD Impact :**\n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n### **HWVPR Impact :  \n**None\n\n', '<void>', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['HCR 1393', 'Evolution', 'ACLOG modification for robustness under lightning test on BIT CT', 'In_Analysis', 'Minor', 'S0', 'S2', '<void>', '06/04/16 11:20', u'**Hardware PN impact :\n\n675CE06Y02\n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\n### **HWVPR Impact :  \n**HWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'- The BIT circuitery is damaged,\n\n- The DP protection is done for an upper current value than expected\n\n', '', '', '', '', '', '', '', ''], ['SACR 1394', 'Evolution', 'AEC command status on the IO state CAN message', 'Under_Modification', 'Minor', 'S1', 'S2', '<void>', '03/05/16 14:41', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'None\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1432', 'Evolution', 'FAILED CLOSED ACMP', 'Under_Modification', 'Minor', 'S1', 'S2', '<void>', '03/05/16 14:43', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5 (FDD impact)\n\n>\n\n> **FDD impact : FDD_GLOBAL_7000_ACLOG issue 2D1\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1437', 'Defect', 'BITE Loss 5s not conform (Timing higher than 6s)', 'Under_Modification', 'Minor', 'S1', 'S2', '<void>', '03/05/16 14:45', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : ET2982-S issue 2\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'The Loss 5s PBIT not conform\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 1440', 'Defect', 'ACLOG modification for robustness under lightning test', 'In_Review', 'Major', 'S1', 'S2', '<void>', '03/05/16 16:32', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : None (none for lightning but new design generate a SW impact :\nCf. SACR 492)  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none (none for lightning but new design generate a HSID\nimpact : Cf. SACR 492)\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'- The Bite circuitery is damaged\n\n- The DP protection is done for an upper current value than expected\n\n', '', '', '', '', '', '', 'Yes', ''], ['SACR 1442', 'Defect', 'CAN Perturbation when ACLOG is switched OFF and still plugged', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '19/05/16 11:23', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'Loss of communication with remaining ACLOG or DCLOG board  \nDegraded communication with EMERLOG if the failure occurs on the Side 2.\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['HCR 1444', 'Defect', 'CAN Perturbation when ACLOG is switched OFF and still plugged', 'In_Analysis', '<void>', 'S1', 'S2', '<void>', '02/05/16 13:49', u'**Harware PN impact :\n\n675CE06Y02\n\n400C06L07Y05\n\n912CE06Y00\n\n910C06L00Y00\n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\nNone\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\nF0214-CE0007720Y00\n\n  \n\n**HWVPR Impact :\n\nNone\n\n', '<void>', u'Loss of communication with remaining ACLOG or DCLOG board  \nDegraded communication with EMERLOG if the failure occurs on the Side 2.\n\n', '', '', '', '', '', '', '', ''], ['SACR 1457', 'Evolution', 'ACEP Evolutions', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '17/05/16 09:48', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> (and ICD impact)\n\n>\n\n> **HSID impact : ET2982-S issue 2\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : ATP101005-09 issue 1\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 1470', 'Defect', 'ACMP re-configuration  inconsistencies (software)', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '13/05/16 09:56', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', 'Power On FTV4', u'- SyCR 1297 : 2 ACMPs will close at the same time in case of loss of communication.\n\n- SyCR 1326: When starting APU GEN while only LGEN is online, the ACMP 1B and 3A loose power for less than 200ms and the APU GEN sees a high in rush current.\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1490', 'Evolution', 'RTCA/DO-160F Category Z', 'Under_Modification', 'Minor', 'S0', 'S2', '<void>', '03/05/16 15:03', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : None  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : None\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', '', 'Yes', ''], ['HCR 1492', 'Evolution', 'RTCA/DO-160F Category Z', 'In_Analysis', 'Minor', 'S0', 'S2', '<void>', '05/04/16 18:01', u'**Harware PN impact : \n\nNone\n\n  \n\n**HWRD Impact :\n\nET3598-S issue 4\n\n  \n\n**HWDD Impact :**\n\nG7000_ACLOG_HWDD_ET4115_E_Issue1D2\n\n  \n\n**Schematics Impact : \n\nNone\n\n  \n\n**HWVPR Impact : \n\n**HWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['SACR 1502', 'Evolution', 'EICAS and Synoptic equation Evolution', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '03/05/16 15:05', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'EICAS and Synoptic equation Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1508', 'Defect', 'HW Overvolatge function Zener diode polarisation', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '03/05/16 15:12', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : None  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : None\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'No limitation\n\n', '', '', '', '', '', '', 'Yes', ''], ['HCR 1509', 'Defect', 'Zener diode polarisation of HW Overvolatge function', 'In_Analysis', '<void>', 'S1', 'S2', '<void>', '02/05/16 11:04', u'**Harware PN impact :\n\n675CE06Y02\n\n400C06L07Y05\n\n912CE06Y00\n\n910C06L00Y00\n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\nNone\n\n  \n\n**Schematics Impact :\n\nF0214-CE0005344Y02\n\nF0214-CE0007720Y00\n\n  \n\n**HWVPR Impact :\n\nNone\n\n', '<void>', u'HW Overvoltage protection precision impacted\n\n', '', '', '', '', '', '', '', ''], ['SACR 1524', 'Evolution', 'Bus Isolation Switch function', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '03/05/16 15:18', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5D3\n\n>\n\n> **HSID impact : None\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1526', 'Defect', 'CAN messages "IO STATE" and "FUSE STATUS" are sent every 500ms instead of 250ms.', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '03/05/16 15:25', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : None\n\n>\n\n> **HSID impact : None\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'None with SyCR 1450 implemented. (Functional impact from SyCR 1526)\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1532', 'Evolution', 'Inhibition of protections and tolerances', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '13/05/16 09:59', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', 'Yes', ''], ['SACR 1544', 'Evolution', 'Removal of Board power supply presence in the CAN IRD and FDD', 'Complementary_Analysis', 'Major', 'S1', 'S2', '<void>', '19/05/16 13:45', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **FDD impact : FDD_GLOBAL_7000_ACLOG issue 2D1\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None (No functional impact, Only PN changes)\n\n>\n\n>  \n\n>\n\n>  \n\n>\n\n>  \n\n>\n\n> **_Complementary analysis : non change\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1596', 'Defect', 'In flight, the ELC contactor is not opened if ATC3 is fail closed', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '03/05/16 15:33', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : None\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'In flight, the ELC contactor is not opened if ATC3 is fail closed. ELC is\nalmost never closed in flight (On jack only)\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1597', 'Evolution', 'Open phase protection reporting', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '13/05/16 10:06', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : None\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'The open phase protection is not reported to the avionic.\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1599', 'Evolution', 'Spare signals on AFDX and CAN busses', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '13/05/16 10:10', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : None\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'Some AFDX messages are invalid\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1605', 'Evolution', 'DSO Monitoring', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '13/05/16 10:13', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1617', 'Evolution', 'ACMP overload protection modification', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '13/05/16 10:18', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None (No functional impact, Only PN changes)\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1619', 'Evolution', 'EDMU listen by the ACLOG for the ACMP command needs to be sent on CAN', 'Under_Modification', 'Major', 'S1', 'S2', '<void>', '13/05/16 10:20', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', 'Yes', '', ''], ['HCR 1628', 'Evolution', 'ACLOG Minor evolution 910CE06Y00 to 910CE06Y01', 'In_Analysis', 'Minor', 'S0', 'S1', '<void>', '23/05/16 08:58', u'**Harware PN impact :\n\n911CE06Y00\n\n912CE06Y00\n\n910C06L00Y00\n\n  \n\n**HWRD Impact : \n\nNone\n\n**HWDD Impact :**\n\nNone\n\n  \n\n**Schematics Impact :\n\nF0214-CE0007720Y00\n\n  \n\n**HWVPR Impact :\n\nNone\n\n', '<void>', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['SACR 1631', 'Defect', 'Current monitoring requirement clean-up', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '19/05/16 09:26', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : No  \nHW impact : Yes\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'N/A (doc impact)\n\n', '', '', '', '', '', '', 'Yes', ''], ['HCR 1632', 'Defect', 'Requirement ASI 115V clean-up', 'In_Analysis', 'Minor', 'S0', 'S2', '<void>', '04/05/16 09:45', u'### **Harware PN impact :  \nNone (Hardware evolution is on HCR1387)  \n  \n**HWRD Impact :   \nET3598-S issue 5  \n  \n**HWDD Impact :**   \nG7000_ACLOG_HWDD_ET4115_E_Issue1D3  \n  \n**Schematics Impact :  \nNone (Schematics evolution is on HCR1387)  \n  \n**HWVPR Impact :  \n**HWVPR_G7000_ACLog_ET3938-V issue 1\n\n', '<void>', u'N/A (doc impact)\n\n', '', '', '', '', '', '', '', ''], ['SACR 1634', 'Evolution', 'Antiparalleling protection raised when AC ext PB and APUGEN PB are pushed exactly at the same time', 'Under_Modification', 'Major', 'S0', 'S2', '<void>', '13/05/16 10:23', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'In theory, antiparalleling protection raised when AC ext PB and APUGEN PB are\npushed exactly at the same time\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1641', 'Evolution', 'Chattering reset when the log is not powered anymore', 'In_Review', 'Major', 'S1', 'S2', '<void>', '18/05/16 14:19', u'**Assembly Board impact : 400CE06L07Y05\n\n> SW impact : Yes  \nHW impact : No\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2923-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : None\n\n', 'Power On FTV4', u'When every power supplies of the logs are flickering (due to short circuit),\ncontactors are locked open.\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1665', 'Defect', 'ACLOG ibit not accepted when request is done by the redundant EDMU', 'In_Review', 'Minor', 'S1', 'S1', '<void>', '25/05/16 16:26', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', 'First Flight FTV1', u'ACLOG ibit not accepted when request is done by the redundant EDMU\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1690', 'Evolution', 'CB status at start up and acknowledge reset', 'In_Analysis', 'Major', 'S1', 'S2', '<void>', '24/05/16 11:19', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', 'Power On FTV4', u'Acknowledge is reset when the AClog is restarting while the EDMU is kept\npowered.\n\n', '', '', '', '', '', 'Yes', '', ''], ['SACR 1691', 'Evolution', 'DSI failure inhibition for ACLOG', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '24/05/16 09:45', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : None  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2924-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['SACR 1692', 'Evolution', 'Application of BITE spec Safety Requirements to SSCS', 'In_Review', 'Minor', 'S1', 'S2', '<void>', '24/05/16 12:46', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : None  \nHW impact : YES\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : ET2924-S issue 5\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', 'Power On FTV4', u'Evolution\n\n', '', '', '', '', '', '', 'Yes', ''], ['HCR 1697', 'Evolution', 'Update of the no compliant BITE Safety impact at HwRD level', 'In_Analysis', '<void>', 'S1', 'S2', '<void>', '24/05/16 10:55', u'### **Hardware PN impact :  \nNone  \n  \n**HWRD Impact :   \nET3598-S issue 5  \n  \n**HWDD Impact :**   \nNone  \n  \n**Schematics Impact :  \nNone  \n  \n**HWVPR Impact :  \nNone\n\n', '<void>', u'Evolution\n\n', '', '', '', '', '', '', '', ''], ['SACR 1703', 'Defect', 'External power protection delay when voltage is lower than 55Vac', 'In_Review', 'Major', 'S1', 'S2', '<void>', '24/05/16 17:12', u'**Assembly Board impact : 400CE06L05Y04\n\n> SW impact : Yes  \nHW impact : None\n\n>\n\n>  \n\n>\n\n> **SSCS impact specification : none\n\n>\n\n> **HSID impact : none\n\n>\n\n> **\n\n>\n\n> **ETPR impact verification : none  \n\n>\n\n> **\n\n>\n\n> **ATP impact : none\n\n', 'Power On FTV4', u'When the external power protection reach the confirmation time, an additional\ntime of 1 sec from the time the voltage is lower than 55V, is needed before\nremoving pin F and opening ELC.  \nOnly applicable if the voltage is lower than 55Vac (pin F absence)\n\n', '', '', '', '', '', 'Yes', '', '']]
        #cr = ['SACR 0168', 'Evolution', 'Safety attribute update according to SAQ 345', 'Under_Verification', 'Minor', 'S0', 'S1', 'EQT_ACLOG_02_04', '28/01/16 15:18', u'REQ_SAFETY attribute of derived requirements.  \n\n', '2nd Release ISTCR', u'N/A\n\n', '', '', '', '', '', '', '', '']
        sheet_name = 'Change Requests'
        wb = Workbook()
        ws = wb.create_sheet(title = sheet_name)
        border=Border(
                        left=Side(border_style=BORDER_THIN),
                        right=Side(border_style=BORDER_THIN),
                        top=Side(border_style=BORDER_THIN),
                        bottom=Side(border_style=BORDER_THIN))
        alignment=Alignment(wrap_text=True,shrink_to_fit=True)
        style_border = Style(border,alignment)
        row = 10
        for cr in tbl_cr:
            if row == 21:
                print cr
                break
            for col_idx in range(1,21):
                Style.setCell(ws,cr,row,col_idx,style_border)
            row += 1

        filename = "Change_Requests_List_%d.xlsx" % floor(time.time())
        wb.save(join("result",filename))
        exit()
    tbl_list_req = [u'SWRD_GLOBAL-ACENM_0001', u'SWRD_GLOBAL-ACENM_0002', u'SWRD_GLOBAL-ACENM_0003', u'SWRD_GLOBAL-ACENM_0004', u'SWRD_GLOBAL-ACENM_0005', u'SWRD_GLOBAL-ACENM_0006', u'SWRD_GLOBAL-ACENM_0007', u'SWRD_GLOBAL-ACENM_0008', u'SWRD_GLOBAL-ACENM_0009', u'SWRD_GLOBAL-ACENM_0010', u'SWRD_GLOBAL-ACENM_0011', u'SWRD_GLOBAL-ACENM_0012', u'SWRD_GLOBAL-ACENM_0013', u'SWRD_GLOBAL-ACENM_0014', u'SWRD_GLOBAL-ACENM_0016', u'SWRD_GLOBAL-ACENM_0017', u'SWRD_GLOBAL-ACENM_0018', u'SWRD_GLOBAL-ACENM_0019', u'SWRD_GLOBAL-ACENM_0020', u'SWRD_GLOBAL-ACENM_0021', u'SWRD_GLOBAL-ACENM_0022', u'SWRD_GLOBAL-ACENM_0023', u'SWRD_GLOBAL-ACENM_0024', u'SWRD_GLOBAL-ACENM_0025', u'SWRD_GLOBAL-ACENM_0026', u'SWRD_GLOBAL-ACENM_0027', u'SWRD_GLOBAL-ACENM_0028', u'SWRD_GLOBAL-ACENM_0032', u'SWRD_GLOBAL-ACENM_0033', u'SWRD_GLOBAL-ACENM_0034', u'SWRD_GLOBAL-ACENM_0035', u'SWRD_GLOBAL-ACENM_0036', u'SWRD_GLOBAL-ACENM_0038', u'SWRD_GLOBAL-ACENM_0039', u'SWRD_GLOBAL-ACENM_0040', u'SWRD_GLOBAL-ACENM_0042', u'SWRD_GLOBAL-ACENM_0044', u'SWRD_GLOBAL-ACENM_0046', u'SWRD_GLOBAL-ACENM_0047', u'SWRD_GLOBAL-ACENM_0048', u'SWRD_GLOBAL-ACENM_0049', u'SWRD_GLOBAL-ACENM_0050', u'SWRD_GLOBAL-ACENM_0051', u'SWRD_GLOBAL-ACENM_0052', u'SWRD_GLOBAL-ACENM_0053', u'SWRD_GLOBAL-ACENM_0054', u'SWRD_GLOBAL-ACENM_0055', u'SWRD_GLOBAL-ACENM_0056', u'SWRD_GLOBAL-ACENM_0057', u'SWRD_GLOBAL-ACENM_0058', u'SWRD_GLOBAL-ACENM_0061', u'SWRD_GLOBAL-ACENM_0062', u'SWRD_GLOBAL-ACENM_0063', u'SWRD_GLOBAL-ACENM_0064', u'SWRD_GLOBAL-ACENM_0065', u'SWRD_GLOBAL-ACENM_0066', u'SWRD_GLOBAL-ACENM_0067', u'SWRD_GLOBAL-ACENM_0068', u'SWRD_GLOBAL-ACENM_0069', u'SWRD_GLOBAL-ACENM_0076', u'SWRD_GLOBAL-ACENM_0077', u'SWRD_GLOBAL-ACENM_0078', u'SWRD_GLOBAL-ACENM_0079', u'SWRD_GLOBAL-ACENM_0080', u'SWRD_GLOBAL-ACENM_0081', u'SWRD_GLOBAL-ACENM_0082', u'SWRD_GLOBAL-ACENM_0083', u'SWRD_GLOBAL-ACENM_0084', u'SWRD_GLOBAL-ACENM_0087', u'SWRD_GLOBAL-ACENM_0088', u'SWRD_GLOBAL-ACENM_0089', u'SWRD_GLOBAL-ACENM_0090', u'SWRD_GLOBAL-ACENM_0094', u'SWRD_GLOBAL-ACENM_0095', u'SWRD_GLOBAL-ACENM_0096', u'SWRD_GLOBAL-ACENM_0097', u'SWRD_GLOBAL-ACENM_0098', u'SWRD_GLOBAL-ACENM_0099', u'SWRD_GLOBAL-ACENM_0100', u'SWRD_GLOBAL-ACENM_0101', u'SWRD_GLOBAL-ACENM_0102', u'SWRD_GLOBAL-ACENM_0103', u'SWRD_GLOBAL-ACENM_0104', u'SWRD_GLOBAL-ACENM_0106', u'SWRD_GLOBAL-ACENM_0107', u'SWRD_GLOBAL-ACENM_0108', u'SWRD_GLOBAL-ACENM_0109', u'SWRD_GLOBAL-ACENM_0110', u'SWRD_GLOBAL-ACENM_0112', u'SWRD_GLOBAL-ACENM_0113', u'SWRD_GLOBAL-ACENM_0114', u'SWRD_GLOBAL-ACENM_0115', u'SWRD_GLOBAL-ACENM_0117', u'SWRD_GLOBAL-ACENM_0118', u'SWRD_GLOBAL-ACENM_0119', u'SWRD_GLOBAL-ACENM_0120', u'SWRD_GLOBAL-ACENM_0121', u'SWRD_GLOBAL-ACENM_0122', u'SWRD_GLOBAL-ACENM_0123', u'SWRD_GLOBAL-ACENM_0124', u'SWRD_GLOBAL-ACENM_0125', u'SWRD_GLOBAL-ACENM_0126', u'SWRD_GLOBAL-ACENM_0127', u'SWRD_GLOBAL-ACENM_0128', u'SWRD_GLOBAL-ACENM_0131', u'SWRD_GLOBAL-ACENM_0132', u'SWRD_GLOBAL-ACENM_0133', u'SWRD_GLOBAL-ACENM_0134', u'SWRD_GLOBAL-ACENM_0135', u'SWRD_GLOBAL-ACENM_0136', u'SWRD_GLOBAL-ACENM_0137', u'SWRD_GLOBAL-ACENM_0138', u'SWRD_GLOBAL-ACENM_0139', u'SWRD_GLOBAL-ACENM_0140', u'SWRD_GLOBAL-ACENM_0141', u'SWRD_GLOBAL-ACENM_0142', u'SWRD_GLOBAL-ACENM_0143', u'SWRD_GLOBAL-ACENM_0145', u'SWRD_GLOBAL-ACENM_0146', u'SWRD_GLOBAL-ACENM_0147', u'SWRD_GLOBAL-ACENM_0148', u'SWRD_GLOBAL-ACENM_0149', u'SWRD_GLOBAL-ACENM_0150', u'SWRD_GLOBAL-ACENM_0151', u'SWRD_GLOBAL-ACENM_0153', u'SWRD_GLOBAL-ACENM_0155', u'SWRD_GLOBAL-ACENM_0156', u'SWRD_GLOBAL-ACENM_0157', u'SWRD_GLOBAL-ACENM_0158', u'SWRD_GLOBAL-ACENM_0159', u'SWRD_GLOBAL-ACENM_0160', u'SWRD_GLOBAL-ACENM_0161', u'SWRD_GLOBAL-ACENM_0162', u'SWRD_GLOBAL-ACENM_0163', u'SWRD_GLOBAL-ACENM_0165', u'SWRD_GLOBAL-ACENM_0166', u'SWRD_GLOBAL-ACENM_0167', u'SWRD_GLOBAL-ACENM_0168', u'SWRD_GLOBAL-ACENM_0169', u'SWRD_GLOBAL-ACENM_0170', u'SWRD_GLOBAL-ACENM_0171', u'SWRD_GLOBAL-ACENM_0172', u'SWRD_GLOBAL-ACENM_0173', u'SWRD_GLOBAL-ACENM_0174', u'SWRD_GLOBAL-ACENM_0175', u'SWRD_GLOBAL-ACENM_0176', u'SWRD_GLOBAL-ACENM_0177', u'SWRD_GLOBAL-ACENM_0178', u'SWRD_GLOBAL-ACENM_0179', u'SWRD_GLOBAL-ACENM_0181', u'SWRD_GLOBAL-ACENM_0182', u'SWRD_GLOBAL-ACENM_0183', u'SWRD_GLOBAL-ACENM_0184', u'SWRD_GLOBAL-ACENM_0185', u'SWRD_GLOBAL-ACENM_0186', u'SWRD_GLOBAL-ACENM_0187', u'SWRD_GLOBAL-ACENM_0188', u'SWRD_GLOBAL-ACENM_0189', u'SWRD_GLOBAL-ACENM_0190', u'SWRD_GLOBAL-ACENM_0193', u'SWRD_GLOBAL-ACENM_0194', u'SWRD_GLOBAL-ACENM_0195', u'SWRD_GLOBAL-ACENM_0196', u'SWRD_GLOBAL-ACENM_0197', u'SWRD_GLOBAL-ACENM_0198', u'SWRD_GLOBAL-ACENM_0199', u'SWRD_GLOBAL-ACENM_0200', u'SWRD_GLOBAL-ACENM_0201', u'SWRD_GLOBAL-ACENM_0202', u'SWRD_GLOBAL-ACENM_0203', u'SWRD_GLOBAL-ACENM_0204', u'SWRD_GLOBAL-ACENM_0205', u'SWRD_GLOBAL-ACENM_0206', u'SWRD_GLOBAL-ACENM_0207', u'SWRD_GLOBAL-ACENM_0208', u'SWRD_GLOBAL-ACENM_0209', u'SWRD_GLOBAL-ACENM_0210', u'SWRD_GLOBAL-ACENM_0211', u'SWRD_GLOBAL-ACENM_0212', u'SWRD_GLOBAL-ACENM_0213', u'SWRD_GLOBAL-ACENM_0214', u'SWRD_GLOBAL-ACENM_0215', u'SWRD_GLOBAL-ACENM_0216', u'SWRD_GLOBAL-ACENM_0217', u'SWRD_GLOBAL-ACENM_0218', u'SWRD_GLOBAL-ACENM_0219', u'SWRD_GLOBAL-ACENM_0220', u'SWRD_GLOBAL-ACENM_0221', u'SWRD_GLOBAL-ACENM_0222', u'SWRD_GLOBAL-ACENM_0223', u'SWRD_GLOBAL-ACENM_0224', u'SWRD_GLOBAL-ACENM_0225', u'SWRD_GLOBAL-ACENM_0226', u'SWRD_GLOBAL-ACENM_0227', u'SWRD_GLOBAL-ACENM_0228', u'SWRD_GLOBAL-ACENM_0229', u'SWRD_GLOBAL-ACENM_0230', u'SWRD_GLOBAL-ACENM_0231', u'SWRD_GLOBAL-ACENM_0232', u'SWRD_GLOBAL-ACENM_0233', u'SWRD_GLOBAL-ACENM_0234', u'SWRD_GLOBAL-ACENM_0235', u'SWRD_GLOBAL-ACENM_0236', u'SWRD_GLOBAL-ACENM_0238', u'SWRD_GLOBAL-ACENM_0239', u'SWRD_GLOBAL-ACENM_0240', u'SWRD_GLOBAL-ACENM_0241', u'SWRD_GLOBAL-ACENM_0242', u'SWRD_GLOBAL-ACENM_0243', u'SWRD_GLOBAL-ACENM_0244', u'SWRD_GLOBAL-ACENM_0245', u'SWRD_GLOBAL-ACENM_0246', u'SWRD_GLOBAL-ACENM_0247', u'SWRD_GLOBAL-ACENM_0248', u'SWRD_GLOBAL-ACENM_0250', u'SWRD_GLOBAL-ACENM_0251', u'SWRD_GLOBAL-ACENM_0252', u'SWRD_GLOBAL-ACENM_0253', u'SWRD_GLOBAL-ACENM_0254', u'SWRD_GLOBAL-ACENM_0255', u'SWRD_GLOBAL-ACENM_0259', u'SWRD_GLOBAL-ACENM_0260', u'SWRD_GLOBAL-ACENM_0262', u'SWRD_GLOBAL-ACENM_0263', u'SWRD_GLOBAL-ACENM_0264', u'SWRD_GLOBAL-ACENM_0267', u'SWRD_GLOBAL-ACENM_0268', u'SWRD_GLOBAL-ACENM_0269', u'SWRD_GLOBAL-ACENM_0270', u'SWRD_GLOBAL-ACENM_0271', u'SWRD_GLOBAL-ACENM_0272', u'SWRD_GLOBAL-ACENM_0274', u'SWRD_GLOBAL-ACENM_0275', u'SWRD_GLOBAL-ACENM_0276', u'SWRD_GLOBAL-ACENM_0277', u'SWRD_GLOBAL-ACENM_0278', u'SWRD_GLOBAL-ACENM_0279', u'SWRD_GLOBAL-ACENM_0280', u'SWRD_GLOBAL-ACENM_0281', u'SWRD_GLOBAL-ACENM_0282', u'SWRD_GLOBAL-ACENM_0283', u'SWRD_GLOBAL-ACENM_0284', u'SWRD_GLOBAL-ACENM_0285', u'SWRD_GLOBAL-ACENM_0286', u'SWRD_GLOBAL-ACENM_0289', u'SWRD_GLOBAL-ACENM_0290', u'SWRD_GLOBAL-ACENM_0291', u'SWRD_GLOBAL-ACENM_0293', u'SWRD_GLOBAL-ACENM_0294', u'SWRD_GLOBAL-ACENM_0296', u'SWRD_GLOBAL-ACENM_0297', u'SWRD_GLOBAL-ACENM_0299', u'SWRD_GLOBAL-ACENM_0301', u'SWRD_GLOBAL-ACENM_0308', u'SWRD_GLOBAL-ACENM_0309', u'SWRD_GLOBAL-ACENM_0310', u'SWRD_GLOBAL-ACENM_0311', u'SWRD_GLOBAL-ACENM_0312', u'SWRD_GLOBAL-ACENM_0313', u'SWRD_GLOBAL-ACENM_0316', u'SWRD_GLOBAL-ACENM_0317', u'SWRD_GLOBAL-ACENM_0318', u'SWRD_GLOBAL-ACENM_0319', u'SWRD_GLOBAL-ACENM_0320', u'SWRD_GLOBAL-ACENM_0321', u'SWRD_GLOBAL-ACENM_0322', u'SWRD_GLOBAL-ACENM_0323', u'SWRD_GLOBAL-ACENM_0324', u'SWRD_GLOBAL-ACENM_0325', u'SWRD_GLOBAL-ACENM_0326', u'SWRD_GLOBAL-ACENM_0327', u'SWRD_GLOBAL-ACENM_0328', u'SWRD_GLOBAL-ACENM_0329', u'SWRD_GLOBAL-ACENM_0330', u'SWRD_GLOBAL-ACENM_0331', u'SWRD_GLOBAL-ACENM_0332', u'SWRD_GLOBAL-ACENM_0333', u'SWRD_GLOBAL-ACENM_0334', u'SWRD_GLOBAL-ACENM_0335', u'SWRD_GLOBAL-ACENM_0336', u'SWRD_GLOBAL-ACENM_0337', u'SWRD_GLOBAL-ACENM_0338', u'SWRD_GLOBAL-ACENM_0339', u'SWRD_GLOBAL-ACENM_0340', u'SWRD_GLOBAL-ACENM_0341', u'SWRD_GLOBAL-ACENM_0342', u'SWRD_GLOBAL-ACENM_0343', u'SWRD_GLOBAL-ACENM_0344', u'SWRD_GLOBAL-ACENM_0345', u'SWRD_GLOBAL-ACENM_0346', u'SWRD_GLOBAL-ACENM_0348', u'SWRD_GLOBAL-ACENM_0349', u'SWRD_GLOBAL-ACENM_0350', u'SWRD_GLOBAL-ACENM_0351', u'SWRD_GLOBAL-ACENM_0352', u'SWRD_GLOBAL-ACENM_0353', u'SWRD_GLOBAL-ACENM_0354', u'SWRD_GLOBAL-ACENM_0355', u'SWRD_GLOBAL-ACENM_0356', u'SWRD_GLOBAL-ACENM_0357', u'SWRD_GLOBAL-ACENM_0358', u'SWRD_GLOBAL-ACENM_0359', u'SWRD_GLOBAL-ACENM_0360', u'SWRD_GLOBAL-ACENM_0361', u'SWRD_GLOBAL-ACENM_0362', u'SWRD_GLOBAL-ACENM_0363', u'SWRD_GLOBAL-ACENM_0364', u'SWRD_GLOBAL-ACENM_0365', u'SWRD_GLOBAL-ACENM_0366', u'SWRD_GLOBAL-ACENM_0367', u'SWRD_GLOBAL-ACENM_0368', u'SWRD_GLOBAL-ACENM_0370', u'SWRD_GLOBAL-ACENM_0371', u'SWRD_GLOBAL-ACENM_0373', u'SWRD_GLOBAL-ACENM_0375', u'SWRD_GLOBAL-ACENM_0376', u'SWRD_GLOBAL-ACENM_0377', u'SWRD_GLOBAL-ACENM_0378', u'SWRD_GLOBAL-ACENM_0384', u'SWRD_GLOBAL-ACENM_0385', u'SWRD_GLOBAL-ACENM_0386', u'SWRD_GLOBAL-ACENM_0387', u'SWRD_GLOBAL-ACENM_0388', u'SWRD_GLOBAL-ACENM_0389', u'SWRD_GLOBAL-ACENM_0390', u'SWRD_GLOBAL-ACENM_0391', u'SWRD_GLOBAL-ACENM_0392', u'SWRD_GLOBAL-ACENM_0394', u'SWRD_GLOBAL-ACENM_0395', u'SWRD_GLOBAL-ACENM_0396', u'SWRD_GLOBAL-ACENM_0397', u'SWRD_GLOBAL-ACENM_0398', u'SWRD_GLOBAL-ACENM_0400', u'SWRD_GLOBAL-ACENM_0401', u'SWRD_GLOBAL-ACENM_0402', u'SWRD_GLOBAL-ACENM_0403', u'SWRD_GLOBAL-ACENM_0405', u'SWRD_GLOBAL-ACENM_0406', u'SWRD_GLOBAL-ACENM_0407', u'SWRD_GLOBAL-ACENM_0408', u'SWRD_GLOBAL-ACENM_0409', u'SWRD_GLOBAL-ACENM_0410', u'SWRD_GLOBAL-ACENM_0411', u'SWRD_GLOBAL-ACENM_0413', u'SWRD_GLOBAL-ACENM_0414', u'SWRD_GLOBAL-ACENM_0415', u'SWRD_GLOBAL-ACENM_0416', u'SWRD_GLOBAL-ACENM_0417', u'SWRD_GLOBAL-ACENM_0418', u'SWRD_GLOBAL-ACENM_0421', u'SWRD_GLOBAL-ACENM_0424', u'SWRD_GLOBAL-ACENM_0425', u'SWRD_GLOBAL-ACENM_0426', u'SWRD_GLOBAL-ACENM_0427', u'SWRD_GLOBAL-ACENM_0428', u'SWRD_GLOBAL-ACENM_0429', u'SWRD_GLOBAL-ACENM_0431', u'SWRD_GLOBAL-ACENM_0432', u'SWRD_GLOBAL-ACENM_0433', u'SWRD_GLOBAL-ACENM_0434', u'SWRD_GLOBAL-ACENM_0435', u'SWRD_GLOBAL-ACENM_0436', u'SWRD_GLOBAL-ACENM_0437', u'SWRD_GLOBAL-ACENM_0438', u'SWRD_GLOBAL-ACENM_0439', u'SWRD_GLOBAL-ACENM_0440', u'SWRD_GLOBAL-ACENM_0441', u'SWRD_GLOBAL-ACENM_0444', u'SWRD_GLOBAL-ACENM_0445', u'SWRD_GLOBAL-ACENM_0446', u'SWRD_GLOBAL-ACENM_0447', u'SWRD_GLOBAL-ACENM_0448', u'SWRD_GLOBAL-ACENM_0449', u'SWRD_GLOBAL-ACENM_0450', u'SWRD_GLOBAL-ACENM_0451', u'SWRD_GLOBAL-ACENM_0452', u'SWRD_GLOBAL-ACENM_0453', u'SWRD_GLOBAL-ACENM_0454', u'SWRD_GLOBAL-ACENM_0455', u'SWRD_GLOBAL-ACENM_0456', u'SWRD_GLOBAL-ACENM_0457', u'SWRD_GLOBAL-ACENM_0458', u'SWRD_GLOBAL-ACENM_0459', u'SWRD_GLOBAL-ACENM_0460', u'SWRD_GLOBAL-ACENM_0461', u'SWRD_GLOBAL-ACENM_0462', u'SWRD_GLOBAL-ACENM_0463', u'SWRD_GLOBAL-ACENM_0464', u'SWRD_GLOBAL-ACENM_0465', u'SWRD_GLOBAL-ACENM_0466', u'SWRD_GLOBAL-ACENM_0467', u'SWRD_GLOBAL-ACENM_0468', u'SWRD_GLOBAL-ACENM_0469', u'SWRD_GLOBAL-ACENM_0470', u'SWRD_GLOBAL-ACENM_0471', u'SWRD_GLOBAL-ACENM_0472', u'SWRD_GLOBAL-ACENM_0473', u'SWRD_GLOBAL-ACENM_0474', u'SWRD_GLOBAL-ACENM_0475', u'SWRD_GLOBAL-ACENM_0476', u'SWRD_GLOBAL-ACENM_0477', u'SWRD_GLOBAL-ACENM_0478', u'SWRD_GLOBAL-ACENM_0479', u'SWRD_GLOBAL-ACENM_0480', u'SWRD_GLOBAL-ACENM_0481', u'SWRD_GLOBAL-ACENM_0482', u'SWRD_GLOBAL-ACENM_0483', u'SWRD_GLOBAL-ACENM_0484', u'SWRD_GLOBAL-ACENM_0485', u'SWRD_GLOBAL-ACENM_0486', u'SWRD_GLOBAL-ACENM_0487', u'SWRD_GLOBAL-ACENM_0488', u'SWRD_GLOBAL-ACENM_0489', u'SWRD_GLOBAL-ACENM_0490', u'SWRD_GLOBAL-ACENM_0491', u'SWRD_GLOBAL-ACENM_0492', u'SWRD_GLOBAL-ACENM_0493', u'SWRD_GLOBAL-ACENM_0494', u'SWRD_GLOBAL-ACENM_0495', u'SWRD_GLOBAL-ACENM_0496', u'SWRD_GLOBAL-ACENM_0497', u'SWRD_GLOBAL-ACENM_0498', u'SWRD_GLOBAL-ACENM_0499', u'SWRD_GLOBAL-ACENM_0500', u'SWRD_GLOBAL-ACENM_0501', u'SWRD_GLOBAL-ACENM_0502', u'SWRD_GLOBAL-ACENM_0503', u'SWRD_GLOBAL-ACENM_0504', u'SWRD_GLOBAL-ACENM_0505', u'SWRD_GLOBAL-ACENM_0506', u'SWRD_GLOBAL-ACENM_0507', u'SWRD_GLOBAL-ACENM_0508', u'SWRD_GLOBAL-ACENM_0509', u'SWRD_GLOBAL-ACENM_0510', u'SWRD_GLOBAL-ACENM_0511', u'SWRD_GLOBAL-ACENM_0512', u'SWRD_GLOBAL-ACENM_0513', u'SWRD_GLOBAL-ACENM_0514', u'SWRD_GLOBAL-ACENM_0515', u'SWRD_GLOBAL-ACENM_0516', u'SWRD_GLOBAL-ACENM_0517', u'SWRD_GLOBAL-ACENM_0518', u'SWRD_GLOBAL-ACENM_0519', u'SWRD_GLOBAL-ACENM_0520', u'SWRD_GLOBAL-ACENM_0521', u'SWRD_GLOBAL-ACENM_0522', u'SWRD_GLOBAL-ACENM_0523', u'SWRD_GLOBAL-ACENM_0524', u'SWRD_GLOBAL-ACENM_0525', u'SWRD_GLOBAL-ACENM_0526', u'SWRD_GLOBAL-ACENM_0527', u'SWRD_GLOBAL-ACENM_0528', u'SWRD_GLOBAL-ACENM_0529', u'SWRD_GLOBAL-ACENM_0530', u'SWRD_GLOBAL-ACENM_0531', u'SWRD_GLOBAL-ACENM_0532', u'SWRD_GLOBAL-ACENM_0533', u'SWRD_GLOBAL-ACENM_0534', u'SWRD_GLOBAL-ACENM_0535', u'SWRD_GLOBAL-ACENM_0536', u'SWRD_GLOBAL-ACENM_0537', u'SWRD_GLOBAL-ACENM_0538', u'SWRD_GLOBAL-ACENM_0539', u'SWRD_GLOBAL-ACENM_0540', u'SWRD_GLOBAL-ACENM_0541', u'SWRD_GLOBAL-ACENM_0542', u'SWRD_GLOBAL-ACENM_0543', u'SWRD_GLOBAL-ACENM_0544', u'SWRD_GLOBAL-ACENM_0545', u'SWRD_GLOBAL-ACENM_0546', u'SWRD_GLOBAL-ACENM_0547', u'SWRD_GLOBAL-ACENM_0548', u'SWRD_GLOBAL-ACENM_0549', u'SWRD_GLOBAL-ACENM_0550', u'SWRD_GLOBAL-ACENM_0551', u'SWRD_GLOBAL-ACENM_0552']
    hlr_vs_llr = {u'SWRD_GLOBAL-ACENM_0360': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0497': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU006_001'], u'SWRD_GLOBAL-ACENM_0496': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU006_001'], u'SWRD_GLOBAL-ACENM_0495': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU006_001'], u'SWRD_GLOBAL-ACENM_0494': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU006_001'], u'SWRD_GLOBAL-ACENM_0493': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU006_001'], u'SWRD_GLOBAL-ACENM_0492': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0491': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU029_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU029_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU029_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU029_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0490': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU028_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU028_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU028_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU028_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0499': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU006_001'], u'SWRD_GLOBAL-ACENM_0498': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU006_001'], u'SWRD_GLOBAL-ACENM_0189': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_009'], u'SWRD_GLOBAL-ACENM_0127': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU014_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU014_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU014_001'], u'SWRD_GLOBAL-ACENM_0126': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_001'], u'SWRD_GLOBAL-ACENM_0125': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_001'], u'SWRD_GLOBAL-ACENM_0124': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006'], u'SWRD_GLOBAL-ACENM_0088': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU056_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU056_001'], u'SWRD_GLOBAL-ACENM_0089': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_010'], u'SWRD_GLOBAL-ACENM_0121': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU002_002'], u'SWRD_GLOBAL-ACENM_0120': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU029_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU029_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU029_002'], u'SWRD_GLOBAL-ACENM_0084': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU056_003'], u'SWRD_GLOBAL-ACENM_0188': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_011', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_015', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_013', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_012', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_009'], u'SWRD_GLOBAL-ACENM_0087': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU056_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU056_001'], u'SWRD_GLOBAL-ACENM_0080': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_003'], u'SWRD_GLOBAL-ACENM_0081': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_003'], u'SWRD_GLOBAL-ACENM_0082': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_011'], u'SWRD_GLOBAL-ACENM_0128': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU010_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU010_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU018_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU010_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU018_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU018_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_004'], u'SWRD_GLOBAL-ACENM_0066': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_001'], u'SWRD_GLOBAL-ACENM_0067': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_001'], u'SWRD_GLOBAL-ACENM_0064': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_001'], u'SWRD_GLOBAL-ACENM_0065': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_001'], u'SWRD_GLOBAL-ACENM_0062': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU054_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU054_008'], u'SWRD_GLOBAL-ACENM_0063': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_009'], u'SWRD_GLOBAL-ACENM_0309': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_004'], u'SWRD_GLOBAL-ACENM_0308': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_017', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020'], u'SWRD_GLOBAL-ACENM_0068': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_001'], u'SWRD_GLOBAL-ACENM_0069': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_001'], u'SWRD_GLOBAL-ACENM_0545': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU024_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019'], u'SWRD_GLOBAL-ACENM_0544': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU021_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU021_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU021_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU009_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU009_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_016', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_014', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_015'], u'SWRD_GLOBAL-ACENM_0468': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU010_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU010_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0469': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU011_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU011_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0017': [u'SWDD_G7000_PPDS_ACENM_CSC048_CSU003_001'], u'SWRD_GLOBAL-ACENM_0540': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU026_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_001'], u'SWRD_GLOBAL-ACENM_0543': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_014', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_013', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_005'], u'SWRD_GLOBAL-ACENM_0462': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0463': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0019': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU024_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU053_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_015', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU056_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU011_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU028_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU013_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_009', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU024_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU010_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU010_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU056_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU005_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_008', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU054_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_006'], u'SWRD_GLOBAL-ACENM_0461': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0466': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_001'], u'SWRD_GLOBAL-ACENM_0548': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU053_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU011_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU011_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU010_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU010_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU011_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU011_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU010_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU010_010'], u'SWRD_GLOBAL-ACENM_0464': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0397': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_004'], u'SWRD_GLOBAL-ACENM_0264': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_006'], u'SWRD_GLOBAL-ACENM_0267': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU027_001'], u'SWRD_GLOBAL-ACENM_0260': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU026_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_010'], u'SWRD_GLOBAL-ACENM_0262': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU026_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU003_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU003_007'], u'SWRD_GLOBAL-ACENM_0263': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU014_004'], u'SWRD_GLOBAL-ACENM_0268': [u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_006', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_004', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_005', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_003', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU007_003', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU007_002', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_002', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_006', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_004', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_005', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU005_002', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_001'], u'SWRD_GLOBAL-ACENM_0269': [u'SWDD_G7000_PPDS_ACENM_CSC048_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU005_002'], u'SWRD_GLOBAL-ACENM_0395': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU008_001'], u'SWRD_GLOBAL-ACENM_0392': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0169': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU006_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_002'], u'SWRD_GLOBAL-ACENM_0168': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU009_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU009_002'], u'SWRD_GLOBAL-ACENM_0394': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU005_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU005_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU005_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU011_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU022_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU016_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU008_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU023_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU013_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU021_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU021_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU021_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU021_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU010_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU010_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU009_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU009_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU009_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU009_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU015_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_009', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU017_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU005_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU005_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU005_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_016', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_014', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_015'], u'SWRD_GLOBAL-ACENM_0163': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU013_001'], u'SWRD_GLOBAL-ACENM_0162': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU012_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001'], u'SWRD_GLOBAL-ACENM_0161': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU011_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_002'], u'SWRD_GLOBAL-ACENM_0160': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU010_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001'], u'SWRD_GLOBAL-ACENM_0167': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU015_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_002'], u'SWRD_GLOBAL-ACENM_0166': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU014_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU014_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001'], u'SWRD_GLOBAL-ACENM_0165': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_012', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU008_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU008_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001'], u'SWRD_GLOBAL-ACENM_0185': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_001'], u'SWRD_GLOBAL-ACENM_0184': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_001'], u'SWRD_GLOBAL-ACENM_0187': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_003'], u'SWRD_GLOBAL-ACENM_0186': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_003'], u'SWRD_GLOBAL-ACENM_0181': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_011', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_013', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_012', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_007'], u'SWRD_GLOBAL-ACENM_0183': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_011', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_013', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_012', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_007'], u'SWRD_GLOBAL-ACENM_0182': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_011', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_013', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_012', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_007'], u'SWRD_GLOBAL-ACENM_0021': [u'SWDD_G7000_PPDS_ACENM_CSC047_CSU001_001'], u'SWRD_GLOBAL-ACENM_0026': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_013', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_010'], u'SWRD_GLOBAL-ACENM_0027': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_013', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_010'], u'SWRD_GLOBAL-ACENM_0509': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_013', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_010'], u'SWRD_GLOBAL-ACENM_0508': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_002'], u'SWRD_GLOBAL-ACENM_0159': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_007'], u'SWRD_GLOBAL-ACENM_0118': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU004_003'], u'SWRD_GLOBAL-ACENM_0119': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU003_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU003_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU003_004'], u'SWRD_GLOBAL-ACENM_0501': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU038_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU038_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU038_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0500': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_004'], u'SWRD_GLOBAL-ACENM_0503': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0502': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU039_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU039_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0505': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_002'], u'SWRD_GLOBAL-ACENM_0504': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0507': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_002'], u'SWRD_GLOBAL-ACENM_0506': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU045_002'], u'SWRD_GLOBAL-ACENM_0396': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_010'], u'SWRD_GLOBAL-ACENM_0336': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_004'], u'SWRD_GLOBAL-ACENM_0334': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_020', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_007', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_006', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_018', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_011', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_016', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_017', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_014', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_015', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_002'], u'SWRD_GLOBAL-ACENM_0335': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_020', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_007', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_006', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_018', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_011', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_016', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_017', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_014', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_015', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_002'], u'SWRD_GLOBAL-ACENM_0332': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_008'], u'SWRD_GLOBAL-ACENM_0333': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_001'], u'SWRD_GLOBAL-ACENM_0330': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014'], u'SWRD_GLOBAL-ACENM_0331': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_001'], u'SWRD_GLOBAL-ACENM_0235': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0234': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0233': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU024_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020'], u'SWRD_GLOBAL-ACENM_0232': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0231': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0339': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0343': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_007'], u'SWRD_GLOBAL-ACENM_0342': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_009', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_018', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_016', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_014'], u'SWRD_GLOBAL-ACENM_0341': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_009', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_018', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_016', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_014'], u'SWRD_GLOBAL-ACENM_0340': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_009', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_018', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_016', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_014'], u'SWRD_GLOBAL-ACENM_0431': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_007'], u'SWRD_GLOBAL-ACENM_0346': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_007'], u'SWRD_GLOBAL-ACENM_0433': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_007'], u'SWRD_GLOBAL-ACENM_0432': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_007'], u'SWRD_GLOBAL-ACENM_0329': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_008', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_001'], u'SWRD_GLOBAL-ACENM_0348': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0439': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_002'], u'SWRD_GLOBAL-ACENM_0438': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_009', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_018', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_016', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_014'], u'SWRD_GLOBAL-ACENM_0378': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_006'], u'SWRD_GLOBAL-ACENM_0376': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_006'], u'SWRD_GLOBAL-ACENM_0078': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_003'], u'SWRD_GLOBAL-ACENM_0250': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_010', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_002'], u'SWRD_GLOBAL-ACENM_0153': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU063_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU063_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU063_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU063_004'], u'SWRD_GLOBAL-ACENM_0099': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001'], u'SWRD_GLOBAL-ACENM_0098': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_009'], u'SWRD_GLOBAL-ACENM_0156': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_004'], u'SWRD_GLOBAL-ACENM_0157': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_001'], u'SWRD_GLOBAL-ACENM_0155': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_005'], u'SWRD_GLOBAL-ACENM_0158': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_001'], u'SWRD_GLOBAL-ACENM_0090': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_004'], u'SWRD_GLOBAL-ACENM_0097': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_009'], u'SWRD_GLOBAL-ACENM_0096': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_006'], u'SWRD_GLOBAL-ACENM_0095': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_004'], u'SWRD_GLOBAL-ACENM_0094': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_004'], u'SWRD_GLOBAL-ACENM_0077': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_003'], u'SWRD_GLOBAL-ACENM_0076': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_003'], u'SWRD_GLOBAL-ACENM_0373': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU028_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU028_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU028_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_012'], u'SWRD_GLOBAL-ACENM_0370': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0371': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0079': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_003'], u'SWRD_GLOBAL-ACENM_0377': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_005'], u'SWRD_GLOBAL-ACENM_0375': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_005'], u'SWRD_GLOBAL-ACENM_0479': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU021_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU021_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0478': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU020_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0471': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_002'], u'SWRD_GLOBAL-ACENM_0470': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_002'], u'SWRD_GLOBAL-ACENM_0473': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU017_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU017_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0472': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0475': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU018_003'], u'SWRD_GLOBAL-ACENM_0474': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_002'], u'SWRD_GLOBAL-ACENM_0477': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0476': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU019_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU019_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU019_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0387': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_001'], u'SWRD_GLOBAL-ACENM_0384': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU027_001'], u'SWRD_GLOBAL-ACENM_0389': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU007_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU007_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU007_002'], u'SWRD_GLOBAL-ACENM_0388': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_003'], u'SWRD_GLOBAL-ACENM_0206': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU020_001'], u'SWRD_GLOBAL-ACENM_0207': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU020_001'], u'SWRD_GLOBAL-ACENM_0204': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU022_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU022_002'], u'SWRD_GLOBAL-ACENM_0205': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU019_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU019_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU019_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU019_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_006'], u'SWRD_GLOBAL-ACENM_0202': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU042_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU042_002'], u'SWRD_GLOBAL-ACENM_0203': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU019_001'], u'SWRD_GLOBAL-ACENM_0200': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU032_002'], u'SWRD_GLOBAL-ACENM_0201': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU042_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU042_001'], u'SWRD_GLOBAL-ACENM_0208': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU020_002'], u'SWRD_GLOBAL-ACENM_0209': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_001'], u'SWRD_GLOBAL-ACENM_0272': [u'SWDD_G7000_PPDS_ACENM_CSC047_CSU003_001'], u'SWRD_GLOBAL-ACENM_0271': [u'SWDD_G7000_PPDS_ACENM_CSC048_CSU007_003', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU007_002'], u'SWRD_GLOBAL-ACENM_0270': [u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_006', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_004', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_005', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_003', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_002', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_006', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_004', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_005', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU006_001'], u'SWRD_GLOBAL-ACENM_0277': [u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_007', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_006', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU001_001'], u'SWRD_GLOBAL-ACENM_0276': [u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_007', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_006', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU001_001'], u'SWRD_GLOBAL-ACENM_0275': [u'SWDD_G7000_PPDS_ACENM_CSC047_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_007', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_006', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_009'], u'SWRD_GLOBAL-ACENM_0279': [u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_009'], u'SWRD_GLOBAL-ACENM_0114': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU005_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU005_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU005_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003'], u'SWRD_GLOBAL-ACENM_0115': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU028_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU028_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU028_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU028_003'], u'SWRD_GLOBAL-ACENM_0112': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU008_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU008_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU008_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU008_003'], u'SWRD_GLOBAL-ACENM_0113': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU006_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU006_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU006_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU006_001'], u'SWRD_GLOBAL-ACENM_0110': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006'], u'SWRD_GLOBAL-ACENM_0039': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_010'], u'SWRD_GLOBAL-ACENM_0038': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_004'], u'SWRD_GLOBAL-ACENM_0035': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001'], u'SWRD_GLOBAL-ACENM_0034': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_004'], u'SWRD_GLOBAL-ACENM_0036': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_007'], u'SWRD_GLOBAL-ACENM_0109': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_010'], u'SWRD_GLOBAL-ACENM_0108': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_003'], u'SWRD_GLOBAL-ACENM_0104': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0107': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_011'], u'SWRD_GLOBAL-ACENM_0106': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0101': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_005'], u'SWRD_GLOBAL-ACENM_0100': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_004'], u'SWRD_GLOBAL-ACENM_0103': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_005'], u'SWRD_GLOBAL-ACENM_0102': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_005'], u'SWRD_GLOBAL-ACENM_0534': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_014', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_012'], u'SWRD_GLOBAL-ACENM_0537': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU024_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU024_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU024_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU024_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_004'], u'SWRD_GLOBAL-ACENM_0530': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_007'], u'SWRD_GLOBAL-ACENM_0531': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU028_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU028_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU028_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU028_003'], u'SWRD_GLOBAL-ACENM_0532': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_009'], u'SWRD_GLOBAL-ACENM_0123': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006'], u'SWRD_GLOBAL-ACENM_0538': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU056_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU056_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU056_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU056_006'], u'SWRD_GLOBAL-ACENM_0539': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU023_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU023_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU023_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU023_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU023_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU023_008', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU023_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU023_007'], u'SWRD_GLOBAL-ACENM_0122': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0325': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU003_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_010', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_011', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_008', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_002'], u'SWRD_GLOBAL-ACENM_0324': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU031_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU031_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006'], u'SWRD_GLOBAL-ACENM_0248': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU028_001'], u'SWRD_GLOBAL-ACENM_0326': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_010', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_011'], u'SWRD_GLOBAL-ACENM_0321': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_008'], u'SWRD_GLOBAL-ACENM_0320': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU003_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU003_002', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU003_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU001_001'], u'SWRD_GLOBAL-ACENM_0323': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU031_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU031_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006'], u'SWRD_GLOBAL-ACENM_0242': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU020_002'], u'SWRD_GLOBAL-ACENM_0243': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU020_002'], u'SWRD_GLOBAL-ACENM_0246': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU021_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU021_002'], u'SWRD_GLOBAL-ACENM_0247': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU021_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU021_004'], u'SWRD_GLOBAL-ACENM_0244': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU033_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006'], u'SWRD_GLOBAL-ACENM_0245': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU033_002'], u'SWRD_GLOBAL-ACENM_0400': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014'], u'SWRD_GLOBAL-ACENM_0402': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_013', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_008'], u'SWRD_GLOBAL-ACENM_0403': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_013', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_006'], u'SWRD_GLOBAL-ACENM_0405': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_005'], u'SWRD_GLOBAL-ACENM_0406': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU019_002'], u'SWRD_GLOBAL-ACENM_0407': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014'], u'SWRD_GLOBAL-ACENM_0408': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014'], u'SWRD_GLOBAL-ACENM_0409': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_001'], u'SWRD_GLOBAL-ACENM_0083': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU056_003'], u'SWRD_GLOBAL-ACENM_0338': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_002'], u'SWRD_GLOBAL-ACENM_0230': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU038_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU039_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU039_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU038_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU038_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0435': [u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_007', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_006', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_004', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU001_001'], u'SWRD_GLOBAL-ACENM_0434': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_009', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_010', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_011', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_008'], u'SWRD_GLOBAL-ACENM_0437': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_009', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_018', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_016', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_014'], u'SWRD_GLOBAL-ACENM_0345': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU030_001'], u'SWRD_GLOBAL-ACENM_0344': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_001'], u'SWRD_GLOBAL-ACENM_0327': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_001'], u'SWRD_GLOBAL-ACENM_0061': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU054_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU054_008'], u'SWRD_GLOBAL-ACENM_0141': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_003'], u'SWRD_GLOBAL-ACENM_0140': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU024_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU024_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU024_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU024_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU024_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006'], u'SWRD_GLOBAL-ACENM_0143': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0142': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU022_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU022_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU021_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU021_002'], u'SWRD_GLOBAL-ACENM_0145': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU009_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU009_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU009_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0301': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU049_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU049_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU049_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU046_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU046_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU046_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU046_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU046_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU046_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU046_006'], u'SWRD_GLOBAL-ACENM_0146': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU009_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU009_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU009_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0398': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_006'], u'SWRD_GLOBAL-ACENM_0547': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_006'], u'SWRD_GLOBAL-ACENM_0040': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU003_001'], u'SWRD_GLOBAL-ACENM_0042': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001'], u'SWRD_GLOBAL-ACENM_0546': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU065_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019'], u'SWRD_GLOBAL-ACENM_0044': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU003_007'], u'SWRD_GLOBAL-ACENM_0046': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_001'], u'SWRD_GLOBAL-ACENM_0047': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_001'], u'SWRD_GLOBAL-ACENM_0048': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU031_001'], u'SWRD_GLOBAL-ACENM_0049': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_009'], u'SWRD_GLOBAL-ACENM_0362': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0364': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0367': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0366': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0444': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_001'], u'SWRD_GLOBAL-ACENM_0445': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU006_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU006_003'], u'SWRD_GLOBAL-ACENM_0440': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_009', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_018', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_016', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_014'], u'SWRD_GLOBAL-ACENM_0441': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_012'], u'SWRD_GLOBAL-ACENM_0391': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_001'], u'SWRD_GLOBAL-ACENM_0460': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU032_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU032_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU032_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0549': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU026_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_005'], u'SWRD_GLOBAL-ACENM_0467': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0215': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_001'], u'SWRD_GLOBAL-ACENM_0214': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_001'], u'SWRD_GLOBAL-ACENM_0217': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0216': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_001'], u'SWRD_GLOBAL-ACENM_0211': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_001'], u'SWRD_GLOBAL-ACENM_0210': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_001'], u'SWRD_GLOBAL-ACENM_0213': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_001'], u'SWRD_GLOBAL-ACENM_0212': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_001'], u'SWRD_GLOBAL-ACENM_0465': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_001'], u'SWRD_GLOBAL-ACENM_0219': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0218': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU027_001'], u'SWRD_GLOBAL-ACENM_0289': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_017', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_016', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_015', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_012'], u'SWRD_GLOBAL-ACENM_0286': [u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_006', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_004', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU002_001'], u'SWRD_GLOBAL-ACENM_0284': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU002_001'], u'SWRD_GLOBAL-ACENM_0285': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_020', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_006', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_007', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_006', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_018', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_011', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_016', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_017', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_014', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_015', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_002'], u'SWRD_GLOBAL-ACENM_0280': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU008_001'], u'SWRD_GLOBAL-ACENM_0281': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_017', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_016', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_015', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_012'], u'SWRD_GLOBAL-ACENM_0051': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_003'], u'SWRD_GLOBAL-ACENM_0138': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU034_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU033_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0139': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU027_003'], u'SWRD_GLOBAL-ACENM_0131': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU019_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU019_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU019_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU019_001'], u'SWRD_GLOBAL-ACENM_0132': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU018_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU018_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU018_002'], u'SWRD_GLOBAL-ACENM_0133': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU018_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU018_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU018_003'], u'SWRD_GLOBAL-ACENM_0134': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU048_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU049_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU049_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU049_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU048_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU048_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU048_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU048_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU048_005'], u'SWRD_GLOBAL-ACENM_0135': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_006'], u'SWRD_GLOBAL-ACENM_0136': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU016_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU016_002'], u'SWRD_GLOBAL-ACENM_0137': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU012_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU012_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU012_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_004'], u'SWRD_GLOBAL-ACENM_0488': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0522': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0521': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0520': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU041_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU041_001'], u'SWRD_GLOBAL-ACENM_0527': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_017', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_016', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_015', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU050_012'], u'SWRD_GLOBAL-ACENM_0526': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU041_005'], u'SWRD_GLOBAL-ACENM_0480': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU022_002'], u'SWRD_GLOBAL-ACENM_0481': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU023_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU023_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0482': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_002'], u'SWRD_GLOBAL-ACENM_0483': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0484': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0485': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0486': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0487': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0310': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_015', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_016', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_017', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_018', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020'], u'SWRD_GLOBAL-ACENM_0311': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU014_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_006'], u'SWRD_GLOBAL-ACENM_0317': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU058_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU058_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_007'], u'SWRD_GLOBAL-ACENM_0318': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU058_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU058_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_007'], u'SWRD_GLOBAL-ACENM_0319': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU058_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU058_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_007'], u'SWRD_GLOBAL-ACENM_0252': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU024_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU053_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_015', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU026_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU056_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU011_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU008_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU025_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU057_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU007_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU028_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU013_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_009', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU005_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU055_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU024_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU010_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU010_007', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_004', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU056_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU005_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU015_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU033_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU026_008', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU054_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU021_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_006'], u'SWRD_GLOBAL-ACENM_0255': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU026_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU010_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU010_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU010_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU003_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU011_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU011_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU011_002'], u'SWRD_GLOBAL-ACENM_0254': [u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_012', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_008', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_006', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_003', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_004', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_011', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU002_002'], u'SWRD_GLOBAL-ACENM_0523': [u'SWDD_G7000_PPDS_ACENM_CSC048_CSU003_001'], u'SWRD_GLOBAL-ACENM_0552': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU059_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_001'], u'SWRD_GLOBAL-ACENM_0550': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU026_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU013_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_010'], u'SWRD_GLOBAL-ACENM_0551': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_013', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU040_010', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU017_001'], u'SWRD_GLOBAL-ACENM_0417': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0416': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_013', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_013', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014'], u'SWRD_GLOBAL-ACENM_0415': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU013_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU013_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_013', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014'], u'SWRD_GLOBAL-ACENM_0413': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0529': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU049_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU049_003'], u'SWRD_GLOBAL-ACENM_0528': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU063_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU063_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU063_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU063_004'], u'SWRD_GLOBAL-ACENM_0355': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0356': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_006'], u'SWRD_GLOBAL-ACENM_0299': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020'], u'SWRD_GLOBAL-ACENM_0259': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU012_004'], u'SWRD_GLOBAL-ACENM_0353': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_014', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_018', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_010'], u'SWRD_GLOBAL-ACENM_0178': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_012', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_006'], u'SWRD_GLOBAL-ACENM_0179': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_006'], u'SWRD_GLOBAL-ACENM_0174': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_003'], u'SWRD_GLOBAL-ACENM_0175': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_003'], u'SWRD_GLOBAL-ACENM_0176': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU009_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU040_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0177': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_007', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_006'], u'SWRD_GLOBAL-ACENM_0170': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_002'], u'SWRD_GLOBAL-ACENM_0171': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001'], u'SWRD_GLOBAL-ACENM_0172': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_001'], u'SWRD_GLOBAL-ACENM_0173': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_001'], u'SWRD_GLOBAL-ACENM_0358': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU018_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_004'], u'SWRD_GLOBAL-ACENM_0418': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU003_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU006_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_013', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_001'], u'SWRD_GLOBAL-ACENM_0196': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_011', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_015', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_013', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_012', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_003'], u'SWRD_GLOBAL-ACENM_0197': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_011', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_015', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_014', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_013', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_012', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_001'], u'SWRD_GLOBAL-ACENM_0194': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_010'], u'SWRD_GLOBAL-ACENM_0195': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_010'], u'SWRD_GLOBAL-ACENM_0193': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU000_010'], u'SWRD_GLOBAL-ACENM_0190': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_010', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_018', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_016', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_017', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_009', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU002_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU017_009'], u'SWRD_GLOBAL-ACENM_0058': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU054_004'], u'SWRD_GLOBAL-ACENM_0057': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU054_004'], u'SWRD_GLOBAL-ACENM_0056': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU053_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU053_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU053_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001'], u'SWRD_GLOBAL-ACENM_0055': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_005'], u'SWRD_GLOBAL-ACENM_0054': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_005'], u'SWRD_GLOBAL-ACENM_0053': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU053_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU053_003'], u'SWRD_GLOBAL-ACENM_0052': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU052_003'], u'SWRD_GLOBAL-ACENM_0198': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU035_006'], u'SWRD_GLOBAL-ACENM_0199': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU032_001'], u'SWRD_GLOBAL-ACENM_0518': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU037_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU037_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU037_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU037_006', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU037_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU037_002'], u'SWRD_GLOBAL-ACENM_0519': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU039_005'], u'SWRD_GLOBAL-ACENM_0512': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0513': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU039_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU038_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU018_017'], u'SWRD_GLOBAL-ACENM_0510': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_013', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU008_010'], u'SWRD_GLOBAL-ACENM_0511': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0516': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_008'], u'SWRD_GLOBAL-ACENM_0517': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU041_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU041_003'], u'SWRD_GLOBAL-ACENM_0514': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_008', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_009', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_018', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_012', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_017', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_016', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_014'], u'SWRD_GLOBAL-ACENM_0515': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_008'], u'SWRD_GLOBAL-ACENM_0451': [u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_001', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_007', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU006_004', u'SWDD_G7000_PPDS_ACENM_CSC040_CSU006_003'], u'SWRD_GLOBAL-ACENM_0455': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0489': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU002_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016'], u'SWRD_GLOBAL-ACENM_0459': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0458': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0220': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0221': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU007_002'], u'SWRD_GLOBAL-ACENM_0222': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU032_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU032_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0223': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU036_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU035_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0224': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU020_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU017_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU021_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU021_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU023_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU023_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU017_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU020_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU019_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU019_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU019_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU018_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU022_002'], u'SWRD_GLOBAL-ACENM_0225': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU017_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU021_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU021_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU023_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU023_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU017_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU019_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU019_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU019_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0226': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU028_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU028_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU028_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU028_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU026_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU027_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU029_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU029_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU029_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU029_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU009_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0227': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU037_006'], u'SWRD_GLOBAL-ACENM_0228': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU010_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU010_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU011_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU011_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0229': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU031_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU030_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU015_001'], u'SWRD_GLOBAL-ACENM_0354': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_003', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_012', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_015', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU000_014'], u'SWRD_GLOBAL-ACENM_0427': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_001', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_006', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_002', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU012_004', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0424': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_014', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU060_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_009', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU027_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU025_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_012', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020'], u'SWRD_GLOBAL-ACENM_0357': [u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_001', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_003', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_004', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_005', u'SWDD_G7000_PPDS_ACENM_CSC032_CSU016_006'], u'SWRD_GLOBAL-ACENM_0350': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0351': [u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_010', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_005', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_007', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_011', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_017', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_018', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_016', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_009', u'SWDD_G7000_PPDS_ACENM_CSC035_CSU001_008'], u'SWRD_GLOBAL-ACENM_0352': [u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_003', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU001_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_015', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_014', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_013', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_010', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_004', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_001', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU024_002', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_019', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_018', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_011', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU020_010'], u'SWRD_GLOBAL-ACENM_0421': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU007_006'], u'SWRD_GLOBAL-ACENM_0294': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_003'], u'SWRD_GLOBAL-ACENM_0297': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_003'], u'SWRD_GLOBAL-ACENM_0296': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_007', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_004'], u'SWRD_GLOBAL-ACENM_0291': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU005_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_004', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_001', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_002', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU043_003'], u'SWRD_GLOBAL-ACENM_0290': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020'], u'SWRD_GLOBAL-ACENM_0293': [u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_003', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_008', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_011', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_005', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_021', u'SWDD_G7000_PPDS_ACENM_CSC034_CSU004_020'], u'SWRD_GLOBAL-ACENM_0429': [u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_020', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_007', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_006', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU002_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_010', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_005', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_009', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_007', u'SWDD_G7000_PPDS_ACENM_CSC033_CSU025_001', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_002', u'SWDD_G7000_PPDS_ACENM_CSC047_CSU004_004', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_012', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_013', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_010', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU000_011', u'SWDD_G7000_PPDS_ACENM_CSC048_CSU007_001', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_018', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_019', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_013', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_016', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_017', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_014', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_015', u'SWDD_G7000_PPDS_ACENM_CSC036_CSU004_005']}
    llr_vs_code = {'CSC019_CSU004': ('CMP/CmpDio/CmpDio_ReadDsi.c', 'Component Layer/Component DIO/Read Selected DSI.docx', 'CmpDio_ReadDsi.c'), 'CSC019_CSU005': ('CMP/CmpDio/CmpDio_WriteDso.c', 'Component Layer/Component DIO/DIO Write 1 DSO delayed.docx', 'CmpDio_WriteDso.c'), 'CSC019_CSU006': ('CMP/CmpDio/CmpDio_WriteImmediateDso.c', 'Component Layer/Component DIO/DIO Write 1 DSO Immediate.docx', 'CmpDio_WriteImmediateDso.c'), 'CSC019_CSU007': ('CMP/CmpDio/CmpDio_UpdateDso.c', 'Component Layer/Component DIO/DIO update DSOs.docx', 'CmpDio_UpdateDso.c'), 'CSC019_CSU000': ('CMP/CmpDio/CmpDio_Init.c', 'Component Layer/Component DIO/Initialize DIO CSC.docx', 'CmpDio_Init.c'), 'CSC019_CSU001': ('CMP/CmpDio/CmpDio_TestDsi.c', 'Component Layer/Component DIO/Test DSI.docx', 'CmpDio_TestDsi.c'), 'CSC019_CSU002': ('CMP/CmpDio/CmpDio_FailedMode.c', 'Component Layer/Component DIO/Set all DSOs to Failed Mode.docx', 'CmpDio_FailedMode.c'), 'CSC019_CSU003': ('CMP/CmpDio/CmpDio_ReadEbi.c', 'Component Layer/Component DIO/DIO Read EBI.docx', 'CmpDio_ReadEbi.c'), 'CSC019_CSU008': ('CMP/CmpDio/CmpDio_TestEbi.c', 'Component Layer/Component DIO/DIO test the EBI.docx', 'CmpDio_TestEbi.c'), 'CSC032_CSU028': ('APL/AplActuation/AplActuation_InitAcNetClosureSeq.c', 'Application Layer/Application Actuation/Initialize AC Network Closure Sequence.docx', 'AplActuation_InitAcNetClosureSeq.c'), 'CSC032_CSU021': ('APL/AplActuation/AplActuation_DsoCtcCmd.c', 'Application Layer/Application Actuation/Compute DSO Contactor Command.docx', 'AplActuation_DsoCtcCmd.c'), 'CSC032_CSU020': ('APL/AplActuation/AplActuation_DsoGcuAck.c', 'Application Layer/Application Actuation/Compute DSO Generator Control Unit acknowledgment.docx', 'AplActuation_DsoGcuAck.c'), 'CSC032_CSU023': ('APL/AplActuation/AplActuation_InitDsoNbpt.c', 'Application Layer/Application Actuation/Initialize Dso no break power transfer.docx', 'AplActuation_InitDsoNbpt.c'), 'CSC032_CSU022': ('APL/AplActuation/AplActuation_InitAcmpNetAcClosureSeq.c', 'Application Layer/Application Actuation/Initialize Acmp Ac network Closure Sequence.docx', 'AplActuation_InitAcmpNetAcClosureSeq.c'), 'CSC032_CSU025': ('APL/AplActuation/AplActuation_InitCheckProtForWaitingSeq.c', 'Application Layer/Application Actuation/Initialize Protection for waiting sequence.docx', 'AplActuation_InitCheckProtForWaitingSeq.c'), 'CSC032_CSU024': ('APL/AplActuation/AplActuation_CheckProtForWaitingSeq.c', 'Application Layer/Application Actuation/Check Protection For Waiting Sequence.docx', 'AplActuation_CheckProtForWaitingSeq.c'), 'CSC032_CSU027': ('APL/AplActuation/AplActuation_InitAcNetOpenSeq.c', 'Application Layer/Application Actuation/Initialize AC Network Open Sequence.docx', 'AplActuation_InitAcNetOpenSeq.c'), 'CSC032_CSU026': ('APL/AplActuation/AplActuation_InitAcmpNetEdmuClosureSeq.c', 'Application Layer/Application Actuation/Initialize Acmp Network Edmu Closure Sequence.docx', 'AplActuation_InitAcmpNetEdmuClosureSeq.c'), 'CSC034_CSU024': ('APL/AplContMonit/AplContMonit_ComputeAcmpOpenLckSte.c', 'Application Layer/Application Continuous Monitoring/Compute Acmp Open Locked State.docx', 'AplContMonit_ComputeAcmpOpenLckSte.c'), 'CSC040_CSU037': ('APL/AplProt/AplProt_InitProtAcEpUnderFreq.c', 'Application Layer/Application Protection/Initialize Ac Ep Under Frequency Protection.docx', 'AplProt_InitProtAcEpUnderFreq.c'), 'CSC040_CSU036': ('APL/AplProt/AplProt_InitProtAcEpUnderVlt.c', 'Application Layer/Application Protection/Initialize Ac Ep Under Voltage Protection.docx', 'AplProt_InitProtAcEpUnderVlt.c'), 'CSC040_CSU035': ('APL/AplProt/AplProt_InitProtAcEpOverVoltage.c', 'Application Layer/Application Protection/Initialize Ac Ep Over Voltage Protection.docx', 'AplProt_InitProtAcEpOverVoltage.c'), 'CSC040_CSU034': ('APL/AplProt/AplProt_InitProtAcEpOverload.c', 'Application Layer/Application Protection/Initialize Ac Ep Overload Protection Protection.docx', 'AplProt_InitProtAcEpOverload.c'), 'CSC040_CSU033': ('APL/AplProt/AplProt_DsoOlAcEp.c', 'Application Layer/Application Protection/Compute Dso Overload AC External Power.docx', 'AplProt_DsoOlAcEp.c'), 'CSC040_CSU032': ('APL/AplProt/AplProt_DsoAcEpPinF.c', 'Application Layer/Application Protection/Compute Dso AC External Power Pin F.docx', 'AplProt_DsoAcEpPinF.c'), 'CSC040_CSU031': ('APL/AplProt/AplProt_UpdateNvmProt.c', 'Application Layer/Application Protection/Update Protection status in NVM.docx', 'AplProt_UpdateNvmProt.c'), 'CSC014_CSU005': ('DRV/DrvSpi/DrvSpi_Read.c', 'Driver Layer/Driver SPI/Read data from SPI driver.docx', 'DrvSpi_Read.c'), 'CSC040_CSU039': ('APL/AplProt/AplProt_InitProtAcEpOpenPhase.c', 'Application Layer/Application Protection/Initialize Ac Ep Open Phase Protection.docx', 'AplProt_InitProtAcEpOpenPhase.c'), 'CSC040_CSU038': ('APL/AplProt/AplProt_InitProtAcEpOverFreq.c', 'Application Layer/Application Protection/Initialize Ac Ep Over Frequency Protection.docx', 'AplProt_InitProtAcEpOverFreq.c'), 'CSC014_CSU004': ('DRV/DrvSpi/DrvSpi_Init.c', 'Driver Layer/Driver SPI/Initialize SPI driver CSC.docx', 'DrvSpi_Init.c'), 'CSC026_CSU001': ('DRV/DrvCpu/DrvCpu_Init.c', 'Driver Layer/Driver CPU/Initialize CPU driver CSC.docx', 'DrvCpu_Init.c'), 'CSC035_CSU039': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2Bite1.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 Bite1 message.docx', 'AplEdmuCom_ComputeACLog2Bite1.c'), 'CSC021_CSU014': ('DPL/DplCanRx/DplCanRx_ConsolidateTimeDate.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate time and date.docx', 'DplCanRx_ConsolidateTimeDate.c'), 'CSC021_CSU015': ('DPL/DplCanRx/DplCanRx_ConsolidateFlightLeg.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate Flight Leg.docx', 'DplCanRx_ConsolidateFlightLeg.c'), 'CSC021_CSU016': ('DPL/DplCanRx/DplCanRx_ConsolidateFlightPhase.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate Flight Phase.docx', 'DplCanRx_ConsolidateFlightPhase.c'), 'CSC021_CSU017': ('DPL/DplCanRx/DplCanRx_ConsolidateAcTail.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate AC Tail.docx', 'DplCanRx_ConsolidateAcTail.c'), 'CSC021_CSU010': ('DPL/DplCanRx/DplCanRx_GetAllMessages.c', 'Data Processing Layer/Data Processing CAN reception/Get All CAN messages.docx', 'DplCanRx_GetAllMessages.c'), 'CSC021_CSU011': ('DPL/DplCanRx/DplCanRx_ReceiveMessage.c', 'Data Processing Layer/Data Processing CAN reception/Receive CAN message.docx', 'DplCanRx_ReceiveMessage.c'), 'CSC021_CSU012': ('DPL/DplCanRx/DplCanRx_CheckRccbCmdValid.c', 'Data Processing Layer/Data Processing CAN reception/Check Rccb Command Valid.docx', 'DplCanRx_CheckRccbCmdValid.c'), 'CSC021_CSU013': ('DPL/DplCanRx/DplCanRx_CheckCanValidity.c', 'Data Processing Layer/Data Processing CAN reception/Check CAN Validity.docx', 'DplCanRx_CheckCanValidity.c'), 'CSC021_CSU018': ('DPL/DplCanRx/DplCanRx_ConsolidateXfrLeft.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate xfr Left.docx', 'DplCanRx_ConsolidateXfrLeft.c'), 'CSC021_CSU019': ('DPL/DplCanRx/DplCanRx_ConsolidateXfrNorm.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate Xfr Normal.docx', 'DplCanRx_ConsolidateXfrNorm.c'), 'CSC034_CSU038': ('APL/AplContMonit/AplContMonit_InitComputeLockoutFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Lockout Failure.docx', 'AplContMonit_InitComputeLockoutFailure.c'), 'CSC034_CSU039': ('APL/AplContMonit/AplContMonit_ComputeLockoutFailure.c', 'Application Layer/Application Continuous Monitoring/Manage Lockout Failure.docx', 'AplContMonit_ComputeLockoutFailure.c'), 'CSC034_CSU034': ('APL/AplContMonit/AplContMonit_InitComputeDsoFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Dso Failure.docx', 'AplContMonit_InitComputeDsoFailure.c'), 'CSC034_CSU035': ('APL/AplContMonit/AplContMonit_ComputeDsoFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Dso Failure.docx', 'AplContMonit_ComputeDsoFailure.c'), 'CSC034_CSU036': ('APL/AplContMonit/AplContMonit_InitComputeWowFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Wow Failure.docx', 'AplContMonit_InitComputeWowFailure.c'), 'CSC034_CSU037': ('APL/AplContMonit/AplContMonit_ComputeWowFailure.c', 'Application Layer/Application Continuous Monitoring/Manage Wow Failure.docx', 'AplContMonit_ComputeWowFailure.c'), 'CSC034_CSU030': ('APL/AplContMonit/AplContMonit_ComputeOvlFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Overload Failures.docx', 'AplContMonit_ComputeOvlFailure.c'), 'CSC034_CSU031': ('APL/AplContMonit/AplContMonit_ComputePwrReadyFailure.c', 'Application Layer/Application Continuous Monitoring/Manage Power Ready Failure.docx', 'AplContMonit_ComputePwrReadyFailure.c'), 'CSC034_CSU032': ('APL/AplContMonit/AplContMonit_InitPwrReadyCntFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Power Ready Counter Failure.docx', 'AplContMonit_InitPwrReadyCntFailure.c'), 'CSC034_CSU033': ('APL/AplContMonit/AplContMonit_ComputeAntiPFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Anti Paralleling Failure.docx', 'AplContMonit_ComputeAntiPFailure.c'), 'CSC025_CSU029': ('DPL/DplNvm/DplNvm_ReadProtectionsCtx.c', 'Data Processing Layer/Data Processing NVM/Read Protections Ctx.docx', 'DplNvm_ReadProtectionsCtx.c'), 'CSC025_CSU028': ('DPL/DplNvm/DplNvm_ReadPinProgCtx.c', 'Data Processing Layer/Data Processing NVM/Read Pin Prog Ctx.docx', 'DplNvm_ReadPinProgCtx.c'), 'CSC025_CSU021': ('DPL/DplNvm/DplNvm_PrepareProtectionsCtxWrite.c', 'Data Processing Layer/Data Processing NVM/Prepare Protections Ctx Write.docx', 'DplNvm_PrepareProtectionsCtxWrite.c'), 'CSC025_CSU020': ('DPL/DplNvm/DplNvm_PreparePinProgCtxWrite.c', 'Data Processing Layer/Data Processing NVM/Prepare Pin Prog Ctx Write.docx', 'DplNvm_PreparePinProgCtxWrite.c'), 'CSC025_CSU023': ('DPL/DplNvm/DplNvm_ProcessWrite.c', 'Data Processing Layer/Data Processing NVM/Nvm Process Write.docx', 'DplNvm_ProcessWrite.c'), 'CSC025_CSU022': ('DPL/DplNvm/DplNvm_PrepareRccbStatusCtxWrite.c', 'Data Processing Layer/Data Processing NVM/Prepare Rccb Status Ctx Write.docx', 'DplNvm_PrepareRccbStatusCtxWrite.c'), 'CSC025_CSU025': ('DPL/DplNvm/DplNvm_ReadBiteHeader.c', 'Data Processing Layer/Data Processing NVM/Read Bite Header context from in Nvm.docx', 'DplNvm_ReadBiteHeader.c'), 'CSC025_CSU024': ('DPL/DplNvm/DplNvm_ReadBiteFailure.c', 'Data Processing Layer/Data Processing NVM/Read Bite Failure in Nvm.docx', 'DplNvm_ReadBiteFailure.c'), 'CSC025_CSU027': ('DPL/DplNvm/DplNvm_ReadHardwareSnPn.c', 'Data Processing Layer/Data Processing NVM/Read Hardware Sn Pn in Nvm.docx', 'DplNvm_ReadHardwareSnPn.c'), 'CSC025_CSU026': ('DPL/DplNvm/DplNvm_ReadCommandFailureCtx.c', 'Data Processing Layer/Data Processing NVM/Read Command Failure Ctx.docx', 'DplNvm_ReadCommandFailureCtx.c'), 'CSC035_CSU037': ('APL/AplEdmuCom/AplEdmuCom_ComputeModuleStatus.c', 'Application Layer/Application EDMU Communication/Compute Module Status message.docx', 'AplEdmuCom_ComputeModuleStatus.c'), 'CSC035_CSU036': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2RccbSts1.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 Rccb Status 1 message.docx', 'AplEdmuCom_ComputeACLog2RccbSts1.c'), 'CSC040_CSU055': ('APL/AplProt/AplProt_InitAcEpPinFState.c', 'Application Layer/Application Protection/Initialize Ac External power Pin F State.docx', 'AplProt_InitAcEpPinFState.c'), 'CSC040_CSU054': ('APL/AplProt/AplProt_InitCheckAcEpAct.c', 'Application Layer/Application Protection/Initialize Checking activation condition AC EP protections.docx', 'AplProt_InitCheckAcEpAct.c'), 'CSC040_CSU056': ('APL/AplProt/AplProt_ComputeAcEpPinFState.c', 'Application Layer/Application Protection/Compute AC external power pin F State.docx', 'AplProt_ComputeAcEpPinFState.c'), 'CSC040_CSU051': ('APL/AplProt/AplProt_InitProtAcEpPineOverVlt.c', 'Application Layer/Application Protection/Initialize Ac Ep Pine OverVoltage Protection.docx', 'AplProt_InitProtAcEpPineOverVlt.c'), 'CSC040_CSU050': ('APL/AplProt/AplProt_InitProtDifferential.c', 'Application Layer/Application Protection/Initialize Differential Protection.docx', 'AplProt_InitProtDifferential.c'), 'CSC076_CSU007': ('DRV/DrvGpio/DrvGpio_GetComponentAddr.c', 'Driver Layer/Driver GPIO/Get GPIO component address.docx', 'DrvGpio_GetComponentAddr.c'), 'CSC076_CSU003': ('DRV/DrvGpio/DrvGpio_Write.c', 'Driver Layer/Driver GPIO/Write Output.docx', 'DrvGpio_Write.c'), 'CSC076_CSU001': ('DRV/DrvGpio/DrvGpio_Init.c', 'Driver Layer/Driver GPIO/Initialize GPIO driver CSC.docx', 'DrvGpio_Init.c'), 'CSC040_CSU052': ('APL/AplProt/AplProt_InitProtAcEpPhaseOrder.c', 'Application Layer/Application Protection/Initialize Ac Ep Phase Order Protection.docx', 'AplProt_InitProtAcEpPhaseOrder.c'), 'CSC074_CSU001': ('CMP/CmpTcb/CmpTcb_Monitoring.c', 'Component Layer/Component TCB/Monitor TCB.docx', 'CmpTcb_Monitoring.c'), 'CSC074_CSU000': ('CMP/CmpTcb/CmpTcb_Init.c', 'Component Layer/Component TCB/Initialize TCB CSC.docx', 'CmpTcb_Init.c'), 'CSC074_CSU002': ('CMP/CmpTcb/CmpTcb_InitMonitoring.c', 'Component Layer/Component TCB/Initialize Monitoring CSU internal data.docx', 'CmpTcb_InitMonitoring.c'), 'CSC033_CSU017': ('APL/AplBite/AplBite_InternalPwrSupplyFailure.c', 'Application Layer/Application Bite/Compute internal power supply failure.docx', 'AplBite_InternalPwrSupplyFailure.c'), 'CSC033_CSU016': ('APL/AplBite/AplBite_PowerRdyFailure.c', 'Application Layer/Application Bite/Compute power ready failure.docx', 'AplBite_PowerRdyFailure.c'), 'CSC033_CSU015': ('APL/AplBite/AplBite_OverloadFailure.c', 'Application Layer/Application Bite/Compute overload failure.docx', 'AplBite_OverloadFailure.c'), 'CSC033_CSU014': ('APL/AplBite/AplBite_OtherDsoFailure.c', 'Application Layer/Application Bite/Compute other DSO failure.docx', 'AplBite_OtherDsoFailure.c'), 'CSC033_CSU013': ('APL/AplBite/AplBite_LockoutFailure.c', 'Application Layer/Application Bite/Compute lockout failure.docx', 'AplBite_LockoutFailure.c'), 'CSC033_CSU011': ('APL/AplBite/AplBite_FoFailure.c', 'Application Layer/Application Bite/Compute failed opened failure.docx', 'AplBite_FoFailure.c'), 'CSC033_CSU010': ('APL/AplBite/AplBite_FcFailure.c', 'Application Layer/Application Bite/Compute failed closed failure.docx', 'AplBite_FcFailure.c'), 'CSC022_CSU000': ('DPL/DplCanTx/DplCanTx_SendMessage.c', 'Data Processing Layer/Data Processing CAN transmission/Manage CAN message transmission.docx', 'DplCanTx_SendMessage.c'), 'CSC033_CSU018': ('APL/AplBite/AplBite_ProtectionFailure.c', 'Application Layer/Application Bite/Compute protection failure.docx', 'AplBite_ProtectionFailure.c'), 'CSC017_CSU002': ('CMP/CmpNvm/CmpNvm_ReadData.c', 'Component Layer/Component NVM/NVM read.docx', 'CmpNvm_ReadData.c'), 'CSC017_CSU003': ('CMP/CmpNvm/CmpNvm_GetStatus.c', 'Component Layer/Component NVM/NVM Get status.docx', 'CmpNvm_GetStatus.c'), 'CSC017_CSU001': ('CMP/CmpNvm/CmpNvm_Write.c', 'Component Layer/Component NVM/NVM write.docx', 'CmpNvm_Write.c'), 'CSC017_CSU006': ('CMP/CmpNvm/CmpNvm_Init.c', 'Component Layer/Component NVM/Initialize NVM CSC.docx', 'CmpNvm_Init.c'), 'CSC017_CSU007': ('CMP/CmpNvm/CmpNvm_InitReadData.c', 'Component Layer/Component NVM/Initialize internal data for read data.docx', 'CmpNvm_InitReadData.c'), 'CSC017_CSU004': ('CMP/CmpNvm/CmpNvm_TestNvm.c', 'Component Layer/Component NVM/NVM Test.docx', 'CmpNvm_TestNvm.c'), 'CSC017_CSU005': ('CMP/CmpNvm/CmpNvm_ReadRequest.c', 'Component Layer/Component NVM/NVM Read Request.docx', 'CmpNvm_ReadRequest.c'), 'CSC020_CSU019': ('DPL/DplAcq/DplAcq_InitComputeWowStatus.c', 'Data Processing Layer/Data Processing acquisition/Initialize Compute Wow Status.docx', 'DplAcq_InitComputeWowStatus.c'), 'CSC035_CSU008': ('APL/AplEdmuCom/AplEdmuCom_SendAnsNvmDlMsgs.c', 'Application Layer/Application EDMU Communication/Send answer to NVM download message.docx', 'AplEdmuCom_SendAnsNvmDlMsgs.c'), 'CSC035_CSU009': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2Analog2.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 Analog2 message.docx', 'AplEdmuCom_ComputeACLog2Analog2.c'), 'CSC035_CSU006': ('APL/AplEdmuCom/AplEdmuCom_SendAnsIdentMsgs.c', 'Application Layer/Application EDMU Communication/Send answer to identification message.docx', 'AplEdmuCom_SendAnsIdentMsgs.c'), 'CSC035_CSU007': ('APL/AplEdmuCom/AplEdmuCom_SendAnsNetworkStatusMsgs.c', 'Application Layer/Application EDMU Communication/Send answer to network status message.docx', 'AplEdmuCom_SendAnsNetworkStatusMsgs.c'), 'CSC035_CSU004': ('APL/AplEdmuCom/AplEdmuCom_SendAnsCbStatusMsgs.c', 'Application Layer/Application EDMU Communication/Send answer to CB status message.docx', 'AplEdmuCom_SendAnsCbStatusMsgs.c'), 'CSC035_CSU005': ('APL/AplEdmuCom/AplEdmuCom_SendAnsEngDataMsgs.c', 'Application Layer/Application EDMU Communication/Send answer to engineering data message.docx', 'AplEdmuCom_SendAnsEngDataMsgs.c'), 'CSC035_CSU002': ('APL/AplEdmuCom/AplEdmuCom_SendAnsAnalogStatusMsgs.c', 'Application Layer/Application EDMU Communication/Send answer to analog status message.docx', 'AplEdmuCom_SendAnsAnalogStatusMsgs.c'), 'CSC035_CSU003': ('APL/AplEdmuCom/AplEdmuCom_SendAnsBiteMsgs.c', 'Application Layer/Application EDMU Communication/Send answer to BITE message.docx', 'AplEdmuCom_SendAnsBiteMsgs.c'), 'CSC020_CSU015': ('DPL/DplAcq/DplAcq_InitComputeDsoFeedback.c', 'Data Processing Layer/Data Processing acquisition/Initialize Dso Compute FeedBack.docx', 'DplAcq_InitComputeDsoFeedback.c'), 'CSC035_CSU001': ('APL/AplEdmuCom/AplEdmuCom_Manager.c', 'Application Layer/Application EDMU Communication/Manage EDMU communication CSC.docx', 'AplEdmuCom_Manager.c'), 'CSC021_CSU029': ('DPL/DplCanRx/DplCanRx_InitCanValidity.c', 'Data Processing Layer/Data Processing CAN reception/Initialize frame validity CSU internal data.docx', 'DplCanRx_InitCanValidity.c'), 'CSC021_CSU028': ('DPL/DplCanRx/DplCanRx_ConsolidateData.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate data.docx', 'DplCanRx_ConsolidateData.c'), 'CSC021_CSU025': ('DPL/DplCanRx/DplCanRx_ConsolidateRequests.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate Requests.docx', 'DplCanRx_ConsolidateRequests.c'), 'CSC021_CSU024': ('DPL/DplCanRx/DplCanRx_ConsolidateIbitRequest.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate IBIT Request.docx', 'DplCanRx_ConsolidateIbitRequest.c'), 'CSC021_CSU027': ('DPL/DplCanRx/DplCanRx_ConsolidateNvmRstRequest.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate NVM Reset Request.docx', 'DplCanRx_ConsolidateNvmRstRequest.c'), 'CSC021_CSU026': ('DPL/DplCanRx/DplCanRx_ConsolidateNvmDlRequest.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate NVM Dl Request.docx', 'DplCanRx_ConsolidateNvmDlRequest.c'), 'CSC021_CSU021': ('DPL/DplCanRx/DplCanRx_ConsolidateXfrEss.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate Xfr Essential.docx', 'DplCanRx_ConsolidateXfrEss.c'), 'CSC021_CSU020': ('DPL/DplCanRx/DplCanRx_ConsolidateXfrRight.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate Xfr Right.docx', 'DplCanRx_ConsolidateXfrRight.c'), 'CSC021_CSU023': ('DPL/DplCanRx/DplCanRx_ConsolidateAcmp3B.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate Acmp3B.docx', 'DplCanRx_ConsolidateAcmp3B.c'), 'CSC021_CSU022': ('DPL/DplCanRx/DplCanRx_ConsolidateAcmp1B.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate Acmp1B.docx', 'DplCanRx_ConsolidateAcmp1B.c'), 'CSC034_CSU005': ('APL/AplContMonit/AplContMonit_Manager.c', 'Application Layer/Application Continuous Monitoring/Manage continuous monitoring CSC.docx', 'AplContMonit_Manager.c'), 'CSC034_CSU004': ('APL/AplContMonit/AplContMonit_Init.c', 'Application Layer/Application Continuous Monitoring/Initialize continuous monitoring CSC.docx', 'AplContMonit_Init.c'), 'CSC034_CSU007': ('APL/AplContMonit/AplContMonit_ComputeBusBarSplySrc.c', 'Application Layer/Application Continuous Monitoring/Compute bus bar supply source.docx', 'AplContMonit_ComputeBusBarSplySrc.c'), 'CSC034_CSU000': ('APL/AplContMonit/AplContMonit_ComputeAcEpPr.c', 'Application Layer/Application Continuous Monitoring/Compute AC External power ready.docx', 'AplContMonit_ComputeAcEpPr.c'), 'CSC034_CSU003': ('APL/AplContMonit/AplContMonit_ComputePowerSrcAvail.c', 'Application Layer/Application Continuous Monitoring/Compute power source availability.docx', 'AplContMonit_ComputePowerSrcAvail.c'), 'CSC034_CSU009': ('APL/AplContMonit/AplContMonit_ComputeSynopticColor.c', 'Application Layer/Application Continuous Monitoring/Compute synoptic colors.docx', 'AplContMonit_ComputeSynopticColor.c'), 'CSC034_CSU008': ('APL/AplContMonit/AplContMonit_ComputeEicasMsg.c', 'Application Layer/Application Continuous Monitoring/Compute EICAS messages.docx', 'AplContMonit_ComputeEicasMsg.c'), 'CSC060_CSU001': ('SRV/SrvMem/SrvMem_MemCpy.c', 'Service Layer/Service Memory/Memory Copy.docx', 'SrvMem_MemCpy.c'), 'CSC072_CSU001': ('CMP/CmpLed/CmpLed_CmdLedInterface.c', 'Component Layer/Component LED/Command the LED interface.docx', 'CmpLed_CmdLedInterface.c'), 'CSC072_CSU000': ('CMP/CmpLed/CmpLed_Init.c', 'Component Layer/Component LED/Initialize LED CSC.docx', 'CmpLed_Init.c'), 'CSC060_CSU004': ('SRV/SrvMem/SrvMem_Crc16bit.c', 'Service Layer/Service Memory/Compute CRC 16bit.docx', 'SrvMem_Crc16bit.c'), 'CSC047_CSU003': ('APL/AplPowerUpMonit/AplPowerUpMonit_TestHwSw.c', 'Application Layer/Application Power Up monitoring/Perform Hardware Software Compatibility test.docx', 'AplPowerUpMonit_TestHwSw.c'), 'CSC047_CSU002': ('APL/AplPowerUpMonit/AplPowerUpMonit_TestPbit.c', 'Application Layer/Application Power Up monitoring/Run all PBIT test.docx', 'AplPowerUpMonit_TestPbit.c'), 'CSC047_CSU001': ('APL/AplPowerUpMonit/AplPowerUpMonit_Startup.c', 'Application Layer/Application Power Up monitoring/Manage the Power up monitoring CSC Startup.docx', 'AplPowerUpMonit_Startup.c'), 'CSC047_CSU000': ('APL/AplPowerUpMonit/AplPowerUpMonit_Init.c', 'Application Layer/Application Power Up monitoring/Initialize power up monitoring CSC.docx', 'AplPowerUpMonit_Init.c'), 'CSC047_CSU004': ('APL/AplPowerUpMonit/AplPowerUpMonit_TestOV.c', 'Application Layer/Application Power Up monitoring/Perform Over Voltage test.docx', 'AplPowerUpMonit_TestOV.c'), 'CSC040_CSU020': ('APL/AplProt/AplProt_Acmp3AUnbalancedProt.c', 'Application Layer/Application Protection/Compute ACMP3A unbalanced protection.docx', 'AplProt_Acmp3AUnbalancedProt.c'), 'CSC040_CSU021': ('APL/AplProt/AplProt_AntiParallAtc1Atc3Prot.c', 'Application Layer/Application Protection/Compute ATC1 ATC3 anti paralleling protection.docx', 'AplProt_AntiParallAtc1Atc3Prot.c'), 'CSC040_CSU022': ('APL/AplProt/AplProt_AntiParallAtc2ElcProt.c', 'Application Layer/Application Protection/Compute ATC2 ELC anti paralleling protection.docx', 'AplProt_AntiParallAtc2ElcProt.c'), 'CSC040_CSU023': ('APL/AplProt/AplProt_CheckAcEpAct.c', 'Application Layer/Application Protection/Check activation condition AC EP protections.docx', 'AplProt_CheckAcEpAct.c'), 'CSC040_CSU024': ('APL/AplProt/AplProt_DifferentialProt.c', 'Application Layer/Application Protection/Compute differential protection.docx', 'AplProt_DifferentialProt.c'), 'CSC040_CSU025': ('APL/AplProt/AplProt_ProtStateComputation.c', 'Application Layer/Application Protection/Compute protection state.docx', 'AplProt_ProtStateComputation.c'), 'CSC040_CSU026': ('APL/AplProt/AplProt_ProtCondition.c', 'Application Layer/Application Protection/Compute activation condition of protection.docx', 'AplProt_ProtCondition.c'), 'CSC040_CSU027': ('APL/AplProt/AplProt_Startup.c', 'Application Layer/Application Protection/Manage protection CSC startup.docx', 'AplProt_Startup.c'), 'CSC040_CSU028': ('APL/AplProt/AplProt_AcEpOverloadProt.c', 'Application Layer/Application Protection/Compute AC external power overload protection.docx', 'AplProt_AcEpOverloadProt.c'), 'CSC040_CSU029': ('APL/AplProt/AplProt_AcEpOpenPhaseProt.c', 'Application Layer/Application Protection/Compute AC external power open phase protection.docx', 'AplProt_AcEpOpenPhaseProt.c'), 'CSC034_CSU011': ('APL/AplContMonit/AplContMonit_ComputeAlcFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Alc Failure.docx', 'AplContMonit_ComputeAlcFailure.c'), 'CSC036_CSU003': ('APL/AplInitiatMonit/AplInitiatMonit_CheckIbitAcceptance.c', 'Application Layer/Application Initiated Monitoring/Check Ibit Acceptance.docx', 'AplInitiatMonit_CheckIbitAcceptance.c'), 'CSC036_CSU002': ('APL/AplInitiatMonit/AplInitiatMonit_Startup.c', 'Application Layer/Application Initiated Monitoring/Manage Initiated monitoring CSC Startup.docx', 'AplInitiatMonit_Startup.c'), 'CSC036_CSU001': ('APL/AplInitiatMonit/AplInitiatMonit_Init.c', 'Application Layer/Application Initiated Monitoring/Initialize initiated monitoring CSC.docx', 'AplInitiatMonit_Init.c'), 'CSC036_CSU000': ('APL/AplInitiatMonit/AplInitiatMonit_Manager.c', 'Application Layer/Application Initiated Monitoring/Manage initiated monitoring CSC.docx', 'AplInitiatMonit_Manager.c'), 'CSC036_CSU006': ('APL/AplInitiatMonit/AplInitiatMonit_InitCheckIbitAcceptance.c', 'Application Layer/Application Initiated Monitoring/Initialize Ibit Acceptance.docx', 'AplInitiatMonit_InitCheckIbitAcceptance.c'), 'CSC036_CSU005': ('APL/AplInitiatMonit/AplInitiatMonit_StoreCmdState.c', 'Application Layer/Application Initiated Monitoring/Store Command State.docx', 'AplInitiatMonit_StoreCmdState.c'), 'CSC036_CSU004': ('APL/AplInitiatMonit/AplInitiatMonit_RestoreCmdState.c', 'Application Layer/Application Initiated Monitoring/Restore Command State.docx', 'AplInitiatMonit_RestoreCmdState.c'), 'CSC035_CSU042': ('APL/AplEdmuCom/AplEdmuCom_InitCheckNvmLoadAcceptance.c', 'Application Layer/Application EDMU Communication/Initialize Ibit and NVM requests acceptance.docx', 'AplEdmuCom_InitCheckNvmLoadAcceptance.c'), 'CSC035_CSU043': ('APL/AplEdmuCom/AplEdmuCom_ResetBuffers.c', 'Application Layer/Application EDMU Communication/Reset CAN Rx Buffers.docx', 'AplEdmuCom_ResetBuffers.c'), 'CSC035_CSU040': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2Color1.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 synoptic color1 message.docx', 'AplEdmuCom_ComputeACLog2Color1.c'), 'CSC035_CSU041': ('APL/AplEdmuCom/AplEdmuCom_CheckIbitNvmAcceptance.c', 'Application Layer/Application EDMU Communication/Check Ibit Nvm Acceptance.docx', 'AplEdmuCom_CheckIbitNvmAcceptance.c'), 'CSC035_CSU046': ('APL/AplEdmuCom/AplEdmuCom_InitComputeEngData.c', 'Application Layer/Application EDMU Communication/Initialize Compute Engeenering Data.docx', 'AplEdmuCom_InitComputeEngData.c'), 'CSC035_CSU047': ('APL/AplEdmuCom/AplEdmuCom_InitSendAnsNvmDlMsgs.c', 'Application Layer/Application EDMU Communication/Initialize Send answer to NVM download message.docx', 'AplEdmuCom_InitSendAnsNvmDlMsgs.c'), 'CSC035_CSU044': ('APL/AplEdmuCom/AplEdmuCom_InitTimeSlotManagement.c', 'Application Layer/Application EDMU Communication/Initialize Time Slot Management.docx', 'AplEdmuCom_InitTimeSlotManagement.c'), 'CSC035_CSU045': ('APL/AplEdmuCom/AplEdmuCom_ComputeEngData.c', 'Application Layer/Application EDMU Communication/Compute Engineering Data message.docx', 'AplEdmuCom_ComputeEngData.c'), 'CSC034_CSU045': ('APL/AplContMonit/AplContMonit_InitFusesCntFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Fuses Counter Failure.docx', 'AplContMonit_InitFusesCntFailure.c'), 'CSC019_CSU012': ('CMP/CmpDio/CmpDio_InitTestEbi.c', 'Component Layer/Component DIO/Initialize Test Ebi.docx', 'CmpDio_InitTestEbi.c'), 'CSC034_CSU049': ('APL/AplContMonit/AplContMonit_ComputeInternalPwrSplyFailure.c', 'Application Layer/Application Continuous Monitoring/Manage Internal Power Supply Failure.docx', 'AplContMonit_ComputeInternalPwrSplyFailure.c'), 'CSC034_CSU048': ('APL/AplContMonit/AplContMonit_ComputeConverterFailure.c', 'Application Layer/Application Continuous Monitoring/Compute internal converter Failures.docx', 'AplContMonit_ComputeConverterFailure.c'), 'CSC034_CSU041': ('APL/AplContMonit/AplContMonit_ComputeAcEpGsMode.c', 'Application Layer/Application Continuous Monitoring/Compute Ac External power and Ground servicing Mode.docx', 'AplContMonit_ComputeAcEpGsMode.c'), 'CSC034_CSU040': ('APL/AplContMonit/AplContMonit_ComputeAcEpGsRequest.c', 'Application Layer/Application Continuous Monitoring/Compute Ac External power and Ground servicing Request.docx', 'AplContMonit_ComputeAcEpGsRequest.c'), 'CSC034_CSU043': ('APL/AplContMonit/AplContMonit_ComputePwrSplyPresFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Power Supply Presence Failure.docx', 'AplContMonit_ComputePwrSplyPresFailure.c'), 'CSC034_CSU042': ('APL/AplContMonit/AplContMonit_DsoPrAcEp.c', 'Application Layer/Application Continuous Monitoring/Compute Dso Power ready Ac External power.docx', 'AplContMonit_DsoPrAcEp.c'), 'CSC064_CSU000': ('CMP/CmpError/CmpError_SwError.c', 'Component Layer/Component Error/Manage Critical SW Error.docx', 'CmpError_SwError.c'), 'CSC034_CSU044': ('APL/AplContMonit/AplContMonit_InitPwrSplyPresCntFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Power Supply Presence Counter Failure.docx', 'AplContMonit_InitPwrSplyPresCntFailure.c'), 'CSC034_CSU047': ('APL/AplContMonit/AplContMonit_InitConverterCntFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Converter Counter Failure.docx', 'AplContMonit_InitConverterCntFailure.c'), 'CSC034_CSU046': ('APL/AplContMonit/AplContMonit_ComputeFusesFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Fuses Failure.docx', 'AplContMonit_ComputeFusesFailure.c'), 'CSC025_CSU038': ('DPL/DplNvm/DplNvm_InitProcessDumpNvm.c', 'Data Processing Layer/Data Processing NVM/Initialize Process Dump Nvm.docx', 'DplNvm_InitProcessDumpNvm.c'), 'CSC025_CSU032': ('DPL/DplNvm/DplNvm_Restore.c', 'Data Processing Layer/Data Processing NVM/Dpl Nvm Restore.docx', 'DplNvm_Restore.c'), 'CSC025_CSU033': ('DPL/DplNvm/DplNvm_InitCheckRequests.c', 'Data Processing Layer/Data Processing NVM/Initialize Check Requests.docx', 'DplNvm_InitCheckRequests.c'), 'CSC025_CSU030': ('DPL/DplNvm/DplNvm_ReadRccbStatusCtx.c', 'Data Processing Layer/Data Processing NVM/Read Rccb Status Ctx.docx', 'DplNvm_ReadRccbStatusCtx.c'), 'CSC025_CSU031': ('DPL/DplNvm/DplNvm_ResetBiteFailures.c', 'Data Processing Layer/Data Processing NVM/Reset Bite Failures.docx', 'DplNvm_ResetBiteFailures.c'), 'CSC025_CSU036': ('DPL/DplNvm/DplNvm_PrepareDumpNvm.c', 'Data Processing Layer/Data Processing NVM/Prepare Dump Nvm.docx', 'DplNvm_PrepareDumpNvm.c'), 'CSC025_CSU037': ('DPL/DplNvm/DplNvm_ProcessDumpNvm.c', 'Data Processing Layer/Data Processing NVM/Process Dump Nvm.docx', 'DplNvm_ProcessDumpNvm.c'), 'CSC025_CSU034': ('DPL/DplNvm/DplNvm_InitManager.c', 'Data Processing Layer/Data Processing NVM/Initialize Manager.docx', 'DplNvm_InitManager.c'), 'CSC025_CSU035': ('DPL/DplNvm/DplNvm_InitResetBiteFailures.c', 'Data Processing Layer/Data Processing NVM/Initialize Reset Bite Failures.docx', 'DplNvm_InitResetBiteFailures.c'), 'CSC020_CSU013': ('DPL/DplAcq/DplAcq_TestAsi.c', 'Data Processing Layer/Data Processing acquisition/Test Asi.docx', 'DplAcq_TestAsi.c'), 'CSC020_CSU012': ('DPL/DplAcq/DplAcq_ReadAsi.c', 'Data Processing Layer/Data Processing acquisition/Read Asi.docx', 'DplAcq_ReadAsi.c'), 'CSC020_CSU011': ('DPL/DplAcq/DplAcq_ComputeTransparency.c', 'Data Processing Layer/Data Processing acquisition/Compute Transparency.docx', 'DplAcq_ComputeTransparency.c'), 'CSC020_CSU010': ('DPL/DplAcq/DplAcq_ComputeStartUpMode.c', 'Data Processing Layer/Data Processing acquisition/Compute Start Up Mode.docx', 'DplAcq_ComputeStartUpMode.c'), 'CSC020_CSU016': ('DPL/DplAcq/DplAcq_InitReadAsi.c', 'Data Processing Layer/Data Processing acquisition/Initialize Read Asi.docx', 'DplAcq_InitReadAsi.c'), 'CSC035_CSU000': ('APL/AplEdmuCom/AplEdmuCom_Init.c', 'Application Layer/Application EDMU Communication/Initialize EDMU communication CSC.docx', 'AplEdmuCom_Init.c'), 'CSC020_CSU014': ('DPL/DplAcq/DplAcq_TestCurrentTransfo.c', 'Data Processing Layer/Data Processing acquisition/Test Current Transfo.docx', 'DplAcq_TestCurrentTransfo.c'), 'CSC018_CSU009': ('DRV/DrvEmios/DrvEmios_GetInputPulsePeriod.c', 'Driver Layer/Driver EMIOS/Get Input Pulse Period.docx', 'DrvEmios_GetInputPulsePeriod.c'), 'CSC018_CSU003': ('DRV/DrvEmios/DrvEmios_GetChannelStatus.c', 'Driver Layer/Driver EMIOS/Get channel status.docx', 'DrvEmios_GetChannelStatus.c'), 'CSC018_CSU002': ('DRV/DrvEmios/DrvEmios_Init.c', 'Driver Layer/Driver EMIOS/Initialize EMIOS driver CSC.docx', 'DrvEmios_Init.c'), 'CSC018_CSU007': ('DRV/DrvEmios/DrvEmios_GetCurrentTime.c', 'Driver Layer/Driver EMIOS/Get Current Time.docx', 'DrvEmios_GetCurrentTime.c'), 'CSC018_CSU006': ('DRV/DrvEmios/DrvEmios_ClearChannelFlag.c', 'Driver Layer/Driver EMIOS/Clear channel flag.docx', 'DrvEmios_ClearChannelFlag.c'), 'CSC018_CSU004': ('DRV/DrvEmios/DrvEmios_CmdOutputPulse.c', 'Driver Layer/Driver EMIOS/Command PWM.docx', 'DrvEmios_CmdOutputPulse.c'), 'CSC016_CSU009': ('DRV/DrvCan/DrvCan_ConfigBuffers.c', 'Driver Layer/Driver CAN/Configure CAN Message Buffers.docx', 'DrvCan_ConfigBuffers.c'), 'CSC016_CSU008': ('DRV/DrvCan/DrvCan_Config.c', 'Driver Layer/Driver CAN/Configure CAN.docx', 'DrvCan_Config.c'), 'CSC016_CSU001': ('DRV/DrvCan/DrvCan_Init.c', 'Driver Layer/Driver CAN/Initialize CAN driver CSC.docx', 'DrvCan_Init.c'), 'CSC040_CSU010': ('APL/AplProt/AplProt_Acmp1BI2tProt.c', 'Application Layer/Application Protection/Compute ACMP1B I2T protection.docx', 'AplProt_Acmp1BI2tProt.c'), 'CSC016_CSU003': ('DRV/DrvCan/DrvCan_Read.c', 'Driver Layer/Driver CAN/Read CAN Message.docx', 'DrvCan_Read.c'), 'CSC016_CSU002': ('DRV/DrvCan/DrvCan_Write.c', 'Driver Layer/Driver CAN/Write CAN Message.docx', 'DrvCan_Write.c'), 'CSC016_CSU004': ('DRV/DrvCan/DrvCan_ClearTxBuffers.c', 'Driver Layer/Driver CAN/Clear Tx buffers.docx', 'DrvCan_ClearTxBuffers.c'), 'CSC016_CSU007': ('DRV/DrvCan/DrvCan_ChangeRxMask.c', 'Driver Layer/Driver CAN/Change Rx Mask.docx', 'DrvCan_ChangeRxMask.c'), 'CSC040_CSU016': ('APL/AplProt/AplProt_Acmp3BUnbalancedProt.c', 'Application Layer/Application Protection/Compute ACMP3B unbalanced protection.docx', 'AplProt_Acmp3BUnbalancedProt.c'), 'CSC033_CSU000': ('APL/AplBite/AplBite_Manager.c', 'Application Layer/Application Bite/Manage Bite CSC.docx', 'AplBite_Manager.c'), 'CSC033_CSU001': ('APL/AplBite/AplBite_Init.c', 'Application Layer/Application Bite/Initialize Bite CSC.docx', 'AplBite_Init.c'), 'CSC033_CSU002': ('APL/AplBite/AplBite_5sPowerCutFailure.c', 'Application Layer/Application Bite/Compute 5s Power Cut failure.docx', 'AplBite_5sPowerCutFailure.c'), 'CSC033_CSU003': ('APL/AplBite/AplBite_AntiParallelingFailure.c', 'Application Layer/Application Bite/Compute antiparalleling failure.docx', 'AplBite_AntiParallelingFailure.c'), 'CSC033_CSU004': ('APL/AplBite/AplBite_AsiFailure.c', 'Application Layer/Application Bite/Compute ASI failure.docx', 'AplBite_AsiFailure.c'), 'CSC033_CSU005': ('APL/AplBite/AplBite_AuxFailure.c', 'Application Layer/Application Bite/Compute auxiliary failure.docx', 'AplBite_AuxFailure.c'), 'CSC033_CSU007': ('APL/AplBite/AplBite_CtcDsoCommandFailure.c', 'Application Layer/Application Bite/Compute contactor DSO command failure.docx', 'AplBite_CtcDsoCommandFailure.c'), 'CSC033_CSU008': ('APL/AplBite/AplBite_CtFailure.c', 'Application Layer/Application Bite/Compute current transformer failure.docx', 'AplBite_CtFailure.c'), 'CSC033_CSU009': ('APL/AplBite/AplBite_DsiFailure.c', 'Application Layer/Application Bite/Compute DSI failure.docx', 'AplBite_DsiFailure.c'), 'CSC023_CSU001': ('DRV/DrvPbridge/DrvPbridge_Init.c', 'Driver Layer/Driver PBRIDGE/Initialize Pbridge driver CSC.docx', 'DrvPbridge_Init.c'), 'CSC020_CSU004': ('DPL/DplAcq/DplAcq_ComputeWowStatus.c', 'Data Processing Layer/Data Processing acquisition/Compute WoW status.docx', 'DplAcq_ComputeWowStatus.c'), 'CSC020_CSU005': ('DPL/DplAcq/DplAcq_ComputeWowInit.c', 'Data Processing Layer/Data Processing acquisition/Compute WoW initialization.docx', 'DplAcq_ComputeWowInit.c'), 'CSC020_CSU006': ('DPL/DplAcq/DplAcq_Init.c', 'Data Processing Layer/Data Processing acquisition/Initialize acquisition CSC.docx', 'DplAcq_Init.c'), 'CSC035_CSU038': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog1Bite1.c', 'Application Layer/Application EDMU Communication/Compute ACLog1 Bite1 message.docx', 'AplEdmuCom_ComputeACLog1Bite1.c'), 'CSC020_CSU000': ('DPL/DplAcq/DplAcq_ComputeAcEpData.c', 'Data Processing Layer/Data Processing acquisition/Compute AC External Power data.docx', 'DplAcq_ComputeAcEpData.c'), 'CSC020_CSU001': ('DPL/DplAcq/DplAcq_ComputeAtcLockout.c', 'Data Processing Layer/Data Processing acquisition/Compute ATC lockout.docx', 'DplAcq_ComputeAtcLockout.c'), 'CSC020_CSU002': ('DPL/DplAcq/DplAcq_ComputeBoardCfg.c', 'Data Processing Layer/Data Processing acquisition/Compute board configuration.docx', 'DplAcq_ComputeBoardCfg.c'), 'CSC035_CSU033': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog1AlarmStatus.c', 'Application Layer/Application EDMU Communication/Compute ACLog1 Alarm Status message.docx', 'AplEdmuCom_ComputeACLog1AlarmStatus.c'), 'CSC035_CSU032': ('APL/AplEdmuCom/AplEdmuCom_ComputeCntorStatus.c', 'Application Layer/Application EDMU Communication/Compute Contactor Status message.docx', 'AplEdmuCom_ComputeCntorStatus.c'), 'CSC035_CSU031': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2NetworkSts1.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 Network Status 1 message.docx', 'AplEdmuCom_ComputeACLog2NetworkSts1.c'), 'CSC035_CSU030': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog1NetworkSts1.c', 'Application Layer/Application EDMU Communication/Compute ACLog1 Network Status 1 message.docx', 'AplEdmuCom_ComputeACLog1NetworkSts1.c'), 'CSC020_CSU008': ('DPL/DplAcq/DplAcq_Manager.c', 'Data Processing Layer/Data Processing acquisition/Manage Acquisition.docx', 'DplAcq_Manager.c'), 'CSC020_CSU009': ('DPL/DplAcq/DplAcq_ComputeDsoFeedback.c', 'Data Processing Layer/Data Processing acquisition/Compute Dso Feedback.docx', 'DplAcq_ComputeDsoFeedback.c'), 'CSC035_CSU035': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog1RccbSts2.c', 'Application Layer/Application EDMU Communication/Compute ACLog1 Rccb Status 2 message.docx', 'AplEdmuCom_ComputeACLog1RccbSts2.c'), 'CSC035_CSU034': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2AlarmStatus.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 Alarm Status message.docx', 'AplEdmuCom_ComputeACLog2AlarmStatus.c'), 'CSC029_CSU002': ('CTL/CtlIsr/CtlIsr_UnhandledIsr.c', 'Control Layer/Control ISR/Manage unhandled ISR.docx', 'CtlIsr_UnhandledIsr.c'), 'CSC029_CSU000': ('CTL/CtlIsr/CtlIsr_MajorTaskIsr.c', 'Control Layer/Control ISR/Manage major task ISR.docx', 'CtlIsr_MajorTaskIsr.c'), 'CSC021_CSU032': ('DPL/DplCanRx/DplCanRx_ConsolidateAcmp3A.c', 'Data Processing Layer/Data Processing CAN reception/Consolidate Acmp3A.docx', 'DplCanRx_ConsolidateAcmp3A.c'), 'CSC021_CSU030': ('DPL/DplCanRx/DplCanRx_InitReceiveMessage.c', 'Data Processing Layer/Data Processing CAN reception/Initialize Receive Message CSU internal data.docx', 'DplCanRx_InitReceiveMessage.c'), 'CSC021_CSU031': ('DPL/DplCanRx/DplCanRx_InitRequestCoherency.c', 'Data Processing Layer/Data Processing CAN reception/Initialize Request Coherency.docx', 'DplCanRx_InitRequestCoherency.c'), 'CSC034_CSU016': ('APL/AplContMonit/AplContMonit_InitLastDsiAcepGsPb.c', 'Application Layer/Application Continuous Monitoring/Initialize last value of Dsi GS and ACEP push button.docx', 'AplContMonit_InitLastDsiAcepGsPb.c'), 'CSC034_CSU017': ('APL/AplContMonit/AplContMonit_InitAcBusAvail.c', 'Application Layer/Application Continuous Monitoring/Initialize Ac Bus Bar Available.docx', 'AplContMonit_InitAcBusAvail.c'), 'CSC034_CSU014': ('APL/AplContMonit/AplContMonit_ComputeAecFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Aec Failure.docx', 'AplContMonit_ComputeAecFailure.c'), 'CSC034_CSU015': ('APL/AplContMonit/AplContMonit_ComputeRccbFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Rccb Failure.docx', 'AplContMonit_ComputeRccbFailure.c'), 'CSC034_CSU012': ('APL/AplContMonit/AplContMonit_ComputeElcFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Elc Failure.docx', 'AplContMonit_ComputeElcFailure.c'), 'CSC034_CSU013': ('APL/AplContMonit/AplContMonit_ComputeAtcFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Atc Failure.docx', 'AplContMonit_ComputeAtcFailure.c'), 'CSC034_CSU010': ('APL/AplContMonit/AplContMonit_ComputeGlcFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Glc Failure.docx', 'AplContMonit_ComputeGlcFailure.c'), 'CSC079_CSU001': ('DRV/DrvXbar/DrvXbar_Init.c', 'Driver Layer/Driver XBAR/Initialize Xbar driver CSC.docx', 'DrvXbar_Init.c'), 'CSC034_CSU018': ('APL/AplContMonit/AplContMonit_ComputeAcBusAvail.c', 'Application Layer/Application Continuous Monitoring/Compute Ac Bus Available.docx', 'AplContMonit_ComputeAcBusAvail.c'), 'CSC034_CSU019': ('APL/AplContMonit/AplContMonit_DsoAcEpAvail.c', 'Application Layer/Application Continuous Monitoring/Compute Dso Ac External power Available.docx', 'AplContMonit_DsoAcEpAvail.c'), 'CSC077_CSU000': ('DRV/DrvInterrupt/DrvInterrupt_Init.c', 'Driver Layer/Driver Interrupt/Initialize interrupt driver CSC.docx', 'DrvInterrupt_Init.c'), 'CSC077_CSU001': ('DRV/DrvInterrupt/DrvInterrupt_DisableInterrupts.c', 'Driver Layer/Driver Interrupt/Disable external interrupts.docx', 'DrvInterrupt_DisableInterrupts.c'), 'CSC077_CSU002': ('DRV/DrvInterrupt/DrvInterrupt_EnableInterrupts.c', 'Driver Layer/Driver Interrupt/Enable external interrupts.docx', 'DrvInterrupt_EnableInterrupts.c'), 'CSC040_CSU000': ('APL/AplProt/AplProt_Manager.c', 'Application Layer/Application Protection/Manage protection CSC.docx', 'AplProt_Manager.c'), 'CSC024_CSU001': ('DRV/DrvPll/DrvPll_Init.c', 'Driver Layer/Driver PLL/Initialize PLL driver CSC.docx', 'DrvPll_Init.c'), 'CSC040_CSU001': ('APL/AplProt/AplProt_Init.c', 'Application Layer/Application Protection/Initialize protection CSC.docx', 'AplProt_Init.c'), 'CSC032_CSU007': ('APL/AplActuation/AplActuation_ComputeAcmp3BCmd.c', 'Application Layer/Application Actuation/Compute ACMP3B command.docx', 'AplActuation_ComputeAcmp3BCmd.c'), 'CSC032_CSU006': ('APL/AplActuation/AplActuation_ComputeAcmp1BCmd.c', 'Application Layer/Application Actuation/Compute ACMP1B command.docx', 'AplActuation_ComputeAcmp1BCmd.c'), 'CSC032_CSU005': ('APL/AplActuation/AplActuation_ComputeAcmp3ACmd.c', 'Application Layer/Application Actuation/Compute ACMP3A command.docx', 'AplActuation_ComputeAcmp3ACmd.c'), 'CSC032_CSU004': ('APL/AplActuation/AplActuation_AcNetOpenSeq.c', 'Application Layer/Application Actuation/Manage AC network open sequence.docx', 'AplActuation_AcNetOpenSeq.c'), 'CSC032_CSU003': ('APL/AplActuation/AplActuation_AcNetClosureSeq.c', 'Application Layer/Application Actuation/Manage AC network closure sequence.docx', 'AplActuation_AcNetClosureSeq.c'), 'CSC032_CSU002': ('APL/AplActuation/AplActuation_AcmpNetOpenSeq.c', 'Application Layer/Application Actuation/Manage ACMP network open sequence.docx', 'AplActuation_AcmpNetOpenSeq.c'), 'CSC032_CSU001': ('APL/AplActuation/AplActuation_AcmpNetEdmuClosureSeq.c', 'Application Layer/Application Actuation/Manage ACMP network EDMU closure sequence.docx', 'AplActuation_AcmpNetEdmuClosureSeq.c'), 'CSC032_CSU000': ('APL/AplActuation/AplActuation_AcmpNetAcClosureSeq.c', 'Application Layer/Application Actuation/Manage ACMP network AC closure sequence.docx', 'AplActuation_AcmpNetAcClosureSeq.c'), 'CSC078_CSU001': ('DRV/DrvMmu/DrvMmu_Init.c', 'Driver Layer/Driver MMU/Initialize MMU driver CSC.docx', 'DrvMmu_Init.c'), 'CSC032_CSU009': ('APL/AplActuation/AplActuation_ComputeAlcCmd.c', 'Application Layer/Application Actuation/Compute ALC command.docx', 'AplActuation_ComputeAlcCmd.c'), 'CSC032_CSU008': ('APL/AplActuation/AplActuation_ComputeAecCmd.c', 'Application Layer/Application Actuation/Compute AEC command.docx', 'AplActuation_ComputeAecCmd.c'), 'CSC015_CSU012': ('DRV/DrvAdc/DrvAdc_Calibrate.c', 'Driver Layer/Driver ADC/Calibrate ADC.docx', 'DrvAdc_Calibrate.c'), 'CSC015_CSU011': ('DRV/DrvAdc/DrvAdc_GetAdcTable.c', 'Driver Layer/Driver ADC/Get adc table.docx', 'DrvAdc_GetAdcTable.c'), 'CSC015_CSU010': ('DRV/DrvAdc/DrvAdc_StartUp.c', 'Driver Layer/Driver ADC/ADC Start up.docx', 'DrvAdc_StartUp.c'), 'CSC075_CSU001': ('DRV/DrvFlash/DrvFlash_Init.c', 'Driver Layer/Driver FLASH/Initialize FLASH driver CSC.docx', 'DrvFlash_Init.c'), 'CSC070_CSU001': ('CMP/CmpWdg/CmpWdg_Refresh.c', 'Component Layer/Component Watchdog/Refresh watchdog.docx', 'CmpWdg_Refresh.c'), 'CSC070_CSU003': ('CMP/CmpWdg/CmpWdg_Init.c', 'Component Layer/Component Watchdog/Initialize Watchdog CSC.docx', 'CmpWdg_Init.c'), 'CSC070_CSU002': ('CMP/CmpWdg/CmpWdg_TestWatchdog.c', 'Component Layer/Component Watchdog/Test the watchdog.docx', 'CmpWdg_TestWatchdog.c'), 'CSC040_CSU046': ('APL/AplProt/AplProt_InitProtAcmp3AUnbalanced.c', 'Application Layer/Application Protection/Initialize Acmp3A Unbalanced Protection.docx', 'AplProt_InitProtAcmp3AUnbalanced.c'), 'CSC040_CSU045': ('APL/AplProt/AplProt_InitProtAcmp3BI2t.c', 'Application Layer/Application Protection/Initialize Acmp3B I2t Protection.docx', 'AplProt_InitProtAcmp3BI2t.c'), 'CSC040_CSU042': ('APL/AplProt/AplProt_InitProtAcmp1BI2t.c', 'Application Layer/Application Protection/Initialize Acmp1B I2t Protection.docx', 'AplProt_InitProtAcmp1BI2t.c'), 'CSC040_CSU043': ('APL/AplProt/AplProt_InitProtAcmp3BUnbalanced.c', 'Application Layer/Application Protection/Initialize Acmp3B Unbalanced Protection.docx', 'AplProt_InitProtAcmp3BUnbalanced.c'), 'CSC040_CSU040': ('APL/AplProt/AplProt_InitProtAcmp1BUnbalanced.c', 'Application Layer/Application Protection/Initialize Acmp1B Unbalanced Protection.docx', 'AplProt_InitProtAcmp1BUnbalanced.c'), 'CSC040_CSU048': ('APL/AplProt/AplProt_InitProtAcmp3AI2t.c', 'Application Layer/Application Protection/Initialize Acmp3A I2t Protection.docx', 'AplProt_InitProtAcmp3AI2t.c'), 'CSC040_CSU049': ('APL/AplProt/AplProt_InitProtAcmpAcBusOverVlt.c', 'Application Layer/Application Protection/Initialize Acmp Ac Bus Over Voltage Protection.docx', 'AplProt_InitProtAcmpAcBusOverVlt.c'), 'CSC028_CSU000': ('CTL/CtlBoot/CtlBoot_Init.c', 'Control Layer/Control Boot/Initialize boot CSC.docx', 'CtlBoot_Init.c'), 'CSC028_CSU001': ('CTL/CtlBoot/CtlBoot_Manager.c', 'Control Layer/Control Boot/Manage boot CSC.docx', 'CtlBoot_Manager.c'), 'CSC034_CSU058': ('APL/AplContMonit/AplContMonit_UpdateNvmRccbStatus.c', 'Application Layer/Application Continuous Monitoring/Update Nvm Rccb Status.docx', 'AplContMonit_UpdateNvmRccbStatus.c'), 'CSC034_CSU059': ('APL/AplContMonit/AplContMonit_UpdateNvmCmdFailure.c', 'Application Layer/Application Continuous Monitoring/Update Nvm Command Failure.docx', 'AplContMonit_UpdateNvmCmdFailure.c'), 'CSC034_CSU052': ('APL/AplContMonit/AplContMonit_ComputeGlcStatus.c', 'Application Layer/Application Continuous Monitoring/Compute Glc Status.docx', 'AplContMonit_ComputeGlcStatus.c'), 'CSC034_CSU053': ('APL/AplContMonit/AplContMonit_ComputeAlcStatus.c', 'Application Layer/Application Continuous Monitoring/Compute Alc Status.docx', 'AplContMonit_ComputeAlcStatus.c'), 'CSC034_CSU050': ('APL/AplContMonit/AplContMonit_5sPowerLossMonitoring.c', 'Application Layer/Application Continuous Monitoring/Manage 5s Power Loss Monitoring.docx', 'AplContMonit_5sPowerLossMonitoring.c'), 'CSC034_CSU051': ('APL/AplContMonit/AplContMonit_Init5sPowerLoss.c', 'Application Layer/Application Continuous Monitoring/Initialize 5s Power Loss monitoring.docx', 'AplContMonit_Init5sPowerLoss.c'), 'CSC034_CSU056': ('APL/AplContMonit/AplContMonit_ComputeAecStatus.c', 'Application Layer/Application Continuous Monitoring/Compute Aec Status.docx', 'AplContMonit_ComputeAecStatus.c'), 'CSC034_CSU057': ('APL/AplContMonit/AplContMonit_ComputeRccbStatus.c', 'Application Layer/Application Continuous Monitoring/Compute Rccb Status.docx', 'AplContMonit_ComputeRccbStatus.c'), 'CSC034_CSU054': ('APL/AplContMonit/AplContMonit_ComputeElcStatus.c', 'Application Layer/Application Continuous Monitoring/Compute Elc Status.docx', 'AplContMonit_ComputeElcStatus.c'), 'CSC034_CSU055': ('APL/AplContMonit/AplContMonit_ComputeAtcStatus.c', 'Application Layer/Application Continuous Monitoring/Compute Atc Status.docx', 'AplContMonit_ComputeAtcStatus.c'), 'CSC025_CSU005': ('DPL/DplNvm/DplNvm_PrepareBiteFailureWrite.c', 'Data Processing Layer/Data Processing NVM/Prepare Bite Failure Write.docx', 'DplNvm_PrepareBiteFailureWrite.c'), 'CSC025_CSU001': ('DPL/DplNvm/DplNvm_Manager.c', 'Data Processing Layer/Data Processing NVM/Manage data processing NVM CSC.docx', 'DplNvm_Manager.c'), 'CSC025_CSU000': ('DPL/DplNvm/DplNvm_Init.c', 'Data Processing Layer/Data Processing NVM/Initialize data processing NVM CSC.docx', 'DplNvm_Init.c'), 'CSC025_CSU009': ('DPL/DplNvm/DplNvm_PrepareCommandFailureCtxWrite.c', 'Data Processing Layer/Data Processing NVM/Prepare Command Failure Ctx Write.docx', 'DplNvm_PrepareCommandFailureCtxWrite.c'), 'CSC025_CSU008': ('DPL/DplNvm/DplNvm_PrepareBiteHeaderWrite.c', 'Data Processing Layer/Data Processing NVM/Prepare Bite Header Write.docx', 'DplNvm_PrepareBiteHeaderWrite.c'), 'CSC071_CSU000': ('CMP/CmpInputCapt/CmpInputCapt_Init.c', 'Component Layer/Component Input Capture/Initialize Input Capture CSC.docx', 'CmpInputCapt_Init.c'), 'CSC071_CSU001': ('CMP/CmpInputCapt/CmpInputCapt_ComputeSignalPeriod.c', 'Component Layer/Component Input Capture/Compute signal period.docx', 'CmpInputCapt_ComputeSignalPeriod.c'), 'CSC060_CSU002': ('SRV/SrvMem/SrvMem_MemSet.c', 'Service Layer/Service Memory/Memory Set.docx', 'SrvMem_MemSet.c'), 'CSC060_CSU003': ('SRV/SrvMem/SrvMem_CheckSum8bit.c', 'Service Layer/Service Memory/Memory CheckSum.docx', 'SrvMem_CheckSum8bit.c'), 'CSC034_CSU063': ('APL/AplContMonit/AplContMonit_ComputeTcbFailure.c', 'Application Layer/Application Continuous Monitoring/Compute Tcb Failure.docx', 'AplContMonit_ComputeTcbFailure.c'), 'CSC019_CSU011': ('CMP/CmpDio/CmpDio_ResetHardProt.c', 'Component Layer/Component DIO/DIO Reset Hard Prot.docx', 'CmpDio_ResetHardProt.c'), 'CSC034_CSU062': ('APL/AplContMonit/AplContMonit_InitComputeTcbFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Compute Tcb Failure.docx', 'AplContMonit_InitComputeTcbFailure.c'), 'CSC013_CSU006': ('DRV/DrvDma/DrvDma_Init.c', 'Driver Layer/Driver DMA/Initialize DMA driver CSC.docx', 'DrvDma_Init.c'), 'CSC013_CSU007': ('DRV/DrvDma/DrvDma_Read.c', 'Driver Layer/Driver DMA/Read data from DMA driver.docx', 'DrvDma_Read.c'), 'CSC013_CSU004': ('DRV/DrvDma/DrvDma_GetLocalDataAddr.c', 'Driver Layer/Driver DMA/Get Local Data address.docx', 'DrvDma_GetLocalDataAddr.c'), 'CSC013_CSU005': ('DRV/DrvDma/DrvDma_GetTransferStatus.c', 'Driver Layer/Driver DMA/Get Transfer status.docx', 'DrvDma_GetTransferStatus.c'), 'CSC013_CSU002': ('DRV/DrvDma/DrvDma_Enable.c', 'Driver Layer/Driver DMA/Enable DMA driver.docx', 'DrvDma_Enable.c'), 'CSC013_CSU001': ('DRV/DrvDma/DrvDma_Disable.c', 'Driver Layer/Driver DMA/Disable DMA driver.docx', 'DrvDma_Disable.c'), 'CSC013_CSU008': ('DRV/DrvDma/DrvDma_Write.c', 'Driver Layer/Driver DMA/Write data to DMA driver.docx', 'DrvDma_Write.c'), 'CSC013_CSU009': ('DRV/DrvDma/DrvDma_Clear.c', 'Driver Layer/Driver DMA/Clear Dma.docx', 'DrvDma_Clear.c'), 'CSC040_CSU008': ('APL/AplProt/AplProt_AcEpUnderVoltageProt.c', 'Application Layer/Application Protection/Compute AC external power under voltage protection.docx', 'AplProt_AcEpUnderVoltageProt.c'), 'CSC040_CSU003': ('APL/AplProt/AplProt_AcEpOverFreqProt.c', 'Application Layer/Application Protection/Compute AC external power over frequency protection.docx', 'AplProt_AcEpOverFreqProt.c'), 'CSC016_CSU010': ('DRV/DrvCan/DrvCan_Freeze.c', 'Driver Layer/Driver CAN/Freeze mode.docx', 'DrvCan_Freeze.c'), 'CSC016_CSU011': ('DRV/DrvCan/DrvCan_SetModuleParameters.c', 'Driver Layer/Driver CAN/Set Module Parameters.docx', 'DrvCan_SetModuleParameters.c'), 'CSC040_CSU006': ('APL/AplProt/AplProt_AcEpOverVoltageProt.c', 'Application Layer/Application Protection/Compute AC external power over voltage protection.docx', 'AplProt_AcEpOverVoltageProt.c'), 'CSC040_CSU007': ('APL/AplProt/AplProt_AcEpPhaseOrderProt.c', 'Application Layer/Application Protection/Compute AC external power phase order protection.docx', 'AplProt_AcEpPhaseOrderProt.c'), 'CSC040_CSU004': ('APL/AplProt/AplProt_AcEpUnderFreqProt.c', 'Application Layer/Application Protection/Compute AC external power under frequency protection.docx', 'AplProt_AcEpUnderFreqProt.c'), 'CSC040_CSU005': ('APL/AplProt/AplProt_AcEpPineOverVoltageProt.c', 'Application Layer/Application Protection/Compute AC external PINE over voltage protection.docx', 'AplProt_AcEpPineOverVoltageProt.c'), 'CSC035_CSU026': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2Analog1.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 Analog1 message.docx', 'AplEdmuCom_ComputeACLog2Analog1.c'), 'CSC035_CSU027': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog1Current1.c', 'Application Layer/Application EDMU Communication/Compute ACLog1 Current1 message.docx', 'AplEdmuCom_ComputeACLog1Current1.c'), 'CSC035_CSU020': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2IOState1.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 IO State1 message.docx', 'AplEdmuCom_ComputeACLog2IOState1.c'), 'CSC035_CSU021': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2IOState2.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 IO State2 message.docx', 'AplEdmuCom_ComputeACLog2IOState2.c'), 'CSC035_CSU022': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2IOState4.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 IO State4 message.docx', 'AplEdmuCom_ComputeACLog2IOState4.c'), 'CSC035_CSU023': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2IOState5.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 IO State5 message.docx', 'AplEdmuCom_ComputeACLog2IOState5.c'), 'CSC035_CSU028': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2Current1.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 Current1 message.docx', 'AplEdmuCom_ComputeACLog2Current1.c'), 'CSC035_CSU029': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog2Current2.c', 'Application Layer/Application EDMU Communication/Compute ACLog2 Current2 message.docx', 'AplEdmuCom_ComputeACLog2Current2.c'), 'CSC021_CSU009': ('DPL/DplCanRx/DplCanRx_ReadXfr.c', 'Data Processing Layer/Data Processing CAN reception/Read XFR status.docx', 'DplCanRx_ReadXfr.c'), 'CSC073_CSU001': ('DRV/DrvEbi/DrvEbi_Init.c', 'Driver Layer/Driver EBI/Initialize EBI driver CSC.docx', 'DrvEbi_Init.c'), 'CSC021_CSU000': ('DPL/DplCanRx/DplCanRx_Init.c', 'Data Processing Layer/Data Processing CAN reception/Initialize CAN reception CSC.docx', 'DplCanRx_Init.c'), 'CSC021_CSU007': ('DPL/DplCanRx/DplCanRx_ReadRccbCmd.c', 'Data Processing Layer/Data Processing CAN reception/Read RCCB command.docx', 'DplCanRx_ReadRccbCmd.c'), 'CSC021_CSU006': ('DPL/DplCanRx/DplCanRx_ReadDateAndTime.c', 'Data Processing Layer/Data Processing CAN reception/Read date and time.docx', 'DplCanRx_ReadDateAndTime.c'), 'CSC021_CSU004': ('DPL/DplCanRx/DplCanRx_DecodeMsg.c', 'Data Processing Layer/Data Processing CAN reception/Decode CAN messages.docx', 'DplCanRx_DecodeMsg.c'), 'CSC034_CSU022': ('APL/AplContMonit/AplContMonit_DsoAcEpOn.c', 'Application Layer/Application Continuous Monitoring/Compute Dso Ac External power On.docx', 'AplContMonit_DsoAcEpOn.c'), 'CSC034_CSU021': ('APL/AplContMonit/AplContMonit_DsoAcBusAvail.c', 'Application Layer/Application Continuous Monitoring/Compute Dso Ac Bus Available.docx', 'AplContMonit_DsoAcBusAvail.c'), 'CSC034_CSU020': ('APL/AplContMonit/AplContMonit_DsoGsModeOn.c', 'Application Layer/Application Continuous Monitoring/Compute Dso Ground service Mode On.docx', 'AplContMonit_DsoGsModeOn.c'), 'CSC034_CSU027': ('APL/AplContMonit/AplContMonit_ResetTripStates.c', 'Application Layer/Application Continuous Monitoring/Reset Trip States.docx', 'AplContMonit_ResetTripStates.c'), 'CSC034_CSU026': ('APL/AplContMonit/AplContMonit_ResetCtcFailure.c', 'Application Layer/Application Continuous Monitoring/Reset Contactor Failure.docx', 'AplContMonit_ResetCtcFailure.c'), 'CSC034_CSU025': ('APL/AplContMonit/AplContMonit_ComputeAcmpTrippedSte.c', 'Application Layer/Application Continuous Monitoring/Compute Acmp Tripped State.docx', 'AplContMonit_ComputeAcmpTrippedSte.c'), 'CSC014_CSU008': ('DRV/DrvSpi/DrvSpi_ComputeRequest.c', 'Driver Layer/Driver SPI/Compute SPI Request.docx', 'DrvSpi_ComputeRequest.c'), 'CSC014_CSU007': ('DRV/DrvSpi/DrvSpi_Write.c', 'Driver Layer/Driver SPI/Write data to SPI driver.docx', 'DrvSpi_Write.c'), 'CSC014_CSU006': ('DRV/DrvSpi/DrvSpi_GetStatus.c', 'Driver Layer/Driver SPI/Read SPI Status.docx', 'DrvSpi_GetStatus.c'), 'CSC034_CSU029': ('APL/AplContMonit/AplContMonit_InitOvlCntFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Overload Counter Failure.docx', 'AplContMonit_InitOvlCntFailure.c'), 'CSC034_CSU028': ('APL/AplContMonit/AplContMonit_ResetTripCauses.c', 'Application Layer/Application Continuous Monitoring/Reset Trip Causes.docx', 'AplContMonit_ResetTripCauses.c'), 'CSC014_CSU003': ('DRV/DrvSpi/DrvSpi_Enable.c', 'Driver Layer/Driver SPI/Enable SPI driver.docx', 'DrvSpi_Enable.c'), 'CSC014_CSU002': ('DRV/DrvSpi/DrvSpi_Disable.c', 'Driver Layer/Driver SPI/Disable SPI driver.docx', 'DrvSpi_Disable.c'), 'CSC014_CSU001': ('DRV/DrvSpi/DrvSpi_Config.c', 'Driver Layer/Driver SPI/Configure SPI driver.docx', 'DrvSpi_Config.c'), 'CSC040_CSU019': ('APL/AplProt/AplProt_AcmpAcBusOverVoltageProt.c', 'Application Layer/Application Protection/Compute ACMP AC bus over voltage protection.docx', 'AplProt_AcmpAcBusOverVoltageProt.c'), 'CSC040_CSU018': ('APL/AplProt/AplProt_Acmp3AI2tProt.c', 'Application Layer/Application Protection/Compute ACMP3A I2T protection.docx', 'AplProt_Acmp3AI2tProt.c'), 'CSC031_CSU000': ('CTL/CtlStartup/CtlStartup_Manager.c', 'Control Layer/Control Start Up/Manage startup CSC.docx', 'CtlStartup_Manager.c'), 'CSC032_CSU010': ('APL/AplActuation/AplActuation_ComputeAtc1Cmd.c', 'Application Layer/Application Actuation/Compute ATC1 command.docx', 'AplActuation_ComputeAtc1Cmd.c'), 'CSC032_CSU011': ('APL/AplActuation/AplActuation_ComputeAtc2Cmd.c', 'Application Layer/Application Actuation/Compute ATC2 command.docx', 'AplActuation_ComputeAtc2Cmd.c'), 'CSC032_CSU012': ('APL/AplActuation/AplActuation_ComputeAtc3Cmd.c', 'Application Layer/Application Actuation/Compute ATC3 command.docx', 'AplActuation_ComputeAtc3Cmd.c'), 'CSC032_CSU013': ('APL/AplActuation/AplActuation_ComputeElcCmd.c', 'Application Layer/Application Actuation/Compute ELC command.docx', 'AplActuation_ComputeElcCmd.c'), 'CSC032_CSU014': ('APL/AplActuation/AplActuation_ComputeGlc1Cmd.c', 'Application Layer/Application Actuation/Compute GLC1 command.docx', 'AplActuation_ComputeGlc1Cmd.c'), 'CSC032_CSU015': ('APL/AplActuation/AplActuation_ComputeGlc2Cmd.c', 'Application Layer/Application Actuation/Compute GLC2 command.docx', 'AplActuation_ComputeGlc2Cmd.c'), 'CSC032_CSU016': ('APL/AplActuation/AplActuation_Init.c', 'Application Layer/Application Actuation/Initialize actuation CSC.docx', 'AplActuation_Init.c'), 'CSC032_CSU017': ('APL/AplActuation/AplActuation_Manager.c', 'Application Layer/Application Actuation/Manage actuation CSC.docx', 'AplActuation_Manager.c'), 'CSC032_CSU018': ('APL/AplActuation/AplActuation_NetworkStatus.c', 'Application Layer/Application Actuation/Compute network Status.docx', 'AplActuation_NetworkStatus.c'), 'CSC032_CSU019': ('APL/AplActuation/AplActuation_DsoNbpt.c', 'Application Layer/Application Actuation/Compute DSO No Break Power Transfer.docx', 'AplActuation_DsoNbpt.c'), 'CSC061_CSU001': ('SRV/SrvDelay/SrvDelay_Us.c', 'Service Layer/Service Delay/Delay in Microseconds.docx', 'SrvDelay_Us.c'), 'CSC015_CSU001': ('DRV/DrvAdc/DrvAdc_Init.c', 'Driver Layer/Driver ADC/Initialize ADC driver CSC.docx', 'DrvAdc_Init.c'), 'CSC015_CSU002': ('DRV/DrvAdc/DrvAdc_SetCfifoOperationMode.c', 'Driver Layer/Driver ADC/Set CFIFO Operation Mode.docx', 'DrvAdc_SetCfifoOperationMode.c'), 'CSC015_CSU003': ('DRV/DrvAdc/DrvAdc_Enable.c', 'Driver Layer/Driver ADC/Enable ADC.docx', 'DrvAdc_Enable.c'), 'CSC015_CSU006': ('DRV/DrvAdc/DrvAdc_Write.c', 'Driver Layer/Driver ADC/Write configuration for CFIFO.docx', 'DrvAdc_Write.c'), 'CSC015_CSU007': ('DRV/DrvAdc/DrvAdc_SetTransferMode.c', 'Driver Layer/Driver ADC/Set the transfer Mode.docx', 'DrvAdc_SetTransferMode.c'), 'CSC040_CSU012': ('APL/AplProt/AplProt_Acmp1BUnbalancedProt.c', 'Application Layer/Application Protection/Compute ACMP1B unbalanced protection.docx', 'AplProt_Acmp1BUnbalancedProt.c'), 'CSC040_CSU014': ('APL/AplProt/AplProt_Acmp3BI2tProt.c', 'Application Layer/Application Protection/Compute ACMP3B I2T protection.docx', 'AplProt_Acmp3BI2tProt.c'), 'CSC033_CSU028': ('APL/AplBite/AplBite_DsoAcLogValid.c', 'Application Layer/Application Bite/Compute DSO Ac Log Valid.docx', 'AplBite_DsoAcLogValid.c'), 'CSC033_CSU029': ('APL/AplBite/AplBite_InitManager.c', 'Application Layer/Application Bite/Initialize Manage bite CSC.docx', 'AplBite_InitManager.c'), 'CSC033_CSU022': ('APL/AplBite/AplBite_WatchdogFailure.c', 'Application Layer/Application Bite/Compute watchdog failure.docx', 'AplBite_WatchdogFailure.c'), 'CSC033_CSU023': ('APL/AplBite/AplBite_WowFailure.c', 'Application Layer/Application Bite/Compute wow failure.docx', 'AplBite_WowFailure.c'), 'CSC033_CSU020': ('APL/AplBite/AplBite_Storage.c', 'Application Layer/Application Bite/Manage failure storage.docx', 'AplBite_Storage.c'), 'CSC033_CSU021': ('APL/AplBite/AplBite_TcbFailure.c', 'Application Layer/Application Bite/Compute thermal circuit breaker failure.docx', 'AplBite_TcbFailure.c'), 'CSC033_CSU026': ('APL/AplBite/AplBite_ComputeEngineeringData.c', 'Application Layer/Application Bite/Compute Engineering Data.docx', 'AplBite_ComputeEngineeringData.c'), 'CSC033_CSU027': ('APL/AplBite/AplBite_DsoHealthLed.c', 'Application Layer/Application Bite/Compute DSO health led.docx', 'AplBite_DsoHealthLed.c'), 'CSC033_CSU024': ('APL/AplBite/AplBite_ComputeFailure.c', 'Application Layer/Application Bite/Compute BITE Failure data.docx', 'AplBite_ComputeFailure.c'), 'CSC033_CSU025': ('APL/AplBite/AplBite_Startup.c', 'Application Layer/Application Bite/Manage Bite CSC Startup.docx', 'AplBite_Startup.c'), 'CSC048_CSU002': ('APL/AplPrelimMonit/AplPrelimMonit_CheckRamIntegrity.c', 'Application Layer/Application Preliminary monitoring/Check the RAM integrity.docx', 'AplPrelimMonit_CheckRamIntegrity.c'), 'CSC048_CSU003': ('APL/AplPrelimMonit/AplPrelimMonit_Startup.c', 'Application Layer/Application Preliminary monitoring/Manage Preliminary monitoring CSC Start Up.docx', 'AplPrelimMonit_Startup.c'), 'CSC048_CSU006': ('APL/AplPrelimMonit/AplPrelimMonit_TestRam.c', 'Application Layer/Application Preliminary monitoring/Perform RAM test.docx', 'AplPrelimMonit_TestRam.c'), 'CSC048_CSU007': ('APL/AplPrelimMonit/AplPrelimMonit_TestRom.c', 'Application Layer/Application Preliminary monitoring/Perform ROM test.docx', 'AplPrelimMonit_TestRom.c'), 'CSC048_CSU005': ('APL/AplPrelimMonit/AplPrelimMonit_TestPinProg.c', 'Application Layer/Application Preliminary monitoring/Perform pin programming test.docx', 'AplPrelimMonit_TestPinProg.c'), 'CSC030_CSU001': ('CTL/CtlSchd/CtlSchd_Init.c', 'Control Layer/Control Scheduler/Initialize scheduler CSC.docx', 'CtlSchd_Init.c'), 'CSC030_CSU000': ('CTL/CtlSchd/CtlSchd_Manager.c', 'Control Layer/Control Scheduler/Manage scheduler CSC.docx', 'CtlSchd_Manager.c'), 'CSC030_CSU003': ('CTL/CtlSchd/CtlSchd_Manager1MS.c', 'Control Layer/Control Scheduler/Manage 1ms task.docx', 'CtlSchd_Manager1MS.c'), 'CSC030_CSU005': ('CTL/CtlSchd/CtlSchd_StartTaskTime.c', 'Control Layer/Control Scheduler/Compute start task time.docx', 'CtlSchd_StartTaskTime.c'), 'CSC030_CSU004': ('CTL/CtlSchd/CtlSchd_EndTaskTime.c', 'Control Layer/Control Scheduler/Compute end task time.docx', 'CtlSchd_EndTaskTime.c'), 'CSC030_CSU007': ('CTL/CtlSchd/CtlSchd_StackChecker.c', 'Control Layer/Control Scheduler/Check stack state.docx', 'CtlSchd_StackChecker.c'), 'CSC030_CSU006': ('CTL/CtlSchd/CtlSchd_ManagePowerCut.c', 'Control Layer/Control Scheduler/Manage Power Cut.docx', 'CtlSchd_ManagePowerCut.c'), 'CSC030_CSU008': ('CTL/CtlSchd/CtlSchd_InitManagePowerCut.c', 'Control Layer/Control Scheduler/Initialize Manage Power Cut.docx', 'CtlSchd_InitManagePowerCut.c'), 'CSC034_CSU067': ('APL/AplContMonit/AplContMonit_InitComputeOwnCtcChattering.c', 'Application Layer/Application Continuous Monitoring/Initialize Own Ctc Chattering.docx', 'AplContMonit_InitComputeOwnCtcChattering.c'), 'CSC034_CSU066': ('APL/AplContMonit/AplContMonit_InitComputeAcmpEdmuCmd.c', 'Application Layer/Application Continuous Monitoring/Initialize Compute Acmp Edmu Command.docx', 'AplContMonit_InitComputeAcmpEdmuCmd.c'), 'CSC034_CSU065': ('APL/AplContMonit/AplContMonit_ComputeAcmpEdmuCmd.c', 'Application Layer/Application Continuous Monitoring/Compute Acmp Edmu Command.docx', 'AplContMonit_ComputeAcmpEdmuCmd.c'), 'CSC034_CSU064': ('APL/AplContMonit/AplContMonit_InitComputeAntiPFailure.c', 'Application Layer/Application Continuous Monitoring/Initialize Compute Anti Paralleling Failure.docx', 'AplContMonit_InitComputeAntiPFailure.c'), 'CSC035_CSU019': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog1IOState5.c', 'Application Layer/Application EDMU Communication/Compute ACLog1 IO State5 message.docx', 'AplEdmuCom_ComputeACLog1IOState5.c'), 'CSC035_CSU018': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog1IOState4.c', 'Application Layer/Application EDMU Communication/Compute ACLog1 IO State4 message.docx', 'AplEdmuCom_ComputeACLog1IOState4.c'), 'CSC034_CSU061': ('APL/AplContMonit/AplContMonit_InitAcEpPrCounter.c', 'Application Layer/Application Continuous Monitoring/Initialize AC External power ready counter.docx', 'AplContMonit_InitAcEpPrCounter.c'), 'CSC034_CSU060': ('APL/AplContMonit/AplContMonit_Startup.c', 'Application Layer/Application Continuous Monitoring/Manage Continuous Monitoring CSC Startup.docx', 'AplContMonit_Startup.c'), 'CSC035_CSU015': ('APL/AplEdmuCom/AplEdmuCom_UpdateMsgs.c', 'Application Layer/Application EDMU Communication/Update Messages.docx', 'AplEdmuCom_UpdateMsgs.c'), 'CSC035_CSU017': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog1IOState2.c', 'Application Layer/Application EDMU Communication/Compute ACLog1 IO State2 message.docx', 'AplEdmuCom_ComputeACLog1IOState2.c'), 'CSC035_CSU016': ('APL/AplEdmuCom/AplEdmuCom_ComputeACLog1IOState1.c', 'Application Layer/Application EDMU Communication/Compute ACLog1 IO State1 message.docx', 'AplEdmuCom_ComputeACLog1IOState1.c'), 'CSC035_CSU011': ('APL/AplEdmuCom/AplEdmuCom_ComputeCBStatus2.c', 'Application Layer/Application EDMU Communication/Compute CB Status 2 message.docx', 'AplEdmuCom_ComputeCBStatus2.c'), 'CSC035_CSU010': ('APL/AplEdmuCom/AplEdmuCom_ComputeCBStatus1.c', 'Application Layer/Application EDMU Communication/Compute CB status 1 message.docx', 'AplEdmuCom_ComputeCBStatus1.c'), 'CSC035_CSU013': ('APL/AplEdmuCom/AplEdmuCom_StopAnswering.c', 'Application Layer/Application EDMU Communication/Stop CAN answering.docx', 'AplEdmuCom_StopAnswering.c'), 'CSC035_CSU012': ('APL/AplEdmuCom/AplEdmuCom_ControlCanActivation.c', 'Application Layer/Application EDMU Communication/Control Can Activation.docx', 'AplEdmuCom_ControlCanActivation.c'), 'CSC025_CSU019': ('DPL/DplNvm/DplNvm_CheckRequests.c', 'Data Processing Layer/Data Processing NVM/Can Check Request.docx', 'DplNvm_CheckRequests.c'), 'CSC022_CSU001': ('DPL/DplCanTx/DplCanTx_ClearBuffers.c', 'Data Processing Layer/Data Processing CAN transmission/Clear CAN message buffers.docx', 'DplCanTx_ClearBuffers.c')}

    test.exportXlsScod(tbl_list_req,hlr_vs_llr,llr_vs_code)
    exit()
    csci_name = "ACENM"
    dir_swrd = "C:\Users\olivier.appere\Desktop\Projets\g7000\SW_ACENM_01_34\SW_ACENM\SwRD"
    dir_swdd = "C:\Users\olivier.appere\Desktop\Projets\g7000\SW_ACENM_01_34\SW_ACENM\SWDD\LLR\Service Layer\Service Memory"
    dir_swdd = "C:\Users\olivier.appere\Desktop\Projets\g7000\SW_ACENM_01_34\SW_ACENM\SWDD\LLR\Application Layer"
    xml_csci = "C:\Users\olivier.appere\Desktop\Projets\g7000\SW_ACENM_01_34\SW_ACENM\Tools\Design\Docs\Template\design_ACENM.csci"
    test._buildSCOD(csci_name)
    exit()
    import html2text
    h = html2text.HTML2Text()
    h.escape_snob = False
    print h.handle("<p>- Hello, world.</p>")
    exit()
    #stdout = "<cell>2642</cell><cell>Defect</cell><cell>Software LED < on board does not blink > at proper frequency</cell><adjustCRcell>SCR_In_Review</cell><cell>Minor</cell><cell>SW_BITE/04</cell><cell>SW_BITE/05</cell><cell></cell><cell>08/12/15 15:19</cell><cell>SW_BITE code</cell><cell>SCR</cell><cell>SW_BITE</cell><cell>EASA - Type 2</cell>"
    stdout = '<cell>2649</cell><cell>Defect</cell><cell>Frequency lower bound monitoring are failed</cell><cell>SCR_In_Review</cell><cell>Medium</cell><cell>SW_BITE/04</cell><cell>SW_BITE/05</cell><cell></cell><cell>10/12/15 12:49</cell><cell>Code</cell><cell></cell><cell><span style="font-size: 10.6667px; line-height: normal; background-color: rgb(255, 255, 255);">ESSNESS:</span><div><span style="font-size: 10.6667px; line-height: normal; background-color: rgb(255, 255, 255);">rotation speed of PPDB fan and TRU fan is badly computed by BITE microcontroller</span><div style="font-size: 10.6667px; line-height: normal; background-color: rgb(255, 255, 255);">- so that<span style="font-size: 10.6667px; line-height: 1;">�CBIT_PPDB_FAN is badly computed (only computed by BITE)</span></div><div style="font-size: 10.6667px; line-height: normal; background-color: rgb(255, 255, 255);"><span style="font-size: 10.6667px; line-height: 1;">- so that CBIT_TRU_FAN is badly computed by BITE (computation by FUNC microcontroller is OK)</span></div></div><div style="font-size: 10.6667px; line-height: normal; background-color: rgb(255, 255, 255);"><span style="font-size: 10.6667px; line-height: 1;">, </span></div><div style="font-size: 10.6667px; line-height: normal; background-color: rgb(255, 255, 255);"><span style="font-size: 10.6667px; line-height: 1;">SDSIO:</span></div><div style="font-size: 10.6667px; line-height: normal; background-color: rgb(255, 255, 255);"><span style="font-size: 10.6667px;">No functional impact if FAN rotation is within normal range or fan is stopped (0Hz)</span><div style="font-size: 10.6667px;"><span style="font-size: 10.6667px;">If 0Hz  strictly lesser than FAN_FREQ  strictly lesser than 100Hz, than FAN failure could be undetected</span><span style="font-size: 13.3333px; line-height: 1;">�.</span></div></div></cell>'
    stdout = "<cell>198</cell><cell>Defect</cell><cell>TCB statuses inverted on the EMERLOG CAN</cell><cell>SACR_Fixed</cell><cell>Major</cell><cell>S0</cell><cell>S0</cell><cell>EQT_EMERLOG_02_02 (SW delivery in baseline SW_BITE_01_21) - correction in the SSCS in EQT_EMERLOG_02_04</cell><cell>11/02/16 18:19</cell><cell>- Assembly Board impact: 402CE06L0201Y02br /SW impact: Yesbr /HW impact: Nonebr /PLD impact: Nonebr /br /- SSCS impact: ET2925-S issue 3br /br /- ICD SPI ICD MCU BITE/FUNC PLDb /bimpact: Nobr /br /- HPID impact: Nobr /br /- HSID impact: Nobr /br /- ETP: nonebr /br /- ATP impact: None</cell><cell>Power On ISTCR</cell><cell>Wrong TCB statuses on the CAN from EMERLOG</cell>"
    stdout = "<cell>198</cell><cell>Defect</cell><cell>TCB statuses inverted on the EMERLOG CAN</cell><cell>SACR_Fixed</cell><cell>Major</cell><cell>S0</cell><cell>S0</cell><cell>EQT_EMERLOG_02_02 (SW delivery in baseline SW_BITE_01_21) - correction in the SSCS in EQT_EMERLOG_02_04</cell><cell>11/02/16 18:19</cell><cell>- Assembly Board impact: 402CE06L0201Y02br /SW impact: Yesbr /HW impact: Nonebr /PLD impact: Nonebr /br /- SSCS impact: ET2925-S issue 3br /br /- ICD SPI ICD MCU BITE/FUNC PLDb /bimpact: Nobr /br /- HPID impact: Nobr /br /- HSID impact: Nobr /br /- ETP: nonebr /br /- ATP impact: None</cell><cell>Power On ISTCR</cell><cell>Wrong TCB statuses on the CAN from EMERLOG</cell>"
    stdout = Tool.replaceBeacon(stdout)
    #stdout = Tool.adjustCR(stdout)
    cr_decod = BuildDoc._parseCRCell(stdout)
    print "cr_decod",cr_decod[9]
    line = ThreadQuery.cleanImpactAnalysis(cr_decod[9])
    print "RESULT:",line
    exit(0)
    list_sacrs = [['SACR ESSNESS 291', 'Fixed', 'PWM accuracy', '715CE06L0007Y02', '']]
    print "DEBUG:",list_sacrs
    patch_sacrs = list_sacrs[0]
    print "DEBUG2:",patch_sacrs
    for sacr_tag,sacr_status,sacr_synopsis,sacr_implm,x in list_sacrs:
        print sacr_tag
    if 0==1:
        syn = ThreadQuery()
        parent_cr = "<tr><td><IMG SRC=../img/changeRequestIcon.gif>SyCR</td><td>PDS</td><td>780</td><td>SyCR_Fixed</td><td>TSBC 1&2 open in BAT configuration</td><td>S1.5.1</td><td></td></tr>"
        print "input",parent_cr
        parent_cr = re.sub("&"," and ",parent_cr)
        parent_decod = syn._parseMultiCRParent(parent_cr)
        # Parent CR;Parent CR status;Parent CR synopsis
        print "output",parent_decod